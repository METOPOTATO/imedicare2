
import calendar
import os
import math
import re
from django.utils.translation import gettext as _

from openpyxl import Workbook,load_workbook
from openpyxl.styles import Color, Font,Border,Side,Alignment,PatternFill
from copy import copy
from django.shortcuts import render
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage

from django.http import JsonResponse, HttpResponseRedirect,HttpResponse
from django.contrib.auth.decorators import login_required
#from django.contrib.staticfiles.templatetags.staticfiles import static
try:
    # Django 2
    from django.contrib.staticfiles.templatetags.staticfiles import static
except ModuleNotFoundError:
    # Django 3
    from django.templatetags.static import static
from django.db.models import Q, Count, F, Min,Sum, Case,When ,Value,CharField,Avg
from django.db.models.functions import Coalesce
import operator
import functools

from django.utils import timezone, translation

from .forms import *
from app.models import *
from Receptionist.models import *
from Doctor.models import *
from Pharmacy.models import *
from KBL.models import *

from dateutil import relativedelta


from django.views import View
from django.views.decorators.csrf import csrf_exempt



# Create your views here.






@login_required
def manage(request):
    #doctor_search_form = DoctorsSearchForm()
    #
    #
    #list_exam_fee = []
    #list_precedures = []
    #list_radiologys = []
    #
    #
    #exam_fees = ExamFee.objects.filter(Q(code = 'E0010') | Q(code = 'E0011'))
    #for exam_fee in exam_fees:
    #    list_exam_fee.append({'code':exam_fee.code,'value':exam_fee.name})
    #
    #precedures = Precedure.objects.filter(code__contains='PM')
    #for precedure in precedures:
    #    list_precedures.append({'code':precedure.code,'value':precedure.name})
    #
    #radiologys = Precedure.objects.filter(code__contains='R', precedure_class_id = 10 )
    #for radiology in radiologys:
    #    list_radiologys.append({'code':radiology.code,'value':radiology.name})
    #
    #return render(request,
    #'Doctor/audit_PM.html',
    #    {
    #        'doctor_search':doctor_search_form,
    #
    #        'list_exam_fee':list_exam_fee,
    #        'list_precedures':list_precedures,
    #        'list_radiologys':list_radiologys,
    #
    #    }
    #)

    #####################################################

    #filters
    ##list_exam_fee = []
    ##list_lab = []
    ##list_precedure = []
    ##list_medicine = []
    ##
    ##
    ##    
    ##exams = ExamFee.objects.all().order_by('name')
    ##for exam in exams:
    ##    list_exam_fee.append({'code':exam.code,'value':exam.name})
    ##
    ##tests = Test.objects.all().order_by('name')
    ##for test in tests:
    ##    list_lab.append({'code':test.code,'value':test.name})
    ##
    ##precedures = Precedure.objects.all().order_by('name')
    ##for precedure in precedures:
    ##    list_precedure.append({'code':precedure.code,'value':precedure.name})
    ##    
    ##medicines = Medicine.objects.all().order_by('name')
    ##for medicine in medicines:
    ##    list_medicine.append({'code':medicine.code,'value':medicine.name})
    ##    



    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })
    #의사 정보 ? 
    doctor = Doctor.objects.values('name_short','id')

    #결제 방법
    payment_method = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_METHOD',upper_commcode ='000014' ).annotate(code = F('commcode'),name = f_name ).values('code','name')

    #결제 상태
    payment_status = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_STATUS',upper_commcode ='000014' ).annotate(code = F('commcode'),name = f_name ).values('code','name')

    return render(request,
        'Manage/Manage.html',
            {
                'depart_medical':depart_medical,
                'doctor' : doctor,

                'payment_method':payment_method,
                'payment_status':payment_status,

                #'payment_search':payment_search_form,
                #'patient_search':patient_search_form,
                #'doctor_search':doctor_search_form,
                #'doctors':Doctor.objects.all(),
                #'medicine_search':medicine_search_form,

                #'list_exam_fee':list_exam_fee, # general will be precedure in template
                #'list_lab':list_lab,
                #'list_precedure':list_precedure,
                #'list_medicine':list_medicine,

            }
        )


def patient(request):
    return render(request,
        'Manage/patient.html',
            {

            }
        )

def payment(reqeust):
    return render(request,
        'Manage/payment.html',
            {

            }
        )



def search_payment(request):
    page_context = request.POST.get('page_context',10) # 페이지 컨텐츠 
    page = request.POST.get('page',1)

    date_type = request.POST.get('date_type')

    date_start = request.POST.get('start')
    date_end = request.POST.get('end')

    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')

    payment_method = request.POST.get('payment_method')
    payment_status = request.POST.get('payment_status')

    category = request.POST.get('patient_type','')
    string = request.POST.get('patient_search','')

    is_vaccine = request.POST.get('is_vaccine','false')
    is_red_invoice = True if request.POST.get('is_red_invoice','false') == 'true' else False

    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
        fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')
    
    
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    #지불 정보
    dict_payment_method = {}
    query_payment_method= COMMCODE.objects.filter(commcode_grp='PAYMENT_METHOD',upper_commcode='000014').annotate(code = F('commcode'),name = fname).values('code','name')
    for data in query_payment_method:
        dict_payment_method.update({
            data['code'] : data['name']
            })


    kwargs = {}

    argument_list = []
    datas = []


    if date_type == 'RECEPTION':

        if category=='':
            argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
            argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
            argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
            argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 
            argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
        elif category=='name':
            argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
            argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
        elif category=='chart':
            argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
        elif category=='date_of_birth':
            argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
        elif category=='phone':
            argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 


        kwargs['progress'] = 'done'
        if depart != '':
            kwargs['depart_id'] = depart
        if doctor != '':
            kwargs['doctor_id'] = doctor
        if payment_method !='':
            kwargs['payment__paymentrecord__method'] = payment_method
        if payment_status !='':
            kwargs['payment__progress'] = payment_status
        if is_red_invoice == True:
            kwargs['need_invoice'] = is_red_invoice
            
        receptions = Reception.objects.filter(
                functools.reduce(operator.or_, argument_list),
                **kwargs ,
                payment__paymentrecord__status = 'paid',
                recorded_date__range = (date_min, date_max), 
            ).prefetch_related(
                'diagnosis__exammanager_set',
                'diagnosis__testmanager_set',
                'diagnosis__preceduremanager_set',
                'diagnosis__medicinemanager_set',
                'payment__paymentrecord_set',
            ).exclude(

                ).order_by("recorded_date").values_list('id',flat=True)


        #중복 제거 시작 2020-10-21
        list_id = list(receptions)
        list_id = list(set(list_id))


        argument_list = []
        for id in list_id:
            argument_list.append( Q(**{'id':id} ) )
        if len(list_id)==0:
            argument_list.append( Q(**{'id':0} ) )

       
        payment_total_paid_amount = 0
        payment_total_subtotal = 0
        payment_total_additional = 0
        payment_total_total = 0
        payment_total_discount = 0
        payment_total_unpaid = 0

        count_page = list_id
        for roof_cnt in range(len(argument_list) // 500  + 1):
            receptions = Reception.objects.filter(
                    functools.reduce(operator.or_, argument_list[ 500 * roof_cnt : 500 * (roof_cnt+1)]),
                ).prefetch_related(
                    'diagnosis__exammanager_set',
                    'diagnosis__testmanager_set',
                    'diagnosis__preceduremanager_set',
                    'diagnosis__medicinemanager_set',
                ).order_by("-id")

            #중복 제거 끝


            #count_page = receptions.values('id')

            #print(len(receptions))
            payment_total = receptions.filter().aggregate(
                Sum('payment__sub_total'),
                Sum('payment__total'),
                Sum('payment__additional'),
                )

            #unpaid 구하기
            real_paid_total = receptions.filter( Q(payment__paymentrecord__status='paid') ).aggregate(
                Sum('payment__paymentrecord__paid'),
                )

            payment_total_paid_amount +=  0 if real_paid_total['payment__paymentrecord__paid__sum'] is None else real_paid_total['payment__paymentrecord__paid__sum']

            payment_total_subtotal += 0 if payment_total['payment__sub_total__sum'] is None else payment_total['payment__sub_total__sum']
            payment_total_additional += 0 if payment_total['payment__additional__sum'] is None else payment_total['payment__additional__sum']
            payment_total_total += 0 if payment_total['payment__total__sum'] is None else payment_total['payment__total__sum']
            payment_total_discount = payment_total_subtotal + payment_total_additional - payment_total_total
            
            # for reception in receptions:
            #     sub_total_ = reception.payment.sub_total
            #     discount_ = reception.payment.discounted_amount
            
            #     if discount_ is 0:
            #         discount_percent_ = reception.payment.discounted
            #         if discount_percent_ is 0:
            #             discount_ = 0
            #         else:
            #             discount_ = discount_percent_ / 100 * sub_total_
            
            #     payment_total_discount += discount_



        #데이터

        receptions = Reception.objects.filter(
                    functools.reduce(operator.or_, argument_list[int(page_context) * (int(page) -1): int(page_context) * int(page)]),
                ).prefetch_related(
                    'diagnosis__exammanager_set',
                    'diagnosis__testmanager_set',
                    'diagnosis__preceduremanager_set',
                    'diagnosis__medicinemanager_set',
                ).order_by("-id")

        datas = []
        for reception in receptions:
            data = {
                'no':reception.id,
                'date':reception.recorded_date.strftime('%Y-%m-%d'),
                'Patient':reception.patient.name_kor,
                'patient_eng':reception.patient.name_eng,
                'date_of_birth':str(reception.patient.get_age()) + '/' + reception.patient.get_gender_simple(),
                'address':reception.patient.address,
                'gender':reception.patient.gender,
                'Depart':reception.depart.name,
                'Doctor_kor':reception.doctor.name_kor,
                'Doctor_eng': reception.doctor.name_eng,
                'red_invoice': reception.need_invoice,
                }

            list_exam_fee = []
            list_lab = []
            list_precedure= []
            list_radiation= []
            list_medicine= []




            #진료 아이템
            ##진료비
            tmp_exam_set = reception.diagnosis.exammanager_set.all()
            for tmp_exam in tmp_exam_set:
                list_exam_fee.append({
                    'checked':tmp_exam.is_checked_discount,
                    'code':tmp_exam.exam.code,
                    'value':tmp_exam.exam.name
                    })

            ##검사
            tmp_test_set = reception.diagnosis.testmanager_set.filter(test__parent_test = None)
            for tmp_test in tmp_test_set:
                list_lab.append({
                    'checked':tmp_test.is_checked_discount,
                    'code':tmp_test.test.code,
                    'value':tmp_test.test.name,
                    })


            ##처치 및 방사선
            tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
            for tmp_precedure in tmp_precedure_set:
                if 'R' in tmp_precedure.precedure.code:
                    list_radiation.append({
                        'checked':tmp_precedure.is_checked_discount,
                        'code':tmp_precedure.precedure.code,
                        'value':tmp_precedure.precedure.name,
                        'amount':tmp_precedure.amount,
                        })
                else:
                    list_precedure.append({
                        'checked':tmp_precedure.is_checked_discount,
                        'code':tmp_precedure.precedure.code,
                        'value':tmp_precedure.precedure.name,
                        'amount':tmp_precedure.amount,
                        })

            ##약
            tmp_medicine_set = reception.diagnosis.medicinemanager_set.all()
            for tmp_medicine in tmp_medicine_set:
                list_medicine.append({
                    'checked':tmp_medicine.is_checked_discount,
                    'code':tmp_medicine.medicine.code,
                    'value':tmp_medicine.medicine.name,
                    'amount':tmp_medicine.amount,
                    })


            #수납 정보
            paid_by = '-'
        
            sub_total = reception.payment.sub_total
            additional = 0 if reception.payment.additional is None else reception.payment.additional
            discount = reception.payment.discounted_amount
            total = reception.payment.total
            unpaid = 0
            
            if discount is 0:
                discount_percent = reception.payment.discounted
                if discount_percent is 0:
                    discount = 0
                else:
                    discount = discount_percent / 100 * sub_total
        
            #수납 상태 및 방법
            #if reception.payment.progress != 'paid':
            list_paid_record = reception.payment.paymentrecord_set.filter(
                status='paid'
                ).exclude(
                status = 'cancel'
                    ).order_by('date')




            all_paid = 0
            
            paid_by = ''
            paid_by_remit = False
            paid_by_card = False
            paid_by_cash = False



            list_used_method = []
            for paid_record in list_paid_record:
                all_paid += paid_record.paid
                list_used_method.append(paid_record.method)

            list_used_method = list(set(list_used_method))
            for tmp_paid_by in list_used_method:
                paid_by += dict_payment_method[tmp_paid_by]

            paid_date = ''
            if list_paid_record.count() != 0:
                paid_date = list_paid_record.first().date.strftime('%Y-%m-%d')

            unpaid = total - all_paid

            data.update({
                'list_exam_fee':list_exam_fee,
                'list_lab':list_lab,
                'list_precedure':list_precedure,
                'list_radiation':list_radiation,
                'list_medicine':list_medicine,

                'paid_by':paid_by,
                'sub_total':sub_total,
                'additional':additional,
                'discount':discount,
                'total':total,
                'unpaid':unpaid,
                'date_paid':paid_date,
                'all_paid':all_paid
                })

            datas.append(data)

    elif date_type == 'PAID':

        if category=='':
            argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
            argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
            argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
            argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
            argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
        elif category=='name':
            argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
            argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        elif category=='chart':
            argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        elif category=='date_of_birth':
            argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
        elif category=='phone':
            argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

        kwargs['payment__reception__progress'] = 'done'
        if depart != '':
            kwargs['payment__reception__depart_id'] = depart
        if doctor != '':
            kwargs['payment__reception__doctor_id'] = doctor
        if payment_method !='':
            kwargs['method'] = payment_method
        if payment_status !='':
            kwargs['payment__progress'] = payment_status

        tmp_receptions = PaymentRecord.objects.filter(
                functools.reduce(operator.or_, argument_list),
                **kwargs ,
                date__range = (date_min, date_max), 
                # payment__reception__recorded_date__range = (date_min, date_max),
                # status = 'paid',
            ).prefetch_related(
                'payment__reception__diagnosis__exammanager_set',
                'payment__reception__diagnosis__testmanager_set',
                'payment__reception__diagnosis__preceduremanager_set',
                'payment__reception__diagnosis__medicinemanager_set',
            ).order_by("-id")

            
        new_argument_list = []
        tmp_list=[]
        for tmp_data in tmp_receptions:
            if tmp_data.payment.reception.id in tmp_list:
                continue
            tmp_list.append(tmp_data.payment.reception.id)
            new_argument_list.append( Q(**{'id':tmp_data.id} ) ) 
        new_argument_list.append( Q(**{'id':0} ) ) 
                
        receptions = PaymentRecord.objects.filter(
                functools.reduce(operator.or_, new_argument_list)
                ).prefetch_related( 
                'payment__reception__diagnosis__exammanager_set',
                'payment__reception__diagnosis__testmanager_set',
                'payment__reception__diagnosis__preceduremanager_set',
                'payment__reception__diagnosis__medicinemanager_set',
            ).order_by("-id")
            

        count_page = receptions.values('id')

        payment_total = receptions.filter().aggregate(
            Sum('payment__sub_total'),
            Sum('payment__total'),
            Sum('payment__additional'),
            )

        #unpaid 구하기
        real_paid_total = receptions.filter( Q(payment__paymentrecord__status='paid') ).aggregate(
            Sum('payment__paymentrecord__paid'),
            )
        payment_total_paid_amount =  0 if real_paid_total['payment__paymentrecord__paid__sum'] is None else real_paid_total['payment__paymentrecord__paid__sum']

        #unpaid 구하기
        # period_paid_total = receptions.filter( Q(payment__paymentrecord__status='paid') ).aggregate(
        #     Sum('payment__paymentrecord__paid'),
        #     )
        # if len(list_id)==0:
        #     payment_total_paid +=  0 if real_paid_total['payment__paymentrecord__paid__sum'] is None else real_paid_total['payment__paymentrecord__paid__sum']


        datas = []

        payment_total_subtotal = 0 if payment_total['payment__sub_total__sum'] is None else payment_total['payment__sub_total__sum']
        payment_total_additional = 0 if payment_total['payment__additional__sum'] is None else payment_total['payment__additional__sum']
        payment_total_total = 0 if payment_total['payment__total__sum'] is None else payment_total['payment__total__sum']
        payment_total_discount = payment_total_subtotal + payment_total_additional - payment_total_total
        payment_total_unpaid = payment_total_total - payment_total_paid_amount


        for reception in receptions[int(page_context) * (int(page) -1): int(page_context) * int(page)]:
            data = {
                'no':reception.payment.reception.id,
                'date':reception.payment.reception.recorded_date.strftime('%Y-%m-%d'),
                'Patient':reception.payment.reception.patient.name_kor,
                'patient_eng':reception.payment.reception.patient.name_eng,
                'date_of_birth':str(reception.payment.reception.patient.get_age()) + '/' + reception.payment.reception.patient.get_gender_simple(),
                'address':reception.payment.reception.patient.address,
                'gender':reception.payment.reception.patient.gender,
                'Depart':reception.payment.reception.depart.name,
                'Doctor_kor':reception.payment.reception.doctor.name_kor,
                'Doctor_eng': reception.payment.reception.doctor.name_eng,

                'date_paid': reception.date.strftime('%Y-%m-%d'),
                }
            

            list_exam_fee = []
            list_lab = []
            list_precedure= []
            list_radiation= []
            list_medicine= []


            #진료 아이템
            ##진료비
            tmp_exam_set = reception.payment.reception.diagnosis.exammanager_set.all()
            for tmp_exam in tmp_exam_set:
                list_exam_fee.append({
                    'checked':tmp_exam.is_checked_discount,
                    'code':tmp_exam.exam.code,
                    'value':tmp_exam.exam.name
                    })

            ##검사
            tmp_test_set = reception.payment.reception.diagnosis.testmanager_set.all()
            for tmp_test in tmp_test_set:
                list_lab.append({
                    'checked':tmp_test.is_checked_discount,
                    'code':tmp_test.test.code,
                    'value':tmp_test.test.name,
                    })


            ##처치 및 방사선
            tmp_precedure_set = reception.payment.reception.diagnosis.preceduremanager_set.all()
            for tmp_precedure in tmp_precedure_set:
                if 'R' in tmp_precedure.precedure.code:
                    list_radiation.append({
                        'checked':tmp_precedure.is_checked_discount,
                        'code':tmp_precedure.precedure.code,
                        'value':tmp_precedure.precedure.name,
                        'amount':tmp_precedure.amount,
                        })
                else:
                    list_precedure.append({
                        'checked':tmp_precedure.is_checked_discount,
                        'code':tmp_precedure.precedure.code,
                        'value':tmp_precedure.precedure.name,
                        'amount':tmp_precedure.amount,
                        })

            ##약
            tmp_medicine_set = reception.payment.reception.diagnosis.medicinemanager_set.all()
            for tmp_medicine in tmp_medicine_set:
                list_medicine.append({
                    'checked':tmp_medicine.is_checked_discount,
                    'code':tmp_medicine.medicine.code,
                    'value':tmp_medicine.medicine.name,
                    'amount':tmp_medicine.amount,
                    })


            #수납 정보
            paid_by = '-'
        
            sub_total = reception.payment.sub_total
            additional = 0 if reception.payment.additional is None else reception.payment.additional
            discount = reception.payment.discounted_amount
            total = reception.payment.total
            unpaid = 0

            if discount is 0:
                discount_percent = reception.payment.discounted
                if discount_percent is 0:
                    discount = 0
                else:
                    discount = discount_percent / 100 * sub_total
        
            #수납 상태 및 방법
            #if reception.payment.progress != 'paid':
            list_paid_record = reception.payment.paymentrecord_set.filter(status='paid')
            all_paid = 0
            
            paid_by = ''
            paid_by_remit = False
            paid_by_card = False
            paid_by_cash = False


            list_used_method = []
            for paid_record in list_paid_record:
                all_paid += paid_record.paid
                list_used_method.append(paid_record.method)

            list_used_method = list(set(list_used_method))
            for tmp_paid_by in list_used_method:
                paid_by += dict_payment_method[tmp_paid_by]

            unpaid = total - all_paid
            print('all_paid: ',all_paid)
            if reception.payment.reception.recorded_date <= date_min or reception.payment.reception.recorded_date >= date_max:
                sub_total = 0
                total = 0
                additional = 0
                discount = 0

            data.update({
                'list_exam_fee':list_exam_fee,
                'list_lab':list_lab,
                'list_precedure':list_precedure,
                'list_radiation':list_radiation,
                'list_medicine':list_medicine,

                'paid_by':paid_by,
                'sub_total':sub_total,
                'additional':additional,
                'discount':discount,
                'total':total,
                'unpaid':unpaid,
                'all_paid':all_paid
                })

            datas.append(data)
    paginator = Paginator(count_page, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    payment_total_unpaid += payment_total_total - payment_total_paid_amount
    context = {
            'datas':datas,
            'page_range_start':paging_data.paginator.page_range.start,
            'page_range_stop':paging_data.paginator.page_range.stop,
            'page_number':paging_data.number,
            'has_previous':paging_data.has_previous(),
            'has_next':paging_data.has_next(),

 

            'payment_total_subtotal':payment_total_subtotal,
            'payment_total_additional':payment_total_additional,
            'payment_total_discount':payment_total_discount,
            'payment_total_total':payment_total_total,
            'payment_total_unpaid':payment_total_unpaid,
            'payment_total_paid_amount':payment_total_paid_amount,
            }
    
    return JsonResponse(context)

    #filter_general = request.POST.get('general')
    #filter_medicine = request.POST.get('medicine')employee_add_edit_get
    #filter_lab = request.POST.get('lab')
    #
    #pup =request.POST.get('pup')
    #paid_by = request.POST.get('paid_by')
    #
    #date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    #date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)
    #
    #
    #
    #
    #
    #if depart != '':
    #    kwargs.update({'depart_id':depart})
    #if doctor != '':
    #    kwargs.update({'doctor_id':doctor})
    #
    #
    #datas = []
    #receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').order_by("-id")
    #
    #
    #page = request.POST.get('page',1)
    #payment_total_total = 0
    #payment_total_paid = 0
    #payment_total_unpaid = 0
    #for reception in receptions:
    #    data = {}
    #    try:
    #        general = []
    #        lab = []
    #        medi = []
    #        scaling = []
    #        panorama = []
    #        total_payment=0
    #        if filter_general == '' and filter_medicine == '' and filter_lab == '':
    #            tmp_exam_set = reception.diagnosis.exammanager_set.all()
    #            for tmp_exam in tmp_exam_set:
    #                if hasattr(tmp_exam,'doctor'):
    #                    general.append({
    #                        'code':tmp_exam.exam.code,
    #                        'value':tmp_exam.name + tmp_exam.exam.doctor.name_kor
    #                        })
    #                else:
    #                    general.append({
    #                        'code':tmp_exam.exam.code,
    #                        'value':tmp_exam.exam.name
    #                        })
    #       
    #            tmp_test_set = reception.diagnosis.testmanager_set.all()
    #            for tmp_test in tmp_test_set:
    #                lab.append({
    #                    'code':tmp_test.test.code,
    #                    'value':tmp_test.test.name,
    #                    })
    #
    #            tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
    #            for tmp_precedure in tmp_precedure_set:
    #                if 'scaling' in tmp_precedure.precedure.name.lower():
    #                    scaling.append({
    #                        'code':tmp_precedure.precedure.code,
    #                        'value':tmp_precedure.precedure.name
    #                        })
    #
    #                elif 'injection' in tmp_precedure.precedure.name.lower():
    #                    general.append({
    #                        'code':tmp_precedure.precedure.code,
    #                        'value':tmp_precedure.precedure.name
    #                        })
    #
    #                elif 'panorama' in tmp_precedure.precedure.name.lower():
    #                    panorama.append({
    #                        'code':tmp_precedure.precedure.code,
    #                        'value':tmp_precedure.precedure.name
    #                        })
    #
    #                else:
    #                    general.append({
    #                        'code':tmp_precedure.precedure.code,
    #                        'value':tmp_precedure.precedure.name
    #                        })
    #        
    #
    #            tmp_medicine_set = reception.diagnosis.medicinemanager_set.all()
    #            for tmp_medicine in tmp_medicine_set:
    #                if tmp_medicine.medicine.medicine_class_id is 31:
    #                    general.append({
    #                        'code':tmp_medicine.medicine.code,
    #                        'value':tmp_medicine.medicine.name + ' x ' + str(tmp_medicine.days * tmp_medicine.amount),
    #                        })
    #                else:
    #                    medi.append({
    #                        'code':tmp_medicine.medicine.code,
    #                        'value':tmp_medicine.medicine.name + ' x ' + str(tmp_medicine.days * tmp_medicine.amount),
    #                        })
    #
    #
    #        else:
    #            if filter_general != '':
    #                if 'E' in filter_general:
    #                    tmp_exam_set = reception.diagnosis.exammanager_set.all()
    #                    res = True
    #                    for tmp_exam_data in tmp_exam_set:
    #                        if 'E_NEW' in filter_general:
    #                            if 'New' not in tmp_exam_data.exam.name:
    #                                res = False
    #                        elif 'E_REP' in filter_general:
    #                            if 'Rep' not in tmp_exam_data.exam.name:
    #                                res = False
    #                        elif 'E_DNT' in filter_general:
    #                            if 'Ora' not in tmp_exam_data.exam.name:
    #                                res = False
    #                        else:
    #                            tmp_exam = ExamFee.objects.get(code = filter_general)
    #                            if tmp_exam.code not in tmp_exam_data.exam.code:
    #                                res = False
    #                        
    #                        if res:
    #                            general.append({
    #                                'code':tmp_exam_data.exam.code,
    #                                'value':tmp_exam_data.exam.name
    #                                })
    #                            total_payment += tmp_exam_data.exam.get_price(reception.recorded_date)
    #                    if res is False or tmp_exam_set.count()==0:
    #                        continue
    #                            
    #
    #                elif 'MR' in filter_general:
    #                    tmp_exam = ExamFee.objects.get(code = filter_general)
    #                    tmp_exam_set = reception.diagnosis.exammanager_set.filter(exam_id = tmp_exam.id)
    #                    if tmp_exam_set.count() == 0:
    #                        continue
    #                    for tmp_exam_data in tmp_exam_set:
    #                        general.append({
    #                            'code':tmp_exam_data.exam.code,
    #                            'value':tmp_exam_data.exam.name
    #                            })
    #                        total_payment += tmp_exam_data.exam.get_price(reception.recorded_date)
    #
    #                elif 'M' in filter_general:
    #                    tmp_medicine = Medicine.objects.get(code = filter_general)
    #                    tmp_medi_set = reception.diagnosis.medicinemanager_set.filter(medicine_id = tmp_medicine.id)
    #
    #                    if tmp_medi_set.count() == 0:
    #                        continue
    #                    for tmp_medi_data in tmp_medi_set:
    #                        general.append({
    #                            'code':tmp_medi_data.medicine.code,
    #                            'value':tmp_medi_data.medicine.name + ' x ' + str(tmp_medi_data.days * tmp_medi_data.amount),
    #                            })
    #                        total_payment += tmp_medi_data.medicine.get_price(reception.recorded_date) * tmp_medi_data.days * tmp_medi_data.amount
    #
    #                else: #P D G R U O OB
    #                    tmp_precedure = Precedure.objects.get(code = filter_general)
    #                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.filter(precedure_id = tmp_precedure.id)
    #
    #                    if tmp_precedure_set.count() == 0:
    #                        continue
    #                    for tmp_precedure_data in tmp_precedure_set:
    #                        general.append({
    #                                'code':tmp_precedure_data.precedure.code,
    #                                'value':tmp_precedure_data.precedure.name
    #                                })
    #                        total_payment += tmp_precedure_data.precedure.get_price(reception.recorded_date)
    #
    #            if filter_medicine != '':
    #                tmp_medicine = Medicine.objects.get(code = filter_medicine)
    #                tmp_set = reception.diagnosis.medicinemanager_set.filter(medicine_id = tmp_medicine.id)
    #
    #                if tmp_set.count() == 0:
    #                        continue
    #                for tmp_data in tmp_set:
    #                    medi.append({
    #                        'code':tmp_data.exam.code,
    #                        'value':tmp_data.medicine.name + ' x ' + str(tmp_data.days * tmp_data.amount),
    #                        })
    #                    total_payment += tmp_precedure_data.precedure.get_price(reception.recorded_date)
    #            if filter_lab != '':
    #                tmp_test = Test.objects.get(code = filter_lab)
    #                tmp_set = reception.diagnosis.testmanager_set.filter(test_id = tmp_test.id)
    #                if tmp_set.count() == 0:
    #                        continue
    #                for tmp_data in tmp_set:
    #                    lab.append({
    #                       'code':tmp_data.exam.code,
    #                        'value':tmp_data.exam.name
    #                        })
    #                    total_payment += tmp_precedure_data.precedure.get_price(reception.recorded_date)
    #            
    #
    #        data.update({'general':general})
    #        data.update({'medi':medi})
    #        data.update({'lab':lab})
    #        data.update({'scaling':scaling})
    #        data.update({'panorama':panorama})
    #
    #        paid_set = reception.payment.paymentrecord_set.all()
    #        paid_sum = 0
    #        
    #        for paid in paid_set:
    #            paid_sum += paid.paid
    #
    #        unpaid_sum = reception.payment.total - paid_sum
    #    
    #        if pup == 'Paid':
    #            if unpaid_sum != 0:
    #                continue
    #        elif pup == 'Unpaid':
    #            if unpaid_sum == 0:
    #                continue
    #
    #        if filter_general == '' and filter_medicine == '' and filter_lab == '':
    #            payment_total_total += reception.payment.total
    #            payment_total_paid += reception.payment.total - unpaid_sum
    #            payment_total_unpaid += unpaid_sum
    #            total_payment = reception.payment.total
    #        else:
    #            payment_total_total += total_payment
    #            paid_sum = 0
    #            unpaid_sum = 0
    #
    #        data.update({
    #            'no':reception.id,
    #            'date':reception.recorded_date.strftime('%d-%b-%y'),
    #            'Patient':reception.patient.name_kor,
    #            'patient_eng':reception.patient.name_eng,
    #            'date_of_birth':str(reception.patient.get_age()) + '/' + reception.patient.get_gender_simple(),
    #            'address':reception.patient.address,
    #            'gender':reception.patient.gender,
    #            'Depart':reception.depart.name,
    #            'Doctor':reception.doctor.get_name(),
    #
    #            'paid_by_cash':'',
    #            'paid_by_card':'',
    #            'paid_by_remit':'',
    #
    #            'total' :total_payment,
    #            'paid':paid_sum,
    #            'unpaid':unpaid_sum,
    #            })
    #
    #        
    #
    #
    #        pay_records = PaymentRecord.objects.filter(payment = reception.payment)
    #
    #        for pay_record in pay_records:
    #            if pay_record.method == 'card':
    #                data.update({'paid_by_card':'card'})
    #            elif pay_record.method == 'cash':
    #                data.update({'paid_by_card':'cash'})
    #            elif pay_record.method == 'remit':
    #                data.update({'paid_by_card':'remit'})
    #
    #        datas.append(data)
    #    except Diagnosis.DoesNotExist:
    #        pass
    #
    #paginator = Paginator(datas, page_context)
    #try:
    #    paging_data = paginator.page(page)
    #except PageNotAnInteger:
    #    paging_data = paginator.page(1)
    #except EmptyPage:
    #    paging_data = paginator.page(paginator.num_pages)
    #
    #
    #context = {
    #           'datas':list(paging_data),
    #           'page_range_start':paging_data.paginator.page_range.start,
    #           'page_range_stop':paging_data.paginator.page_range.stop,
    #           'page_number':paging_data.number,
    #           'has_previous':paging_data.has_previous(),
    #           'has_next':paging_data.has_next(),
    #
    #           #for graph
    #           'days':(date_max - date_min).days +1 ,
    #
    #           'payment_total_total':payment_total_total,
    #           'payment_total_paid':payment_total_paid,
    #           'payment_total_unpaid':payment_total_unpaid,
    #
    #           }
    #
    #return JsonResponse(context)

@login_required
def doctor_profit(request):

    page_context = request.POST.get('page_context',10) # 페이지 컨텐츠 
    kwargs = {}
    datas = []

    doctor = request.POST.get('doctor')

    date_start = request.POST.get('start_date')
    date_end = request.POST.get('end_date')


    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)
    
    page = request.POST.get('page',1)
    
    if doctor != '':
        kwargs.update({'doctor_id':doctor})

    context = {}

    if request.user.is_admin or request.user.doctor.depart.name == 'PM' :
        total_amount = 0
        total_additional = 0
        amount_discount = 0

        filter_search = request.POST.get('search')

        receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').select_related('payment').filter(payment__progress='paid').order_by("-id")

        count_page = receptions.values('id')

        #receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').select_related('diagnosis').select_related('payment').order_by("-id")
        for reception in receptions[int(page_context) * (int(page) -1): int(page_context) * int(page)]:
            exams = []
            precedures = []
            radiographys = []

            data={}
            sub_total = 0


            if filter_search == '':
                tmp_exam_set = reception.diagnosis.exammanager_set.all()
                tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
            else:
                if 'E' in filter_search:
                    tmp_exam_set = reception.diagnosis.exammanager_set.prefetch_related('exam').filter(exam__code=filter_search)
                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.none()
                elif 'PM' in filter_search:
                    tmp_exam_set = reception.diagnosis.exammanager_set.none()
                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.all().prefetch_related('precedure').filter(precedure__code=filter_search)

                elif 'R' in filter_search:
                    tmp_exam_set = reception.diagnosis.exammanager_set.none()
                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.all().prefetch_related('precedure').filter(precedure__code__icontains='R')

                if tmp_exam_set.count() is 0 and tmp_precedure_set.count() is 0:
                    continue
                
            for tmp_exam in tmp_exam_set:
                    exams.append({
                        'code':tmp_exam.exam.code,
                        'value':tmp_exam.exam.name,
                        })
                    sub_total += tmp_exam.exam.get_price(reception.recorded_date)
                    total_amount += sub_total
            for precedure_set in tmp_precedure_set:
                    if 'R' in precedure_set.precedure.code:
                        radiographys.append({
                            'code':precedure_set.precedure.code,
                            'value':precedure_set.precedure.name,
                            'amount':precedure_set.amount,
                            })
                        sub_total += precedure_set.precedure.get_price(reception.recorded_date)
                        total_amount += sub_total
                    else:
                        precedures.append({
                            'code':precedure_set.precedure.code,
                            'value':precedure_set.precedure.name,
                            })
                        sub_total += precedure_set.precedure.get_price(reception.recorded_date)
                        total_amount += sub_total

            total_additional += 0 if reception.payment.additional is None else reception.payment.additional
            
            if reception.payment.discounted is None:
                discount = reception.payment.discounted_amount if filter_search == '' else 0
            else:
                discount = reception.payment.discounted / 100 * reception.payment.sub_total
            amount_discount += 0 if discount is None else discount



            data.update({
                'exams':exams,
                'precedures':precedures,
                'radiographys':radiographys,
                
                'subtotal':reception.payment.sub_total if filter_search == '' else sub_total,
                'subtotal':reception.payment.sub_total if filter_search == '' else sub_total,
                'discount':discount if filter_search == '' else 0,
                'total':reception.payment.total if filter_search == '' else sub_total,
                'additional':reception.payment.additional if filter_search == '' else 0,

  
                'no':reception.id,
                'date':reception.recorded_date.strftime('%d-%b-%y'),
                'Patient':reception.patient.name_kor,
                'patient_eng':reception.patient.name_eng,
                'date_of_birth':str(reception.patient.get_age()) + '/' + reception.patient.get_gender_simple(),
                'address':reception.patient.address,
                'gender':reception.patient.gender,
                'Depart':reception.depart.name,
                'Doctor':reception.doctor.get_name(),
                })

            
            pay_records = PaymentRecord.objects.filter(payment = reception.payment)
            for pay_record in pay_records:
                if pay_record.method == 'card':
                    data.update({'paid_by_card':'card'})
                elif pay_record.method == 'cash':
                    data.update({'paid_by_cash':'cash'})
                elif pay_record.method == 'remit':
                    data.update({'paid_by_remit':'remit'})



            datas.append(data)

        context.update({
            'total_amount':total_amount,
            'total_additional':total_additional,
            })
    else:
        filter_exam_fee = request.POST.get('exam_fee')
        filter_test = request.POST.get('test')
        filter_precedure = request.POST.get('precedure')
        filter_medicine = request.POST.get('medicine')

        pup =request.POST.get('pup')
        paid_by = request.POST.get('paid_by')


        receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').select_related('payment').filter(payment__progress='paid').order_by("-id")
        count_page = receptions.values('id')
        
        
        amount_exam_fee = 0
        amount_test = 0
        amount_precedure = 0
        amount_medicine = 0

        
        
        for reception in receptions[int(page_context) * (int(page) -1): int(page_context) * int(page)]:
            data = {}
            try:
                exam_fee = []
                test = []
                precedure = []
                medi = []

                
                #필터링 없을 때
                if filter_exam_fee == '' and filter_test == '' and filter_precedure == '' and filter_medicine=='':
                   
                    tmp_exam_set = reception.diagnosis.exammanager_set.all()
                    for tmp_exam in tmp_exam_set:
                        exam_fee.append({
                            'code':tmp_exam.exam.code,
                            'value':tmp_exam.exam.name
                            })

                    tmp_test_set = reception.diagnosis.testmanager_set.all()
                    for tmp_test in tmp_test_set:
                        test.append({
                            'code':tmp_test.test.code,
                            'value':tmp_test.test.name,
                            })

                    tmp_precedure_set = reception.diagnosis.preceduremanager_set.all()
                    for tmp_precedure in tmp_precedure_set:
                        precedure.append({
                            'code':tmp_precedure.precedure.code,
                            'value':tmp_precedure.precedure.name
                            })
            

                    tmp_medicine_set = reception.diagnosis.medicinemanager_set.all()
                    for tmp_medicine in tmp_medicine_set:
                        medi.append({
                            'code':tmp_medicine.medicine.code,
                            'value':tmp_medicine.medicine.name + ' x ' + str(tmp_medicine.days * tmp_medicine.amount),
                            })

                paid_set = reception.payment.paymentrecord_set.all()
                paid_sum = 0
                for paid in paid_set:
                    paid_sum  += paid.paid

                data.update({
                    'no':reception.id,
                    'date':reception.recorded_date.strftime('%d-%m-%Y'),
                    'Patient':reception.patient.name_kor,
                    'patient_eng':reception.patient.name_eng,
                    'date_of_birth':str(reception.patient.get_age()) + '/' + reception.patient.get_gender_simple(),
                    'address':reception.patient.address,
                    'gender':reception.patient.gender,
                    'Depart':reception.depart.name,
                    'Doctor':reception.doctor.get_name(),

                    'list_exam_fee':exam_fee,
                    'list_test':test,
                    'list_precedure':precedure,
                    'list_medi':medi,

                    'paid_by_cash':'',
                    'paid_by_card':'',
                    'paid_by_remit':'',

                    'sub_total':reception.payment.sub_total,
                    'total' :reception.payment.total,
                    'discount':reception.payment.discounted_amount,
                    })


                for pay_record in paid_set:
                    if pay_record.method == 'card':
                        data.update({'paid_by_card':'card'})
                    elif pay_record.method == 'cash':
                        data.update({'paid_by_card':'cash'})
                    elif pay_record.method == 'remit':
                        data.update({'paid_by_card':'remit'})

                datas.append(data)
            except Diagnosis.DoesNotExist:
                pass


        if (date_max - date_min).days == 0:  #단일 날짜는 당일을 Subtotal / 선택된 달의 금액을 Total
             first_day = datetime.datetime.strptime(date_start, "%Y-%m-%d").date().replace(day=1)
             last_day = (first_day + relativedelta.relativedelta(months=1)) - datetime.timedelta(seconds=1)
             
             receptions = Reception.objects.filter(**kwargs ,recorded_date__range = (first_day, last_day), progress = 'done').order_by("-id").select_related('payment')
             monthly_total = 0
             for reception in receptions:
                 monthly_total += reception.payment.total
             context.update({
                'monthly_total':monthly_total,
                })
        #else: #범위 날짜는 SubTotal 무시 /범위 계산을 Total로

    query_total = Reception.objects.filter(**kwargs ,recorded_date__range = (date_min, date_max), progress = 'done').select_related('payment').filter(payment__progress='paid')


    
    query_total = receptions.aggregate(
        amount_sub_total=Sum('payment__sub_total'),
        #amount_discount = Sum('payment__discounted_amount'),
        amount_total = Sum('payment__total'),
        )

    context.update({
        'amount_sub_total':query_total['amount_sub_total'] if query_total['amount_sub_total'] is not None else 0,
        'amount_discount':amount_discount,
        'amount_total':query_total['amount_total'] if query_total['amount_total'] is not None else 0,
        })
    
    paginator = Paginator(count_page, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)
    

    context.update({
                'datas':datas,
                'page_range_start':paging_data.paginator.page_range.start,
                'page_range_stop':paging_data.paginator.page_range.stop,
                'page_number':paging_data.number,
                'has_previous':paging_data.has_previous(),
                'has_next':paging_data.has_next(),

                #for graph
                'days':(date_max - date_min).days +1 ,


                })
    return JsonResponse(context)



def audit_excel(request):

    date_type = request.GET.get('date_type')

    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')

    depart = request.GET.get('depart')
    doctor = request.GET.get('doctor')

    payment_method = request.GET.get('payment_method')
    payment_status = request.GET.get('payment_status')
 
    category = request.GET.get('patient_type','')
    string = request.GET.get('patient_search','')

    is_vaccine = request.GET.get('is_vaccine','false')

    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
        fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')
    
    

    
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    #지불 정보
    dict_payment_method = {}
    query_payment_method= COMMCODE.objects.filter(commcode_grp='PAYMENT_METHOD',upper_commcode='000014').annotate(code = F('commcode'),name = fname).values('code','name')
    for data in query_payment_method:
        dict_payment_method.update({
            data['code'] : data['name']
            })


    kwargs = {}

    argument_list = []
    datas = []



    if category=='':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 


    kwargs['progress'] = 'done'
    if depart != '':
        kwargs['depart_id'] = depart
    if doctor != '':
        kwargs['doctor_id'] = doctor
    if payment_method !='':
        kwargs['payment__paymentrecord__method'] = payment_method
    if payment_status !='':
        kwargs['payment__progress'] = payment_status



    receptions = Reception.objects.filter(
            functools.reduce(operator.or_, argument_list),
            **kwargs ,
            payment__paymentrecord__status = 'paid',
            # payment__paymentrecord__paid__gt = 0,
            recorded_date__range = (date_min, date_max), 
        ).prefetch_related(
            'diagnosis__exammanager_set',
            'diagnosis__testmanager_set',
            'diagnosis__preceduremanager_set',
            'diagnosis__medicinemanager_set',
            'payment__paymentrecord_set',
        ).exclude(

            ).order_by("recorded_date").values_list('id',flat=True)


    list_id = list(receptions)
    list_id = list(set(list_id))


    argument_list = []
    for id in list_id:
        argument_list.append( Q(**{'id':id} ) )
    if len(list_id)==0:
        argument_list.append( Q(**{'id':0} ) )

    # receptions = Reception.objects.filter(
    #         functools.reduce(operator.or_, argument_list),
    #     ).prefetch_related(
    #         'diagnosis__exammanager_set',
    #         'diagnosis__testmanager_set',
    #         'diagnosis__preceduremanager_set',
    #         'diagnosis__medicinemanager_set',
    #     ).order_by("-id")

    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="AUDIT REPORT ' + date_start + '_' + date_end +'.xlsx"'
    
    #기본 양식 경로
    path = static('excel_form/I-MEDICARE_REPORT.xlsx')

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/audit_report.xlsx') #Workbook()
    # wb = load_workbook('/home/light/Desktop/Projects/imedicare2/static/excel_form/audit_report.xlsx')
    ws = wb.active# grab the active worksheet

    #선택한 날짜
    ws['B3'] = date_start + ' - ' + date_end

    #처음 시작 A7 ~ W7
    current_row = 7
    data_num = 1
    # print('argument_list: ', len(argument_list))
    for roof_cnt in range(len(argument_list) // 500  + 1):
        # print('roof_cnt: ', roof_cnt)
        receptions = Reception.objects.filter(
                functools.reduce(operator.or_, argument_list[ 500 * roof_cnt : 500 * (roof_cnt+1)]),
            ).prefetch_related(
                'diagnosis__exammanager_set',
                'diagnosis__testmanager_set',
                'diagnosis__preceduremanager_set',
                'diagnosis__medicinemanager_set',
                'payment__paymentrecord_set',
            ).order_by("recorded_date")  
        # print('receptions.count(): ', receptions.count())                  
        for i in range(receptions.count()):
            recorded_date = receptions[i].recorded_date
            #기본 정보
            ws['A' + str(current_row)] = data_num
            ws['B' + str(current_row)] = receptions[i].recorded_date.strftime('%Y-%m-%d')
            ws['C' + str(current_row)] = receptions[i].patient.get_chart_no()
            ws['D' + str(current_row)] = receptions[i].patient.name_kor + ' / ' + receptions[i].patient.name_eng
            ws['E' + str(current_row)] = receptions[i].depart.name
            ws['F' + str(current_row)] = receptions[i].doctor.name_short



            #진료 아이템
            ##진료비
            temp_row = current_row
            highest = 1
            count = 1
            tmp_exam_set = receptions[i].diagnosis.exammanager_set.all()
            for tmp_exam in tmp_exam_set:
                ws['G' + str(temp_row)] = tmp_exam.exam.code
                ws['H' + str(temp_row)] = tmp_exam.exam.name
                ws['I' + str(temp_row)] = "-"
                ws['J' + str(temp_row)] = tmp_exam.exam.price
                ws['K' + str(temp_row)] = "-"
                ws['L' + str(temp_row)] = "-"
                ws['M' + str(temp_row)] = tmp_exam.exam.get_price(recorded_date)
                ws['N' + str(temp_row)] = "Exam fee"
                temp_row += 1
                count += 1
    
            if tmp_exam_set.count() > highest:
                highest = tmp_exam_set.count() 

            ##약
            tmp_medicine_set = receptions[i].diagnosis.medicinemanager_set.all()
            for tmp_medicine in tmp_medicine_set:
                quantity = int(tmp_medicine.days) * int(tmp_medicine.amount)
                unit = tmp_medicine.medicine.get_price()
                price = quantity * unit

                print(quantity)
                print(unit)
                print(price)
                ws['G' + str(temp_row)] = tmp_medicine.medicine.code
                ws['H' + str(temp_row)] = tmp_medicine.medicine.name
                ws['I' + str(temp_row)] = tmp_medicine.amount
                ws['J' + str(temp_row)] = tmp_medicine.medicine.price
                ws['K' + str(temp_row)] = tmp_medicine.amount
                ws['L' + str(temp_row)] = tmp_medicine.medicine.unit_vie + '/' + tmp_medicine.medicine.unit
                ws['M' + str(temp_row)] = tmp_medicine.amount * tmp_medicine.days * tmp_medicine.medicine.get_price(recorded_date)
                ws['N' + str(temp_row)] = "Medicine"
                temp_row += 1
                count += 1

            if tmp_medicine_set.count() > highest:
                highest = tmp_medicine_set.count() 


            ##검사
            tmp_test_set = receptions[i].diagnosis.testmanager_set.all()
            for tmp_test in tmp_test_set:
                ws['G' + str(temp_row)] = tmp_test.test.code
                ws['H' + str(temp_row)] = tmp_test.test.name
                ws['I' + str(temp_row)] = ""
                ws['J' + str(temp_row)] = tmp_test.test.price
                ws['K' + str(temp_row)] = "-"
                ws['L' + str(temp_row)] = "-"
                ws['M' + str(temp_row)] = tmp_test.test.get_price(recorded_date)
                ws['N' + str(temp_row)] = "Lab"
                temp_row += 1
                count += 1
            
            if tmp_test_set.count() > highest:
                highest = tmp_test_set.count() 

            ##처치 및 방사선
            tmp_precedure_set = receptions[i].diagnosis.preceduremanager_set.all()
            for tmp_precedure in tmp_precedure_set:
                if 'R' in tmp_precedure.precedure.code:
                    ws['G' + str(temp_row)] = tmp_precedure.precedure.code
                    ws['H' + str(temp_row)] = tmp_precedure.precedure.name
                    ws['I' + str(temp_row)] = tmp_precedure.amount
                    ws['J' + str(temp_row)] = tmp_precedure.precedure.price
                    ws['K' + str(temp_row)] = tmp_precedure.amount
                    ws['L' + str(temp_row)] = "-"
                    ws['M' + str(temp_row)] = tmp_precedure.precedure.get_price(recorded_date)
                    ws['N' + str(temp_row)] = "Radiation"
                    temp_row += 1
                    count += 1
            
                else:
                    ws['G' + str(temp_row)] = tmp_precedure.precedure.code
                    ws['H' + str(temp_row)] = tmp_precedure.precedure.name
                    ws['I' + str(temp_row)] = tmp_precedure.amount
                    ws['J' + str(temp_row)] = tmp_precedure.precedure.price
                    ws['K' + str(temp_row)] = tmp_precedure.amount
                    ws['L' + str(temp_row)] = "-"
                    ws['M' + str(temp_row)] = tmp_precedure.precedure.get_price(recorded_date)
                    ws['N' + str(temp_row)] = "Procedure"
                    temp_row += 1
                    count += 1
            
            if tmp_precedure_set.filter(precedure__code__icontains='R').count() > highest:
                highest = tmp_precedure_set.filter(precedure__code__icontains='R').count()
            
            if tmp_precedure_set.exclude(precedure__code__icontains='R').count() > highest:
                highest = tmp_precedure_set.exclude(precedure__code__icontains='R').count()
            
            # real_paid_total = receptions.filter(id = receptions[i].id, 
            #                                     payment__paymentrecord__status='paid',
            #                                     payment__paymentrecord__paid__gt=0 ).prefetch_related(
            #         'payment__paymentrecord_set',
            #         )
            real_paid_total = receptions[i].payment.paymentrecord_set.filter(status__icontains='paid',
                                                                             paid__gt=0 )                   
            # print('real_paid_total: ', real_paid_total)
            # payment_total_paid_amount =  0 if real_paid_total['payment__paymentrecord__paid__sum'] is None else real_paid_total['payment__paymentrecord__paid__sum']            
            paid_sum = real_paid_total.aggregate(Sum('paid')).get('paid__sum')
            
            if paid_sum is None:
                paid_sum = 0             

            ws['O' + str(current_row)] = receptions[i].payment.sub_total
            if receptions[i].payment.discounted != 0:
                ws['P' + str(current_row)] = receptions[i].payment.sub_total / 100 * receptions[i].payment.discounted
            else:
                ws['P' + str(current_row)] = receptions[i].payment.discounted_amount
            ws['Q' + str(current_row)] = receptions[i].payment.additional
            ws['R' + str(current_row)] = receptions[i].payment.total
            ws['S' + str(current_row)] = receptions[i].payment.total - paid_sum
            ws['T' + str(current_row)] = paid_sum
            ws['W' + str(current_row)] = 'Yes' if receptions[i].need_invoice else 'NO'

            payment_record = receptions[i].payment.paymentrecord_set.filter(status = 'paid')

            if payment_record:
                ws['U' + str(current_row)] = receptions[i].payment.paymentrecord_set.first().method
            else:
                ws['U' + str(current_row)] = '-'
                
            if payment_record:
                str_record = ''
                for record_data in payment_record:
                    str_record += record_data.memo + "\n"
                ws['V' + str(current_row)] = str_record


            if count != 0:
                ws.merge_cells('A' + str(current_row) + ':A' + str(current_row + count-1))
                ws.merge_cells('B' + str(current_row) + ':B' + str(current_row + count-1))
                ws.merge_cells('C' + str(current_row) + ':C' + str(current_row + count-1))
                ws.merge_cells('D' + str(current_row) + ':D' + str(current_row + count-1))
                ws.merge_cells('E' + str(current_row) + ':E' + str(current_row + count-1))
                ws.merge_cells('F' + str(current_row) + ':F' + str(current_row + count-1))
                # ws.merge_cells('L' + str(current_row) + ':L' + str(current_row + count-1))
                # ws.merge_cells('M' + str(current_row) + ':M' + str(current_row + count-1))
                # ws.merge_cells('N' + str(current_row) + ':N' + str(current_row + count-1))
                ws.merge_cells('O' + str(current_row) + ':O' + str(current_row + count-1))
                ws.merge_cells('P' + str(current_row) + ':P' + str(current_row + count-1))
                ws.merge_cells('Q' + str(current_row) + ':Q' + str(current_row + count-1))
                ws.merge_cells('R' + str(current_row) + ':R' + str(current_row + count-1))
                ws.merge_cells('S' + str(current_row) + ':S' + str(current_row + count-1))
                ws.merge_cells('T' + str(current_row) + ':T' + str(current_row + count-1))
                ws.merge_cells('U' + str(current_row) + ':U' + str(current_row + count-1))
                ws.merge_cells('V' + str(current_row) + ':V' + str(current_row + count-1))
                ws.merge_cells('W' + str(current_row) + ':W' + str(current_row + count-1))
            current_row += count 
            data_num +=1

    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A7:V' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin


    wb.save(response)
    return response


def rec_report_excel(request):
    date_start = datetime.datetime.today().strftime("%Y-%m-%d")

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.max)

    kwargs = {}
    #kwargs['progress'] = 'done'

    datas = []
    #receptions = Reception.objects.filter(
    #        **kwargs ,
    #        recorded_date__range = (date_min, date_max), 
    #    ).prefetch_related(
    #        'diagnosis__exammanager_set',
    #        'diagnosis__testmanager_set',
    #        'diagnosis__preceduremanager_set',
    #        'diagnosis__medicinemanager_set',
    #        'payment__paymentrecord_set',
    #    ).order_by("-id")

    #paid_query = PaymentRecord.objects.filter(
    #    **kwargs ,
    #    recorded_date__range = (date_min, date_max), 
    #    )


    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="RECEPTION REPORT ' + date_start +'.xlsx"'

    #엑셀 파일 불러오기
    # try:
    #     wb = load_workbook('/home/imedicare/Cofee/static/excel_form/reception_report.xlsx') #Workbook()
    # except:
    #     wb = load_workbook('/Users/light/Desktop/Work/imdc/imedicare2/static/excel_form/reception_report.xlsx') #Workbook()
    # wb = load_workbook('/home/light/Desktop/Projects/imedicare2/static/excel_form/reception_report.xlsx') #Workbook()
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/reception_report.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet

    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                    left=Side(border_style="thin", color="000000") ,
                    right=Side(border_style="thin", color="000000") ,
                    bottom=Side(border_style="thin", color="000000") )

    font_base = Font(
        name = 'Times New Roman',
        bold = False,
        size = 14,
        )
    alignment = Alignment(
        horizontal = 'center',
        vertical = 'center',
        wrap_text=True
        )
    font_bold = Font(
        name = 'Times New Roman',
        bold = True,
        size = 14,
        )

    number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'
    date_format = 'Date'
    general_format = 'General'

    list_id = []
    recept_tmp = Reception.objects.filter(
            recorded_date__range = (date_min, date_max),
            **kwargs ,
            progress = 'done',
        ).values_list('id',flat=True)

    paymentrecord_tmp = PaymentRecord.objects.filter(
            date__range = (date_min, date_max) ,
            **kwargs ,
            status = 'paid',
        ).values_list("payment__reception__id",flat=True)

    list_id = list(recept_tmp) + list(paymentrecord_tmp)
    list_id = list(set(list_id))


    argument_list = []
    for id in list_id:
        argument_list.append( Q(**{'id':id} ) )
    if len(list_id)==0:
        argument_list.append( Q(**{'id':0} ) )

    receptions = Reception.objects.filter(
            functools.reduce(operator.or_, argument_list),
            progress='done',
        ).prefetch_related(
            'diagnosis__exammanager_set',
            'diagnosis__testmanager_set',
            'diagnosis__preceduremanager_set',
            'diagnosis__medicinemanager_set',
        ).order_by("-id")



    #선택한 날짜
    ws['A1'] = 'BÁO CÁO DOANH THU PHÒNG KHÁM NGÀY ' + date_min.strftime('%d.%m.%Y')
    
    current_row = 7
    data_num = 1

    for reception in receptions:
        ws['A' + str(current_row)] = data_num
        ws['B' + str(current_row)] = ''
        
        ws['C' + str(current_row)] = date_start
        ws['D' + str(current_row)] = reception.depart.name
        ws['E' + str(current_row)] = reception.patient.name_kor + ' / ' + reception.patient.name_eng
        ws['F' + str(current_row)] = reception.diagnosis.diagnosis
        
        paymentrecords = None
        if reception.recorded_date.strftime('%Y-%m-%d') != date_start:#과거 방문 / 당일 지급
            ws['G' + str(current_row)] = '0'
            ws['H' + str(current_row)] = '0'
            ws['I' + str(current_row)] = '0'
            ws['J' + str(current_row)] = '0'
            ws['K' + str(current_row)] = date_start

            paymentrecords = reception.payment.paymentrecord_set.filter(
                date__range = (date_min, date_max),
                status='paid'
                )

            total = paymentrecords.filter(method = 'cash').annotate(sum = Coalesce(Sum('paid'),0))
            if len(total)!=0:
                ws['W' + str(current_row)] = total[0].sum

            total = paymentrecords.filter(method = 'BIDV_CARD').annotate(sum = Coalesce(Sum('paid'),0))
            if len(total)!=0:
                ws['X' + str(current_row)] = total[0].sum

            total = paymentrecords.filter(method = 'BIDV_TRANS').annotate(sum = Coalesce(Sum('paid'),0))
            if len(total)!=0:
                ws['Y' + str(current_row)] = total[0].sum

            total = paymentrecords.filter(method = 'VP_CARD').annotate(sum = Coalesce(Sum('paid'),0))
            if len(total)!=0:
                ws['Z' + str(current_row)] = total[0].sum

            total = paymentrecords.filter(method = 'VP_TRANS').annotate(sum = Coalesce(Sum('paid'),0))
            if len(total)!=0:
                ws['AA' + str(current_row)] = total[0].sum            

            total = paymentrecords.filter(method = 'TP_CARD').annotate(sum = Coalesce(Sum('paid'),0))
            if len(total)!=0:
                ws['AB' + str(current_row)] = total[0].sum

            total = paymentrecords.filter(method = 'TP_TRANS').annotate(sum = Coalesce(Sum('paid'),0))
            if len(total)!=0:
                ws['AC' + str(current_row)] = total[0].sum

            total = paymentrecords.filter(method = 'SHINHAN_TRANS').annotate(sum = Coalesce(Sum('paid'),0))
            if len(total)!=0:
                ws['AD' + str(current_row)] = total[0].sum                         

        else:#당일 방문
            ws['G' + str(current_row)] = reception.payment.sub_total

            if reception.payment.discounted != 0:
                ws['H' + str(current_row)] = reception.payment.sub_total / 100 * reception.payment.discounted
            else:
                ws['H' + str(current_row)] = reception.payment.discounted_amount

            ws['I' + str(current_row)] = reception.payment.additional
            #ws['J' + str(current_row)] = reception.payment.total

            if reception.payment.paymentrecord_set.count() == 0:#미납
                ws['K' + str(current_row)] = ''
            else:#지불
                paymentrecords = reception.payment.paymentrecord_set.filter(
                    date__range = (date_min, date_max),
                    status='paid'
                    )

                ws['K' + str(current_row)] = reception.recorded_date.strftime('%Y-%m-%d')

                total = paymentrecords.filter(method = 'cash').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['L' + str(current_row)] = total[0].sum

                total = paymentrecords.filter(method = 'BIDV_CARD').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['M' + str(current_row)] = total[0].sum

                total = paymentrecords.filter(method = 'VP_CARD').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['N' + str(current_row)] = total[0].sum
                
                total = paymentrecords.filter(method = 'TP_CARD').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['O' + str(current_row)] = total[0].sum

                total = paymentrecords.filter(method = 'BIDV_TRANS').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['P' + str(current_row)] = total[0].sum

                total = paymentrecords.filter(method = 'VP_TRANS').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['Q' + str(current_row)] = total[0].sum
                
                total = paymentrecords.filter(method = 'TP_TRANS').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['Q' + str(current_row)] = total[0].sum
                
                total = paymentrecords.filter(method = 'SHINHAN_TRANS').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['S' + str(current_row)] = total[0].sum

                total = paymentrecords.filter(method = 'KOOKMIN').annotate(sum = Coalesce(Sum('paid'),0))
                if len(total)!=0:
                    ws['T' + str(current_row)] = total[0].sum

        #ws['Q' + str(current_row)] = '=SUM(L' + str(current_row) + ':P' + str(current_row) + ')'
        #ws['R' + str(current_row)] = '=+J' + str(current_row) + '-Q' + str(current_row) + ')'
        #ws['V' + str(current_row)] = '=SUM(S' + str(current_row) + ':U' + str(current_row) + ')'

        
       #ws['L' + str(current_row)].number_format  = number_format
       #ws['M' + str(current_row)].number_format  = number_format
       #ws['N' + str(current_row)].number_format  = number_format
       #ws['O' + str(current_row)].number_format  = number_format
       #ws['P' + str(current_row)].number_format  = number_format
       #ws['Q' + str(current_row)].number_format  = number_format
       #ws['R' + str(current_row)].number_format  = number_format
       #ws['S' + str(current_row)].number_format  = number_format
       #ws['T' + str(current_row)].number_format  = number_format
       #ws['U' + str(current_row)].number_format  = number_format
       #ws['V' + str(current_row)].number_format  = number_format
       

        
        ws['AF' + str(current_row)] = reception.payment.progress
        ws['AG' + str(current_row)] = reception.doctor.name_eng
        #if paymentrecords:
        #    ws['Y' + str(current_row)] = paymentrecords.memo

        if paymentrecords: 
            str_record = ''
            for record_data in paymentrecords.order_by("-date"):
                str_record += record_data.memo + "\n"
            else:
                str_record = reception.payment.memo
            ws['AH' + str(current_row)] = str_record
        else:
            ws['AH' + str(current_row)] = reception.payment.memo
        pay_time = reception.payment.pay_time

        datetime_str = '01/01/20 00:00:00'
        if pay_time == datetime.datetime.strptime(datetime_str, '%m/%d/%y %H:%M:%S'):
            ws['AI' + str(current_row)] = ''
        else:
            ws['AI' + str(current_row)] = reception.payment.pay_time

        data_num += 1 
        current_row += 1
    
    #rows = ws['A7:X' + str(current_row + 2)]
    #for row in rows:
    #    for cell in row:
    #        cell.font = font_base
    #        cell.border = border_thin
    #
    #ws.merge_cells('D' + str(current_row +2) + ':F' + str(current_row +2))
    #ws['D' + str(current_row +2 )] = 'Total'
    #ws['G' + str(current_row +2 )] = '=SUM(G7:G' + str(current_row) + ')'
    #ws['H' + str(current_row +2 )] = '=SUM(H7:H' + str(current_row) + ')'
    #ws['I' + str(current_row +2 )] = '=SUM(I7:I' + str(current_row) + ')'
    #ws['J' + str(current_row +2 )] = '=SUM(J7:J' + str(current_row) + ')'
    #ws['K' + str(current_row +2 )] = '=SUM(K7:K' + str(current_row) + ')'
    #ws['L' + str(current_row +2 )] = '=SUM(L7:L' + str(current_row) + ')'
    #ws['M' + str(current_row +2 )] = '=SUM(M7:M' + str(current_row) + ')'
    #ws['N' + str(current_row +2 )] = '=SUM(N7:N' + str(current_row) + ')'
    #ws['O' + str(current_row +2 )] = '=SUM(O7:O' + str(current_row) + ')'
    #ws['P' + str(current_row +2 )] = '=SUM(P7:P' + str(current_row) + ')'
    #ws['Q' + str(current_row +2 )] = '=SUM(Q7:Q' + str(current_row) + ')'
    #ws['R' + str(current_row +2 )] = '=SUM(R7:R' + str(current_row) + ')'
    #ws['S' + str(current_row +2 )] = '=SUM(S7:S' + str(current_row) + ')'
    #ws['T' + str(current_row +2 )] = '=SUM(T7:T' + str(current_row) + ')'
    #ws['U' + str(current_row +2 )] = '=SUM(U7:U' + str(current_row) + ')'
    #ws['V' + str(current_row +2 )] = '=SUM(V7:V' + str(current_row) + ')'
    #ws['W' + str(current_row +2 )] = '=SUM(W7:W' + str(current_row) + ')'
    #
    #array = ws['A' + str(current_row + 2) + ':W' + str(current_row + 2)]
    #for row in array:
    #    for cell in row:
    #        cell.number_format = number_format
    #        cell.font = font_bold
    #
    #
    #
    #rows = ws['H' + str(current_row + 4)] = 'ENT'
    #rows = ws['I' + str(current_row + 4)] = 'DERM'
    #rows = ws['J' + str(current_row + 4)] = 'IM'
    #rows = ws['K' + str(current_row + 4)] = 'PM'
    #rows = ws['L' + str(current_row + 4)] = 'PS'
    #rows = ws['M' + str(current_row + 4)] = 'DENTAL'
    #rows = ws['N' + str(current_row + 4)] = 'OBGYN'
    #
    #ws.merge_cells('D' + str(current_row + 4) + ':F' + str(current_row +4))
    #ws['D' + str(current_row +4 )] = 'DOANH THU'
    #
    #ws.merge_cells('D' + str(current_row + 5) + ':F' + str(current_row +5))
    #ws['D' + str(current_row +5 )] = 'Tổng doanh thu dịch vụ (1)/ Total Real revenue'
    #ws['G' + str(current_row +5 )] = '=H' + str(current_row +6 ) + '+H' + str(current_row +7 )
    #ws['G' + str(current_row +5 )] = '=I' + str(current_row +6 ) + '+I' + str(current_row +7 )
    #ws['G' + str(current_row +5 )] = '=J' + str(current_row +6 ) + '+J' + str(current_row +7 )
    #ws['G' + str(current_row +5 )] = '=K' + str(current_row +6 ) + '+K' + str(current_row +7 )
    #ws['G' + str(current_row +5 )] = '=L' + str(current_row +6 ) + '+L' + str(current_row +7 )
    #ws['G' + str(current_row +5 )] = '=M' + str(current_row +6 ) + '+M' + str(current_row +7 )
    #ws['G' + str(current_row +5 )] = '=N' + str(current_row +6 ) + '+N' + str(current_row +7 )
    #
    #ws.merge_cells('D' + str(current_row + 6) + ':F' + str(current_row +6))
    #ws['D' + str(current_row +6 )] = 'Tổng tiền giảm giá (2)/ Total discount'
    #ws['G' + str(current_row +6 )] = '=SUM(H' + str(current_row +6 ) + ':N' + str(current_row +6 ) + ')'
    #ws['H' + str(current_row +6 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 4) + ',H7:H' + str(current_row) + ')'
    #ws['I' + str(current_row +6 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 4) + ',H7:H' + str(current_row) + ')'
    #ws['J' + str(current_row +6 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 4) + ',H7:H' + str(current_row) + ')'
    #ws['K' + str(current_row +6 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 4) + ',H7:H' + str(current_row) + ')'
    #ws['L' + str(current_row +6 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 4) + ',H7:H' + str(current_row) + ')'
    #ws['M' + str(current_row +6 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 4) + ',H7:H' + str(current_row) + ')'
    #ws['N' + str(current_row +6 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 4) + ',H7:H' + str(current_row) + ')'
    #
    #ws.merge_cells('D' + str(current_row + 7) + ':F' + str(current_row +7))
    #ws['D' + str(current_row +7 )] = 'Doanh thu thực tế (1)-(2)/ Due amount'
    #ws['G' + str(current_row +7 )] = '=SUM(H' + str(current_row +7 ) + ':N' + str(current_row +7 ) + ')'
    #ws['H' + str(current_row +7 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 4) + ',J7:J' + str(current_row) + ')'
    #ws['N' + str(current_row +7 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 4) + ',J7:J' + str(current_row) + ')'    
    #ws['I' + str(current_row +7 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 4) + ',J7:J' + str(current_row) + ')'
    #ws['J' + str(current_row +7 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 4) + ',J7:J' + str(current_row) + ')'
    #ws['K' + str(current_row +7 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 4) + ',J7:J' + str(current_row) + ')'
    #ws['L' + str(current_row +7 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 4) + ',J7:J' + str(current_row) + ')'
    #ws['M' + str(current_row +7 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 4) + ',J7:J' + str(current_row) + ')'    
    #
    #
    #ws.merge_cells('E' + str(current_row + 8) + ':E' + str(current_row +8))
    #ws['E' + str(current_row +8 )] = 'CASH (Tiền mặt)'
    #
    #ws.merge_cells('E' + str(current_row + 9) + ':E' + str(current_row +10))
    #ws['E' + str(current_row +9 )] = 'POS (Quẹt thẻ)'
    #ws['F' + str(current_row +9 )] = 'BIDV'
    #ws['F' + str(current_row +10 )] = 'VP BANK'
    #
    #ws.merge_cells('E' + str(current_row + 11) + ':E' + str(current_row +12))
    #ws['E' + str(current_row +11 )] = 'TRANFER (Chuyển khoản)'
    #ws['F' + str(current_row +11 )] = 'BIDV'
    #ws['F' + str(current_row +12 )] = 'VP BANK'
    #
    #ws.merge_cells('E' + str(current_row + 13) + ':F' + str(current_row +13))
    #ws['E' + str(current_row +13 )] = 'Unpaid amount/Chưa thanh toán'
    #
    #
    #
    #ws.merge_cells('D' + str(current_row + 8) + ':D' + str(current_row +13))
    #ws['D' + str(current_row +8 )] = 'Phương thức thanh toán/ Payment method'
    #
    #ws['G' + str(current_row +8 )] = '=SUM(H' + str(current_row +8 ) + ':N' + str(current_row +8 ) + ')'
    #ws['H' + str(current_row +8 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 4) + ',L7:L' + str(current_row) + ')'
    #ws['I' + str(current_row +8 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 4) + ',L7:L' + str(current_row) + ')'
    #ws['J' + str(current_row +8 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 4) + ',L7:L' + str(current_row) + ')'
    #ws['K' + str(current_row +8 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 4) + ',L7:L' + str(current_row) + ')'
    #ws['L' + str(current_row +8 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 4) + ',L7:L' + str(current_row) + ')'
    #ws['M' + str(current_row +8 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 4) + ',L7:L' + str(current_row) + ')'    
    #ws['N' + str(current_row +8 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 4) + ',L7:L' + str(current_row) + ')'    
    #
    #ws['G' + str(current_row +9 )] = '=SUM(H' + str(current_row +9 ) + ':N' + str(current_row +9 ) + ')'
    #ws['H' + str(current_row +9 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 4) + ',M7:M' + str(current_row) + ')'
    #ws['I' + str(current_row +9 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 4) + ',M7:M' + str(current_row) + ')'
    #ws['J' + str(current_row +9 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 4) + ',M7:M' + str(current_row) + ')'
    #ws['K' + str(current_row +9 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 4) + ',M7:M' + str(current_row) + ')'
    #ws['L' + str(current_row +9 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 4) + ',M7:M' + str(current_row) + ')'
    #ws['M' + str(current_row +9 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 4) + ',M7:M' + str(current_row) + ')' 
    #ws['N' + str(current_row +9 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 4) + ',M7:M' + str(current_row) + ')'    
    #
    #ws['G' + str(current_row +10 )] = '=SUM(H' + str(current_row +10 ) + ':N' + str(current_row +10 ) + ')'
    #ws['H' + str(current_row +10 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 4) + ',N7:N' + str(current_row) + ')'
    #ws['I' + str(current_row +10 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 4) + ',N7:N' + str(current_row) + ')'
    #ws['J' + str(current_row +10 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 4) + ',N7:N' + str(current_row) + ')'
    #ws['K' + str(current_row +10 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 4) + ',N7:N' + str(current_row) + ')'
    #ws['L' + str(current_row +10 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 4) + ',N7:N' + str(current_row) + ')'
    #ws['M' + str(current_row +10 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 4) + ',N7:N' + str(current_row) + ')' 
    #ws['N' + str(current_row +10 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 4) + ',N7:N' + str(current_row) + ')'  
    #
    #ws['G' + str(current_row +11 )] = '=SUM(H' + str(current_row +11 ) + ':N' + str(current_row +11 ) + ')'
    #ws['H' + str(current_row +11 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 4) + ',O7:O' + str(current_row) + ')'
    #ws['I' + str(current_row +11 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 4) + ',O7:O' + str(current_row) + ')'
    #ws['J' + str(current_row +11 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 4) + ',O7:O' + str(current_row) + ')'
    #ws['K' + str(current_row +11 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 4) + ',O7:O' + str(current_row) + ')'
    #ws['L' + str(current_row +11 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 4) + ',O7:O' + str(current_row) + ')'
    #ws['M' + str(current_row +11 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 4) + ',O7:O' + str(current_row) + ')' 
    #ws['N' + str(current_row +11 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 4) + ',O7:O' + str(current_row) + ')' 
    #
    #ws['G' + str(current_row +12 )] = '=SUM(H' + str(current_row +12 ) + ':N' + str(current_row +12 ) + ')'
    #ws['H' + str(current_row +12 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 4) + ',P7:P' + str(current_row) + ')'
    #ws['I' + str(current_row +12 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 4) + ',P7:P' + str(current_row) + ')'
    #ws['J' + str(current_row +12 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 4) + ',P7:P' + str(current_row) + ')'
    #ws['K' + str(current_row +12 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 4) + ',P7:P' + str(current_row) + ')'
    #ws['L' + str(current_row +12 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 4) + ',P7:P' + str(current_row) + ')'
    #ws['M' + str(current_row +12 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 4) + ',P7:P' + str(current_row) + ')' 
    #ws['N' + str(current_row +12 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 4) + ',P7:P' + str(current_row) + ')' 
    #
    #ws['G' + str(current_row +13 )] = '=SUM(H' + str(current_row +13 ) + ':N' + str(current_row +13 ) + ')'
    #ws['H' + str(current_row +13 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 4) + ',R7:R' + str(current_row) + ')'
    #ws['I' + str(current_row +13 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 4) + ',R7:R' + str(current_row) + ')'
    #ws['J' + str(current_row +13 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 4) + ',R7:R' + str(current_row) + ')'
    #ws['K' + str(current_row +13 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 4) + ',R7:R' + str(current_row) + ')'
    #ws['L' + str(current_row +13 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 4) + ',R7:R' + str(current_row) + ')'
    #ws['M' + str(current_row +13 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 4) + ',R7:R' + str(current_row) + ')' 
    #ws['N' + str(current_row +13 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 4) + ',R7:R' + str(current_row) + ')' 
    #
    #
    #
    #ws.merge_cells('D' + str(current_row + 15) + ':F' + str(current_row +15))
    #ws['D' + str(current_row +15 )] = 'CÔNG NỢ'
    #
    #rows = ws['H' + str(current_row + 15)] = 'ENT'
    #rows = ws['I' + str(current_row + 15)] = 'DERM'
    #rows = ws['J' + str(current_row + 15)] = 'IM'
    #rows = ws['K' + str(current_row + 15)] = 'PM'
    #rows = ws['L' + str(current_row + 15)] = 'PS'
    #rows = ws['M' + str(current_row + 15)] = 'DENTAL'
    #rows = ws['N' + str(current_row + 15)] = 'OBGYN'
    #
    #ws.merge_cells('D' + str(current_row + 16) + ':F' + str(current_row +16))
    #ws['D' + str(current_row +16 )] = 'Tổng số tiền thanh toán công nợ/ Total Paid debit'
    #
    #ws.merge_cells('D' + str(current_row + 17) + ':D' + str(current_row +19))
    #ws['D' + str(current_row +17 )] = 'Phương thức thanh toán/ Payment method'
    #
    #ws.merge_cells('E' + str(current_row + 17) + ':F' + str(current_row +17))
    #ws['E' + str(current_row +17 )] = 'Tiền mặt/ Cash'
    #
    #ws.merge_cells('E' + str(current_row + 18) + ':E' + str(current_row +19))
    #ws['E' + str(current_row +18 )] = 'BIDV'
    #
    #ws['F' + str(current_row +19 )] = 'POS (Quẹt thẻ)'
    #ws['F' + str(current_row +18 )] = 'TRANFER (Chuyển khoản)'
    #
    #ws['G' + str(current_row +16 )] = '=SUM(G' + str(current_row +17 ) + ':G' + str(current_row +19 ) + ')'
    #ws['H' + str(current_row +16 )] = '=SUM(H' + str(current_row +17 ) + ':H' + str(current_row +19 ) + ')'
    #ws['I' + str(current_row +16 )] = '=SUM(I' + str(current_row +17 ) + ':I' + str(current_row +19 ) + ')'
    #ws['J' + str(current_row +16 )] = '=SUM(J' + str(current_row +17 ) + ':J' + str(current_row +19 ) + ')'
    #ws['K' + str(current_row +16 )] = '=SUM(K' + str(current_row +17 ) + ':K' + str(current_row +19 ) + ')'
    #ws['L' + str(current_row +16 )] = '=SUM(L' + str(current_row +17 ) + ':L' + str(current_row +19 ) + ')'
    #ws['M' + str(current_row +16 )] = '=SUM(M' + str(current_row +17 ) + ':M' + str(current_row +19 ) + ')'
    #ws['N' + str(current_row +16 )] = '=SUM(N' + str(current_row +17 ) + ':N' + str(current_row +19 ) + ')'
    #
    #ws['G' + str(current_row +17 )] = '=SUM(H' + str(current_row +17 ) + ':N' + str(current_row +17 ) + ')'
    #ws['H' + str(current_row +17 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 15) + ',S7:S' + str(current_row) + ')'
    #ws['I' + str(current_row +17 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 15) + ',S7:S' + str(current_row) + ')'
    #ws['J' + str(current_row +17 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 15) + ',S7:S' + str(current_row) + ')'
    #ws['K' + str(current_row +17 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 15) + ',S7:S' + str(current_row) + ')'
    #ws['L' + str(current_row +17 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 15) + ',S7:S' + str(current_row) + ')'
    #ws['M' + str(current_row +17 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 15) + ',S7:S' + str(current_row) + ')'
    #ws['N' + str(current_row +17 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 15) + ',S7:S' + str(current_row) + ')'
    #
    #ws['G' + str(current_row +18 )] = '=SUM(H' + str(current_row +18 ) + ':N' + str(current_row +18 ) + ')'
    #ws['H' + str(current_row +18 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 15) + ',T7:T' + str(current_row) + ')'
    #ws['I' + str(current_row +18 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 15) + ',T7:T' + str(current_row) + ')'
    #ws['J' + str(current_row +18 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 15) + ',T7:T' + str(current_row) + ')'
    #ws['K' + str(current_row +18 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 15) + ',T7:T' + str(current_row) + ')'
    #ws['L' + str(current_row +18 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 15) + ',T7:T' + str(current_row) + ')'
    #ws['M' + str(current_row +18 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 15) + ',T7:T' + str(current_row) + ')'
    #ws['N' + str(current_row +18 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 15) + ',T7:T' + str(current_row) + ')'
    #
    #ws['G' + str(current_row +19 )] = '=SUM(H' + str(current_row +19 ) + ':N' + str(current_row +19 ) + ')'
    #ws['H' + str(current_row +19 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'H' + str(current_row + 15) + ',U7:U' + str(current_row) + ')'
    #ws['I' + str(current_row +19 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'I' + str(current_row + 15) + ',U7:U' + str(current_row) + ')'
    #ws['J' + str(current_row +19 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'J' + str(current_row + 15) + ',U7:U' + str(current_row) + ')'
    #ws['K' + str(current_row +19 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'K' + str(current_row + 15) + ',U7:U' + str(current_row) + ')'
    #ws['L' + str(current_row +19 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'L' + str(current_row + 15) + ',U7:U' + str(current_row) + ')'
    #ws['M' + str(current_row +19 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'M' + str(current_row + 15) + ',U7:U' + str(current_row) + ')'
    #ws['N' + str(current_row +19 )] = '=SUMIF(D7:D' + str(current_row) + ',' + 'N' + str(current_row + 15) + ',U7:U' + str(current_row) + ')'
    #
    #
    #tmp_arr = []
    #tmp_arr.append(ws['D' + str(current_row + 4) + ':N' + str(current_row + 13)])
    #tmp_arr.append(ws['D' + str(current_row + 15) + ':N' + str(current_row + 19)])
    #for array in tmp_arr:
    #    for row in array:
    #        for cell in row:
    #            cell.border = border_thin
    #
    #tmp_arr = []
    #tmp_arr.append(ws['D' + str(current_row + 4) + ':N' + str(current_row + 6)])
    #tmp_arr.append(ws['D' + str(current_row + 6) + ':G' + str(current_row + 13)])
    #
    #tmp_arr.append(ws['D' + str(current_row + 15) + ':N' + str(current_row + 16)])
    #tmp_arr.append(ws['D' + str(current_row + 15) + ':G' + str(current_row + 19)])
    #for array in tmp_arr:
    #    for row in array:
    #        for cell in row:
    #            cell.font = font_bold
    #
    #tmp_arr = []
    #tmp_arr.append(ws['D' + str(current_row + 5) + ':N' + str(current_row + 13)])
    #tmp_arr.append(ws['D' + str(current_row + 15) + ':N' + str(current_row + 19)])
    #tmp_arr.append(ws['G7:J' + str(current_row)])
    #for array in tmp_arr:
    #    for row in array:
    #        for cell in row:
    #            cell.number_format = number_format
    #         
    #array = ws['A7:Y' + str(current_row + 19)]
    #for row in array:
    #    for cell in row:
    #        cell.alignment = alignment
    #
    #
    #no_border = Border(top=Side(border_style=None) ,
    #                left=Side(border_style=None) ,
    #                right=Side(border_style=None) ,
    #                bottom=Side(border_style="none", color="000000") )
    #
    #
    #array = ws['D' + str(current_row + 4) + ':F' + str(current_row + 4)]
    #for row in array:
    #    for cell in row:
    #        cell.border = no_border
    #        cell.fill = PatternFill(start_color="C6E0B4",end_color="C6E0B4", fill_type = "solid")
    #            
    #array = ws['D' + str(current_row + 15) + ':F' + str(current_row + 15)]
    #for row in array:
    #    for cell in row:
    #        cell.border = no_border
    #        cell.fill = PatternFill(start_color="FFE699",end_color="FFE699", fill_type = "solid")


    
    wb.save(response)
    return response



#고객 정보 다운로드
def cumstomer_management_excel(request):
    depart = request.GET.get('depart')
    date_start = request.GET.get('start')
    date_end = request.GET.get('end')


    #엑셀
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Customer_Datas_' + datetime.datetime.now().strftime('%Y_%m_%d') +'.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/customer_excel_form.xlsx') #Workbook()
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/customer_excel_form.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end,"%Y-%m-%d").date(), datetime.time.max)

    kwargs={}
    if depart != '':
        kwargs['depart_id']=depart

    patients = Patient.objects.filter(date_registered__range = (date_min, date_max))
    count_page = patients.values_list('id')

    current_row = 2
    for patient in patients:
        sub_query = Reception.objects.filter(
            **kwargs,
            patient_id = patient.id
            ).exclude(progress='deleted').prefetch_related('payment__paymentrecord_set')
        visits = sub_query.count()
        if depart != '' and visits == 0: # 과를 선택했을때, 방문수도 0이면 안보여줌
            continue

        total_amount = sub_query.aggregate(total_price=Sum('payment__paymentrecord__paid'))
       

        ws['A' + str(current_row)] = patient.get_chart_no()
        ws['B' + str(current_row)] = patient.name_kor
        ws['C' + str(current_row)] = patient.name_eng
        ws['D' + str(current_row)] = patient.date_of_birth.strftime('%Y-%m-%d')
        ws['E' + str(current_row)] = patient.gender
        ws['F' + str(current_row)] = patient.nationality
        ws['G' + str(current_row)] = patient.phone
        ws['H' + str(current_row)] = patient.address
        ws['I' + str(current_row)] = patient.date_registered.strftime('%Y-%m-%d')
        ws['J' + str(current_row)] = visits
        ws['K' + str(current_row)] = 0 if total_amount['total_price'] == None else total_amount['total_price']

        
        current_row += 1

    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A2:K' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin

    wb.save(response)
    return response


@login_required
def debit_report_excel(request):
    
    start_date = request.GET.get('date_start')
    end_date = request.GET.get('date_end')

    date_min = datetime.datetime.combine(datetime.datetime.strptime(start_date, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end_date, "%Y-%m-%d").date(), datetime.time.max)


    #엑셀
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="EX THEO DOI CONG NO_' + start_date + '_' + end_date +'.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/EX THEO DOI CONG NO.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet

    ws['D1'] = 'Date : ' + start_date + ' ~ ' + end_date


    receptions = Reception.objects.filter(
            recorded_date__range = (date_min, date_max), 
            progress='done',
        ).prefetch_related().order_by("-id")



    current_row = 7
    writing_number = 1
    for data in receptions:
        payment_records = data.payment.paymentrecord_set.filter(status='paid')
        
        if payment_records.count() != 0 :
            if payment_records.first().date.date() == data.recorded_date.date():
                if payment_records.first().paid == data.payment.total:
                    continue

        ws['A' + str(current_row)] = writing_number
        ws['B' + str(current_row)] = data.recorded_date.strftime('%Y-%m-%d')
        ws['C' + str(current_row)] = data.depart.name
        ws['D' + str(current_row)] = data.patient.get_name_kor_eng()
        ws['E' + str(current_row)] = "{:06d}".format(data.patient.id)
        ws['F' + str(current_row)] = data.patient.phone
        ws['G' + str(current_row)] = ''
        
        ws['H' + str(current_row)] = data.payment.sub_total
        if data.payment.reception.payment.discounted != 0:
            ws['I' + str(current_row)] = data.payment.sub_total / 100 * data.payment.discounted
        else:
            ws['I' + str(current_row)] = data.payment.discounted_amount
        ws['J' + str(current_row)] = data.payment.total

        cul_bigin = 11
        for payment_record in payment_records:
            if payment_record.paid == 0:
                continue
            ws.cell(row = current_row, column = cul_bigin ).value = payment_record.date.strftime('%Y-%m-%d')
            ws.cell(row = current_row, column = cul_bigin + 1 ).value = payment_record.paid
            cul_bigin += 2



        ws['U' + str(current_row)] = '=+J'+str(current_row)+'-L'+str(current_row)+'-N'+str(current_row)+'-P'+str(current_row)+'-R'+str(current_row)+'-T'+str(current_row)
        

        current_row += 1
        writing_number += 1


    ws.merge_cells('C' + str(current_row+2) + ':G' + str(current_row+2))
    ws['C' + str(current_row+2)] = 'TỔNG CỘNG/ TOTAL'
    ws['H' + str(current_row+2)] = '=+SUM(H7:H' + str(current_row) + ')'
    ws['I' + str(current_row+2)] = '=+SUM(I7:I' + str(current_row) + ')'
    ws['J' + str(current_row+2)] = '=+SUM(J7:J' + str(current_row) + ')'
    ws['L' + str(current_row+2)] = '=+SUM(L7:L' + str(current_row) + ')'
    ws['O' + str(current_row+2)] = '=+SUM(O7:O' + str(current_row) + ')'
    ws['R' + str(current_row+2)] = '=+SUM(R7:R' + str(current_row) + ')'
    ws['T' + str(current_row+2)] = '=+SUM(T7:T' + str(current_row) + ')'
    ws['U' + str(current_row+2)] = '=+SUM(U7:U' + str(current_row) + ')'

    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                    left=Side(border_style="thin", color="000000") ,
                    right=Side(border_style="thin", color="000000") ,
                    bottom=Side(border_style="thin", color="000000") )

    font_base = Font(
        name = 'Times New Roman',
        bold = False,
        size = 12,
        )
    alignment = Alignment(
        horizontal = 'center',
        vertical = 'center',
        wrap_text=True
        )
    font_bold = Font(
        name = 'Times New Roman',
        bold = True,
        size = 12,
        )
    number_format = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

    tmp_arr=[]
    tmp_arr.append(ws['A7:U' + str(current_row+2)])
    for array in tmp_arr:
        for row in array:
            for cell in row:
                cell.number_format = number_format
                cell.alignment = alignment
                cell.font= font_base
                cell.border= border_thin
    
    tmp_arr=[]
    tmp_arr.append(ws['A' + str(current_row+2) + ':U' + str(current_row+2)])
    for array in tmp_arr:
        for row in array:
            for cell in row:
                cell.font = font_bold

    wb.save(response)
    return response


@login_required
def test_statistics_excel(request):
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    depart = request.GET.get('depart')
    doctor = request.GET.get('doctor')
    category = request.GET.get('patient_type','')
    string = request.GET.get('patient_search','')    

    kwargs = {}
    if depart is not None and depart != '':
        kwargs['depart'] = depart # 기본 
    if doctor  is not None and doctor != '':
        kwargs['payment__reception__doctor_id'] = doctor            

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Lab_statistics_' + start_date + '_' + end_date +'.xlsx"'
    
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Lab_statistics.xlsx') #Workbook()

    ws = wb.get_sheet_by_name('Theo so luong')# grab the active worksheet

    ws['E2'] = 'Ngày : ' + start_date 
    ws['F2'] = 'đến ngày : ' + end_date

    date_min = datetime.datetime.combine(datetime.datetime.strptime(start_date,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end_date,"%Y-%m-%d").date(), datetime.time.max)

    depart_list = Depart.objects.filter().values('id','name')
    
    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    font_base = Font(
        name = 'Arial',
        bold = False,
        size = 11,
        )
    
    ##Test
    test_class_query = TestClass.objects.all()
    test_class_dict = {}
    for test_class in test_class_query:
        test_class_dict[str(test_class.id)] = test_class.name

    current_row = 7
    writing_number = 1

    tests = Test.objects.all().order_by('code')
    for test in tests:
        sub_query = Reception.objects.filter(
            recorded_date__range = (date_min, date_max), 
            diagnosis__testmanager__test = test.id,
        ).exclude(
            progress='deleted'
        ).prefetch_related(
            'diagnosis__testmanager_set',
        )

        if not sub_query:
            continue


        for depart in depart_list:
            sub_query_depart = sub_query.filter(depart_id = depart['id'])
            if sub_query_depart.count() == 0:
                continue

            ws['A' + str(current_row)] = writing_number
            ws['B' + str(current_row)] = depart['name']
            ws['C' + str(current_row)] = test_class_dict[str(test.test_class_id)]
            ws['D' + str(current_row)] = test.code
            ws['E' + str(current_row)] = test.name
            ws['F' + str(current_row)] = sub_query_depart.count()
            ws['G' + str(current_row)] = test.get_price()
            ws['H' + str(current_row)] = '=+F' + str(current_row) + '*G' + str(current_row)
            ws['G' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
            ws['H' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'



            writing_number +=1 
            current_row +=1




    rows = ws['A7:H' + str(current_row + 3)]
    for row in rows:
        for cell in row:
            cell.border = border_thin
            cell.font = font_base

    ws['E' + str(current_row + 3)] = 'Tổng'
    ws['E' + str(current_row + 3)].font = Font(bold = True)
    ws['F' + str(current_row + 3)] = '=+SUM(F7:F' + str(current_row) + ')'
    ws['F' + str(current_row + 3)].font = Font(bold = True)
    ws['H' + str(current_row + 3)] = '=+SUM(H7:H' + str(current_row) + ')'
    ws['H' + str(current_row + 3)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    ws['H' + str(current_row + 3)].font = Font(bold = True)
    ##########################################################################################################
    ws = wb.get_sheet_by_name('Theo ngay')# grab the active worksheet
    current_row = 7
    writing_number = 1
    ws['E2'] = 'Ngày : ' + start_date 
    ws['G2'] = 'đến ngày : ' + end_date

    tests = Test.objects.all().order_by('code')
    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    argument_list.append( Q(**{'id':0} ) )     
    
    sub_query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs ,
        recorded_date__range = (date_min, date_max), 
    ).exclude(
        progress='deleted'
        ).order_by('recorded_date')
 
    for data in sub_query:

        try: 
            test_records = data.diagnosis.testmanager_set.all()
        except:
            test_records = None
        if not test_records:
            continue
        for test_record in test_records:
            testInfo = Test.objects.get(id = test_record.test_id)
            ws['A' + str(current_row)] = writing_number
            ws['B' + str(current_row)] = data.id
            ws['C' + str(current_row)] = data.recorded_date.strftime('%Y-%m-%d')
            ws['D' + str(current_row)] = data.patient.get_name_kor_eng()
            ws['E' + str(current_row)] = data.depart.name
            ws['F' + str(current_row)] = data.doctor.name_eng + '/' + data.doctor.name_kor
            ws['G' + str(current_row)] = testInfo.test_class.name
            ws['H' + str(current_row)] = testInfo.code
            ws['I' + str(current_row)] = testInfo.name
            ws['J' + str(current_row)] = 1
            ws['K' + str(current_row)] = testInfo.get_price(data.recorded_date)
            ws['K' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
            ws['L' + str(current_row)] = testInfo.get_price(data.recorded_date)
            ws['L' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'

            writing_number +=1 
            current_row +=1      

    rows = ws['A7:L' + str(current_row + 3)]
    for row in rows:
        for cell in row:
            cell.border = border_thin
            cell.font = font_base              

    ws['K' + str(current_row + 3)] = 'Tổng'
    ws['K' + str(current_row + 3)].font = Font(bold = True)
    ws['L' + str(current_row + 3)] = '=+SUM(L7:L' + str(current_row) + ')'
    ws['L' + str(current_row + 3)].font = Font(bold = True)
    # ws['H' + str(current_row + 3)] = '=+SUM(H7:H' + str(current_row) + ')'
    ws['L' + str(current_row + 3)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    # ws['H' + str(current_row + 3)].font = Font(bold = True)

    wb.save(response)
    return response

@login_required
def procedure_statistics_excel(request):

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    depart = request.GET.get('depart')
    doctor = request.GET.get('doctor')
    category = request.GET.get('patient_type','')
    string = request.GET.get('patient_search','')    

    kwargs = {}
    if depart is not None and depart != '':
        kwargs['depart'] = depart # 기본 
    if doctor  is not None and doctor != '':
        kwargs['payment__reception__doctor_id'] = doctor             

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="procedure_statistics_' + start_date + '_' + end_date +'.xlsx"'
    
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Proc_statistics.xlsx') #Workbook()
    # wb = load_workbook('/Users/light/Desktop/Work/imdc/imedicare2/static/excel_form/Proc_statistics.xlsx')
    ws = wb.get_sheet_by_name('Theo so luong')# grab the active worksheet

    ws['E2'] = 'Ngày : ' + start_date 
    ws['F2'] = 'đến ngày : ' + end_date

    date_min = datetime.datetime.combine(datetime.datetime.strptime(start_date,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end_date,"%Y-%m-%d").date(), datetime.time.max)

    depart_list = Depart.objects.values('id','name')
    
    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    font_base = Font(
        name = 'Arial',
        bold = False,
        size = 11,
        )
    
    procedure_class_query = PrecedureClass.objects.all()
    procedure_class_dict = {}
    procedure_code_dict = {}
    for procedure_class in procedure_class_query:
        procedure_class_dict[str(procedure_class.id)] = procedure_class.name
        procedure_code_dict[str(procedure_class.id)] = procedure_class.code

    current_row = 7
    writing_number = 1
    
    procedures = Precedure.objects.all().order_by('code')
    for procedure in procedures:

        sub_query = Reception.objects.filter(
            **kwargs,
            recorded_date__range = (date_min, date_max), 
            diagnosis__preceduremanager__precedure = procedure.id,
        ).exclude(
            progress='deleted'
        ).prefetch_related(
            'diagnosis__preceduremanager_set',
        )
        
        if not sub_query:
            continue
        for depart in depart_list:
            sub_query_depart = sub_query.filter(depart_id = depart['id'])
            if sub_query_depart.count() == 0:
                continue
            
            ws['A' + str(current_row)] = writing_number
            ws['B' + str(current_row)] = depart['name']
            ws['C' + str(current_row)] = procedure_class_dict[str(procedure.precedure_class_id)]
            ws['D' + str(current_row)] = procedure.code
            ws['E' + str(current_row)] = procedure.name
            ws['F' + str(current_row)] = procedure.name_vie
            ws['G' + str(current_row)] = sub_query_depart.count()
            try:
                ws['H' + str(current_row)] = procedure.get_price()
            except Exception as e:
                print(procedure.code)
                print(e)
            ws['I' + str(current_row)] = '=+G' + str(current_row) + '*H' + str(current_row)
            ws['H' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
            ws['I' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'

            writing_number +=1 
            current_row +=1

    rows = ws['A7:I' + str(current_row + 3)]
    for row in rows:
        for cell in row:
            cell.border = border_thin
            cell.font = font_base

    ws['F' + str(current_row + 3)] = 'Tổng/ Total'
    ws['F' + str(current_row + 3)].font = Font(bold = True)
    ws['G' + str(current_row + 3)] = '=+SUM(G7:G' + str(current_row) + ')'
    ws['G' + str(current_row + 3)].font = Font(bold = True)
    ws['I' + str(current_row + 3)] = '=+SUM(I7:I' + str(current_row) + ')'
    ws['I' + str(current_row + 3)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    ws['I' + str(current_row + 3)].font = Font(bold = True)
    ##########################################################################################################
    ws = wb.get_sheet_by_name('Theo ngay')# grab the active worksheet
    current_row = 7
    writing_number = 1
    ws['E2'] = 'Ngày : ' + start_date 
    ws['G2'] = 'đến ngày : ' + end_date

    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    argument_list.append( Q(**{'id':0} ) )     
    
    sub_query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs ,
        recorded_date__range = (date_min, date_max), 
    ).exclude(
        progress='deleted'
        ).order_by('recorded_date')
 
    for data in sub_query:

        try: 
            preceduremanager_set = data.diagnosis.preceduremanager_set.all()
        except:
            preceduremanager_set = None
        if not preceduremanager_set:
            continue

        for precedure_record in preceduremanager_set:
            procedures = Precedure.objects.get(id = precedure_record.precedure_id)
            ws['A' + str(current_row)] = writing_number
            ws['B' + str(current_row)] = data.id
            ws['C' + str(current_row)] = data.recorded_date.strftime('%Y-%m-%d')
            ws['D' + str(current_row)] = data.patient.get_name_kor_eng()
            ws['E' + str(current_row)] = data.depart.name
            ws['F' + str(current_row)] = data.doctor.name_eng + '/' + data.doctor.name_kor
            ws['G' + str(current_row)] = procedures.precedure_class.name
            ws['H' + str(current_row)] = procedures.code
            ws['I' + str(current_row)] = procedures.name
            ws['J' + str(current_row)] = procedures.name_vie
            ws['K' + str(current_row)] = 1
            ws['L' + str(current_row)] = procedures.get_price(data.recorded_date)
            ws['L' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
            ws['M' + str(current_row)] = procedures.get_price(data.recorded_date)
            ws['M' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'

            writing_number +=1 
            current_row +=1      

    rows = ws['A7:M' + str(current_row + 3)]
    for row in rows:
        for cell in row:
            cell.border = border_thin
            cell.font = font_base              

    ws['K' + str(current_row + 3)] = 'Tổng'
    ws['K' + str(current_row + 3)].font = Font(bold = True)
    ws['M' + str(current_row + 3)] = '=+SUM(M7:M' + str(current_row) + ')'
    ws['M' + str(current_row + 3)].font = Font(bold = True)
    # ws['H' + str(current_row + 3)] = '=+SUM(H7:H' + str(current_row) + ')'
    ws['M' + str(current_row + 3)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    # ws['H' + str(current_row + 3)].font = Font(bold = True)

    wb.save(response)
    return response


@login_required
def medicine_statistics_excel(request):

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    is_vaccine = request.POST.get('is_vaccine','false')
    depart = request.GET.get('depart')
    doctor = request.GET.get('doctor')
    category = request.GET.get('patient_type','')
    string = request.GET.get('patient_search','')  

    vac_kwargs={}

    if is_vaccine == 'true':
        vac_kwargs['code__icontains'] = 'VC'   

    kwargs = {}
    if depart is not None and depart != '':
        kwargs['depart'] = depart # 기본 
    if doctor  is not None and doctor != '':
        kwargs['payment__reception__doctor_id'] = doctor             

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="medicine_statistics_' + start_date + '_' + end_date +'.xlsx"'
    
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/medicine_statistics.xlsx') #Workbook()

    ws = wb.get_sheet_by_name('Top 100 export')# grab the active worksheet

    ws['E2'] = 'Ngày : ' + start_date 
    ws['F2'] = 'đến ngày : ' + end_date

    date_min = datetime.datetime.combine(datetime.datetime.strptime(start_date,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end_date,"%Y-%m-%d").date(), datetime.time.max)

    depart_list = Depart.objects.values('id','name')
    
    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    font_base = Font(
        name = 'Arial',
        bold = False,
        size = 11,
        )
    
    # procedure_class_query = PrecedureClass.objects.all()
    # procedure_class_dict = {}
    # procedure_code_dict = {}
    # for procedure_class in procedure_class_query:
    #     procedure_class_dict[str(procedure_class.id)] = procedure_class.name
    #     procedure_code_dict[str(procedure_class.id)] = procedure_class.code

    # current_row = 7
    # writing_number = 1
    
    # procedures = Precedure.objects.all().order_by('code')
    # for procedure in procedures:

    #     sub_query = Reception.objects.filter(
    #         **kwargs,
    #         recorded_date__range = (date_min, date_max), 
    #         diagnosis__preceduremanager__precedure = procedure.id,
    #     ).exclude(
    #         progress='deleted'
    #     ).prefetch_related(
    #         'diagnosis__preceduremanager_set',
    #     )
        
    #     if not sub_query:
    #         continue
    #     for depart in depart_list:
    #         sub_query_depart = sub_query.filter(depart_id = depart['id'])
    #         if sub_query_depart.count() == 0:
    #             continue
            
    #         ws['A' + str(current_row)] = writing_number
    #         ws['B' + str(current_row)] = depart['name']
    #         ws['C' + str(current_row)] = procedure_class_dict[str(procedure.precedure_class_id)]
    #         ws['D' + str(current_row)] = procedure.code
    #         ws['E' + str(current_row)] = procedure.name
    #         ws['F' + str(current_row)] = procedure.name_vie
    #         ws['G' + str(current_row)] = sub_query_depart.count()
    #         ws['H' + str(current_row)] = procedure.get_price()
    #         ws['I' + str(current_row)] = '=+G' + str(current_row) + '*H' + str(current_row)
    #         ws['H' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    #         ws['I' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'

    #         writing_number +=1 
    #         current_row +=1

    # rows = ws['A7:I' + str(current_row + 3)]
    # for row in rows:
    #     for cell in row:
    #         cell.border = border_thin
    #         cell.font = font_base

    # ws['F' + str(current_row + 3)] = 'Tổng/ Total'
    # ws['F' + str(current_row + 3)].font = Font(bold = True)
    # ws['G' + str(current_row + 3)] = '=+SUM(G7:G' + str(current_row) + ')'
    # ws['G' + str(current_row + 3)].font = Font(bold = True)
    # ws['I' + str(current_row + 3)] = '=+SUM(I7:I' + str(current_row) + ')'
    # ws['I' + str(current_row + 3)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    # ws['I' + str(current_row + 3)].font = Font(bold = True)
    ##########################################################################################################
    ws = wb.get_sheet_by_name('BC xuất bán')# grab the active worksheet
    current_row = 7
    writing_number = 1
    ws['E2'] = 'Ngày : ' + start_date 
    ws['G2'] = 'đến ngày : ' + end_date

    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    argument_list.append( Q(**{'id':0} ) )     
    
    sub_query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs ,
        recorded_date__range = (date_min, date_max), 
    ).exclude(
        progress='deleted'
        ).order_by('recorded_date')
 
    for data in sub_query:

        try: 
            medicine_set = data.diagnosis.medicinemanager_set.all()
        except:
            medicine_set = None
        if not medicine_set:
            continue

        for medicine_log in medicine_set:

            try: 
                medicine = Medicine.objects.get( **vac_kwargs, id = medicine_log.medicine_id)
            except Medicine.DoesNotExist:
                medicine = None
            if not medicine:
                continue                 
            # medicine = Medicine.objects.get(id = medicine_log.medicine_id)
            ws['A' + str(current_row)] = writing_number
            ws['B' + str(current_row)] = data.id
            ws['C' + str(current_row)] = data.recorded_date.strftime('%Y-%m-%d')
            ws['D' + str(current_row)] = data.patient.get_name_kor_eng()
            ws['E' + str(current_row)] = data.depart.name
            ws['F' + str(current_row)] = data.doctor.name_eng + '/' + data.doctor.name_kor
            ws['G' + str(current_row)] = medicine.medicine_class.name
            ws['H' + str(current_row)] = medicine.code
            ws['I' + str(current_row)] = medicine.name 
            ws['J' + str(current_row)] = medicine.unit
            ws['K' + str(current_row)] = medicine_log.amount * medicine_log.days
            ws['L' + str(current_row)] = medicine.get_price(data.recorded_date)
            ws['L' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
            ws['M' + str(current_row)] = medicine.get_price(data.recorded_date) * (medicine_log.amount * medicine_log.days)
            ws['M' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'

            ws['N' + str(current_row)] = data.patient.phone
            ws['O' + str(current_row)] = data.diagnosis.diagnosis
            writing_number +=1 
            current_row +=1      

    rows = ws['A7:M' + str(current_row + 3)]
    for row in rows:
        for cell in row:
            cell.border = border_thin
            cell.font = font_base              

    ws['K' + str(current_row + 3)] = 'Tổng'
    ws['K' + str(current_row + 3)].font = Font(bold = True)
    ws['M' + str(current_row + 3)] = '=+SUM(M7:M' + str(current_row) + ')'
    ws['M' + str(current_row + 3)].font = Font(bold = True)
    # ws['H' + str(current_row + 3)] = '=+SUM(H7:H' + str(current_row) + ')'
    ws['M' + str(current_row + 3)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    # ws['H' + str(current_row + 3)].font = Font(bold = True)
    ##########################################################################################################
    ws = wb.get_sheet_by_name('BC nhập vào')# grab the active worksheet
    current_row = 7
    writing_number = 1
    ws['E2'] = 'Ngày : ' + start_date 
    ws['G2'] = 'đến ngày : ' + end_date

    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    argument_list.append( Q(**{'id':0} ) )     
    
    sub_query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs ,
        recorded_date__range = (date_min, date_max), 
    ).exclude(
        progress='deleted'
        ).order_by('recorded_date')

    medicine_class_query = MedicineClass.objects.all()

    current_row = 7
    writing_number = 1      
 
    for medimedicine_class in medicine_class_query:
        try: 
            medicines = Medicine.objects.filter(medicine_class = medimedicine_class).order_by('code')  
        except Medicine.DoesNotExist:
            medicine = None
        if not medicine:
            continue    
        
        for medicine in medicines:
            medicineLogs = MedicineLog.objects.filter(date__range = (date_min, date_max), medicine = medicine, type = 'add', changes__gt = 0).order_by('date')        
            if medicineLogs.count() == 0:
                continue        
            count = 1
            ws['A' + str(current_row)] = writing_number
            ws['B' + str(current_row)] = medimedicine_class.name
            ws['C' + str(current_row)] = medicine.code
            ws['D' + str(current_row)] = medicine.name + '/' + medicine.name_vie
            ws['E' + str(current_row)] = medicine.company    
            temp_row = current_row

            for medicineLog in medicineLogs:
                # medicine = Medicine.objects.get(id = medicine.medicine_id)
                ws['F' + str(temp_row)] = medicineLog.date.strftime('%Y-%m-%d')
                ws['G' + str(temp_row)] = medicine.unit
                ws['H' + str(temp_row)] = medicineLog.changes
                ws['I' + str(temp_row)] = medicine.get_price(medicineLog.date)
                ws['I' + str(temp_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
                ws['J' + str(temp_row)] = medicine.get_price(medicineLog.date) * (medicineLog.changes)
                ws['J' + str(temp_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
                count += 1
                temp_row += 1
            if count != 0:
                ws.merge_cells('A' + str(current_row) + ':A' + str(current_row + count-1))
                ws.merge_cells('B' + str(current_row) + ':B' + str(current_row + count-1))
                ws.merge_cells('C' + str(current_row) + ':C' + str(current_row + count-1))
                ws.merge_cells('D' + str(current_row) + ':D' + str(current_row + count-1))
                ws.merge_cells('E' + str(current_row) + ':E' + str(current_row + count-1))

            writing_number +=1 
            current_row += count    

    rows = ws['A7:J' + str(current_row + 3)]
    for row in rows:
        for cell in row:
            cell.border = border_thin
            cell.font = font_base              

    ws['K' + str(current_row + 3)] = 'Tổng'
    ws['K' + str(current_row + 3)].font = Font(bold = True)
    ws['M' + str(current_row + 3)] = '=+SUM(M7:M' + str(current_row) + ')'
    ws['M' + str(current_row + 3)].font = Font(bold = True)
    # ws['H' + str(current_row + 3)] = '=+SUM(H7:H' + str(current_row) + ')'
    ws['M' + str(current_row + 3)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    # ws['H' + str(current_row + 3)].font = Font(bold = True)

    ##########################################################################################################
    ws = wb.get_sheet_by_name('lợi nhuận - hao phí')# grab the active worksheet
    current_row = 7
    writing_number = 1
    ws['E2'] = 'Ngày : ' + start_date 
    ws['G2'] = 'đến ngày : ' + end_date

    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    argument_list.append( Q(**{'id':0} ) )     

    current_row = 7
    writing_number = 1      
 
    medicines = Medicine.objects.filter(**vac_kwargs).order_by('code')  
    for medicine in medicines:
        medicineLogs_add = MedicineLog.objects.filter(date__range = (date_min, date_max), medicine = medicine, type = 'add', changes__gt = 0).aggregate(Sum('changes')) 
        if medicineLogs_add['changes__sum'] is None:
            medicineLogs_add = 0
        else:
            medicineLogs_add = medicineLogs_add['changes__sum']   

        medicineLogs_dec = MedicineLog.objects.filter(date__range = (date_min, date_max), medicine = medicine, type = 'dec').aggregate(Sum('changes'))   
        if medicineLogs_dec['changes__sum'] is None:
            medicineLogs_dec = 0
        else:
            medicineLogs_dec = medicineLogs_dec['changes__sum']   

        if medicineLogs_add == 0 and medicineLogs_dec == 0:
            continue        

        ws['A' + str(current_row)] = writing_number
        ws['B' + str(current_row)] = medicine.medicine_class.name
        ws['C' + str(current_row)] = medicine.code
        ws['D' + str(current_row)] = medicine.name + '/' + medicine.name_vie
        ws['E' + str(current_row)] = medicine.company    
        ws['F' + str(current_row)] = medicine.unit + '/' + medicine.unit_vie
        ws['G' + str(current_row)] = medicineLogs_add
        ws['H' + str(current_row)] = medicineLogs_dec
        ws['I' + str(current_row)] = medicineLogs_add - medicineLogs_dec
        ws['J' + str(current_row)] = medicine.price
        ws['J' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
        ws['K' + str(current_row)] = medicine.price_input
        ws['K' + str(current_row)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
        current_row += 1
        writing_number +=1  

    # ws.delete_rows(current_row + 3, 5434 - current_row)
    rows = ws['A7:J' + str(current_row + 3)]
    
    for row in rows:
        for cell in row:
            cell.border = border_thin
            cell.font = font_base              

    # ws['K' + str(current_row + 3)] = 'Tổng'
    # ws['K' + str(current_row + 3)].font = Font(bold = True)
    # ws['M' + str(current_row + 3)] = '=+SUM(M7:M' + str(current_row) + ')'
    # ws['M' + str(current_row + 3)].font = Font(bold = True)
    # # ws['H' + str(current_row + 3)] = '=+SUM(H7:H' + str(current_row) + ')'
    # ws['M' + str(current_row + 3)].number_format = '_-* #,##0 _₫_-;-* #,##0 _₫_-;_-* "-"?? _₫_-;_-@_-'
    # # ws['H' + str(current_row + 3)].font = Font(bold = True)

    wb.save(response)
    return response    


@login_required
def KBL_project_excel(request):

    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    project_type=request.POST.get("project_type",'')
    project_status=request.POST.get("project_status",'')
    project_in_charge=request.POST.get("project_in_charge",'')

    string=request.POST.get("string",'')


    kwargs = {}
    argument_list = [] 
    #if string != '':
    argument_list.append( Q(**{'customer_name__icontains':string} ) )
    argument_list.append( Q(**{'project_name__icontains':string} ) )


    if project_type != '':
        kwargs['type'] = project_type
    if project_status !='':
        kwargs['progress'] = project_status
    #if project_in_charge !='':
    #    kwargs['in_charge'] = project_status

    project_query = Project_Manage.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs,
        start_date__range = (start_date + ' 00:00:00', end_date + ' 23:59:59'),
        use_yn = 'Y',
        ).order_by('-id')
    


    #엑셀
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Project_excel_' + start_date + '_' + end_date +'.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Project_excel.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet

    ws['D1'] = 'Date : ' + start_date + ' ~ ' + end_date

    current_row = 4
    tmp_index = 1
    for data in project_query:
        writing_number = 0

        ws['A' +  str(current_row)] = tmp_index
        ws['B' +  str(current_row)] = data.customer_name
        ws['C' +  str(current_row)] = data.project_name
        ws['D' +  str(current_row)] = data.level


        ws['H' +  str(current_row)] = data.priority

        ws['J' +  str(current_row)] = data.in_charge1

        #서브 쿼리
        sub_query_cnt = 0

        if Project_Manage_Detail.objects.filter(
            use_yn = 'Y',
            project = data.id,
            ).count() != 0:

            sub_query  = Project_Manage_Detail.objects.filter(
                use_yn = 'Y',
                project = data.id,)

            for sub_data in sub_query:
                sub_string = sub_data.date + '\n - ' + sub_data.project_details

                ws['E' +  str(current_row + sub_query_cnt)] = sub_data.type

                ws['G' +  str(current_row + sub_query_cnt)] = sub_string

                ws['I' +  str(current_row + sub_query_cnt)] = sub_data.note
                
                sub_query_cnt += 1


        if Work_Permit_Manage.objects.filter(
            use_yn = 'Y',
            project = data.id,
            ).count() != 0:

            sub_query  = Work_Permit_Manage.objects.filter(
                use_yn = 'Y',
                project = data.id,)

            for sub_data in sub_query:
                sub_string = 'Employment Apporval Application Date : ' + sub_data.EA_application_date
                sub_string += '\nEmployment Apporval Expected Date : ' + sub_data.EA_exp_date
                sub_string += '\nWork Permit Application Date : ' + sub_data.WP_application_date
                sub_string += '\nWork Permit Expected Date : ' + sub_data.WP_exp_date

                ws['E' +  str(current_row + sub_query_cnt)] = sub_data.employee_name
                ws['F' +  str(current_row + sub_query_cnt)] = sub_data.expected_date + '\n' + sub_data.status

                ws['G' +  str(current_row + sub_query_cnt)] = sub_string

                ws['I' +  str(current_row + sub_query_cnt)] = sub_data.note

                sub_query_cnt += 1

        if Visa_Manage.objects.filter(
            use_yn = 'Y',
            project = data.id,
            ).count() != 0:

            sub_query  = Visa_Manage.objects.filter(
                use_yn = 'Y',
                project = data.id,
                )

            for sub_data in sub_query:
                sub_string = 'Date of Entry : ' + sub_data.date_entry
                sub_string += '\nDate Ordered : ' + sub_data.date_ordered
                sub_string += '\nReceipt Application : ' + sub_data.date_receipt_application
                sub_string += '\nSubmit Document : ' + sub_data.date_subbmit_doc
                sub_string += '\nReceipt Document : ' + sub_data.date_receipt_doc
                sub_string += '\nExpected Date: ' + sub_data.date_expected

                ws['E' +  str(current_row + sub_query_cnt)] = '[' + sub_data.type + ']' + sub_data.employee_name
                ws['F' +  str(current_row + sub_query_cnt)] = sub_data.date_expected + '\n' + sub_data.status

                ws['G' +  str(current_row + sub_query_cnt)] = sub_string

                #ws['I' +  str(current_row + sub_query_cnt)] = sub_data.note

                sub_query_cnt += 1

        if sub_query_cnt == 0:
            tmp_merge_area = 0
        else:
            tmp_merge_area = sub_query_cnt - 1

        ws.merge_cells('A' + str(current_row) + ':A' + str(current_row + tmp_merge_area))
        ws.merge_cells('B' + str(current_row) + ':B' + str(current_row + tmp_merge_area))
        ws.merge_cells('C' + str(current_row) + ':C' + str(current_row + tmp_merge_area))
        ws.merge_cells('D' + str(current_row) + ':D' + str(current_row + tmp_merge_area))
        ws.merge_cells('H' + str(current_row) + ':H' + str(current_row + tmp_merge_area))
        ws.merge_cells('J' + str(current_row) + ':J' + str(current_row + tmp_merge_area))








        if sub_query_cnt == 0:
            current_row += 1
        else:
            current_row += sub_query_cnt
        tmp_index += 1


    #테두리
    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )
    font_base = Font(
        name = 'Times New Roman',
        bold = False,
        size = 8,
        )
    alignment = Alignment(
        horizontal = 'center',
        vertical = 'center',
        wrap_text=True
        )

    rows = ws['A4:J' + str(current_row + sub_query_cnt)]
    for row in rows:
        for cell in row:
            cell.border = border_thin
            cell.font = font_base
            cell.alignment = alignment


    wb.save(response)
    return response




@login_required
def PM_SPECIAL1_excel(request):
    year = request.GET.get('year')
    month = request.GET.get('month')


    #엑셀
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="PM STATISTICS_' + year + '_' + month +'.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/PM_statistics_form.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet


    reception_query = Reception.objects.filter(
        depart = 7,
        recorded_date__year = int(year),
        recorded_date__month = int(month),
        progress = 'done',
        ).select_related('payment')

    ws['C1'] = year + ' 년 ' + month + ' 월'

    #신환 재진
    new_count = 0
    repeat_count = 0
    for data in reception_query:
        is_new = Reception.objects.filter(
            depart = 7,
            recorded_date__lt = data.recorded_date,
            patient_id = data.patient_id,
            progress = 'done',
            ).count()

        if is_new == 0:
            new_count += 1
        else:
            repeat_count += 1

    ws['G3'] = new_count
    ws['H3'] = repeat_count
    ws['I3'] = new_count + repeat_count

    #일별 수입
    list_by_day = [] 
    day_of_month = calendar.monthrange(int(year),int(month))[1]

    
    current_row = 3
    for day in range(1, day_of_month + 1):
        str_day = "{:04d}-{:02d}-{:02d}".format(int(year),int(month),day)

        sub_query = reception_query.filter(
            recorded_date__day = day,
            ).aggregate(
                day_amount = Sum('payment__sub_total'),
                day_count = Count('depart_id'),
                )

        ws['A' + str(current_row)] = str_day
        ws['B' + str(current_row)] = '0' if sub_query['day_count'] is None else sub_query['day_count']
        ws['C' + str(current_row)] = '0' if sub_query['day_amount'] is None else sub_query['day_amount']

        current_row += 1


    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A3:C' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin


    wb.save(response)
    return response



@login_required
def exam_inventory_excel(request):
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ExamFee_list.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/ExamFee_list.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet


    examfee_query = ExamFee.objects.filter(
        use_yn = 'Y',
        ).order_by('code')


    current_row = 2
    for data in examfee_query:
        
        doctor = None
        if data.doctor_id != None:
            doctor = Doctor.objects.get(id = data.doctor_id)

        ws['A' + str(current_row)] = data.id
        ws['B' + str(current_row)] = data.code
        ws['C' + str(current_row)] = data.name
        ws['D' + str(current_row)] = '' if doctor is None else doctor.name_eng
        ws['E' + str(current_row)] = data.get_price()

        current_row += 1


    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A2:E' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin


    wb.save(response)
    return response


    
@login_required
def test_inventory_excel(request):
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Test_list.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Test_list.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet


    test_query = Test.objects.filter(
        use_yn = 'Y',
        ).order_by('code')


    current_row = 2
    for data in test_query:
        
        test_class = TestClass.objects.get(id = data.test_class_id)
        interval_query = TestReferenceInterval.objects.filter(
            test_id = data.id,
            use_yn = 'Y'
        )

        str_interval = ''
        for interval in interval_query:
            if interval.name != '':
                str_interval += interval.name + " : "

            if interval.minimum != '':
                str_interval += str(interval.minimum) + interval.unit

            if interval.minimum != '' or interval.maximum != '':
                str_interval += ' - '

            if interval.maximum != '':
                str_interval += str(interval.maximum) + interval.unit
                
            str_interval += '\n'
        

        ws['A' + str(current_row)] = data.id
        ws['B' + str(current_row)] = data.code
        ws['C' + str(current_row)] = test_class.name
        ws['D' + str(current_row)] = data.name
        ws['E' + str(current_row)] = data.name_vie
        ws['F' + str(current_row)] = data.get_price()
        ws['G' + str(current_row)] = str_interval

        current_row += 1


    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A2:G' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin


    wb.save(response)
    return response


@login_required
def procedure_inventory_excel(request):
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Procedure_list.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Procedure_list.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet


    precedure_query = Precedure.objects.filter(
        use_yn = 'Y',
        ).order_by('code')


    current_row = 2
    for data in precedure_query:
        
        precedure_class = PrecedureClass.objects.get(id = data.precedure_class_id)

        str_type = 'General'
        if data.type == 'PKG':
            str_type = 'Package'

        ws['A' + str(current_row)] = data.id
        ws['B' + str(current_row)] = data.code
        ws['C' + str(current_row)] = precedure_class.name
        ws['D' + str(current_row)] = data.name
        ws['E' + str(current_row)] = data.name_vie
        ws['F' + str(current_row)] = data.get_price()
        ws['G' + str(current_row)] = str_type

        current_row += 1


    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A2:G' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin


    wb.save(response)
    return response

@login_required
def medicine_inventory_excel(request):
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Medicine_list.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Medicine_list_2.xlsx') #Workbook()
    # wb = load_workbook('/Users/light/Desktop/Work/imdc/imedicare2/static/excel_form/Medicine_list_2.xlsx')
    ws = wb.active# grab the active worksheet


    precedure_query = Medicine.objects.filter(
        type = 'PHARM',
        use_yn = 'Y',
        ).order_by('code')


    current_row = 2
    for data in precedure_query:
        
        medicine_class = MedicineClass.objects.get(id = data.medicine_class_id)

        str_type = 'Medicine'
        if data.vaccine_code != '':
            str_type = 'Vaccine'

        ws['A' + str(current_row)] = data.id
        ws['B' + str(current_row)] = data.code
        ws['C' + str(current_row)] = medicine_class.name
        ws['D' + str(current_row)] = data.name
        ws['E' + str(current_row)] = data.name_vie
        ws['F' + str(current_row)] = data.unit
        ws['G' + str(current_row)] = data.unit_vie
        ws['H' + str(current_row)] = data.company
        ws['I' + str(current_row)] = data.country
        ws['J' + str(current_row)] = data.country_vie
        ws['K' + str(current_row)] = data.ingredient
        ws['L' + str(current_row)] = data.ingredient_vie

        ws['M' + str(current_row)] = data.get_price()
        ws['N' + str(current_row)] = str_type
        ws['O' + str(current_row)] = data.inventory_count
        current_row += 1


    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A2:N' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin


    wb.save(response)
    return response

@login_required
def expendables_inventory_excel(request):
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Expendable_list.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Expendable_list.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet


    precedure_query = Medicine.objects.filter(
        type = 'TOOL',
        use_yn = 'Y',
        ).order_by('code')


    current_row = 2
    for data in precedure_query:
        
        #str_type = 'Medicine'
        #if data.vaccine_code != '':
        #    str_type = 'Vaccine'

        ws['A' + str(current_row)] = data.id
        ws['B' + str(current_row)] = data.code
        ws['C' + str(current_row)] = data.name
        ws['D' + str(current_row)] = data.name_vie
        ws['E' + str(current_row)] = data.unit
        ws['F' + str(current_row)] = data.unit_vie
        ws['G' + str(current_row)] = data.company
        ws['H' + str(current_row)] = data.country
        ws['I' + str(current_row)] = data.country_vie

        ws['J' + str(current_row)] = data.get_price()

        current_row += 1


    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    rows = ws['A2:J' + str(current_row)]
    for row in rows:
        for cell in row:
            cell.border = border_thin


    wb.save(response)
    return response

def search_medicine(request):
    filter = request.POST.get('filter')
    string = request.POST.get('string')

    date_start = request.POST.get('start_end_date').split(' - ')[0]
    date_end = request.POST.get('start_end_date').split(' - ')[1]

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    
    #if filter != '':
    #    kwargs.update({'depart_id':depart})


    datas_dict= {}
    receptions = Reception.objects.filter(recorded_date__range = (date_min, date_max), progress = 'done').order_by("-id")
    for reception in receptions:
        if hasattr(reception,'diagnosis'):
            medicine_manager_set = MedicineManager.objects.filter(diagnosis_id = reception.diagnosis.id)
            for medicine_manager in medicine_manager_set:
                medicine = Medicine.objects.get(pk = medicine_manager.medicine_id)
            
                if medicine.code in datas_dict.keys():
                    datas_dict[medicine.code]['sales'] += medicine_manager.amount * medicine_manager.days
                    datas_dict[medicine.code]['total_salse'] += medicine_manager.amount * medicine_manager.days * medicine.get_price(reception.recorded_date)

                else:
                    datas_dict.update({
                        medicine.code:{
                            'code':medicine.code,
                            'name':medicine.name,
                            'ingredient':'' if medicine.ingredient is None else medicine.ingredient,
                            'company':'' if medicine.company is None else medicine.company,
                            'count':medicine.inventory_count,
                            'price':medicine.get_price(reception.recorded_date),
                            'sales' : medicine_manager.amount * medicine_manager.days,
                            'total_salse' : medicine_manager.amount * medicine_manager.days * medicine.get_price(reception.recorded_date)
                            }
                        })

        
    id = 1
    total_amount = 0
    datas=[]
    datas_dict_sorted = sorted(datas_dict.items())
    for data_dict in datas_dict_sorted:
        data_dict[1].update({
            'id':id,
            })
        datas.append(data_dict)
        total_amount += data_dict[1]['total_salse']
        id += 1

    page = request.POST.get('page',1)
    page_context = request.POST.get('context_in_page',10)

    paginator = Paginator(datas, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)
    

    context = {
            'datas':list(paging_data),
            'total_amount':total_amount,

            'page_range_start':paging_data.paginator.page_range.start,
            'page_range_stop':paging_data.paginator.page_range.stop,
            'page_number':paging_data.number,
            'has_previous':paging_data.has_previous(),
            'has_next':paging_data.has_next(),
        }

    return JsonResponse(context)





@login_required
def temp_doctor_audit(request):
    doctor_search_form = DoctorsSearchForm()

    
    list_exam_fee = []
    list_precedures = []
    list_radiologys = []


    exam_fees = ExamFee.objects.filter(Q(code = 'E0010') | Q(code = 'E0011'))
    for exam_fee in exam_fees:
        list_exam_fee.append({'code':exam_fee.code,'value':exam_fee.name})

    precedures = Precedure.objects.filter(code__contains='PM')
    for precedure in precedures:
        list_precedures.append({'code':precedure.code,'value':precedure.name})

    radiologys = Precedure.objects.filter(code__contains='R', precedure_class_id = 10 )
    for radiology in radiologys:
        list_radiologys.append({'code':radiology.code,'value':radiology.name})

    return render(request,
    'Doctor/audit_PM.html',
        {
            'doctor_search':doctor_search_form,

            'list_exam_fee':list_exam_fee,
            'list_precedures':list_precedures,
            'list_radiologys':list_radiologys,

        }
    )


@login_required
def search_patient(request):


    return JsonResponse(context)


@login_required
def inventory_examfee(request):

    doctor_list = Doctor.objects.filter(
        user__is_active = True
    )


    return render(request,
    'Manage/inventory_examfee.html',
            {
                'doctor_list':doctor_list,
            },
        )

@login_required
def examfee_search(request):
    string = request.POST.get('string')
    filter = request.POST.get('filter')
    #class_id = request.POST.get('class_id')

    kwargs = {}


    argument_list = [] 
    if string !='':
        argument_list.append( Q(**{'name__icontains':string} ) )
   
    if len(argument_list) == 0:
         argument_list.append( Q(**{'name__icontains':''} ) )



    query_datas = ExamFee.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs,
        use_yn = 'Y',
        ).order_by("id")


    datas=[]
    for query_data in query_datas:
        doctor = None
        if query_data.doctor_id != None:
            doctor = Doctor.objects.get(id = query_data.doctor_id)

        data = {
                'id' : query_data.id,
                'code': query_data.code,
                'doctor': '' if doctor is None else doctor.name_eng,
                'name' : query_data.name,
                'price' : query_data.get_price(),
                
            } 
        datas.append(data)


    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page');
    paginator = Paginator(datas, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    context = {
        #'datas':datas,
        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        
        }
    return JsonResponse(context)


@login_required
def exam_add_edit_get(request):

    id = request.POST.get('id');

    exam = ExamFee.objects.get(id=id)

    return JsonResponse({
        'id':exam.id,
        'name':exam.name,
        'price':exam.get_price(),
        'doctor_id':exam.doctor_id,
        'result':True,
        })


@login_required
def exam_add_edit_set(request):

    id = request.POST.get('id');
    name = request.POST.get('name');
    price = request.POST.get('price');
    doctor = request.POST.get('doctor');

    if int(id) == 0 :
        data = ExamFee()
        data.price = price
        

        last_code = ExamFee.objects.filter(
            code__istartswith='E',
        ).order_by('code').last()
        CODE = 'E'
        
        if last_code == None:
           data.code = CODE + str('0001')
        else:
            temp_code = last_code.code.split(CODE)
            data.code = CODE + str('%04d' % (int(temp_code[1]) + 1))
    else:
        data = ExamFee.objects.get(id=id)
        now = datetime.datetime.now()
        now = now - datetime.timedelta(seconds = 1) 
        str_now = now.strftime('%Y%m%d%H%M%S')
        try: 
            old_price = Pricechange.objects.get(type="ExamFee",country='VI',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price.price != int(price):
                old_price.date_end = str_now
                old_price.save()

                new_price = Pricechange(type="ExamFee",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price):
                new_price = Pricechange(type="ExamFee",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()
    print(price)
    data.price = price

    data.doctor_id = doctor
    data.name = name
    data.save()


    return JsonResponse({
        'result':True,
        })



@login_required
def inventory_test(request):
    class_datas=[]
    test_class = TestClass.objects.all()
    for tmp_class in test_class:
        class_datas.append({
            'id':tmp_class.id,
            'name': tmp_class.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
            })


    return render(request,
    'Manage/inventory_test.html',
            {
                'test_class':class_datas,
            },
        )


@login_required
def test_search(request):
    string = request.POST.get('string')
    filter = request.POST.get('filter')
    class_id = request.POST.get('class_id')

    kwargs = {}
    if class_id != '':
        kwargs.update({
            'test_class_id':class_id,
            })

    argument_list = [] 
    if string !='':
        argument_list.append( Q(**{'name__icontains':string} ) )
        argument_list.append( Q(**{'name_vie__icontains':string} ) )
   


    if string == '' :
        query_datas = Test.objects.filter(**kwargs).select_related('test_class').select_related('parent_test').exclude(use_yn = 'N').order_by("id")
    #elif filter == 'name':
    #    query_datas = Test.objects.filter( Q(name__icontains = string) | Q(name_vie__icontains = string)).select_related('test_class').exclude(use_yn = 'N').order_by("name")
    else:
        query_datas = Test.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).select_related('test_class').select_related('parent_test').exclude(use_yn = 'N').order_by("id")


    datas=[]
    for query_data in query_datas:
        print(query_data.__dict__)
        data = {
                'id' : query_data.id,
                'code': query_data.code,
                'name' : query_data.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'class':query_data.test_class.name,
                'price' : query_data.get_price(),
                'price_dollar' : query_data.get_price_dollar(),
                'parent_test': query_data.parent_test.code if query_data.parent_test else ''
            }
        
        datas.append(data)


    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page');
    paginator = Paginator(datas, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    context = {
        #'datas':datas,
        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        
        }
    return JsonResponse(context)


@login_required
def test_add_edit_get(request):
    id = request.POST.get('id');

    test = Test.objects.get(id=id)


    return JsonResponse({
        'id':test.id,
        'name':test.name,
        'name_vie':test.name_vie,
        'price':test.get_price(),
        'price_dollar':test.get_price_dollar(),
        'precedure_class_id':test.test_class_id,
        'result':True,
        'parent_test': test.parent_test.code if test.parent_test else ''
        })


@login_required
def test_add_edit_set(request):
    id = request.POST.get('id');
    type = request.POST.get('type');
    test_class = request.POST.get('test_class');
    name = request.POST.get('name');
    name_vie = request.POST.get('name_vie');
    price = request.POST.get('price');
    price_dollar = request.POST.get('price_dollar');

    parent_test_code = request.POST.get('parent_test');
    parent_test_obj = Test.objects.filter(code=parent_test_code).first()

    if int(id) == 0 :
        data = Test()
        data.price = price
        

        last_code = Test.objects.last()
        test_class = int(test_class)
        CODE = 'L'
            
        if last_code == None:
           data.code = CODE + str('0001')
        else:
            temp_code = last_code.code.split(CODE)
            data.code = CODE + str('%04d' % (int(temp_code[1]) + 1))
    else:
        data = Test.objects.get(id=id)
        now = datetime.datetime.now()
        now = now - datetime.timedelta(seconds = 1) 
        str_now = now.strftime('%Y%m%d%H%M%S')
        try: 
            old_price = Pricechange.objects.get(type="Test",country='VI',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price.price != int(price):
                old_price.date_end = str_now
                old_price.save()

                new_price = Pricechange(type="Test",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price):
                new_price = Pricechange(type="Test",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()
                
        try:
            old_price_dollar = Pricechange.objects.get(type="Test",country='US',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price_dollar.price != int(price_dollar):
                old_price_dollar.date_end = str_now
                old_price_dollar.save()

                new_price = Pricechange(type="Test",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price_dollar):
                new_price = Pricechange(type="Test",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

    data.test_class_id = test_class
    data.name = name
    data.name_vie = name_vie
    data.parent_test = parent_test_obj
    data.save()


    return JsonResponse({
        'result':True,
        })



    return JsonResponse({})

@login_required
def test_add_edit_delete(request):

    id = request.POST.get('id')
    test = Test.objects.get(id=id)
    test.use_yn = 'N'
    test.save()

    return JsonResponse({
        'result':True,
        })

@login_required
def test_get_interval_list(request):
    test_id = request.POST.get('test_id')

    query = TestReferenceInterval.objects.filter(test_id = test_id,use_yn='Y')
    datas=[]
    for data in query:
        datas.append({
            'id':data.id,
            'minimum':data.minimum,
            'maximum':data.maximum,
            'unit':data.unit,
            'unit_vie':data.unit_vie,
            'name':data.name,
            'name_vie':data.name_vie,
            'sign':data.sign,
            })

    return JsonResponse({
        'result':True,
        'datas':datas,
        })

@login_required
def test_get_interval(request):
    id = request.POST.get('id')
    data = TestReferenceInterval.objects.get(id = id)

    return JsonResponse({
        'result':True,
        'id':data.id,
        'minimum':data.minimum,
        'maximum':data.maximum,
        'unit':data.unit,
        'unit_vie':data.unit_vie,
        'name':data.name,
        'name_vie':data.name_vie,
        'sign':data.sign,
        })

@login_required
def test_save_interval(request):
    selected_test = request.POST.get('selected_test')
    id = request.POST.get('id','')
    remark = request.POST.get('remark')
    remark_vi = request.POST.get('remark_vi')
    unit = request.POST.get('unit')
    unit_vi = request.POST.get('unit_vi')
    minimum = request.POST.get('minimum')
    maximum = request.POST.get('maximum')
    sign = request.POST.get('sign')

    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if id != '':
        test_range = TestReferenceInterval.objects.get(id = id)

    else:
        test_range = TestReferenceInterval()
        test_range.date_register = now
        test_range.registrant = request.user.id


    test_range.test_id = selected_test
    test_range.name = remark
    test_range.name_vi = remark_vi
    test_range.unit = unit
    test_range.unit_vi = unit_vi
    test_range.minimum = minimum
    test_range.maximum = maximum
    test_range.sign = sign


    test_range.date_modify = now
    test_range.modifier = request.user.id

    test_range.save()

    return JsonResponse({
        'result':True,
        })

@login_required
def test_delete_interval(request):
    id = request.POST.get('id')
    test_range = TestReferenceInterval.objects.get(id = id)
    test_range.use_yn = 'N'

    test_range.date_modify = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    test_range.modifier = request.user.id

    test_range.save()

    return JsonResponse({
        'result':True,
        })


@login_required
def list_database_test_class_get(request):

    datas = []
    query = TestClass.objects.all()
    for data in query:
        datas.append({
            'id':data.id,
            'name':data.name,
            'name_vie':data.name_vie,
            })


    return JsonResponse({
        'result':True,
        'datas':datas,
        })

@login_required
def add_edit_test_class_menu_get(request):
    id = request.POST.get('id')

    query = TestClass.objects.get(pk = id)


    return JsonResponse({
        'result':True,
        'name':query.name,
        'name_vie':query.name_vie,
        })


@login_required
def add_edit_test_class_menu_save(request):
    id = request.POST.get('id','')
    name = request.POST.get('class_name')
    name_vie = request.POST.get('class_name_vie')

    if id == '':
        query = TestClass()
    else:
        query = TestClass.objects.get(pk = id)
        

    query.name = name
    query.name_vie = name_vie
    query.save()

    return JsonResponse({
        'result':True,
        })

@login_required
def inventory_precedure(request):

    class_datas=[]
    precedure_class = PrecedureClass.objects.all()
    for tmp_class in precedure_class: 
        class_datas.append({
            'id':tmp_class.id,
            'name': tmp_class.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
            })


    return render(request,
    'Manage/inventory_precedure.html',
            {
                'precedure_class':class_datas,
            },
        )


def precedure_search(request):
    string = request.POST.get('string')
    filter = request.POST.get('filter')
    class_id = request.POST.get('class_id')

    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page',30);


    kwargs = {}
    if class_id != '':
        kwargs.update({
            'precedure_class_id':class_id,
            })

    argument_list = [] 
    if string !='':
        argument_list.append( Q(**{'name__icontains':string} ) )
        argument_list.append( Q(**{'name_vie__icontains':string} ) )
        argument_list.append( Q(**{'code__icontains':string} ) )   


    if string == '' :
        query_datas = Precedure.objects.filter(**kwargs).select_related('precedure_class').exclude(use_yn = 'N').order_by("id")
    #elif filter == 'name':
    #    query_datas = Test.objects.filter( Q(name__icontains = string) | Q(name_vie__icontains = string)).select_related('test_class').exclude(use_yn = 'N').order_by("name")
    else:
        query_datas = Precedure.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).select_related('precedure_class').exclude(use_yn = 'N').order_by("id")




    datas=[]
    for query_data in query_datas:
        data = {
                'id' : query_data.id,
                'code': query_data.code,
                'name' : query_data.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'class':query_data.precedure_class.name,
                'price' : query_data.get_price(),
            }
        datas.append(data)


    paginator = Paginator(datas, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    context = {
        #'datas':datas,
        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        
        }
    return JsonResponse(context)


@login_required
def precedure_add_edit_get(request):
    id = request.POST.get('id');

    precedure = Precedure.objects.get(id=id)


    return JsonResponse({
        'id':precedure.id,
        'name':precedure.name,
        'name_vie':precedure.name_vie,
        'price':precedure.get_price(),
        'price_dollar':precedure.get_price_dollar(),
        'precedure_class_id':precedure.precedure_class_id,
        'type':precedure.type,
        'count':precedure.count,
        'result':True,
        })


@login_required
def precedure_add_edit_set(request):
    id = request.POST.get('id');
    type = request.POST.get('type');
    precedure_class = request.POST.get('precedure_class');
    name = request.POST.get('name');
    name_vie = request.POST.get('name_vie');
    price = request.POST.get('price');
    price_dollar = request.POST.get('price_dollar');
    is_package = request.POST.get('is_package');
    count = request.POST.get('count');

    if int(id) == 0 :
        data = Precedure()
        data.price = price
        data.price_dollar = price_dollar

        
        #if precedure_class ==1: #D
        #    CODE = 'D'
        #elif precedure_class == 2: #CT
        #    CODE = 'CT'
        #elif precedure_class == 3: #ENT
        #    CODE = 'ENT'
        #elif precedure_class == 4: #GE
        #    CODE = 'GE'
        #elif precedure_class == 5: #Radi
        #    CODE = 'R'
        #elif precedure_class == 6: #U
        #    CODE = 'U'
        #elif precedure_class == 7: #P
        #    CODE = 'P'
        #elif precedure_class == 8: #T
        #    CODE = 'T'
        #elif precedure_class == 10: #PM
        #    CODE = 'PM'
        #elif ( precedure_class >= 11 and precedure_class <= 30 ) or precedure_class == 9: #: #DERM
        #    CODE = 'DM' 
        #elif ( precedure_class >= 31 and precedure_class <=40 ) or precedure_class == 42: #PS
        #    CODE = 'PS'
        #elif precedure_class == 41: #MRI
        #    CODE = 'MRI'
        #elif precedure_class == 44: #IM
        #    CODE = 'IM'
        #elif precedure_class == 45: #Vaccin
        #    CODE = 'VC'
        #elif precedure_class == 46: #Vaccin
        #    CODE = 'EC'
        last_code = Precedure.objects.filter(precedure_class = precedure_class).order_by('code').last()
        precedure_class = int(precedure_class)
        CODE = PrecedureClass.objects.get(pk = precedure_class).code
        print(CODE)
        if CODE == 'DM' or CODE == 'PS':
            last_code = Precedure.objects.filter(code__istartswith = CODE ).order_by('code').last()
            temp_code = last_code.code.split(CODE)
            data.code = CODE + str('%04d' % (int(temp_code[1]) + 1))
        elif CODE == 'PM':
            last_code = Precedure.objects.filter(code__istartswith = CODE ).order_by('code').last()
            temp_code = last_code.code.split(CODE)
            data.code = CODE + str('%03d' % (int(temp_code[1]) + 1))
        elif CODE == 'D3' or CODE == 'D4':
            last_code = Precedure.objects.filter(code__istartswith = CODE, code__regex=r'D+[0-9]*$').order_by('code').last()
            temp_code = last_code.code.split(CODE)
            data.code = CODE + str('%04d' % (int(temp_code[1]) + 1))
        else:
            if last_code == None:
                data.code = CODE + str('0001')
            else:
                prefix_code = last_code.code[0:3]
                if prefix_code == 'DM':
                    last_code = Precedure.objects.filter(code__istartswith = CODE, code__range = ('D00001','D99999') ).order_by('code').last()
                    temp_code = last_code.code.split(CODE)
                    data.code = CODE + str('%04d' % (int(temp_code[1]) + 1))
                else:
                    temp_code = last_code.code.split(CODE)
                    data.code = CODE + str('%04d' % (int(temp_code[1]) + 1)) 
    else:
        data = Precedure.objects.get(id=id)
        str_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

        try: 
            old_price = Pricechange.objects.get(type="Precedure",country='VI',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price.price != int(price):
                old_price.date_end = str_now
                old_price.save()

                new_price = Pricechange(type="Precedure",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price):
                new_price = Pricechange(type="Precedure",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()
                
        try:
            old_price_dollar = Pricechange.objects.get(type="Precedure",country='US',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price_dollar.price != int(price_dollar):
                old_price_dollar.date_end = str_now
                old_price_dollar.save()

                new_price = Pricechange(type="Precedure",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price_dollar):
                new_price = Pricechange(type="Precedure",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()
        






    data.precedure_class_id = precedure_class
    data.name = name
    data.name_vie = name_vie
    data.type = is_package
    data.count = count
    data.save()


    return JsonResponse({
        'result':True,
        })


@login_required
def precedure_add_edit_delete(request):
    id = request.POST.get('id')
    precedure = Precedure.objects.get(id=id)
    precedure.use_yn = 'N'
    precedure.save()

    return JsonResponse({
        'result':True,
        })


@login_required
def list_database_precedure_class_get(request):

    
    datas = []
    query = PrecedureClass.objects.all()
    for data in query:
        datas.append({
            'id':data.id,
            'name':data.name,
            'name_vie':data.name_vie,
            })


    return JsonResponse({
        'result':True,
        'datas':datas,
        })

@login_required
def add_edit_precedure_class_menu_get(request):

    id = request.POST.get('id')

    query = PrecedureClass.objects.get(pk = id)

    return JsonResponse({
        'result':True,
        'code':query.code,
        'name':query.name,
        'name_vie':query.name_vie,
        })



@login_required
def add_edit_precedure_class_menu_save(request):
    id = request.POST.get('id','')
    name = request.POST.get('class_name')
    name_vie = request.POST.get('class_name_vie')
    class_code = request.POST.get('class_code')

    if id == '':
        query = PrecedureClass()
    else:
        query = PrecedureClass.objects.get(pk = id)
        

    query.name = name
    query.name_vie = name_vie
    query.code = class_code
    query.save()

    return JsonResponse({
        'result':True,
        })


@login_required
def inventory__menu(request):

    #과 가져오기
    depart_datas = []
    depart_query = Depart.objects.all()
    for tmp_depart in depart_query: 
        depart_datas.append({
            'id':tmp_depart.id,
            'name': tmp_depart.name,
            })

    #검사 클래스
    test_class_datas = []
    test_class_query = TestClass.objects.all()
    for tmp_class in test_class_query: 
        test_class_datas.append({
            'id':tmp_class.id,
            'name': tmp_class.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
            })

    #처치 클래스
    precedure_class_datas=[]
    precedure_class_query = PrecedureClass.objects.all()
    for tmp_class in precedure_class_query: 
        precedure_class_datas.append({
            'id':tmp_class.id,
            'name': tmp_class.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
            })

    #약 클래스
    medicine_class_datas=[]
    medicine_class_query = MedicineClass.objects.all()
    for tmp_class in medicine_class_query: 
        medicine_class_datas.append({
            'id':tmp_class.id,
            'name': tmp_class.get_name_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
            })


    return render(request,
    'Manage/inventory__menu.html',
            {
                'depart_list':depart_datas,

                'test_class':test_class_datas,
                'precedure_class':precedure_class_datas,
                'medicine_class':medicine_class_datas,
            },
        )


@login_required
def inventory__menu_get(request):
    id = request.POST.get('id')
     
    test_list = []
    depart_test = Depart_TestMenu.objects.filter(depart = id)
    for data in depart_test:
        test_list.append({
            'id':data.test_class_id
            })

    precedure_list = []
    depart_precedure = Depart_PrecedureMenu.objects.filter(depart = id)
    for data in depart_precedure:
        precedure_list.append({
            'id':data.precedure_class_id
            })

    medicine_list = []
    depart_medicine = Depart_MedicineMenu.objects.filter(depart = id)
    for data in depart_medicine:
        medicine_list.append({
            'id':data.medicine_class_id
            })

    return JsonResponse({
        'result':True,
        'test_list':test_list,
        'precedure_list':precedure_list,
        'medicine_list':medicine_list,
        })


@login_required
def inventory__menu_set(request):

    #depart_test = Depart_TestClass.objects.filter(depart = id)
    #depart_precedure = Depart_PrecedureMenu.objects.filter(depart = id)
    #depart_medicine = Depart_MedicineClass.objects.filter(depart = id)


    depart_id = request.POST.get('depart_id')
    test_array = request.POST.getlist('test_array[]')
    precedure_array = request.POST.getlist('precedure_array[]')
    medicine_array = request.POST.getlist('medicine_array[]')



    Depart_TestMenu.objects.filter(depart = depart_id).delete()
    Depart_PrecedureMenu.objects.filter(depart = depart_id).delete()
    Depart_MedicineMenu.objects.filter(depart = depart_id).delete()

    now = datetime.datetime.now()
    
    for data in test_array: 
        data_split = data.split('test_')
   
        Depart_TestMenu(
            depart_id = depart_id, 
            test_class_id = data_split[1],
            creator = request.user,
            created_date = now,
            ).save()

    for data in precedure_array: 
        data_split = data.split('precedure_')
   
        Depart_PrecedureMenu(
            depart_id = depart_id, 
            precedure_class_id = data_split[1],
            creator = request.user,
            created_date = now,
            ).save()

    for data in medicine_array: 
        data_split = data.split('medicine_')
   
        Depart_MedicineMenu(
            depart_id = depart_id, 
            medicine_class_id = data_split[1],
            creator = request.user,
            created_date = now,
            ).save()



    return JsonResponse({
        'result':True,
        })


@login_required
def medicine_search(request):
    string = request.POST.get('string')
    filter = request.POST.get('filter')
    class_id = request.POST.get('class_id')
    

    kwargs = {}
    if class_id != '':
        kwargs.update({
            'medicine_class_id':class_id,
            'type':"TOOL",
            })

    
    datas=[]
    if string == '':
        argument_list_expriry_date = []
        argument_list_expriry_date.append( Q(**{'medicine__code__icontains':'T'} ) )

        expiry_date = datetime.datetime.now() + datetime.timedelta(days=180)
        #medicine_tmp = MedicineLog.objects.filter(type='add',expiry_date__lte = expiry_date ).select_related('medicine').exclude(medicine__use_yn='N' ,expiry_date=None,tmp_count=0).order_by('expiry_date')
        medicine_tmp= MedicineLog.objects.filter( medicine__type="TOOL", type='add', expiry_date__lte = expiry_date, tmp_count__gte= 0 ).select_related('medicine').values(
            'medicine_id',
            ).annotate(Count('medicine_id')).order_by('expiry_date')
       
        #medicine_tmp = MedicineLog.objects.filter( type='add', expiry_date__lte = expiry_date).exclude(tmp_count__lt = 0,medicine__use_yn='N',expiry_date=None).order_by('expiry_date').select_related('medicine')
        for tmp in medicine_tmp:
            medicine = Medicine.objects.get(id = tmp['medicine_id'])
            data = {
                    'id' : medicine.id,
                    'code': medicine.code,
                    'name' : medicine.name,
                    'company' : '' if medicine.company is None else medicine.company,
                    'unit' : '' if medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                    'price' : medicine.get_price(),
                    'count' : medicine.inventory_count,
                    'alaert_expiry':True,
                }
            datas.append(data)
            
        medicines = Medicine.objects.filter(**kwargs, type="TOOL").exclude(use_yn='N').order_by('id')
    else:
        argument_list = [] 
        #argument_list = [] 
        argument_list.append( Q(**{'name__icontains':string} ) )
        argument_list.append( Q(**{'name_vie__icontains':string} ) )
        argument_list.append( Q(**{'name_display__icontains':string} ) )

        medicines = Medicine.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs,type="TOOL").exclude(use_yn='N').order_by("id")#.select_related('medicine_class').exclude(use_yn = 'N').order_by("name")



    
    for medicine in medicines:
        data = {
                'id' : medicine.id,
                'code': medicine.code,
                'name' : medicine.name,
                'company' : '' if medicine.company is None else medicine.company,
                'unit' : '' if medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'price' : medicine.get_price(),
                'count' : medicine.inventory_count,
            }

        datas.append(data)


    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page');
    paginator = Paginator(datas, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    

    context = {
        #'datas':datas,
        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        
        }
    return JsonResponse(context)


@login_required
def set_data_control(request):
    medicine_id = request.POST.get('medicine_id');


    medicine = Medicine.objects.get(pk = medicine_id)

    context = {
        'name':medicine.name,
        'company':medicine.company,
        'country':medicine.country,
        'ingredient':medicine.ingredient,
        'unit':medicine.unit,
        'price':medicine.get_price(),
        }
    return JsonResponse(context)


@login_required
def save_data_control(request):
    selected_option = request.POST.get('selected_option');
    name = request.POST.get('name');
    price = request.POST.get('price');
    company = request.POST.get('company');
    ingredient = request.POST.get('ingredient');
    unit = request.POST.get('unit');
    changes = request.POST.get('changes') if request.POST.get('changes') is not '' else 0;


    if selected_option == 'new':
        medicine = Medicine()
        log = MedicineLog(type='new')
        code = Medicine.objects.all().last()
        code = 'M{:04d}'.format(code.id)
        medicine.price = price
    else:
        medicine = Medicine.objects.get(pk = selected_option )
        if int(changes) < 0:
            log = MedicineLog(type='dec')
        else:
            log = MedicineLog(type='add')

    medicine.name = name
    #medicine.price = price
    medicine.company = company
    medicine.ingredient = ingredient
    medicine.unit = unit

    medicine.inventory_count += int(changes)
    medicine.save()
    log.medicine = medicine
    log.changes = int(changes)
    log.save()
     
    context = {'result':True}
    return JsonResponse(context)


@login_required
def inventory_medical_tool(request):
    medicine_search_form = MedicineSearchForm()

    price_multiple_level = COMMCODE.objects.filter(commcode_grp = 'MED_MULTI_CODE').values('commcode','se1','se2').order_by('commcode_grp')


    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        medicine_class = MedicineClass.objects.all().annotate(name_display = F('name_vie')).values('id','name_display')
    else:
        medicine_class = MedicineClass.objects.all().annotate(name_display = F('name')).values('id','name_display')

    type = ["Medical Tool"]
    

    return render(request,
    'Manage/inventory_medical_tool.html',
            {
                'medicinesearch':medicine_search_form,
                'price_multiple_level':price_multiple_level,
                'medicine_class':medicine_class,
                'type':type,

            },
        )


@login_required
def medicine_add_edit_get(request):
    id = request.POST.get('id')

    try:
        medicine = Medicine.objects.get(id=id)

        context = {
            'result':True,
            'id':medicine.id,
            'code':medicine.code,
            'name':medicine.name,
            'name_vie':medicine.name_vie,
            'unit':medicine.unit,
            'unit_vie':medicine.unit_vie,
            'company':medicine.company,

            'type':'Medical Tool' if 'T' in medicine.code else 'Medical Tool',

            'price':medicine.get_price(),
            'price_input':medicine.get_price_input(),
            'price_dollar':medicine.get_price_dollar(),
            'multiple_level':medicine.multiple_level,

            'inventory_count':medicine.inventory_count,
            'medicine_class_id':medicine.medicine_class_id,

            }
    except Medicine.DoesNotExist:
        context = {'result':False}




    return JsonResponse(context)


@login_required
def medicine_add_edit_set(request):
    id = int(request.POST.get('id'))
    code = request.POST.get('code')
    type = request.POST.get('type')
    medicine_class = request.POST.get('medicine_class')
    name = request.POST.get('name')
    name_vie = request.POST.get('name_vie')
    unit = request.POST.get('unit')
    unit_vie = request.POST.get('unit_vie')
    company = request.POST.get('company')
    price_input = request.POST.get('price_input')
    multiple_level = request.POST.get('multiple_level')
    price = request.POST.get('price')   
    price_dollar = request.POST.get('price_dollar')

    if int(id) == 0 :
        data = Medicine()
        if type in 'Medical Tool':
            last_code =Medicine.objects.filter(code__icontains="T").last()
            temp_code = last_code.code.split('T')
            code = 'T' + str('%04d' % (int(temp_code[1]) + 1))

        data.code = code
        data.price = int(price)
        data.price_input = int(price_input)
        data.price_dollar = int(price_dollar)
        data.type = 'TOOL'
       
    else:
        data = Medicine.objects.get(id=id)
        str_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

        try: 
            old_price = Pricechange.objects.get(type="Medicine",country='VI',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price.price != int(price):
                old_price.date_end = str_now
                old_price.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price):
                new_price = Pricechange(type="Medicine",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()


        try:
            old_price_input = Pricechange.objects.get(type="Medicine",country='VI',type2='INPUT',code=data.code, date_end="99999999999999")
            
            if old_price_input.price != int(price_input):
                old_price_input.date_end = str_now
                old_price_input.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=data.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price_input != int(price_input):
                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=data.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        try:
            old_price_dollar = Pricechange.objects.get(type="Medicine",country='US',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price_dollar.price != int(price_dollar):
                old_price_dollar.date_end = str_now
                old_price_dollar.save()

                new_price = Pricechange(type="Medicine",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price_dollar != int(price_dollar):
                new_price = Pricechange(type="Medicine",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()


    data.medicine_class_id = medicine_class
    data.name = name
    data.name_vie = name_vie
    data.unit = unit
    data.unit_vie = unit_vie
    data.company = company
    data.multiple_level = multiple_level
    

    data.save()

    context = {'result':True}




    return JsonResponse(context)


@login_required
def medicine_add_edit_check_code(request):
    code = request.POST.get('code')
    id = request.POST.get('id')
    try:
        medicine = Medicine.objects.get(code=code)
        res = "N"
        
        if medicine.id == int(id):
            res = 'Same'
    except Medicine.DoesNotExist:
        res = "Y"

    context = {'result':res}

    return JsonResponse(context)


@login_required
def medicine_add_edit_delete(request):
    id = request.POST.get('id')

    medicine = Medicine.objects.get(id=id)
    medicine.use_yn = 'N'
    medicine.save()

    log = MedicineLog()
    log.medicine = medicine
    log.type='del'
    log.save()



    return JsonResponse({'result':True})


@login_required
def get_inventory_history(request):
    id = request.POST.get('id')
    date = request.POST.get('date')


    #date_min = datetime.datetime.combine(datetime.datetime.strptime(date, "%Y-%m-%d").date(), datetime.time.min)
    #date_max = datetime.datetime.combine(datetime.datetime.strptime(date, "%Y-%m-%d").date(), datetime.time.max)

    #medicine_logs = MedicineLog.objects.filter(date__range = (date_min, date_max), medicine_id = id).values('date','changes','type','memo').order_by('-date')
    medicine_logs = MedicineLog.objects.filter(medicine_id = id).values('date','changes','type','memo').order_by('-date')
    datas = []
    for medicine_log in medicine_logs:
        data = {
            'date':medicine_log['date'].strftime('%Y-%m-%d %H:%M:%S'),
            'changes':medicine_log['changes'],
            'type':medicine_log['type'],
            'memo':'' if medicine_log['memo'] is None else medicine_log['memo'],
            }
        datas.append(data)

    medicine = Medicine.objects.get(id=id)
    return JsonResponse({
        'datas':datas,
        'count':medicine.inventory_count,
        })


@login_required
def save_database_add_medicine(request):
    id = request.POST.get('id')
    registration_date= request.POST.get('registration_date')
    expiry_date= request.POST.get('expiry_date')
    changes= request.POST.get('changes')
    memo= request.POST.get('memo')
    check= request.POST.get('check')
    
    if int(check)==0 :
        medicine = Medicine.objects.get(id=id)
        count = medicine.inventory_count
        medicine.inventory_count =  count + int(changes)
        medicine.save()

        medicine_logs = MedicineLog()
        medicine_logs.changes = changes
        medicine_logs.medicine = medicine
    else:
        medicine_logs = MedicineLog.objects.get(id = check)
        medicine_logs.date = datetime.datetime.strptime(registration_date, '%Y-%m-%d')

    medicine_logs.memo = memo
    if expiry_date != '' :
        medicine_logs.expiry_date = datetime.datetime.strptime(expiry_date, '%Y-%m-%d')
    
 
    medicine_logs.tmp_count = changes
    medicine_logs.type='add'
    medicine_logs.save()




    input_price = request.POST.get('input_price',None)
    if input_price:
        try:
            str_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            old_price_input = Pricechange.objects.get(type="Medicine",country='VI',type2='INPUT',code=medicine.code, date_end="99999999999999")
            
            if old_price_input.price != int(input_price):
                old_price_input.date_end = str_now
                old_price_input.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=medicine.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if medicine.price_input != int(input_price):
                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=medicine.code)
                new_price.price = input_price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

    return JsonResponse({'result':True})


@login_required
def get_expiry_date(request):

    id = request.POST.get('id')
    medicine = Medicine.objects.get(id=id)

    medicine_logs = MedicineLog.objects.filter(medicine = medicine, type='add').exclude(tmp_count__lte = 0).order_by('expiry_date').values('id','date','tmp_count','expiry_date')
    
    datas = []
    for medicine_log in medicine_logs:
        
        data = {
            'id':medicine_log['id'],
            'date':medicine_log['date'].strftime('%Y-%m-%d'),
            'expiry_date':0 if medicine_log['expiry_date'] is None else medicine_log['expiry_date'].strftime('%Y-%m-%d'),
            'tmp_count':medicine_log['tmp_count'] if medicine_log['tmp_count'] is not None else 0,
            }
        datas.append(data)

    return JsonResponse({
        'result':True,
        'datas':datas,
        })


@login_required
def get_edit_database_add_medicine(request):
    id = request.POST.get('id')

    data = MedicineLog.objects.values('id','date','expiry_date','memo','tmp_count').get(id=id)

    return JsonResponse({
        'result':True,
        'data':{
            'id':data['id'],
            'date':data['date'].strftime('%Y-%m-%d'),
            'expiry_date':'' if data['expiry_date'] is None else data['expiry_date'].strftime('%Y-%m-%d'),
            'memo':data['memo'],
            'tmp_count':data['tmp_count'],
            },
        })


@login_required
def save_database_disposal_medicine(request):
    id = request.POST.get('id')
    disposial = request.POST.get('disposial')
    memo = request.POST.get('memo')


    disposal_data = MedicineLog.objects.get(id=id)
    if disposal_data.tmp_count < int(disposial):
        return JsonResponse({
            'result':False,
            'msg':1, # tmp 카운트가 더 적을 경우
                             })
    else:
        medicine = Medicine.objects.get(id = disposal_data.medicine_id)
        disposal_data.tmp_count -= int(disposial)
        
        disposal_data.save()

        medicine_log = MedicineLog()
        medicine_log.type = 'dec'
        medicine_log.changes = disposial
        medicine_log.memo = memo
        medicine_log.medicine_id = disposal_data.medicine_id
        medicine_log.save()

        
        medicine.inventory_count -= int(disposial)
        medicine.save()



    return JsonResponse({'result':True,})



#기안서
@login_required
def draft(request):

    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
        user_name = F('name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
        user_name = F('name_ko')
    else:
        f_name = F('commcode_name_en')
        user_name = F('name_en')

    draft_request_type = None



    list_query_depart = []
    
    list_depart = []
    query_depart= COMMCODE.objects.filter(upper_commcode = '000002',commcode_grp = 'DEPART_ADMIN',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')
    for data in query_depart:
        list_depart.append({
            'code':data['code'],
            'name':data['name'],
            'id':data['id']
            })


    url = 'Manage/Draft.html'
    if not request.session['is_KBL']:#테스트서버
        url = 'Manage/Draft.html'
        #부서 - IMEDICARE
        query_depart = COMMCODE.objects.filter(upper_commcode = '000002',commcode_grp = 'DEPART_CLICINC',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id').order_by('code')
        for data in query_depart:
            if data['code'] == 'DOCTOR':
                temp_commcode = COMMCODE.objects.get(id = data['id'])
                data['code'] += '_' + temp_commcode.se1
            list_depart.append({
                'code':data['code'],
                'name':data['name'],
                'id':data['id']
                })
            list_query_depart.append( Q(**{'depart':data['code']} ) ) 


        #기안서 종류 - IMEDICARE
        draft_type = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'DRAFT_TYPE_IMEDI',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name').order_by('seq')

        #기안서 요청 형식 - PAYMENT
        draft_request_type =COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'REQUEST_FORM',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')
        draft_payment_type =COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'PAYMENT_TYPE',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')

        #서류번호
        draft_doc_num = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'DOC_NUM_IMEDI',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')

    else:#if request.session['is_KBL']:#경천애인
        #부서 - KBL
        query_depart= COMMCODE.objects.filter(upper_commcode = '000002',commcode_grp = 'DEPART_KBL', use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id').order_by('code')
        for data in query_depart:
            list_depart.append({
                'code':data['code'],
                'name':data['name'],
                'id':data['id']
                })
            list_query_depart.append( Q(**{'depart':data['code']} ) ) 

        #기안서 종류 - KBL
        draft_type = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'DRAFT_TYPE',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name').order_by('seq')
        #서류번호
        draft_doc_num = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'DOC_NUM_KBL',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')
        
        #기안서 요청 형식 - PAYMENT
        draft_request_type =COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'REQUEST_FORM',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')
        draft_payment_type =COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'PAYMENT_TYPE',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')


    #기안서 상태
    draft_status = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp = 'DRAFT_STATUS',use_yn="Y").annotate(code = F('commcode'),name = f_name ).values('code','name','id')

    #set user data
    user_dict = {}
    users = User.objects.filter(
        functools.reduce(operator.or_, list_query_depart)
        ).annotate(name = user_name).values('id','name','depart','depart_doctor')
    for user in users:
        depart_name = user['depart']
        if user['depart'] == 'DOCTOR':
            depart_name = user['depart'] + '_' + user['depart_doctor']
        
        user_dict.update({
            user['id'] : user['name']
            #{
            #    'user_id':user['user_id'],
            #    'name':user['name'],
            #    }
            })


    file_form = board_file_form()


    return render(request,
        url,
            {
                'draft_type':draft_type,

                'list_depart':list_depart,
                'user_dict':user_dict,

                'draft_status':draft_status,

                'file_form':file_form,

                'draft_request_type':draft_request_type,
                'draft_payment_type':draft_payment_type,

                'draft_doc_num':draft_doc_num,
            }
        )


#기안서 검색
def draft_search(request):
    start = request.POST.get('start','')
    end = request.POST.get('end','')


    string = request.POST.get('string','')

    type = request.POST.get('type','')
    requester = request.POST.get('requester','')
    status = request.POST.get('status','')

    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
        f_user = F('name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
        f_user = F('name_ko')
    else:
        f_name = F('commcode_name_en')
        f_user = F('name_en')



    type_dict = {}
    for data in COMMCODE.objects.filter(commcode_grp='DRAFT_TYPE_IMEDI',upper_commcode='000007' ).annotate(code = F('commcode'),name = f_name ).values('code','name'):
        type_dict.update({
            data['code'] : data['name']
            })
    for data in COMMCODE.objects.filter(commcode_grp='DRAFT_TYPE',upper_commcode='000007' ).annotate(code = F('commcode'),name = f_name ).values('code','name'):
        type_dict.update({
            data['code'] : data['name']
            })




    date_min = datetime.datetime.combine(datetime.datetime.strptime(start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end, "%Y-%m-%d").date(), datetime.time.max)



    kwargs={}
    if string != '':
        kwargs['string__icontains']=string
    if type != '':
        kwargs['type']=type
    if requester != '':
        kwargs['requester']=requester
    if status != '':
        kwargs['status']=status

    draft_list=[]
    if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
        draft_query = Draft.objects.filter(date_registered__range = (date_min, date_max),**kwargs,use_yn='Y',is_KBL='Y').order_by('-date_registered')
    else:
        draft_query = Draft.objects.filter(date_registered__range = (date_min, date_max),**kwargs,use_yn='Y',is_KBL='N').order_by('-date_registered')

    for draft in draft_query:
        if draft.depart == '':
            depart = ''
        else:
            depart = COMMCODE.objects.filter(id = draft.depart).annotate(name = f_name ).values('name')[:1]

        status_val = ''
        if draft.status !='':
            status_val = COMMCODE.objects.filter(upper_commcode = '000007',commcode_grp='DRAFT_STATUS', commcode=draft.status).annotate(name = f_name ).values('name')
            status_val = status_val[0]['name']

        user_creator = User.objects.filter(id = draft.creator).annotate(name = f_user ).values('name')[:1]

        
        if draft.date_in_charge == '0000-00-00 00:00:00':
            in_charge = None
        else:
            in_charge = draft.date_in_charge[0:10]
            in_charge_id = draft.user_id_in_charge

        if draft.date_leader == '0000-00-00 00:00:00':
            leader = None
        else:
            leader = draft.date_leader[0:10]
            leader_id = draft.user_id_leader

        if draft.date_accounting == '0000-00-00 00:00:00':
            accounting = None
        else:
            accounting = draft.date_accounting[0:10]
            accounting_id = draft.user_id_accounting

        if draft.date_ceo == '0000-00-00 00:00:00':
            ceo = None
        else:
            ceo = draft.date_ceo[0:10]
            ceo_id = draft.user_id_ceo


        
        draft_list.append({
            'id':draft.id,
            'status':draft.status,
            'type':draft.type,
            'type_val':draft.type,
            'title':draft.title,
            'depart':'' if depart == '' else depart[0]['name'],
            'requester':draft.request_user,
            'RQSTR':user_creator[0]['name'],
            'RQSTD_DATE':draft.date_registered[0:10],
            
            'in_charge_id':draft.creator,

            'in_charge':in_charge,
            'leader':leader,
            'accounting':accounting,
            'ceo': ceo,
            
            'status_val':status_val,
            })

    page = request.POST.get('page',1)
    context_in_page = request.POST.get('page_context');
    paginator = Paginator(draft_list, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    return JsonResponse({
        'type_dict':type_dict,

        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        })


#기안서 불러오기
def draft_get_data(request):
    form_id = request.POST.get('id')

    query_data = Draft.objects.get(id = form_id)

    return JsonResponse({
        'result':True,

        'type':query_data.type,
        'depart':query_data.depart,
        'request_user':query_data.request_user,
        'cost_name':query_data.cost_name,
        'title':query_data.title,
        'contents':query_data.contents,
        'consultation':query_data.consultation,
        'additional':query_data.additional,
        'status':query_data.status,

        'consultation_depart1':query_data.consultation_depart1,
        'consultation_incharge1':query_data.consultation_incharge1,
        'consultation_content1':query_data.consultation_content1,
        'consultation_depart2':query_data.consultation_depart2,
        'consultation_incharge2':query_data.consultation_incharge2,
        'consultation_content2':query_data.consultation_content2,

        'payment_type':query_data.payment_type,
        'request_type':query_data.request_type,
        'doc_type':query_data.doc_type,
        })






#기안서 양식
def draft_get_form(request):

    form_id = request.POST.get('form_id')

    path = 'static/draft/' + form_id
    file = open(path,'rt',encoding='UTF-8')
    data = file.read()

    return JsonResponse({'result':True,
                         'data':data,
                         })

#기안서 저장
def draft_save(request):

    id = request.POST.get("id",'')

    new_edit_type=request.POST.get("new_edit_type",'')
    new_edit_payment_type=request.POST.get("new_edit_payment_type",'')
    new_edit_request_type=request.POST.get("new_edit_request_type",'')
    new_edit_depart=request.POST.get("new_edit_depart",'')
    new_edit_cost_name=request.POST.get("new_edit_cost_name",'')
    new_edit_name=request.POST.get("new_edit_name",'')
    new_edit_title=request.POST.get("new_edit_title",'')
    new_edit_content=request.POST.get("new_edit_content",'')
    new_edit_consultation=request.POST.get("new_edit_consultation",'')
    new_edit_MORE_CMNTS=request.POST.get("new_edit_MORE_CMNTS",'')

    consultation_depart1=request.POST.get("consultation_depart1",'')
    consultation_incharge1=request.POST.get("consultation_incharge1",'')
    consultation_content1=request.POST.get("consultation_content1",'')
    consultation_depart2=request.POST.get("consultation_depart2",'')
    consultation_incharge2=request.POST.get("consultation_incharge2",'')
    consultation_content2=request.POST.get("consultation_content2",'')

    new_edit_status=request.POST.get("new_edit_status",'')
    new_edit_code=request.POST.get("new_edit_code",'')

    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if id != '':
        draft = Draft.objects.get(pk = id)
    else:
        draft = Draft()
        draft.creator = request.user.id
        draft.date_registered = now

    

    draft.type = new_edit_type
    draft.payment_type = new_edit_payment_type
    draft.request_type = new_edit_request_type
    draft.cost_name = new_edit_cost_name
    draft.depart = new_edit_depart
    
    draft.title = new_edit_title
    draft.request_user = new_edit_name
    draft.contents = new_edit_content
    draft.consultation = new_edit_consultation
    draft.additional = new_edit_MORE_CMNTS
    draft.status = new_edit_status

    draft.consultation_depart1 = consultation_depart1
    draft.consultation_incharge1 = consultation_incharge1
    draft.consultation_content1 = consultation_content1
    draft.consultation_depart2 = consultation_depart2
    draft.consultation_incharge2 = consultation_incharge2
    draft.consultation_content2 = consultation_content2



    #서류 번호 - 변경시, 새로 등록시 추가
    if draft.doc_type is not new_edit_code:
        tmp_draft_query = Draft.objects.filter(
            date_registered__istartswith = draft.date_registered[0:10],
            doc_type__icontains = new_edit_code,
            ).order_by('doc_num').last()

        if tmp_draft_query is None:
            draft.doc_num = draft.date_registered[8:10] + draft.date_registered[5:7]
        else:
            num_split = tmp_draft_query.doc_num.split('-')
            if len(num_split) == 1 :
                draft.doc_num = draft.date_registered[8:10] + draft.date_registered[5:7] + '-1'
            else:
                draft.doc_num = draft.date_registered[8:10] + draft.date_registered[5:7] + '-' + str( int(num_split[1]) + 1 )

    draft.doc_type = new_edit_code

    draft.modifier = request.user.id
    draft.date_last_modified = now

    if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
        draft.is_KBL='Y'

    draft.save()


        
    try:
        alert_query = AlertLog.objects.get(
            page_type="DRAFT",
            content_type="CONTENT",
            content_id = id,
            )
    except AlertLog.DoesNotExist:
        alert_query = AlertLog(
            page_type="DRAFT",
            content_type="CONTENT",
            content_id = id,
            creator = request.user.id,
            )

    
    alert_query.status=new_edit_status

    alert_query.save()





    return JsonResponse({'result':True,})


#기안서 삭제
def draft_delete(request):
    id = request.POST.get('id')

    draft = Draft.objects.get(id = id )
    draft.use_yn = "N"
    draft.save()


    alert_query = AlertLog.objects.filter(
        page_type="DRAFT",
        content_type="CONTENT",
        content_id = id,
        )
    for data in alert_query:
        data.use_yn = 'N'
        data.save()


    return JsonResponse({'result':True,})


#기안서 파일 리스트
def draft_list_file(request):
    id = request.POST.get('id')

    list_file = []
    query_file = Board_File.objects.filter(board_id = id,board_type='DRAFT').order_by('registered_date')

    for file in query_file:
        list_file.append({
            'id':file.id,
            'url':file.file.url,
            'name':file.title,
            'origin_name':file.origin_name,
            'date':file.registered_date.strftime("%Y-%m-%d"),
            'creator':file.user,
            'memo':file.memo,
            });

    return JsonResponse({
        'result':True,
        'datas':list_file,
        })

#기안서 파일 정보 불러오기
def draft_get_file(request):
    id = request.POST.get('id')

    query_file = Board_File.objects.get(id = id)

    file_name = query_file.file.url,

    return JsonResponse({
        'result':True,
        'title':query_file.title,
        'memo':query_file.memo,
        'origin_name':query_file.origin_name,
        })

#기안서 파일 저장
def draft_save_file(request):
    id = request.POST.get('id')

    if request.method == 'POST':

        selected_file_id = request.POST.get('selected_file_id','')#파일 ID
        selected_file_list = request.POST.get('selected_file_list','')#기안서 ID 
        new_edit_file_name = request.POST.get('new_edit_file_name','')#문서 이름
        new_edit_file_remark = request.POST.get('new_edit_file_remark','')#문서 설명

        form = board_file_form(request.POST, request.FILES)
        files = request.FILES.getlist('file') 

        if form.is_valid():
           if selected_file_id != '': #수정
                file_instance = Board_File.objects.get(id = selected_file_id)
           else:
                file_instance = Board_File()


           #    old_data
           #file save
           for f in files:
               if file_instance.file:
                    if os.path.isfile(file_instance.file.path):
                        os.remove(file_instance.file.path)

               file_instance.file = f
               file_instance.origin_name = f._name
               pass

           file_instance.board_id = selected_file_list
           file_instance.user = request.user.user_id
           file_instance.board_type = 'DRAFT'
           file_instance.title = new_edit_file_name
           file_instance.memo = new_edit_file_remark
           file_instance.save()
           #form.save()
           return JsonResponse({'error': False, 'message': 'Uploaded Successfully'})
        else:
           return JsonResponse({'error': True, 'errors': form.errors})

    else:
        form = board_file_form()
        return render(request, 'django_image_upload_ajax.html', {'form': form})



#기안서 파일 삭제
def draft_delete_file(request):

    id = request.POST.get('id')
    Board_File.objects.get(id=id).delete()



    return JsonResponse({
        'result':True,
        })



#기안서 승인
def check_appraove(request):

    id = request.POST.get('id')
    type = request.POST.get('type')
    val = request.POST.get('val')



    draft = Draft.objects.get(id = id )

    if val=='true':
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
        user_id = request.user.id
        name_en = request.user.name_en
        name_ko = request.user.name_ko
        name_vi = request.user.name_vi
    else:
        now = '0000-00-00 00:00:00'
        user_id = ''
        name_en = ''
        name_ko = ''
        name_vi = ''
    


    if type == 'incharge':
        draft.date_in_charge = now
        draft.user_id_in_charge = user_id
        draft.name_en_in_charge = name_en
        draft.name_ko_in_charge = name_ko
        draft.name_vi_in_charge = name_vi


    elif type == 'leader':
        draft.date_leader = now
        draft.user_id_leader= user_id
        draft.name_en_leader= name_en
        draft.name_ko_leader= name_ko
        draft.name_vi_leader= name_vi



    elif type == 'accounting':
        draft.date_accounting = now
        draft.user_id_accounting= user_id
        draft.name_en_accounting= name_en
        draft.name_ko_accounting= name_ko
        draft.name_vi_accounting= name_vi

    elif type == 'ceo':
        draft.date_ceo = now
        draft.user_id_ceo= user_id
        draft.name_en_ceo= name_en
        draft.name_ko_ceo= name_ko
        draft.name_vi_ceo= name_vi

    else:
            
        return JsonResponse({
            'result':False ,
            })

    draft.save()



    
    return JsonResponse({
        'result':True,
        })




#기안서 프린트
def draft_print(request,id=None):

    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    else:
        f_name = F('commcode_name_en')


    draft = Draft.objects.get(id = id)
    #type = COMMCODE.objects.get(commcode = draft.type ,commcode_grp='DRAFT_TYPE',upper_commcode='000007' )

    file_list= []
    file_list_query = Board_File.objects.filter(board_id=id, board_type='DRAFT').values('title').order_by('registered_date')


    #파일
    file_rowspan = len(file_list_query)
    if file_rowspan == 0 :
        file_rowspan = 1
    else:
        tmp = 1
        for file in file_list_query:

            file_list.append({
                'id':tmp,
                'title':file['title'],
                })
            tmp +=1

    
    #url 및 context

    try:
        depart = COMMCODE.objects.get(id = draft.depart)
        depart_vi = depart.commcode_name_vi
        depart_ko = depart.commcode_name_ko
    except ValueError or COMMCODE.DoesNotExist: 
        depart_vi = ''
        depart_ko = ''

    context = {}
    if request.session['is_KBL'] is True:
        url = 'Draft_Form/basic_form_KBL.html'
        type = COMMCODE.objects.get(commcode = draft.type ,commcode_grp='DRAFT_TYPE',upper_commcode='000007' )
    

        
        context.update({
            
            'incharge_name':draft.name_en_in_charge,
            'incharge_date':'' if draft.date_in_charge == '0000-00-00 00:00:00' else draft.date_in_charge[0:10],
            'leader_name':draft.name_en_leader,
            'leader_date':'' if draft.date_leader == '0000-00-00 00:00:00' else draft.date_leader[0:10],
            'account_name':draft.name_en_accounting,
            'account_date':'' if draft.date_accounting == '0000-00-00 00:00:00' else draft.date_accounting[0:10],
            'ceo_name':draft.name_en_ceo,
            'ceo_date':'' if draft.date_ceo == '0000-00-00 00:00:00' else draft.date_ceo[0:10],

            })
        
    else:
        type = COMMCODE.objects.get(commcode = draft.type ,commcode_grp='DRAFT_TYPE_IMEDI',upper_commcode='000007' )

        if draft.type == 'PAYMENT':
            url = 'Draft_Form/basic_form_payment.html'
            
            #요청 형식
            request_type_query = COMMCODE.objects.get(upper_commcode = '000007',commcode_grp = 'REQUEST_FORM',commcode = draft.request_type,)
            payment_type_query = COMMCODE.objects.get(upper_commcode = '000007',commcode_grp = 'PAYMENT_TYPE',commcode = draft.payment_type,)

            request_type_v = request_type_query.commcode_name_vi + '\n' + request_type_query.commcode_name_ko
            payment_type = payment_type_query.commcode_name_vi + '<br/>' + payment_type_query.commcode_name_ko



            context.update({
                'request_type_vi':request_type_query.commcode_name_vi,
                'request_type_ko':request_type_query.commcode_name_ko,
                'payment_type_vi':payment_type_query.commcode_name_vi,
                'payment_type_ko':payment_type_query.commcode_name_ko,

            })
        elif draft.type == 'PAYMENTAPPROVAL':
            url = 'Draft_Form/basic_form_payment_approval.html'
        else:
            url = 'Draft_Form/basic_form.html'
            

    
    context.update({
        'type_vi':type.commcode_name_vi,
        'type_ko':type.commcode_name_ko,

        'date_registered':draft.date_registered[0:10],

        'status':draft.status,

        'title':draft.title,
        'contents':draft.contents,
        'cost_name':draft.cost_name,
        'request_user':draft.request_user,
        'file_rowspan':file_rowspan,
        'file_list':file_list,

        'doc_num':draft.doc_num,
        'doc_type':draft.doc_type,

        
                
        'depart_vi':depart_vi,
        'depart_ko':depart_ko,
        })


    #diagnostic = reception.diagnosis.diagnosis
    return render(request,url,context,)

    
    return JsonResponse({
        'result':True,
        })




#고객 관리
def customer_manage(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')



    departs = Depart.objects.all()

    patient_mark = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PT_INFO',use_yn="Y").values('commcode','se1').order_by('seq')

    list_funnels = []
    funnels = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PATIENTS_FUNNELS',use_yn="Y").annotate(name = f_name ).values('commcode','name',)
    for data in funnels:
        list_funnels.append({
            'code':data['commcode'],
            'name':data['name'],
            })


    return render(request,
        'Manage/CRM.html',
            {
                'departs':departs,
                'patient_mark':patient_mark,
                'list_funnels':list_funnels,
            }
        )


def customer_manage_get_patient_list(request):
    page_context = request.POST.get('context_in_page',10)
    page = request.POST.get('page',1)
    

    depart = request.POST.get('depart')
    category = request.POST.get('category')
    string = request.POST.get('string')
    date_start = request.POST.get('start')
    date_end = request.POST.get('end')

    kwargs={}
    kwargs_sub={}
    argument_list = [] 

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end,"%Y-%m-%d").date(), datetime.time.max)
    if depart != '':
        kwargs['depart_id']=depart


    if category=='':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 


    query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs,
        patient__date_registered__range = (date_min, date_max)
        ).select_related(
            'patient'
            )
    
    if len(kwargs):#과 필터 선택 시
        query = query.exclude(
            progress='deleted'
            ).order_by('patient_id').values('patient_id').annotate(
                Count('patient_id'
                    #Case(
                    #        When(patient_id__gt=0, then=Value('patient_id') ),
                    #        When(patient_id=0, then=Value(0) ),
                    #        default = Value(0),
                    #    )
                    )
                )
    else:
        query = query.order_by('patient_id').values('patient_id').annotate(
                Count('patient_id')
                )
        
    count_page = query
    datas=[]
    for data in query[int(page_context) * (int(page) -1): int(page_context) * int(page)]:

        patient = Patient.objects.get(pk = data['patient_id'])

        receptions = Reception.objects.filter(
            **kwargs,
            patient_id = data['patient_id'],
            # progress='done',
                ).exclude(progress='deleted')
        paid_total = 0
        for reception in receptions:
            try:
                paid_total += reception.payment.sub_total
            except:
                pass
        #total_amount = receptions.aggregate(total_price=Sum('payment__paymentrecord__paid'))

        datas.append({
            'id':patient.id,
            'chart':patient.get_chart_no(),
            'name_kor':patient.name_kor,
            'name_eng':patient.name_eng,
            'gender':patient.get_gender_simple(),
            'date_of_birth':patient.date_of_birth.strftime('%Y-%m-%d'),
            'phonenumber':patient.phone,
            'age' : patient.get_age(),
            'address':patient.address,
        
            'memo':patient.memo,
            'date_registered':patient.date_registered.strftime('%Y-%m-%d'),
        
            'visits':receptions.count(),
            'paid_total':paid_total,
            })

    #paginator = Paginator(datas , page_context)
    paginator = Paginator(count_page , page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    return JsonResponse({
        'result':True,
        'datas':datas,
        #'datas':list(paging_data),

        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),


        })

#고객 관리
def profile_status(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')



    departs = Depart.objects.all()

    patient_mark = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PT_INFO',use_yn="Y").values('commcode','se1').order_by('seq')

    list_funnels = []
    funnels = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PATIENTS_FUNNELS',use_yn="Y").annotate(name = f_name ).values('commcode','name',)
    for data in funnels:
        list_funnels.append({
            'code':data['commcode'],
            'name':data['name'],
            })


    return render(request,
        'Manage/profile_status.html',
            {
                'departs':departs,
                'patient_mark':patient_mark,
                'list_funnels':list_funnels,
            }
        )


def search_profile_status(request):
    page_context = request.POST.get('context_in_page',10)
    page = request.POST.get('page',1)
    
    depart = request.POST.get('depart')
    category = request.POST.get('category')
    date_start = request.POST.get('start')
    date_end = request.POST.get('end')

    profile_status = request.POST.get('status')  
    string = request.POST.get('string')
    invoice_insurance = request.POST.get('invoice_insurance')  

    kwargs={}
    kwargs_sub={}
    argument_list = [] 
    count = 0
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end,"%Y-%m-%d").date(), datetime.time.max)
    if depart != '':
        kwargs['depart_id']=depart

    if profile_status != '':
        kwargs['profile_status']=profile_status
        # argument_list.append( Q(**{'profile_status':profile_status} ) )
    if invoice_insurance == 'Invoice':
        kwargs['need_invoice'] = True
    elif invoice_insurance == 'Insurance':     
        kwargs['need_insurance'] = True
    elif invoice_insurance == 'Invoice_Insurance':     
        kwargs['need_insurance'] = True   
        kwargs['need_invoice'] = True     

    if category=='':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 

    query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs,
        recorded_date__range = (date_min, date_max)
        )
    
    # if len(kwargs):#과 필터 선택 시
    #     query = query.exclude(
    #         progress='deleted'
    #         ).order_by('patient_id').values('patient_id', 'depart').annotate(
    #             Count('patient_id'
    #                 #Case(
    #                 #        When(patient_id__gt=0, then=Value('patient_id') ),
    #                 #        When(patient_id=0, then=Value(0) ),
    #                 #        default = Value(0),
    #                 #    )
    #                 )
    #             )
    # else:
    #     query = query.order_by('patient_id').values('patient_id', 'depart').annotate(
    #             Count('patient_id')
    #             )
        
    count_page = query
    datas=[]
    for data in query[int(page_context) * (int(page) -1): int(page_context) * int(page)]:
        patient = Patient.objects.get(pk = data.patient_id)     
        count_ = 0
        if int(page) != 1:
            count_ = 10 * int(page)
        count += 1
        count_ += count
        receptions = Reception.objects.filter(
            **kwargs,
            patient_id = data.patient_id,
            progress='done',
                ).exclude(progress='deleted')
        paid_total = 0
        for reception in receptions:
            paid_total += reception.payment.sub_total
        #total_amount = receptions.aggregate(total_price=Sum('payment__paymentrecord__paid'))
        need_insurance = 'No'
        need_invoice = 'No'
        tax_number = ''  
        company = ''  
        company_address = ''
        recommend = ''

        if data.need_invoice == True:
            need_invoice = 'Yes'
            try:
                taxInvoice = TaxInvoice.objects.get(patient_id = data.patient_id)
                tax_number = taxInvoice.number
                company = taxInvoice.company_name 
                company_address =  taxInvoice.address
                recommend =  taxInvoice.recommend
            except TaxInvoice.DoesNotExist:          
                nothing = 1
        if data.need_insurance == True:
            need_insurance = 'Yes'  

        
        # try:
        #     recommend = data.diagnosis.recommendation
        #     if not recommend:
        #         recommend = ''
        # except:          
        #     nothing = 1      
        # name_eng =  re.sub("\n|\r", "", patient.name_eng)     

        datas.append({
            'id':count_,
            'chart':patient.get_chart_no(),
            'name_kor': patient.name_kor, # re.sub(r'[^\w.]', '', patient.name_kor),
            'name_eng':patient.name_eng,
            'gender':patient.get_gender_simple(),
            'recorded_date':data.recorded_date.strftime('%Y-%m-%d'),
            'phonenumber':patient.phone, # re.sub(r'[^\w.]', '', patient.phone),
            'age' : patient.get_age(),
            'address':patient.address,
            'email':patient.email,
            'memo':patient.memo,
            'date_registered':patient.date_registered.strftime('%Y-%m-%d'),
            'depart':data.depart.name,
            'red_invoice': need_invoice,
            'tax_number':tax_number,
            'company':company,
            'company_address':company_address,
            'need_insurance':need_insurance,
            'recommend':recommend,
            'status':data.profile_status,
            'rec_id':data.id,
            'patient_id':data.patient_id,
            })

    #paginator = Paginator(datas , page_context)
    paginator = Paginator(count_page , page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    return JsonResponse({
        'result':True,
        'datas':datas,
        #'datas':list(paging_data),

        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),


        })


def customer_manage_get_patient_info(request):
    patient_id = request.POST.get('patient_id')

    context={}
    patient = Patient.objects.get(pk=int(patient_id))

    
    context.update({
        'id':patient.id,
        'chart':patient.get_chart_no(),
        'name_kor':patient.name_kor,
        'name_eng':patient.name_eng,
        'date_of_birth':patient.date_of_birth,
        'gender':patient.gender,
        'email':patient.email,
        'nationality':patient.nationality,
        'phone':patient.phone,
        'address':patient.address,
        'memo':patient.memo,
        'marking':patient.marking,
        'funnel':patient.funnel,
        'funnel_etc':patient.funnel_etc,
        'passport':patient.passport
        })
    return JsonResponse(context)


def customer_manage_get_patient_visit(request):

    depart = request.POST.get('depart')
    patient_id = request.POST.get('patient_id')

    kwargs={}
    if depart != '':
        kwargs['depart_id']=depart

    context={}
    receptions = Reception.objects.filter(**kwargs,patient_id=int(patient_id),progress='done').exclude(progress='deleted').order_by('recorded_date')



    datas = []
    for reception in receptions:
        payment = Payment.objects.get(reception_id = reception.id)
        #payment = Payment.objects.filter(reception_id = reception.id).aggregate(paid_sum=Sum('paymentrecord__paid'))


        datas.append({
            'reception_id':reception.id,
            'depart':reception.depart.name,
            'doctor':reception.doctor.name_short,
            'paid':payment.sub_total,
            #'paid':'0' if payment['paid_sum'] is None else payment['paid_sum'],
            'date_visited':reception.recorded_date.strftime('%Y-%m-%d %H:%M'),
            })

    context.update({
        'datas':datas,
        })
    return JsonResponse(context)


def customer_manage_get_patient_visit_history(request):

    reception_id = request.POST.get('reception_id')

    reception = Reception.objects.get(id = reception_id)
    res = True
    data = {}
    try:
        diagnosis = Diagnosis.objects.get(reception_id = reception.id)

        exam_set = ExamManager.objects.filter(diagnosis_id = diagnosis.id)
        test_set = TestManager.objects.filter(diagnosis_id = diagnosis.id, test__parent_test = None)
        precedure_set = PrecedureManager.objects.filter(diagnosis_id = diagnosis.id)
        medicine_set = MedicineManager.objects.filter(diagnosis_id = diagnosis.id)

        exams = []
        for data in exam_set:
            exam = {}
            exam.update({
                    'name':data.exam.name,
                })
            exams.append(exam)

        tests = []
        for data in test_set:
            test = {}
            test.update({
                    'name':data.test.name,
                    'amount':data.amount,
                    'days':data.days,
                    'memo':data.memo,
                })
            tests.append(test)

        precedures = []
        for data in precedure_set:
            precedure = {}
            precedure.update({
                    'name':data.precedure.name,
                    'amount':data.amount,
                    'days':data.days,
                    'memo':data.memo,
                })
            precedures.append(precedure)

        medicines = []
        for data in medicine_set:
            medicine = {}
            medicine.update({
                    'name':data.medicine.name,
                    'amount':data.amount,
                    'days':data.days,
                    'memo':data.memo,
                    'unit':data.medicine.unit,
                })
            medicines.append(medicine)

        data = {'date':diagnosis.recorded_date.strftime('%Y-%m-%d'),
        'day':diagnosis.recorded_date.strftime('%a'),
        'subjective':reception.chief_complaint,
        'objective':diagnosis.objective_data,
        'assessment':diagnosis.assessment,
        'plan':diagnosis.plan,
        'diagnosis':diagnosis.diagnosis,
        'ICD': diagnosis.ICD,
        'icd_code': diagnosis.ICD_code,
        'recommendation':diagnosis.recommendation,
                

        'doctor':reception.doctor.name_kor,

        'exams':exams,
        'tests':tests,
        'precedures':precedures,
        'medicines':medicines,
        'amount':'null',
        }

    except Diagnosis.DoesNotExist:
        res = False

    return JsonResponse({
        'result':res,

        'data':data,
    })



def customer_manage_get_patient_sms_info(request):
    patient_id = request.POST.get('patient_id')
        
    context={}
    patient = Patient.objects.get(pk=int(patient_id))

    context.update({
        'id':patient.id,
        'name_kor':patient.name_kor,
        'name_eng':patient.name_eng,
        'phone':patient.phone,
        })
    return JsonResponse(context)


def customer_manage_get_patient_save(request):

    patient_id = request.POST.get('patient_id')
    basic_info_name_kor = request.POST.get('basic_info_name_kor')
    basic_info_name_eng = request.POST.get('basic_info_name_eng')
    basic_info_dob = request.POST.get('basic_info_dob')
    basic_info_address = request.POST.get('basic_info_address')
    basic_info_phone = request.POST.get('basic_info_phone')
    basic_info_gender = request.POST.get('basic_info_gender')
    patient_nationality = request.POST.get('patient_nationality')
    basic_info_email = request.POST.get('basic_info_email')
    basic_info_passport = request.POST.get('basic_info_passport')
    basic_info_memo = request.POST.get('basic_info_memo')
    basic_info_mark = request.POST.get('basic_info_mark')
    basic_info_funnel = request.POST.get('basic_info_funnel')
    basic_info_funnel_etc = request.POST.get('basic_info_funnel_etc')

     
    patient = Patient.objects.get( pk=patient_id )

    patient.name_kor = basic_info_name_kor
    patient.name_eng = basic_info_name_eng
    patient.date_of_birth = basic_info_dob
    patient.address = basic_info_address 
    patient.phone = basic_info_phone
    patient.gender = basic_info_gender
    patient.nationality = patient_nationality
    patient.email = basic_info_email
    patient.passport = basic_info_passport
    patient.memo = basic_info_memo
    patient.marking = basic_info_mark
    patient.funnel = basic_info_funnel
    patient.funnel_etc = basic_info_funnel_etc
 
    patient.save()

    return JsonResponse({
        'result':True,
        })




def get_vaccine_history_list(request):
    
    patient_id = request.POST.get('patient_id')

    history_query = VaccineHistory.objects.filter(
        patient_id = patient_id,
        use_yn="Y",
        ).order_by('vaccine_name','round','vaccine_date')

    list_history = []
    for data in history_query:
        reservation = ''
        re_reservation = ''
        if data.reception != None:
            if data.reception.reservation != None:
                reservation = data.reception.reservation.reservation_date.strftime("%Y-%m-%d")
                if data.reception.reservation.re_reservation_date != None:
                    re_reservation = data.reception.reservation.re_reservation_date.strftime("%Y-%m-%d")

        list_history.append({
            'id':data.id,
            'vaccine_name':data.vaccine_name,
            'medicine_name':data.medicine_name,
            'round':data.round,
            'vaccine_date':data.vaccine_date.strftime("%Y-%m-%d"),
            'hospital':data.vaccine_hospital,
            'memo':data.memo,

            'reservation':reservation,#data.memo,
            're_reservation':re_reservation,#data.memo,
            })

    return JsonResponse({
        'result':True,
        'list_history':list_history,
        })



def get_vaccine_history(request):

    id = request.POST.get('id')

    history_query = VaccineHistory.objects.get(id = id )



    return JsonResponse({
        'result':True,
        'vaccine_name':history_query.vaccine_name,
        'medicine_name':history_query.medicine_name,
        'round':history_query.round,
        'vaccine_date':history_query.vaccine_date.strftime("%Y-%m-%d"),
        'vaccine_hospital':history_query.vaccine_hospital,
        'memo':history_query.memo,
        })


def vaccine_history_save(request):
    
    id = request.POST.get('id')
    patient_id = request.POST.get('patient_id')
    vaccine_name = request.POST.get('vaccine_name')
    medicine_name = request.POST.get('medicine_name')
    round = request.POST.get('round')
    vaccine_date = request.POST.get('vaccine_date')
    vaccine_hospital = request.POST.get('hospotal')
    memo = request.POST.get('memo')

    if id != '':
        history_query = VaccineHistory.objects.get(id = id )
    else:
        history_query = VaccineHistory()
        history_query.creator = request.user.id




    vaccine_date = datetime.datetime.combine(datetime.datetime.strptime(vaccine_date,"%Y-%m-%d").date(), datetime.time.min)
 
    history_query.patient_id = patient_id
    history_query.vaccine_name = vaccine_name
    history_query.medicine_name = medicine_name
    history_query.round = round
    history_query.vaccine_date = vaccine_date
    history_query.vaccine_hospital = vaccine_hospital
    history_query.memo = memo
    history_query.modifier = request.user.id
    history_query.modified_date = datetime.datetime.now()

    history_query.save()

    return JsonResponse({
        'result':True,
        'vaccine_name':history_query.vaccine_name,
        'medicine_name':history_query.medicine_name,
        'round':history_query.round,
        'vaccine_date':history_query.vaccine_date.strftime("%Y-%m-%d"),
        'memo':history_query.memo,
        })


def vaccine_history_delete(request):
    id = request.POST.get('id')

    history_query = VaccineHistory.objects.get(id = id )
    
    history_query.use_yn = 'N'
    history_query.modifier = request.user.id
    history_query.modified_date = datetime.datetime.now()

    history_query.save()

    return JsonResponse({
        'result':True,
        })

def manage_employee(request):


    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
        fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')



    #직급
    list_rank = []
    query_rank= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='RANK', upper_commcode='000002').annotate(code = F('commcode'),name = fname).values('code','name')
    for data in query_rank:
        list_rank.append({
            'id':data['code'],
            'name':data['name']
            })

    #관리자 파트
    list_depart = []
    query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_ADMIN',upper_commcode='000002').annotate(code = F('commcode'),name = fname).values('code','name')
    for data in query_depart:
            list_depart.append({
                'id':data['code'],
                'name':data['name']
                })


    if request.session['is_KBL']:
        #부서 - KBL
        query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL',upper_commcode='000002').annotate(code = F('commcode'),name = fname).values('code','name')
        for data in query_depart:
            list_depart.append({
                'id':data['code'],
                'name':data['name']
                })

        #직급 - KBL
        query_rank= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='RANK_KBL',upper_commcode='000002').annotate(code = F('commcode'),name = fname).values('code','name')
        for data in query_rank:
            list_rank.append({
                'id':data['code'],
                'name':data['name']
                })

    else:
        #부서 - 병원
        query_depart= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC',upper_commcode='000002').annotate(code = F('commcode'),name = fname).values('code','name','id')
        for data in query_depart:
            if data['code'] == 'DOCTOR':
                temp_commcode = COMMCODE.objects.get(id = data['id'])
                data['code'] += '_' + temp_commcode.se1
            list_depart.append({
                'id':data['code'],
                'name':data['name']
                })



    #사원 구분 
    list_division = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='EMPLOYEE DIVISION',upper_commcode='000003').annotate(code = F('commcode'),name = fname).values('code','name')
    for data in query_division:
        list_division.append({
            'id':data['code'],
            'name':data['name']
            })

    #재직 상태
    list_status = []
    query_status= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='EMPLOYEE STATUS',upper_commcode='000004').annotate(code = F('commcode'),name = fname).values('code','name')
    for data in query_status:
        list_status.append({
            'id':data['code'],
            'name':data['name']
            })

    return render(request,
        'Manage/employee_manage.html',
            {
                'list_rank':list_rank,
                'list_depart':list_depart,
                'list_division':list_division,
                'list_status':list_status,
            }
        )


def employee_search(request):
    string = request.POST.get('string')
    division_type = request.POST.get('division_type')
    depart_filter = request.POST.get('depart_filter')

    print('string: ', string)
    kwargs = {}
    # kwargs['is_active'] = True # 기본 
    if division_type is not '':
        kwargs['division_type'] = division_type
    if depart_filter is not '':
        if 'DOCTOR' in depart_filter:
            split_depart = depart_filter.split('_')

            kwargs['depart'] = 'DOCTOR'
            kwargs['depart_doctor'] = split_depart[1]
        else:
            kwargs['depart'] = depart_filter

    argument_list = []
    argument_list.append( Q(**{'user_id__icontains':string} ) )
    argument_list.append( Q(**{'name_en__icontains':string} ) )
    argument_list.append( Q(**{'name_ko__icontains':string} ) )
    argument_list.append( Q(**{'name_vi__icontains':string} ) )

    ex_list= []
    ex_list.append( Q(**{'depart':''} ) ) 
    if depart_filter == '':
        if not request.session['is_KBL']:
            depart_query= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL').values('commcode')
        else:
            depart_query= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC').values('commcode')

        for data in depart_query:
            ex_list.append( Q(**{'depart':data['commcode']} ) )
    

    query_user = User.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs
        ).exclude(
            functools.reduce(operator.or_, ex_list),
            )


    list_user = []
    for user in query_user:
       list_user.append({
               'id':user.id,
               'division_type':'' if user.division_type is None else user.division_type,
               'depart':'' if user.depart is None else user.depart,
               'rank':'' if user.rank is None else user.rank,
               'user_id':'' if user.user_id is None else user.user_id,
               'name_ko':'' if user.name_ko is None else user.name_ko,
               'name_en':'' if user.name_en is None else user.name_en,
               'name_vi':'' if user.name_vi is None else user.name_vi,
               'gender':'' if user.gender is None else user.gender,
               'date_of_birth':'' if user.date_of_birth is None else user.date_of_birth,
               'phone_number1':'' if user.phone_number1 is None else user.phone_number1,
               'phone_number2':'' if user.phone_number2 is None else user.phone_number2,
               'email':'' if user.email is None else user.email,
               'date_of_employment':'' if user.date_of_employment is None else user.date_of_employment,
               'status':'' if user.status is None else user.status,
               
           })
    
    page_context = request.POST.get('context_in_page',10)
    page = request.POST.get('page',1)
    paginator = Paginator(list_user, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)



    return JsonResponse({
        'result':True,
        'datas':list(paging_data),

        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),


        })




@login_required
def employee_check_id(request):
       
    user_id = request.POST.get('user_id')

    try:
        user = User.objects.get(user_id = user_id)
        result = False
    except User.DoesNotExist:
        result = True

    return JsonResponse({
        'result':result,
        })



@login_required
def employee_add_edit(request):

    id = request.POST.get('id')

    user_id = request.POST.get('user_id')
    password = request.POST.get('password')
    name_ko = request.POST.get('name_ko')
    name_en = request.POST.get('name_en')
    name_vi = request.POST.get('name_vi')
    gender = request.POST.get("gender")
    phone1 = request.POST.get('phone1')
    phone2 = request.POST.get('phone2')
    date_of_birth = request.POST.get('date_of_birth')
    email = request.POST.get('email')
    address = request.POST.get('address')

    rank = request.POST.get('rank')
    depart = request.POST.get('depart')
    division = request.POST.get('division')

    status = request.POST.get('status')
    date_of_employment = request.POST.get('date_of_employment')
    remark = request.POST.get('remark')

    is_new = False
    try:
        user = User.objects.get(pk = id)
        user.lastest_modified_date = datetime.datetime.now()

    except User.DoesNotExist:
        user = User()
        user.set_password(password)
        is_new = True
    

    user.user_id = user_id
    #기본 정보
    
    user.name_ko = name_ko
    user.name_en = name_en
    user.name_vi = name_vi
    user.gender = gender
    user.phone_number1 = phone1
    user.phone_number2 = phone2
    user.date_of_birth = date_of_birth
    user.email = email
    user.address = address

    user.status = status
    user.date_of_employment = date_of_employment
    user.memo = remark


    #권한 설정
    user.rank = rank
    user.depart = depart
    user.division_type = division

    if user.depart == 'ADMIN':
        user.is_superuser = True
        user.is_staff= True

    #권한 설정 #superuser / staff / 
    user.save()

    
    User_Menu.objects.filter(user_id = user.id,menu = 'doctor').delete()
    User_Menu.objects.filter(user_id = user.id,menu = 'med_rep').delete()

    if "DOCTOR" in depart: #의사 권한은 별도로 ... 
        str_split = depart.split('_')
        
        commcode = COMMCODE.objects.get(upper_commcode = '000002', commcode_grp='DEPART_CLICINC', commcode='DOCTOR',se1 = str_split[1])
       
        user.depart_doctor = str_split[1]
        user.depart = str_split[0]


        try:#의사 정보 변경
            doctor = Doctor.objects.get(user_id=user.id)
        except Doctor.DoesNotExist:#의사 신규
            doctor = Doctor()

        depart = Depart.objects.get(name = user.depart_doctor)
        doctor.depart = depart
        doctor.name_kor = name_en
        doctor.name_eng = name_en
        doctor.name_short = name_en
        doctor.user_id = user.id

        user.save()
        doctor.save()

        #의사는 페이지 강제 추가
        User_Menu(user_id = user.id, menu = 'doctor',creator = request.user,created_date =datetime.datetime.now() ,seq = 100,).save()
        User_Menu(user_id = user.id, menu = 'med_rep',creator = request.user,created_date =datetime.datetime.now() ,seq = 100,).save()

    
    #draft 와 Board는 모든 계정 강제 추가
    if is_new:
        User_Menu(user_id = user.id, menu = 'draft',creator = request.user,created_date =datetime.datetime.now() ,seq = 1500,).save()
        User_Menu(user_id = user.id, menu = 'board_board',creator = request.user,created_date =datetime.datetime.now() ,seq = 1600, grp = 'BOARD').save()
        User_Menu(user_id = user.id, menu = 'board_coboard',creator = request.user,created_date =datetime.datetime.now() ,seq = 1600, grp = 'BOARD').save()

    return JsonResponse({
        'result':True,
        })

@login_required
def employee_add_edit_get(request):
     
    id = request.POST.get('id')

    user = User.objects.get(pk = id)

    if user.depart == 'DOCTOR':
        depart = user.depart + "_" + user.depart_doctor


    else:
        depart = user.depart

    return JsonResponse({
        'result':True,

        "user_id" : user.user_id,
        "name_ko" : user.name_ko,
        "name_en" : user.name_en,
        "name_vi" : user.name_vi,
        "gender" : user.gender,
        "phone1" : user.phone_number1,
        "phone2" : user.phone_number2,
        "date_of_birth" : user.date_of_birth,
        "email" : user.email,
        "address" : user.address,

        "rank" : user.rank,
        "depart" : depart,
        "division" : user.division_type,

        "status" : user.status,
        "date_of_employment" : user.date_of_employment,
        "remark" : user.memo,

        })




@login_required
def employee_delete(request):
    id = request.POST.get('id')

    date_of_resignation = request.POST.get('date_of_resignation')
    resignation_reason = request.POST.get('resignation_reason')

    user = User.objects.get(pk = id)

    user.is_active = False

    user.status = 'RESIGNED'
    user.date_of_resignation = date_of_resignation
    user.resignation_memo = resignation_reason

    user.lastest_modified_date = datetime.datetime.now()

    user.save()

    return JsonResponse({
        'result':True,
        
        })


@login_required
def employee_change_password(request):
    print('employee_change_password: ')
    id = request.POST.get('id')
    password = request.POST.get('password')

    user = User.objects.get(pk = id)
    user.set_password(password)
    user.save()

    return JsonResponse({
        'result':True,
        })



@login_required
def employee_menu_get(request):

    user_id = request.POST.get('user_id')

    datas = []
    menu_query = User_Menu.objects.filter(user = user_id).values('menu')
    for data in menu_query:
        datas.append(data['menu'])


    return JsonResponse({
        'result':True,
        'datas':datas,
        })


@login_required
def employee_menu_save(request):
      
    user_id = request.POST.get('user_id')
    checked_array = request.POST.getlist('checked_array[]')

    User_Menu.objects.filter(user_id = user_id).delete()
    now = datetime.datetime.now()

    for data in checked_array: 
        data_split = data.split('menu_')
        commcode = COMMCODE.objects.get(upper_commcode ='000012',commcode_grp = 'MENU_LIST_IMDI',commcode = data_split[1])
        seq = int(commcode.se2)
        if commcode.se3 is not '':
            seq += int(commcode.se3)/10

        User_Menu(
            user_id = user_id, 
            menu = data_split[1],
            creator = request.user,
            created_date = now,
            seq = seq,
            grp = commcode.se4
            ).save()

    #draft 와 Board는 모든 계정 강제 추가
    User_Menu(user_id = user_id, menu = 'draft',creator = request.user,created_date =now ,seq = 1500,).save()
    User_Menu(user_id = user_id, menu = 'board_board',creator = request.user,created_date =now ,seq = 1600, grp = 'BOARD').save()
    User_Menu(user_id = user_id, menu = 'board_coboard',creator = request.user,created_date =now ,seq = 1600, grp = 'BOARD').save()


    return JsonResponse({
        'result':True,
        })

@login_required
def board_list(request,id=None):

    if id == None:
        id = request.POST.get('selected_content',None)
    
    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
         fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')
    #초기 - Division
    list_division = COMMCODE.objects.filter(upper_commcode = '000005',commcode_grp = 'BOARD_DIVISION')

    dict_division = {}
    for division in list_division.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_division.update({
            division['code'] : division['name']
            })



    #content
    content = None;
    #게시글 오픈 유무
    if id is not '':
        try:
            read_page = Board_Contents.objects.get(id = id )
            creator = User.objects.get(id = int(read_page.creator))
            #조회수 카운트
            ##자기 자신을 제외한 유저에만 해당
            if creator.id != request.user.id:
                date_min = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.min)
                date_max = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.max)

                is_new = Board_View_Log.objects.filter(registered_date__range = (date_min, date_max),user_id = request.user.id, board_id = read_page.id, ).count()
                if is_new == 0 : #오늘 안봤으면 추가
                    new_view = Board_View_Log()
                    new_view.board_id = read_page.id
                    new_view.user_id = request.user.id
                    new_view.save()

                    read_page.view_count += 1
                    read_page.save()

            #페이지 상세 정보 
            
            
            comments_count = Board_Comment.objects.filter(content_id = id, use_yn = 'Y').count()
            content = {
                'id':read_page.id,
                'title':read_page.title,
                'contents':read_page.contents,
                'creator':creator.user_id,
                'date':read_page.created_date.strftime('%Y-%m-%d %H:%M'),
                'views': read_page.view_count,
                'comments_count':comments_count,
                }

            query_file = Board_File.objects.filter(board_id = id,board_type='BASIC')

            list_file = []
            for file in query_file:
                list_file.append({
                    'id':file.id,
                    'file_name':file.file.url,
                    'origin_name':file.origin_name,
                    })
            content.update({
                'list_file': list_file,
                'list_file_count':query_file.count(),
                })

        except Board_Contents.DoesNotExist:
            content = None;



   

    #search filter
    kwargs = {}
    kwargs['use_yn'] = 'Y' # 기본 
    kwargs['board_type'] = 'BASIC' #일반 게시판

    kwargs['is_KBL'] = 'N'
    if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
        kwargs['is_KBL'] = 'Y'


    search_string = request.POST.get('search_string','')
    view_division_filter = request.POST.get('view_division_filter','')
    if view_division_filter != '':
        kwargs['options'] = view_division_filter

    users = User.objects.all()
    if search_string == '':
        query = Board_Contents.objects.filter(**kwargs).order_by('-top_seq','-created_date')

    else:
        argument_list = [] 

        argument_list.append( Q(**{'title__icontains':search_string} ) ) 
        argument_list.append( Q(**{'contents__icontains':search_string} ) ) 

        #list_search_user = users.filter()
        query = Board_Contents.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).order_by('-created_date')


    #목록 정렬
    contents_list = []
    for item in query:
        comment_count = Board_Comment.objects.filter(content_id = item.id, use_yn = 'Y').count()
        creator = users.get(id = int(item.creator))
        file_count = Board_File.objects.filter(board_id = item.id,board_type='BASIC').count()


        contents_list.append({
            'id':item.id,
            'is_notice':item.top_seq,
            'division':dict_division[item.options],
            'title':item.title,
            'creator':creator.user_id,
            'date':item.created_date.strftime('%Y-%m-%d %H:%M'),
            'comment_count':comment_count,
            'is_file':False if file_count is 0 else True,
            'view_count':item.view_count,
            })


    #페이지네이션
    page = request.POST.get('page',1)
    view_contents_count = request.POST.get('view_contents_count',10);
    paginator = Paginator(contents_list, view_contents_count)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    return render(request,
        'board/list.html',
            {
                'content_count':query.count(),
                'content':content,
                #'contents_list':contents_list,
                'contents_list':paging_data,

                'search_string':search_string,
                'view_division_filter':view_division_filter,
                'view_contents_count':view_contents_count,
                'dict_division':dict_division,

                'page':int(page),
                'page_range':list( range(paging_data.paginator.page_range.start, paging_data.paginator.page_range.stop) ) ,
                'page_range_start':paging_data.paginator.page_range.start,
                'page_range_stop':paging_data.paginator.page_range.stop,
                'page_number':paging_data.number,
                'has_previous':paging_data.has_previous(),
                'has_next':paging_data.has_next(),
            }
        )



@login_required
def board_create_edit(request,id=None):
    
    division_selected= None
    option_err=''
    is_top= '0'
    #language
    lang = ''
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        lang = 'vi-VN'
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        lang = 'ko-KR'
    else:
        lang = 'en-US'
    
    list_file = []
    if id is None: # 새글
        load_contents = Board_Contents()
        form = board_form()

        file_form = board_file_form()

        load_contents.creator = request.user.id
        if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
            load_contents.is_KBL = 'Y' 
    else: # 글 수정 
        load_contents = Board_Contents.objects.get(id = id)
        form = board_form(instance = load_contents)
        division_selected = load_contents.options

        is_top = load_contents.top_seq

        file_form = board_file_form()
        query_file = Board_File.objects.filter(board_id = id,board_type='BASIC')
        for file in query_file:
            list_file.append({
                'id':file.id,
                'file_name':file.file.url,
                'origin_name':file.origin_name,
                })


    #저장
    if request.method == 'POST':
        form = board_form(request.POST)
        file_form = board_file_form(request.POST, request.FILES)   
        files = request.FILES.getlist('file') 


        select_valid = False
        division_selected = request.POST.get('select_division',None)
        if division_selected is not None:
            select_valid = True


        if form.is_valid() and file_form.is_valid() and select_valid:
            load_contents.board_type = 'BASIC'
            load_contents.title = form.cleaned_data['title']
            load_contents.contents = form.cleaned_data['contents']
            load_contents.options = division_selected

            is_notice = request.POST.get('top_seq','off')
            load_contents.top_seq = 0
            if is_notice == 'on':
                load_contents.top_seq = 1
   

            
            load_contents.lastest_modifier = request.user.id
            load_contents.lastest_modified_date = datetime.datetime.now()

            load_contents.save()

            #file save
            for f in files:
                file_instance = Board_File(file=f, board_id = load_contents.pk)
                file_instance.origin_name = f.name
                file_instance.board_type = 'BASIC'
                file_instance.user = request.user.user_id

                file_instance.save()
 

            if id is None:
                return HttpResponseRedirect('./../' + str(load_contents.pk))
            else:
                return HttpResponseRedirect('./../../' + str(load_contents.pk))
        else:
            pass#form = board_form(instance=profile)




    
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    list_division = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_DIVISION').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_division.append({
            'id':data['code'],
            'name':data['name']
            })
    

    return render(request,
        'board/create_edit.html',
            {
                'form':form,
                'file_form':file_form,
                'lang':lang,

                'list_file': list_file if list_file else None,
                'list_file_count':len(list_file),

                'list_division':list_division,
                'division_selected':division_selected, #division
                'option_err':option_err,
                'is_top':is_top
            }
        )



@login_required
def board_delete(request,id):

    try:
        content = Board_Contents.objects.get(id = id)
        content.use_yn = 'N'
        content.save()

    except Board_Contents.DoesNotExist:
        pass

    try:
        alert_query = AlertLog.objects.get(
            page_type="COWORK",
            content_type="COMMENT",
            content_id = id,
            )

        for data in alert_query:
            data.use_yn = 'N'
            data.save()
    except AlertLog.DoesNotExist:
        pass





    return HttpResponseRedirect('./../')


@login_required
def board_delete_file(request):
    id = request.POST.get('id')

    file = Board_File.objects.get(pk = id)


    file.delete()


    return JsonResponse({
        'result':True,
        })



@login_required
def board_comment_get(request):
    content_id = request.POST.get('content_id')

    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
         fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')

    # - Status
    dict_status ={}
    query_status= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_STATUS',upper_commcode ='000005' )
    for data in query_status.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_status.update({
            data['code'] : data['name']
            })

    #set user data
    user_dict = {}
    users = User.objects.all().values('id','user_id')
    for user in users:
        user_dict[user['id']] = user['user_id']

    list_comment = []
    query = Board_Comment.objects.filter(content_id = content_id,use_yn = 'Y').order_by('orderno').values()

    
    for data in query:
        comment = {}
        incharge = ''
        status = ''
        if data['in_charge'] != '':
            incharge = User.objects.get(pk = data['in_charge']).user_id
        if data['progress'] != '':
            status = dict_status[data['progress']]

        comment.update({
            'id':data['id'],
            'user_id':data['creator'],
            'user':user_dict[ int(data['creator']) ],
            'comment':data['comment'],
            'datetime':data['created_date'].strftime('%Y-%m-%d %H:%M'),
            'depth':data['depth'],

            'in_charge':incharge,
            'start_date':'' if data['start_date'] == '0000-00-00 00:00:00' else data['start_date'],
            'end_date':'' if data['end_date'] == '0000-00-00 00:00:00' else data['end_date'],
            'expected_date':'' if data['expected_date'] == '0000-00-00 00:00:00' else data['expected_date'],
            'status':status,
            })

        if request.user.id is int(data['creator']):
            comment.update({'is_creator':True})
        else:
            comment.update({'is_creator':False})

        list_comment.append(comment)




    return JsonResponse({
        'result':True,
        'list_comment':list_comment,
        'conntents_count':query.count(),
        })


def board_comment_set_seq(input_data):
    



    return

def board_comment_add(request,comment_id=None):
    content_id = request.POST.get('content_id')
    comment = request.POST.get('comment','')

    new_comment = Board_Comment()
    
    if upper_id != '':#대댓글
        upper = Board_Comment.objects.get(pk = upper_id)
        new_comment.depth = upper.depth + 1
        
        try:
            check = Board_Comment.objects.filter(content_id = content_id, orderno__gt = upper.orderno, depth__lte = upper.depth,).order_by('orderno')[:1]
            
            new_comment.orderno = check[0].orderno
        except IndexError:
            check = Board_Comment.objects.filter(content_id = content_id, ).order_by('-orderno')[:1]
           
            new_comment.orderno = check[0].orderno + 1
        

        #저장 전 sequence 자리 비우기
        query_set = Board_Comment.objects.filter(content_id = content_id, orderno__gte = new_comment.orderno ,).order_by('orderno')
        
        for query in query_set:
            query.orderno += 1
            
            query.save()

          
    else:#첫번째 뎁스 댓글
        check = Board_Comment.objects.filter(content_id = content_id, ).order_by('orderno')
        
        if check.count() != 0:
            check_last = check.last()
            new_comment.orderno = check_last.orderno + 1

        new_comment.in_charge = select_user
        new_comment.start_date = start_date
        new_comment.end_date = due_date
        new_comment.expected_date = expected_date
        new_comment.progress = status


    new_comment.content_id = content_id
    new_comment.comment = comment
    new_comment.creator = request.user.id
    new_comment.lastest_modifier = request.user.id
    new_comment.last_modified_date = datetime.datetime.now()

    new_comment.save()


    

    return JsonResponse({
        'result':True,
        })






def board_comment_edit(request,id):
    content_content = request.POST.get('comment')

    comment = Board_Comment.objects.get(id = id)
    comment.comment = content_content
    comment.lastest_modifier = request.user.id
    comment.last_modified_date = datetime.datetime.now()
    comment.save()

    return JsonResponse({
        'result':True,
        })


def board_comment_delete(request,id):#Comment ID
    content_id = request.POST.get('content_id')

    comment = Board_Comment.objects.get(id=id)
    comment.use_yn = 'N'
    comment.save()


    return JsonResponse({
        'result':True,
        })








@login_required
def board_work_list(request,id=None):
    def_date = '0000-00-00 00:00:00'


    if id == None:
        id = request.POST.get('selected_content',None)
    
    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
        fname = F('commcode_name_ko')
        user_name = F('name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
        user_name = F('name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')
        user_name = F('name_vi')


    #초기 - Division
    list_division = COMMCODE.objects.filter(upper_commcode = '000005',commcode_grp = 'BOARD_WORK_DIVISION')

    dict_division = {}
    for division in list_division.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_division.update({
            division['code'] : division['name']
            })

    # - Depart
    dict_depart = {}
    list_query_depart = []
    if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
        ##경천
        query_depart_kbl= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL',upper_commcode ='000002' )
        for data in query_depart_kbl.annotate(code = F('commcode'),name = fname).values('code','name'):
            dict_depart.update({
                data['code'] : data['name']
                })
            list_query_depart.append( Q(**{'depart':data['code']} ) ) 

    else:
        ##IMEDI
        query_depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC',upper_commcode ='000002' )
        for data in query_depart_medical.annotate(code = F('commcode'),name = fname).values('code','name','id'):
            list_query_depart.append( Q(**{'depart':data['code']} ) ) 
            if data['code'] == 'DOCTOR':
                temp_commcode = COMMCODE.objects.get(id = data['id'])
                data['code'] += '_' + temp_commcode.se1
            dict_depart.update({
                data['code'] : data['name']
                })
            
    ##관리자
    query_depart_admin= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_ADMIN',upper_commcode ='000002' )
    for data in query_depart_admin.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_depart.update({
            data['code'] : data['name']
            })
        list_query_depart.append( Q(**{'depart':data['code']} ) ) 

    # - Status
    dict_status ={}
    query_status = COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_STATUS',upper_commcode ='000005' )
    for data in query_status.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_status.update({
            data['code'] : data['name']
            })

    #set user data
    user_dict = {}
    users = User.objects.filter(
        functools.reduce(operator.or_, list_query_depart)
        ).annotate(name = user_name).values('id','user_id','name','depart','depart_doctor')
    for user in users:
        depart_name = user['depart']
        if user['depart'] == 'DOCTOR':
            depart_name = user['depart'] + '_' + user['depart_doctor']
        
        user_dict.update({
            user['id'] : {
                'user_id':user['user_id'],
                'name':user['name'],
                'depart':dict_depart[depart_name],
                }
            })



    #content
    content = None;
    #게시글 오픈 유무
    if id != '' and id != None:
        try:
            read_page = Board_Contents.objects.get(id = id )
            creator = User.objects.get(id = int(read_page.creator))
            #조회수 카운트
            ##자기 자신을 제외한 유저에만 해당
            if creator.id != request.user.id:
                #date_min = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.min)
                #date_max = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.max)

                is_new = Board_View_Log.objects.filter(user_id = request.user.id, board_id = read_page.id).count()
                if is_new == 0 : #오늘 안봤으면 추가
                    new_view = Board_View_Log()
                    new_view.board_id = read_page.id
                    new_view.user_id = request.user.id
                    new_view.save()

                    read_page.view_count += 1
                    read_page.save()

            #페이지 상세 정보 

            ##요청 받은 과의 사용자 지정
            selected_depart_list=[]
            request_user_dict = {}
            if read_page.depart_to1 != '':
                selected_depart_list.append( Q(**{'depart':read_page.depart_to1} ) ) 
            if read_page.depart_to2 != '':
                selected_depart_list.append( Q(**{'depart':read_page.depart_to2} ) ) 
            if read_page.depart_to3 != '':
                selected_depart_list.append( Q(**{'depart':read_page.depart_to3} ) ) 
            if read_page.depart_to4 != '':
                selected_depart_list.append( Q(**{'depart':read_page.depart_to4} ) ) 


            
            t_user_name = F('name_en')
            if current_language == 'ko':
                t_user_name = F('name_ko')
            elif current_language == 'vi':
                t_user_name = F('name_vi')

            tmp_users= users.filter(functools.reduce(operator.or_, selected_depart_list))
            for user in tmp_users:
                if not dict_depart[user['depart']] in request_user_dict:
                    request_user_dict[dict_depart[user['depart']]] = []
                request_user_dict[dict_depart[user['depart']]].append({
                    'id':user['id'],
                    'user_id':user['user_id'],
                    'name':user['name'],
                    'depart':user['depart'],
                    })

            ##데이터 가공
            comments_query= Board_Comment.objects.filter(content_id = id,).order_by('orderno')
            list_comment_id = comments_query.values('id')

            content = {
                'id':read_page.id,
                'title':read_page.title,
                'contents':read_page.contents,
                'creator_id':int(read_page.creator),
                'creator':user_dict[int(read_page.creator)]['name'],
                'depart':user_dict[int(read_page.creator)]['depart'],
                'date':read_page.created_date.strftime('%Y-%m-%d %H:%M'),
                'views': read_page.view_count,
                'comments_count':comments_query.filter(use_yn = 'Y').count(),
                #'expected_date':'' if read_page.date_to_be_done == def_date else read_page.date_to_be_done,
                'due_date':'' if read_page.date_done == def_date else read_page.date_done,
                
                'request_user_dict':request_user_dict,

                }
            selected_status = read_page.status
            query_file = Board_File.objects.filter(board_id = id,board_type='COWORK')

            list_file = []
            for file in query_file:
                list_file.append({
                    'id':file.id,
                    'file_name':file.file.url,
                    'origin_name':file.origin_name,
                    })
            content.update({
                'list_file': list_file,
                'list_file_count':query_file.count(),
                'selected_status':selected_status,
                })

            list_comment = []
            comment_top = 0
            #코멘트
            for data in comments_query:
                comment = {}
                incharge = ''
                in_charge_id=''
                status = ''
                status_id = ''

                if data.in_charge != '':
                    incharge = User.objects.get(pk = data.in_charge).user_id
                    in_charge_id = int(data.in_charge)
                if data.progress != '':
                    status = dict_status[data.progress]
                    status_id = data.progress

                if data.depth == 0:
                    comment_top = data.id

                    
                start_date = ''
                expected_date = ''
                end_date = ''

                if data.start_date != '0000-00-00 00:00:00':
                    start_date = datetime.datetime.strptime(data.start_date,'%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %I %p')
                if data.expected_date != '0000-00-00 00:00:00':
                    expected_date = datetime.datetime.strptime(data.expected_date,'%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %I %p')
                if data.end_date != '0000-00-00 00:00:00':
                    end_date = datetime.datetime.strptime(data.end_date,'%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %I %p')


                comment.update({
                    'id':data.id,
                    'user_id':int(data.creator),
                    'user':user_dict[ int(data.creator) ]['user_id'],
                    'name':user_dict[ int(data.creator) ]['name'],
                    'depart':user_dict[ int(data.creator) ]['depart'],
                    'comment':data.comment,
                    'datetime':data.created_date.strftime('%Y-%m-%d %H:%M'),
                    'depth':range(data.depth),

                    
                    'emoji1':data.emoji1,
                    'emoji2':data.emoji2,
                    'emoji3':data.emoji3,
                    'emoji4':data.emoji4,

                    'in_charge':incharge,
                    'in_charge_id':in_charge_id,
                    'start_date':start_date,
                    'end_date':end_date,
                    'expected_date':expected_date,
                    'status':status, 
                    'status_id':status_id,

                    'use_yn':True if data.use_yn == 'Y' else False,

                    'comment_top':comment_top,
                    })

                if request.user.id is int(data.creator):
                    comment.update({'is_creator':True})
                else:
                    comment.update({'is_creator':False})

                list_comment.append(comment)

            content.update({
                'list_comment':list_comment,
                })

            
            for comment_id in list_comment_id:
                alert_query = AlertLog.objects.filter(
                        page_type="COWORK",
                        content_type="COMMENT",
                        content_id = comment_id['id'],
                        user_id=request.user.id,
                    )
                for alert in alert_query:
                    alert.check_yn ='Y'
                    alert.save()


             
        except (Board_Contents.DoesNotExist,ValueError):
            content = None;



    #search filter
    argument_list = [] 
    kwargs = {}
    kwargs['use_yn'] = 'Y' # 기본 
    kwargs['board_type'] = 'COWORK' #협업 게시판

    search_string = request.POST.get('search_string','')
    view_division_filter = request.POST.get('view_division_filter','')
    if view_division_filter != '':
        kwargs['options'] = view_division_filter

    view_status_filter = request.POST.get('view_status_filter','')
    if view_status_filter != '':
        kwargs['status'] = view_status_filter

    view_depart_filter = request.POST.get('view_depart_filter','')
    if view_depart_filter != '':
        kwargs['depart_from'] = view_depart_filter
        


    kwargs['is_KBL'] = 'N'
    if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
        kwargs['is_KBL'] = 'Y'

    user_depart = request.user.depart
    if user_depart == 'DOCTOR':
        user_depart += '_' + request.user.depart_doctor

    #모든 파트 ALL 넣기
    argument_list.append( Q(**{'depart_to1':'ALL'} ) ) 
    argument_list.append( Q(**{'depart_to2':'ALL'} ) ) 
    argument_list.append( Q(**{'depart_to3':'ALL'} ) ) 
    argument_list.append( Q(**{'depart_to4':'ALL'} ) ) 
    if request.user.depart != 'ADMIN':
        argument_list.append( Q(**{'depart_to1':user_depart} ) ) 
        argument_list.append( Q(**{'depart_to2':user_depart} ) ) 
        argument_list.append( Q(**{'depart_to3':user_depart} ) ) 
        argument_list.append( Q(**{'depart_to4':user_depart} ) ) 
        argument_list.append( Q(**{'depart_from':user_depart} ) ) 

    if search_string != '':
        argument_list.append( Q(**{'title__icontains':search_string} ) ) 
        argument_list.append( Q(**{'contents__icontains':search_string} ) ) 
    
    if request.user.depart == 'ADMIN' and search_string == '':
        query = Board_Contents.objects.filter(**kwargs).order_by('-created_date')
    else:
        query = Board_Contents.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).order_by('-created_date')
    users = User.objects.all()

    #목록 정렬
    contents_list = []
    for item in query:
        comment_count = Board_Comment.objects.filter(content_id = item.id, use_yn = 'Y').count()
        creator = users.get(id = int(item.creator))
        file_count = Board_File.objects.filter(board_id = item.id,board_type='COWORK').count()


        #댓글 불러오기
        list_comment = []
        comment_query = Board_Comment.objects.filter(content_id = item.id,use_yn = 'Y').order_by('orderno').values()
        no=0
        
        comment_top=0
        for data in comment_query:
            
            incharge = ''
            in_charge_id=''
            status = ''
            status_id = ''

            if data['in_charge'] != '':
                incharge = User.objects.get(pk = data['in_charge']).user_id
                in_charge_id = int(data['in_charge'])
            if data['progress'] != '':
                status = dict_status[data['progress']]
                status_id = data['progress']

            if data['depth'] == 0:
                comment_top = data['id']

                    
            start_date = ''
            expected_date = ''
            end_date = ''

            if data['start_date']!= '0000-00-00 00:00:00':
                start_date = datetime.datetime.strptime(data['start_date'],'%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %I %p')
            if data['expected_date'] != '0000-00-00 00:00:00':
                expected_date = datetime.datetime.strptime(data['expected_date'],'%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %I %p')
            if data['end_date'] != '0000-00-00 00:00:00':
                end_date = datetime.datetime.strptime(data['end_date'],'%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %I %p')

            comment = {}
            comment.update({
                'no':no,
                'id':data['id'],
                'user_id':data['creator'],
                'user':user_dict[ int(data['creator']) ],
                'name':user_dict[ int(data['creator']) ]['name'],
                'depart':user_dict[ int(data['creator']) ]['depart'],
                'comment':data['comment'],
                'datetime':data['created_date'].strftime('%Y-%m-%d %H:%M'),
                'depth':data['depth'],
                'orderno':data['orderno'],

                
                'emoji1':data['emoji1'],
                'emoji2':data['emoji2'],
                'emoji3':data['emoji3'],
                'emoji4':data['emoji4'],

                'in_charge':incharge,
                'in_charge_id':in_charge_id,
                'start_date':start_date,
                'end_date':end_date,
                'expected_date':expected_date,
                'status':status, 
                'status_id':status_id,

                'comment_top':comment_top,
                })
            no +=1

            if request.user.id is int(data['creator']):
                comment.update({'is_creator':True})
            else:
                comment.update({'is_creator':False})

            list_comment.append(comment)


        contents_list.append({
            'id':item.id,
            'is_notice':item.top_seq,
            'division':dict_division[item.options],
            'title':item.title,
            'creator':creator.user_id,
            'date':item.created_date.strftime('%Y-%m-%d'),
            'depart_from':dict_depart[item.depart_from],
            

            'depart_to1':'' if item.depart_to1 == '' else 'ALL' if item.depart_to1 == 'ALL' else dict_depart[item.depart_to1],
            'depart_to2':'' if item.depart_to2 == '' else 'ALL' if item.depart_to2 == 'ALL' else dict_depart[item.depart_to2],
            'depart_to3':'' if item.depart_to3 == '' else 'ALL' if item.depart_to3 == 'ALL' else dict_depart[item.depart_to3],
            'depart_to4':'' if item.depart_to4 == '' else 'ALL' if item.depart_to4 == 'ALL' else dict_depart[item.depart_to4],

            #'depart_user_to1':_('All') if item.depart_user_to1 == 'all' else '' if item.depart_user_to1 == '' else user_dict[int(item.depart_user_to1)],
            #'depart_user_to2':_('All') if item.depart_user_to2 == 'all' else '' if item.depart_user_to2 == '' else user_dict[int(item.depart_user_to2)],
            #'depart_user_to4':_('All') if item.depart_user_to3 == 'all' else '' if item.depart_user_to3 == '' else user_dict[int(item.depart_user_to3)],
            #'depart_user_to3':_('All') if item.depart_user_to4 == 'all' else '' if item.depart_user_to4 == '' else user_dict[int(item.depart_user_to4)],

            'status':item.status,
            'is_file':False if file_count is 0 else True,
            'view_count':item.view_count,
            
            'list_comment':list_comment,
            'count_comment':len(list_comment),
            #'expected_date':'' if item.date_to_be_done == def_date else item.date_to_be_done,
            #'due_date':'' if item.date_done == def_date else item.date_done,

            #'comment_count':comment_count,
            })


    #페이지네이션
    page = request.POST.get('page',1)
    view_contents_count = request.POST.get('view_contents_count',10);
    paginator = Paginator(contents_list, view_contents_count)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    return render(request,
        'board_work/list.html',
            {
                'content_count':query.count(),
                'content':content,
                #'contents_list':contents_list,
                'contents_list':paging_data,

                'search_string':search_string,
                'view_division_filter':view_division_filter,
                'view_depart_filter':view_depart_filter,
                'view_status_filter':view_status_filter,
                'view_contents_count':view_contents_count,
                'dict_division':dict_division,
                'dict_status':dict_status,
                'dict_depart':dict_depart,


                'page':int(page),
                'page_range':list( range(paging_data.paginator.page_range.start, paging_data.paginator.page_range.stop) ) ,
                'page_range_start':paging_data.paginator.page_range.start,
                'page_range_stop':paging_data.paginator.page_range.stop,
                'page_number':paging_data.number,
                'has_previous':paging_data.has_previous(),
                'has_next':paging_data.has_next(),
            }
        )



@login_required
def board_work_create_edit(request,id=None):
    option_err=''
    is_top= '0'
    division_selected = None
    depart_to_selected = None
    depart_selected1 = ''
    depart_selected2 = ''
    depart_selected3 = ''
    depart_selected4 = ''
    status_selected = ''

    depart_user_selected1 = ''
    depart_user_selected2 = ''
    depart_user_selected3 = ''
    depart_user_selected4 = ''
    #language
    
    list_file = []
    if id is None: # 새글
        load_contents = Board_Contents()
        form = board_form()

        file_form = board_file_form()

        load_contents.creator = request.user.id
        load_contents.depart_from = request.user.depart
        if request.user.depart =='DOCTOR':
            load_contents.depart_from = request.user.depart + "_" + request.user.depart_doctor

            
        if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
            load_contents.is_KBL = 'Y'
    else: # 글 수정 
        load_contents = Board_Contents.objects.get(id = id)
        form = board_form(instance = load_contents)
        division_selected = load_contents.options
        depart_selected1 = load_contents.depart_to1
        depart_selected2 = load_contents.depart_to2
        depart_selected3 = load_contents.depart_to3
        depart_selected4 = load_contents.depart_to4

        depart_user_selected1 = load_contents.depart_user_to1
        depart_user_selected2 = load_contents.depart_user_to2
        depart_user_selected3 = load_contents.depart_user_to3
        depart_user_selected4 = load_contents.depart_user_to4

        status_selected = load_contents.status


        is_top = load_contents.top_seq

        file_form = board_file_form()
        query_file = Board_File.objects.filter(board_id = id,board_type = 'COWORK')
        for file in query_file:
            list_file.append({
                'id':file.id,
                'file_name':file.file.url,
                'origin_name':file.origin_name,
                })


    #저장
    if request.method == 'POST':
        form = board_form(request.POST)
        file_form = board_file_form(request.POST, request.FILES)   
        files = request.FILES.getlist('file') 

        division_select_valid = False
        division_selected = request.POST.get('select_division',None)
        if division_selected is not None:
            division_select_valid = True

        #depart_to_select_valid = False
        #depart_to_selected = request.POST.get('select_depart_to',None)
        #if depart_to_selected is not None:
        #    depart_to_select_valid = True

        
        depart_to_selected1 = request.POST.get('select_depart_to1','')
        depart_to_selected2 = request.POST.get('select_depart_to2','')
        depart_to_selected3 = request.POST.get('select_depart_to3','')
        depart_to_selected4 = request.POST.get('select_depart_to4','')

        #depart_user_to1 = request.POST.get('select_user_to1','')
        #depart_user_to2 = request.POST.get('select_user_to2','')
        #depart_user_to3 = request.POST.get('select_user_to3','')
        #depart_user_to4 = request.POST.get('select_user_to4','')

        satus_selected = request.POST.get('select_status','SUBMIT')

        if form.is_valid() and file_form.is_valid() and division_select_valid:
            load_contents.board_type = 'COWORK'
            load_contents.title = form.cleaned_data['title']
            load_contents.contents = form.cleaned_data['contents']
            load_contents.options = division_selected
            load_contents.status = satus_selected
            load_contents.depart_to1 = depart_to_selected1
            load_contents.depart_to2 = depart_to_selected2
            load_contents.depart_to3 = depart_to_selected3
            load_contents.depart_to4 = depart_to_selected4

            #load_contents.depart_user_to1 = depart_user_to1
            #load_contents.depart_user_to2 = depart_user_to2
            #load_contents.depart_user_to3 = depart_user_to3
            #load_contents.depart_user_to4 = depart_user_to4


            load_contents.lastest_modifier = request.user.id
            load_contents.lastest_modified_date = datetime.datetime.now()
            
            load_contents.save()

            #file save
            for f in files:
                file_instance = Board_File(file=f, board_id = load_contents.pk)
                file_instance.origin_name = f.name
                file_instance.board_type = 'COWORK'
                file_instance.user = request.user.user_id
                file_instance.save()
 

            if id is None:
                return HttpResponseRedirect('./../' + str(load_contents.pk))
            else:
                return HttpResponseRedirect('./../../' + str(load_contents.pk))
        else:
            pass#form = board_form(instance=profile)




    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    
    #구분
    list_division = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_WORK_DIVISION').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_division.append({
            'id':data['code'],
            'name':data['name']
            })
    #부서
    ##경천
    list_depart = []
    if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
        query_depart_kbl= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_KBL',upper_commcode ='000002' ).annotate(code = F('commcode'),name = f_name ).values('code','name')
        for data in query_depart_kbl:
            list_depart.append({
                'id':data['code'],
                'name':data['name']
                })
    else:
        ##IMEDI
        list_depart_medical = []
        query_depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('commcode'),name = f_name ).values('code','name','id')
        for data in query_depart_medical:
            if data['code'] == 'DOCTOR':
                temp_commcode = COMMCODE.objects.get(id = data['id'])
                data['code'] += '_' + temp_commcode.se1
            list_depart.append({
                'id':data['code'],
                'name':data['name']
                }) 

    ##관리자
    query_depart_admin= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='DEPART_ADMIN',upper_commcode ='000002' ).annotate(code = F('commcode'),name = f_name).values('code','name')
    for data in query_depart_admin:
            list_depart.append({
                'id':data['code'],
                'name':data['name']
                })

    #게시판 상태
    list_status = []
    query_status= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='BOARD_STATUS',upper_commcode ='000005' ).annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_status:
        list_status.append({
            'id':data['code'],
            'name':data['name']
            })





    return render(request,
        'board_work/create_edit.html',
            {
                'form':form,
                'file_form':file_form,

                'list_file': list_file if list_file else None,
                'list_file_count':len(list_file),

                'list_division':list_division,

                'list_depart':list_depart,
                'list_status':list_status,
                
                'option_err':option_err,

                'division_selected':division_selected, #division
                'depart_selected1':depart_selected1,
                'depart_selected2':depart_selected2,
                'depart_selected3':depart_selected3,
                'depart_selected4':depart_selected4,
                'status_selected':status_selected,

                'depart_user_selected1':depart_user_selected1,
                'depart_user_selected2':depart_user_selected2,
                'depart_user_selected3':depart_user_selected3,
                'depart_user_selected4':depart_user_selected4,

            }
        )


def board_work_comment_add(request):
    content_id = request.POST.get('content_id')
    comment = request.POST.get('comment','')
    upper_id = request.POST.get('upper_id','')
    top_id = request.POST.get('top_id','')

    select_user = request.POST.get('select_user','')
    start_date = request.POST.get('start_date','0000-00-00 00:00:00')
    expected_date = request.POST.get('expected_date','0000-00-00 00:00:00')
    due_date = request.POST.get('due_date','0000-00-00 00:00:00')
    status = request.POST.get('status','')
    
    new_comment = Board_Comment()

    if upper_id != '':#대댓글
        upper = Board_Comment.objects.get(pk = upper_id)
        new_comment.depth = upper.depth + 1
        
        try:
            check = Board_Comment.objects.filter(content_id = content_id, orderno__gt = upper.orderno, depth__lte = upper.depth,).order_by('orderno')[:1]
            
            new_comment.orderno = check[0].orderno
        except IndexError:
            check = Board_Comment.objects.filter(content_id = content_id, ).order_by('-orderno')[:1]
           
            new_comment.orderno = check[0].orderno + 1
        

        #저장 전 sequence 자리 비우기
        query_set = Board_Comment.objects.filter(content_id = content_id, orderno__gte = new_comment.orderno ,).order_by('orderno')
        
        for query in query_set:
            query.orderno += 1
            
            query.save()

        
        #Co-Board top 댓글
        if top_id != '':
            try:
                top_comment = Board_Comment.objects.get(pk=top_id)

                top_comment.in_charge = select_user
                top_comment.start_date = start_date
                top_comment.end_date = due_date
                top_comment.expected_date = expected_date
                top_comment.progress = status
                 
                top_comment.lastest_modifier = request.user.id
                top_comment.last_modified_date = datetime.datetime.now()
                top_comment.save()
                
            except Board_Comment.DoesNotExist:
                pass


          
    else:#첫번째 뎁스 댓글
        check = Board_Comment.objects.filter(content_id = content_id, ).order_by('orderno')
        
        def_date = '0000-00-00 00:00:00'

        #expected_date = request.POST.get('expected_date',def_date)
        #due_date = request.POST.get('due_date',def_date)
        #select_status = request.POST.get('select_status')

        #content = Board_Contents.objects.get(id = content_id)
        #content.status = select_status
        #content.date_to_be_done = expected_date
        #content.date_done = due_date
        #content.save()

        if check.count() != 0: # 완전 처음이 아닐때 
            check_last = check.last()
            new_comment.orderno = check_last.orderno + 1


    new_comment.content_id = content_id
    new_comment.comment = comment
    new_comment.creator = request.user.id
    new_comment.in_charge = select_user
    new_comment.start_date = start_date
    new_comment.end_date = due_date
    new_comment.expected_date = expected_date
    new_comment.progress = status

    new_comment.save() 

    def_date = '0000-00-00 00:00:00'
    if select_user != '':
        AlertLog(
            page_type="COWORK",
            content_type="COMMENT",
            content_id=new_comment.pk,
            pointed_date=def_date if expected_date == '' else expected_date,
            user_id=select_user,
            creator=request.user.id,
            status=status,
            ).save()




    return JsonResponse({
        'result':True,
        })


@login_required
def board_work_comment_edit(request,id=None):
    if id==None:
        return JsonResponse({'result':False,})

    
    comment = request.POST.get('comment','')

    top_id = request.POST.get('top_id','')

    select_user = request.POST.get('select_user','')
    start_date = request.POST.get('start_date','0000-00-00 00:00:00')
    expected_date = request.POST.get('expected_date','0000-00-00 00:00:00')
    due_date = request.POST.get('due_date','0000-00-00 00:00:00')
    status = request.POST.get('status','')


    #Co-Board top 댓글
    if top_id != '':
        try:
            top_comment = Board_Comment.objects.get(pk=top_id)

            top_comment.in_charge = select_user
            top_comment.start_date = start_date
            top_comment.end_date = due_date
            top_comment.expected_date = expected_date
            top_comment.progress = status
                 
            top_comment.lastest_modifier = request.user.id
            top_comment.last_modified_date = datetime.datetime.now()
            top_comment.save()
                
        except Board_Comment.DoesNotExist:
            pass




    comment_query = Board_Comment.objects.get(id = id)
    comment_query.comment = comment

    comment_query.lastest_modifier = request.user.id
    comment_query.last_modified_date = datetime.datetime.now()
    comment_query.save()

    
    try:
        alert_query = AlertLog.objects.get(
            page_type="COWORK",
            content_type="COMMENT",
            content_id = id,
            )
    except AlertLog.DoesNotExist:
        alert_query = AlertLog(
            page_type="COWORK",
            content_type="COMMENT",
            content_id = id,
            creator = request.user.id,
            )

    alert_query.pointed_date = def_date if expected_date == '' else expected_date
    alert_query.status=status
    alert_query.user_id=select_user
    alert_query.save()
     



    return JsonResponse({
        'result':True,
        })


#
def reverse_truefalse(val):
    res = val
    if res == True:
        res = False
    else:
        res = True
    return res 
    

@login_required
def board_work_comment_emoji(request):
    comment_id = request.POST.get('comment_id')
    number = request.POST.get('number')

    comment = Board_Comment.objects.get(id = comment_id)
    res = False
    if number == '1':
        comment.emoji1 = reverse_truefalse(comment.emoji1)
        comment.emoji1_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        res = comment.emoji1
    elif number == '2':
        comment.emoji2 = reverse_truefalse(comment.emoji2)
        comment.emoji2_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        res = comment.emoji2
    elif number == '3':
        comment.emoji3 = reverse_truefalse(comment.emoji3)
        comment.emoji3_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        res = comment.emoji3
    elif number == '4':
        comment.emoji4 = reverse_truefalse(comment.emoji4)
        comment.emoji4_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        res = comment.emoji4

    comment.save()

    return JsonResponse({
        'result':res,
        })



@login_required
def board_info_list(request,id=None):

    if id == None:
        id = request.POST.get('selected_content',None)
    
    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
         fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')
    #초기 - Division
    list_division = COMMCODE.objects.filter(upper_commcode = '000005',commcode_grp = 'INFOBOARD_DIVISION')

    dict_division = {}
    for division in list_division.annotate(code = F('commcode'),name = fname).values('code','name'):
        dict_division.update({
            division['code'] : division['name']
            })



    #content
    content = None;
    #게시글 오픈 유무
    if id is not '':
        try:
            read_page = Board_Contents.objects.get(id = id )
            creator = User.objects.get(id = int(read_page.creator))
            #조회수 카운트
            ##자기 자신을 제외한 유저에만 해당
            if creator.id != request.user.id:
                date_min = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.min)
                date_max = datetime.datetime.combine(datetime.datetime.now().date(), datetime.time.max)

                is_new = Board_View_Log.objects.filter(registered_date__range = (date_min, date_max),user_id = request.user.id, board_id = read_page.id, ).count()
                if is_new == 0 : #오늘 안봤으면 추가
                    new_view = Board_View_Log()
                    new_view.board_id = read_page.id
                    new_view.user_id = request.user.id
                    new_view.save()

                    read_page.view_count += 1
                    read_page.save()

            #페이지 상세 정보 
            
            
            comments_count = Board_Comment.objects.filter(content_id = id, use_yn = 'Y').count()
            content = {
                'id':read_page.id,
                'title':read_page.title,
                'contents':read_page.contents,
                'creator':creator.user_id,
                'date':read_page.created_date.strftime('%Y-%m-%d %H:%M'),
                'views': read_page.view_count,
                'comments_count':comments_count,
                }

            query_file = Board_File.objects.filter(board_id = id,board_type='INFOBOARD')

            list_file = []
            for file in query_file:
                list_file.append({
                    'id':file.id,
                    'file_name':file.file.url,
                    'origin_name':file.origin_name,
                    })
            content.update({
                'list_file': list_file,
                'list_file_count':query_file.count(),
                })

        except Board_Contents.DoesNotExist:
            content = None;



   

    #search filter
    kwargs = {}
    kwargs['use_yn'] = 'Y' # 기본 
    kwargs['board_type'] = 'INFOBOARD' #일반 게시판

    kwargs['is_KBL'] = 'N'
    if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
        kwargs['is_KBL'] = 'Y'


    search_string = request.POST.get('search_string','')
    view_division_filter = request.POST.get('view_division_filter','')
    if view_division_filter != '':
        kwargs['options'] = view_division_filter

    users = User.objects.all()
    if search_string == '':
        query = Board_Contents.objects.filter(**kwargs).order_by('-top_seq','-created_date')

    else:
        argument_list = [] 

        argument_list.append( Q(**{'title__icontains':search_string} ) ) 
        argument_list.append( Q(**{'contents__icontains':search_string} ) ) 

        #list_search_user = users.filter()
        query = Board_Contents.objects.filter(functools.reduce(operator.or_, argument_list),**kwargs).order_by('-created_date')


    #목록 정렬
    contents_list = []
    for item in query:
        comment_count = Board_Comment.objects.filter(content_id = item.id, use_yn = 'Y').count()
        creator = users.get(id = int(item.creator))
        file_count = Board_File.objects.filter(board_id = item.id,board_type='INFOBOARD').count()


        contents_list.append({
            'id':item.id,
            'is_notice':item.top_seq,
            'division':dict_division[item.options],
            'title':item.title,
            'creator':creator.user_id,
            'date':item.created_date.strftime('%Y-%m-%d %H:%M'),
            'comment_count':comment_count,
            'is_file':False if file_count is 0 else True,
            'view_count':item.view_count,
            })


    #페이지네이션
    page = request.POST.get('page',1)
    view_contents_count = request.POST.get('view_contents_count',10);
    paginator = Paginator(contents_list, view_contents_count)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    return render(request,
        'board_info/list.html',
            {
                'content_count':query.count(),
                'content':content,
                #'contents_list':contents_list,
                'contents_list':paging_data,

                'search_string':search_string,
                'view_division_filter':view_division_filter,
                'view_contents_count':view_contents_count,
                'dict_division':dict_division,

                'page':int(page),
                'page_range':list( range(paging_data.paginator.page_range.start, paging_data.paginator.page_range.stop) ) ,
                'page_range_start':paging_data.paginator.page_range.start,
                'page_range_stop':paging_data.paginator.page_range.stop,
                'page_number':paging_data.number,
                'has_previous':paging_data.has_previous(),
                'has_next':paging_data.has_next(),
            }
        )



@login_required
def board_info_create_edit(request,id=None):
    
    division_selected= None
    option_err=''
    is_top= '0'
    #language
    lang = ''
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        lang = 'vi-VN'
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        lang = 'ko-KR'
    else:
        lang = 'en-US'
    
    list_file = []
    if id is None: # 새글
        load_contents = Board_Contents()
        form = board_form()

        file_form = board_file_form()

        load_contents.creator = request.user.id
        if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
            load_contents.is_KBL = 'Y' 
    else: # 글 수정 
        load_contents = Board_Contents.objects.get(id = id)
        form = board_form(instance = load_contents)
        division_selected = load_contents.options

        is_top = load_contents.top_seq

        file_form = board_file_form()
        query_file = Board_File.objects.filter(board_id = id,board_type='INFOBOARD')
        for file in query_file:
            list_file.append({
                'id':file.id,
                'file_name':file.file.url,
                'origin_name':file.origin_name,
                })


    #저장
    if request.method == 'POST':
        form = board_form(request.POST)
        file_form = board_file_form(request.POST, request.FILES)   
        files = request.FILES.getlist('file') 


        select_valid = False
        division_selected = request.POST.get('select_division',None)
        if division_selected is not None:
            select_valid = True


        if form.is_valid() and file_form.is_valid() and select_valid:
            load_contents.board_type = 'INFOBOARD'
            load_contents.title = form.cleaned_data['title']
            load_contents.contents = form.cleaned_data['contents']
            load_contents.options = division_selected

            is_notice = request.POST.get('top_seq','off')
            if is_notice == 'on':
                load_contents.top_seq = 1
   

            
            load_contents.lastest_modifier = request.user.id
            load_contents.lastest_modified_date = datetime.datetime.now()

            load_contents.save()

            #file save
            for f in files:
                file_instance = Board_File(file=f, board_id = load_contents.pk)
                file_instance.origin_name = f.name
                file_instance.board_type = 'INFOBOARD'
                file_instance.user = request.user.user_id

                file_instance.save()
 

            if id is None:
                return HttpResponseRedirect('./../' + str(load_contents.pk))
            else:
                return HttpResponseRedirect('./../../' + str(load_contents.pk))
        else:
            pass#form = board_form(instance=profile)




    
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    list_division = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='INFOBOARD_DIVISION').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_division.append({
            'id':data['code'],
            'name':data['name']
            })
    

    return render(request,
        'board_info/create_edit.html',
            {
                'form':form,
                'file_form':file_form,
                'lang':lang,

                'list_file': list_file if list_file else None,
                'list_file_count':len(list_file),

                'list_division':list_division,
                'division_selected':division_selected, #division
                'option_err':option_err,
                'is_top':is_top
            }
        )



@login_required
def board_info_delete(request,id):

    try:
        content = Board_Contents.objects.get(id = id)
        content.use_yn = 'N'
        content.save()

    except Board_Contents.DoesNotExist:
        pass


    return HttpResponseRedirect('./../')


@login_required
def board_info_delete_file(request):
    id = request.POST.get('id')

    file = Board_File.objects.get(pk = id)


    file.delete()


    return JsonResponse({
        'result':True,
        })



def sms_send_sms(request):

    type = request.POST.get('type','')
    company = request.POST.get('company','')
    receiver = request.POST.get('receiver','')

    phone = request.POST.get('phone','')
    contents = request.POST.get('contents','')

    is_KBL = request.POST.get('is_KBL',None)

    sms_send = sms_history()

    sms_send.type = type
    sms_send.company = company
    sms_send.receiver = receiver

    sms_send.sender = request.user.user_id
    sms_send.phone = phone
    sms_send.contents = contents
    sms_send.date_of_registered = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    if is_KBL:
        sms_send.is_KBL = 'Y'
    sms_send.save()

    return JsonResponse({
        'res':True,
        'id':sms_send.id,
        })


def sms_recv_result(request):

    context = {}
    msg_id = request.GET.get('msg_id',None)
    if msg_id is not None:
        result = request.GET.get('result','')
        code = request.GET.get('code','')


        res_sms = sms_history.objects.get(id = msg_id)


        res_sms.status = result
        res_sms.res_code = code
        res_sms.date_of_recieved = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        res_sms.save()


    return JsonResponse({
        'res':True,
        })


def sms_history_index(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')

    list_type = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='SMS_TYPE',upper_commcode ='000008').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_type.append({
            'id':data['code'],
            'name':data['name']
            })

    list_search = []
    query_division= COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='SMS_SEARCH',upper_commcode ='000008').annotate(code = F('commcode'),name = f_name ).values('code','name')
    for data in query_division:
        list_search.append({
            'id':data['code'],
            'name':data['name']
            })


    return render(request,
    'Manage/sms_history.html',
        {
            'list_type':list_type,
            'list_search':list_search,
        }
    )


def sms_history_search(request):

    start = request.POST.get('start')
    end = request.POST.get('end')

    type = request.POST.get('type','')
    option = request.POST.get('option','')
    string = request.POST.get('string','')

    kwargs = {}
    if type != '':
        kwargs['type'] = type
    #if option != '':
    #    kwargs['type'] == type
    #if type != '':
    #    kwargs['type'] == type
    kwargs['is_KBL'] = 'N'
    if request.META['SERVER_PORT'] == '8888' or request.META['SERVER_PORT'] == '22222':#경천애인
        kwargs['is_KBL'] = 'Y'

    argument_list = [] 
    if string !='':
        if option != 'COMPANY':
            argument_list.append( Q(**{'receiver__icontains':string} ) )
        if option != 'PERSONAL':
            argument_list.append( Q(**{'company_name__icontains':string} ) )
        

    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    list_sms_history = []

    if len(argument_list) != 0:
        query_sms = sms_history.objects.filter(functools.reduce(operator.or_, argument_list) ,**kwargs, date_of_registered__range = (date_min, date_max)).order_by('-id')
    else:
        query_sms = sms_history.objects.filter(**kwargs, date_of_registered__range = (date_min, date_max)).order_by('-id')

    for query in query_sms:
        list_sms_history.append({
            'id':query.id,
            'type':query.type,
            'company':query.company_name,
            'receiver':query.receiver,
            'phone':query.phone,
            'datetime':query.date_of_registered[0:16],
            'contents':query.contents,
            'sender':query.sender,
            'status':query.status,
            'remark':query.res_code,
            })


    #페이지네이션
    page = request.POST.get('page',1)
    view_contents_count = request.POST.get('context_in_page',10);

    paginator = Paginator(list_sms_history, view_contents_count)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    return JsonResponse({
        'res':True,

        'datas':list(paging_data), 

        'page':int(page),
        'page_range':list( range(paging_data.paginator.page_range.start, paging_data.paginator.page_range.stop) ) ,
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        })


def sms_history_get(request):

    history_id = request.POST.get('id')
         
    context={}
    history = sms_history.objects.get(pk=int(history_id))

    context.update({
        'name':history.receiver,
        'phone':history.phone,
        'contents':history.contents,
        })

    return JsonResponse(context)



def statistics_test(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    doctor = Doctor.objects.values('name_short','id')
    
    return render(request,
    'statistics/statistics_test.html',
        {
            'depart_medical':depart_medical,
            'doctor' : doctor,
        }
    )
def statistics_medicine(request):

    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    doctor = Doctor.objects.values('name_short','id')
    return render(request,
    'statistics/statistics_medicine.html',
        {
            'depart_medical':depart_medical,
            'doctor' : doctor,
        }
    )

def statistics_procedure(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    doctor = Doctor.objects.values('name_short','id')
    return render(request,
    'statistics/statistics_procedure.html',
        {
            'depart_medical':depart_medical,
            'doctor' : doctor,
        }
    )

def statistics_package(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request,
    'statistics/statistics_package.html',
        {
            'depart_medical':depart_medical,
        }
    )

def statistics_depart(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request, 
    'statistics/statistics_depart.html',
        {
            'depart_medical':depart_medical,
        }
    )


def statistics_search(request):
    type = request.POST.get('type')

    start = request.POST.get('start')
    end = request.POST.get('end')
    depart = request.POST.get('depart')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    data_list = []

 
    total_revenue = 0
    total_purchace = 0
    


    if type == 'TEST':
        tests = Test.objects.all().order_by('code')
        for test in tests:
            price_sum = 0

            sub_query = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                diagnosis__testmanager__test = test.id,
            ).exclude(
                progress='deleted'
                ).prefetch_related(
                'diagnosis__testmanager_set',
            )

            if not sub_query:
                continue

            for data in sub_query:
                price_sum = test.get_price(data.recorded_date)


            data_list.append({
                'id':test.id,
                'code':test.code,
                'name':test.name,
                'name_vi':test.name_vie,
                'count':sub_query.count(),
                'price_sum':price_sum,
                })

            total_revenue += price_sum
    elif type == 'PROCEDURE':
        procedures = Precedure.objects.all().order_by('code')
        for procedure in procedures:
            price_sum = 0
            count = 0

            sub_query = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                diagnosis__preceduremanager__precedure= procedure.id,
                progress='done'
                )
            #).exclude(
            #    #progress='deleted'
            #    ).prefetch_related(
            #    'diagnosis__preceduremanager_set',
            #)
            if not sub_query:
                continue

            for data in sub_query:
                preceduremanager_set = PrecedureManager.objects.filter(
                    diagnosis_id = data.diagnosis.id,
                    precedure_id = procedure.id,
                    diagnosis__reception__recorded_date__range = (date_min, date_max),
                    diagnosis__reception__progress='done',
                    )
                #price_sum = procedure.get_price(data.recorded_date)
                if preceduremanager_set.count() != 0:
                    for set_data in preceduremanager_set:
                        count += set_data.amount
                        price_sum += procedure.get_price(data.recorded_date) * set_data.amount

            count = sub_query.count()
            price_sum = procedure.get_price(data.recorded_date) * count

            data_list.append({
                'id':procedure.id,
                'code':procedure.code,
                'name':procedure.name,
                'name_vi':procedure.name_vie,
                'count':count,
                'price_sum':price_sum,
                })

            total_revenue += price_sum

    elif type =='MEDICINE':
        
        vac_kwargs={}
        is_vaccine = request.POST.get('is_vaccine','false')
        if is_vaccine == 'true':
            vac_kwargs['code__icontains'] = 'VC'

        medicines= Medicine.objects.filter(
            **vac_kwargs,
            ).order_by('code')

        for medicine in medicines:
           vac_kwargs={}
        is_vaccine = request.POST.get('is_vaccine','false')
        if is_vaccine == 'true':
            vac_kwargs['code__icontains'] = 'VC'

        medicines= Medicine.objects.filter(
            **vac_kwargs,
            ).order_by('code')

        #갯수가 많으면 에러남;
        #500개씩 쪼개서 쿼리문 작성
        


        tmp_roof_cnt = len(medicines) // 500 
        
        for roof_cnt in range(tmp_roof_cnt + 1):
            
            for medicine in medicines[ 500 * roof_cnt : 500 * (roof_cnt+1)]:
                price_sum = 0

                count = 0

                # 1000개의 쿼리가 넘어가면 sql lib 자체적으로 막음.
                ## 해결방안 - count 먼저 처리 후 500개씩 나눠서 처리
                sub_query_cnt = Reception.objects.filter(
                    **kwargs ,
                    recorded_date__range = (date_min, date_max), 
                    diagnosis__medicinemanager__medicine_id= medicine.id,
                
                    progress='done',
                ).exclude(
                    #progress='deleted'
                    ).prefetch_related(
                    'diagnosis__medicinemanager_set',
                ).count()


                for roof_cnt_sub in range(sub_query_cnt // 500  + 1):
                    
                    sub_query = Reception.objects.filter(
                        **kwargs ,
                        recorded_date__range = (date_min, date_max), 
                        diagnosis__medicinemanager__medicine_id= medicine.id,
                        progress='done',
                    ).exclude(
                        #progress='deleted'
                        ).prefetch_related(
                        'diagnosis__medicinemanager_set',
                    )[ 500 * roof_cnt_sub : 500 * (roof_cnt_sub+1)]

                    for data in sub_query: 
                        medicine_set_roof_cnt = MedicineLog.objects.filter(
                            diagnosis_id = data.diagnosis.id,                             
                            type='dec',
                            medicine_id=medicine.id
                            ).count()

                        for sub_medicine_set_roof_cnt in range(sub_query_cnt // 500  + 1):
                            #출고한 약
                            medicine_set = MedicineLog.objects.filter(
                                diagnosis_id = data.diagnosis.id, 
                                type='dec',
                                medicine_id=medicine.id
                                )[ 500 * sub_medicine_set_roof_cnt : 500 * (sub_medicine_set_roof_cnt+1)]

                            for set_data in medicine_set:
                                count += set_data.changes
                                if(data.payment.paymentrecord_set.first() is not None):
                                    price_sum += medicine.get_price(data.payment.paymentrecord_set.first().date) * set_data.changes
                                
                                #count += set_data.amount * set_data.days
                                #price_sum += medicine.get_price(data.recorded_date) * set_data.amount * set_data.days
                            #회수한 약
                            withdraw_set = MedicineLog.objects.filter(
                                diagnosis_id = data.diagnosis.id, 
                                type='add',
                                memo='withdraw',
                                medicine_id=medicine.id,
                                )[ 500 * sub_medicine_set_roof_cnt : 500 * (sub_medicine_set_roof_cnt+1)]

                            for set_data in withdraw_set:
                                count -= set_data.changes
                                price_sum -= medicine.get_price(data.payment.paymentrecord_set.first().date) * set_data.changes
   
                    
                #input
                beginning_count = 0
                beginning_unit_price = 0

                input = 0
                input_price = 0
             
                consuming_count = 0

                medicine_log = None
                if len(kwargs) == 0:


                    #입고
                    medicine_log = MedicineLog.objects.filter(
                        date__range = (date_min, date_max), 
                        type='add',
                        medicine_id = medicine.id
                        ).exclude(
                            memo = 'withdraw'
                            )
                    #입고 및 출고 없으면 출력 안함
                    if not sub_query_cnt and not medicine_log:
                        continue

                    for data in medicine_log:
                        input += data.changes
                        input_price += medicine.get_price_input(data.date) * data.changes

                    #beginning
                    beginning_date = date_min - datetime.timedelta(1)  
                
                    add_count = MedicineLog.objects.filter(
                        date__range = (date_min, date_max), 
                        type='add',
                        medicine_id = medicine.id
                        ).aggregate(Sum('changes'))

                

                    if add_count['changes__sum'] is None:
                        add_count = 0
                    else:
                        add_count = add_count['changes__sum'] 

                    dec_count = MedicineLog.objects.filter(
                        date__range = (date_min, date_max), 
                        type='add',
                        medicine_id = medicine.id
                        ).aggregate(Sum('changes'))

                    if dec_count['changes__sum'] is None:
                        dec_count = 0
                    else:
                        dec_count = dec_count['changes__sum'] 

                    beginning_count = medicine.inventory_count# + dec_count - add_count
                
                    #changed to get average
                    #beginning_unit_price = medicine.get_price(beginning_date)
                    average_query =  Pricechange.objects.filter(
                        type='Medicine',
                        type2="OUTPUT",
                        country="VI",
                        code=medicine.code,
                        )
                    if average_query.count() == 0:
                        beginning_unit_price = medicine.get_price(beginning_date)
                    else:
                        average_total = average_query.aggregate(
                            Sum('price'),
                            )

                        beginning_unit_price = round( ( average_total['price__sum'] + medicine.get_price(beginning_date) ) / (average_query.count() + 1) )

                if input == 0 and count == 0:
                    continue

                data_list.append({
                    'id':medicine.id,
                    'code':medicine.code,
                    'name':medicine.name,
                    'name_vi':medicine.name_vie,
                    'unit':medicine.unit,

                    'beginning_count':beginning_count,
                    'beginning_unit_price':beginning_unit_price,

                    'input':input,
                    'input_average':0 if input is 0 else math.trunc(input_price / input),
                    'input_price':input_price,

                    'count':count,
                    'price_average':0 if count is 0 else math.trunc(price_sum / count),
                    'price_sum':price_sum,

                    })

                total_revenue += price_sum
                total_purchace +=input_price

    elif type == 'DEPART':
        departs = Depart.objects.all()
        
        for depart in departs:
            
            views_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                depart_id = depart.id 
                ).exclude(progress='deleted').count()


            revenue_total = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                depart_id = depart.id 
                ).exclude(
                progress='deleted'
                ).prefetch_related('payment__paymentrecord_set').aggregate(total_price=Sum('payment__paymentrecord__paid'))
                 
            data_list.append({
                'id':depart.id,
                'name':depart.name,
                'name_vi':depart.full_name_vie,
                'count':views_count,
                'price_sum':0 if revenue_total['total_price'] is None else revenue_total['total_price'],
                })




    return JsonResponse({
        'result':True,
        'datas':data_list,

        'total_revenue':total_revenue,
        'total_purchace':total_purchace,
        })


def search_lab(request):
    type = request.POST.get('type')

    start = request.POST.get('start')
    end = request.POST.get('end')
    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')
    category = request.POST.get('patient_type','')
    string = request.POST.get('patient_search','')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 
    if doctor != '':
        kwargs['payment__reception__doctor_id'] = doctor        


    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    data_list = []

 
    total_revenue = 0
    total_purchace = 0
    
    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    # for test in tests:
    #     argument_list.append( Q(**{'diagnosis__testmanager__test':test.id} ) ) 
    argument_list.append( Q(**{'id':0} ) )     
    count = 0
    # for test in tests:
    price_sum = 0

    sub_query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs ,
        recorded_date__range = (date_min, date_max), 
    ).exclude(
        progress='deleted'
        ).order_by('recorded_date')

    # reception = Reception.objects.filter(
    #     **kwargs ,
    #     recorded_date__range = (date_min, date_max), 
    #     diagnosis = test.diagnosis_id,
    # ).exclude(
    #     progress='deleted'
    #     )
    # print('reception: ', reception)

    # if not sub_query:
    #     continue
    

    for data in sub_query:
        try: 
            test_records = data.diagnosis.testmanager_set.all()
        except:
            test_records = None
        if not test_records:
            continue
        for test_record in test_records:
            count += 1
            testInfo = Test.objects.get(id = test_record.test_id)
            price_sum = testInfo.get_price(data.recorded_date)
            data_list.append({
                'no':count,
                'id':data.id,
                'date':data.recorded_date.strftime('%Y-%m-%d'),
                'patient_name':data.patient.get_name_kor_eng(),
                'depart':data.depart.name,
                'doctor':data.doctor.name_eng + '/' + data.doctor.name_kor,
                'class':testInfo.test_class.name,
                'service_code':testInfo.code,
                'item_name':testInfo.name,
                'quantity':1,
                'price':testInfo.get_price(data.recorded_date),

                # 'code':test.code,
                # 'name':test.name,
                # 'name_vi':test.name_vie,
                # 'count':sub_query.count(),
                'price_sum':testInfo.get_price(data.recorded_date)
                })

            total_revenue += price_sum

    return JsonResponse({
        'result':True,
        'datas':data_list,

        'total_revenue':total_revenue,
        'total_purchace':total_purchace,
        })

def statistics_search_medicine(request):
    start = request.POST.get('start')
    end = request.POST.get('end')
    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')
    category = request.POST.get('patient_type','')
    string = request.POST.get('patient_search','')
    is_vaccine = request.POST.get('is_vaccine','false')
    kwargs = {}
    vac_kwargs={}
    if depart!='':
        kwargs['depart'] = depart # 기본 
    if doctor != '':
        kwargs['payment__reception__doctor_id'] = doctor    

    if is_vaccine == 'true':
        vac_kwargs['code__icontains'] = 'VC'         

    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)

    data_list = []

    total_revenue = 0
    total_purchace = 0
    
    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    argument_list.append( Q(**{'id':0} ) )     
    count = 0
    price_sum = 0

    sub_query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs ,
        recorded_date__range = (date_min, date_max), 
    ).exclude(
        progress='deleted'
        ).order_by('recorded_date')


    for data in sub_query:
        try: 
            medicine_set = data.diagnosis.medicinemanager_set.all()
        except:
            medicine_set = None
        if not medicine_set:
            continue
        for medicine_log in medicine_set:
            try: 
                medicine = Medicine.objects.get( **vac_kwargs, id = medicine_log.medicine_id)
            except Medicine.DoesNotExist:
                medicine = None
            if not medicine:
                continue                
            count += 1
            
            price_sum = medicine.get_price(data.recorded_date) * (medicine_log.amount * medicine_log.days)
            data_list.append({
                'no':count,
                'id':data.id,
                'date':data.recorded_date.strftime('%Y-%m-%d'),
                'patient_name':data.patient.get_name_kor_eng(),
                'depart':data.depart.name,
                'doctor':data.doctor.name_eng + '/' + data.doctor.name_kor,
                'class':medicine.medicine_class.name,
                'service_code':medicine.code,
                'item_name':medicine.name,
                'item_name_vi':medicine.name_vie,
                'unit':medicine.unit,
                'quantity':medicine_log.amount * medicine_log.days,
                'price':medicine.get_price(data.recorded_date),
                'price_sum':price_sum
                })

            total_revenue += price_sum
            total_purchace += medicine_log.amount * medicine_log.days

    return JsonResponse({
        'result':True,
        'datas':data_list,

        'total_revenue':total_revenue,
        'total_purchace':total_purchace,
        })

def search_procedure(request):
    type = request.POST.get('type')

    start = request.POST.get('start')
    end = request.POST.get('end')
    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')
    category = request.POST.get('patient_type','')
    string = request.POST.get('patient_search','')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 
    if doctor != '':
        kwargs['payment__reception__doctor_id'] = doctor        


    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    data_list = []

 
    total_revenue = 0
    total_purchace = 0
    
    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    # for test in tests:
    #     argument_list.append( Q(**{'diagnosis__testmanager__test':test.id} ) ) 
    argument_list.append( Q(**{'id':0} ) )     
    count = 0
    # for test in tests:
    price_sum = 0

    sub_query = Reception.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs ,
        recorded_date__range = (date_min, date_max), 
    ).exclude(
        progress='deleted'
        ).order_by('recorded_date')

    # reception = Reception.objects.filter(
    #     **kwargs ,
    #     recorded_date__range = (date_min, date_max), 
    #     diagnosis = test.diagnosis_id,
    # ).exclude(
    #     progress='deleted'
    #     )
    # print('reception: ', reception)

    # if not sub_query:
    #     continue
    

    for data in sub_query:
        try: 
            preceduremanager_set = data.diagnosis.preceduremanager_set.all()
        except:
            preceduremanager_set = None
        if not preceduremanager_set:
            continue
        for precedure_record in preceduremanager_set:
            count += 1
            procedures = Precedure.objects.get(id = precedure_record.precedure_id)
            price_sum = procedures.get_price(data.recorded_date)
            data_list.append({
                'no':count,
                'id':data.id,
                'date':data.recorded_date.strftime('%Y-%m-%d'),
                'patient_name':data.patient.get_name_kor_eng(),
                'depart':data.depart.name,
                'doctor':data.doctor.name_eng + '/' + data.doctor.name_kor,
                'class':procedures.precedure_class.name,
                'service_code':procedures.code,
                'item_name':procedures.name,
                'item_name_vi':procedures.name_vie,
                'quantity':1,
                'price':procedures.get_price(data.recorded_date),
                'price_sum':procedures.get_price(data.recorded_date)
                })

            total_revenue += price_sum

    return JsonResponse({
        'result':True,
        'datas':data_list,

        'total_revenue':total_revenue,
        'total_purchace':total_purchace,
        })

def statistics_search_pkg(request):
    type = request.POST.get('type')

    start = request.POST.get('start')
    end = request.POST.get('end')
    depart = request.POST.get('depart')
    category = request.POST.get('patient_type','')
    string = request.POST.get('patient_search','')
    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page',30);

    argument_list = []
    if category=='':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    data_list = []

 
    total_revenue = 0
    total_purchace = 0

    packages = Precedure.objects.filter(type = 'PKG').order_by('code')
    
    packages = Package_Manage.objects.filter(functools.reduce(operator.or_, argument_list),
                **kwargs ,
                date_bought__range = (date_min, date_max),
                itme_round = '1').order_by('date_bought', 'patient_id', 'depart' )

    for package in packages:
        price_sum = 0
        count = 0

        procedure = Precedure.objects.get(type = 'PKG', id = package.precedure.id)

        patient = Patient.objects.get(id = package.patient.id)

        data_list.append({
            'id':package.id,
            'code':package.precedure_mng_id,
            'name':patient.name_eng,
            'pkg_name':package.precedure_name,
            'price':procedure.get_price(datetime.datetime.strptime(package.date_bought, '%Y-%m-%d %H:%M:%S')),
            'count':procedure.count,
            'price_sum':procedure.get_price(datetime.datetime.strptime(package.date_bought, '%Y-%m-%d %H:%M:%S')),
            'memo':'',
            'gender':patient.gender,
            'phone':patient.phone,
            'date_bought':package.date_bought,
            })

        total_revenue += procedure.get_price(datetime.datetime.strptime(package.date_bought, '%Y-%m-%d %H:%M:%S'))

    paginator = Paginator(data_list, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    context = {
        'result':True,
        'datas':list(paging_data),
        # 'page_range_start':paging_data.paginator.page_range.start,
        # 'page_range_stop':paging_data.paginator.page_range.stop,
        # 'page_number':paging_data.number,
        # 'has_previous':paging_data.has_previous(),
        # 'has_next':paging_data.has_next(),
        'total_revenue':total_revenue,
        'total_purchace':total_purchace,        
        }


    return JsonResponse(context)



def statistics_customer_info(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request, 
    'statistics/statistics_customer_info.html',
        {
            'depart_medical':depart_medical,
        }
    )


def search_customer_info(request):

    start = request.POST.get('start')
    end = request.POST.get('end')
    depart = request.POST.get('depart')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)



    data_list_gender = []
    data_list_nation = []
    data_list_payment_method = []
    data_list_age = []

    #gender
    gender_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                payment__paymentrecord__status='paid',
            ).exclude(
                progress='deleted'
            ).select_related(
                'patient'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).values(
                'patient__gender'
            ).annotate(
                gender_count = Count('patient__gender'),
                total_price = Sum('payment__paymentrecord__paid')
            )

    for data in gender_count:
        data_list_gender.append({
            'name':data['patient__gender'],
            'count':data['gender_count'],
            'price_sum':data['total_price'],
            })

    #Nationality
    nation_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                payment__paymentrecord__status='paid',
            ).exclude(
                progress='deleted'
            ).select_related(
                'patient'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).values(
                'patient__nationality'
            ).annotate(
                nation_count= Count('patient__nationality'),
                total_price = Sum('payment__paymentrecord__paid')
            )

    for data in nation_count:
        data_list_nation.append({
            'name':data['patient__nationality'],
            'count':data['nation_count'],
            'price_sum':data['total_price'],
            })

    #payment_method
    payment_method_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                payment__paymentrecord__status='paid',
            ).exclude(
                progress='deleted'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).values(
                'payment__paymentrecord__method'
            ).annotate(
                method_count= Count('payment__paymentrecord__method'),
                total_price = Sum('payment__paymentrecord__paid')
            )

    for data in payment_method_count:
        if data['payment__paymentrecord__method'] == None:
            continue

        data_list_payment_method.append({
            'name':data['payment__paymentrecord__method'],
            'count':data['method_count'],
            'price_sum':0 if data['total_price'] is None else data['total_price'],
            })



 #age
    today = datetime.datetime.now()

   
    age_list=[
            {#0~9세
                'name':'0~9',
                'date_start':today,
                'date_end': (today - datetime.timedelta(days=365.25 * 10 ) ).date(),
            },
            {#10~19세
                'name':'10~19',
                'date_start':(today - datetime.timedelta(days=365.25 * 10 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 20 ) ).date(),
            },
            {#20~29세
                'name':'20~29',
                'date_start':(today - datetime.timedelta(days=365.25 * 20 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 30 ) ).date(),
            },
            {#30~39세
                'name':'30~39',
                'date_start':(today - datetime.timedelta(days=365.25 * 30 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 40 ) ).date(),
            },
            {#40~49세
                'name':'40~49',
                'date_start':(today - datetime.timedelta(days=365.25 * 40 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 50 ) ).date(),
            },
            {#50~59세
                'name':'50~59',
                'date_start':(today - datetime.timedelta(days=365.25 * 50 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 60 ) ).date(),
            },
            {#60~69세
                'name':'60~69',
                'date_start':(today - datetime.timedelta(days=365.25 * 60 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 70 ) ).date(),
            },
            {#70~79세
                'name':'70~79',
                'date_start':(today - datetime.timedelta(days=365.25 * 70 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 80 ) ).date(),
            },
            {#80세 이상
                'name':'80~',
                'date_start':(today - datetime.timedelta(days=365.25 * 80 ) ).date(),
                'date_end': (today - datetime.timedelta(days=365.25 * 150 ) ).date(),
            },
        ]



    for age in age_list: 
        age_count = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                patient__date_of_birth__lte = age['date_start'].strftime("%Y-%m-%d"),
                patient__date_of_birth__gt = age['date_end'].strftime("%Y-%m-%d"),
                payment__paymentrecord__status='paid',
            ).exclude(
                progress='deleted'
            ).select_related(
                'patient'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        data_list_age.append({
            'name':age['name'],
            'count':age_count['count'],
            'price_sum':0 if age_count['total_price'] is None else age_count['total_price'],
            })


        
    #top20
    query_top20 = Reception.objects.filter(
                **kwargs ,
                recorded_date__range = (date_min, date_max), 
                payment__paymentrecord__status='paid',
            ).exclude(
                progress='deleted'
            ).prefetch_related(
                'payment__paymentrecord_set'
            ).values(
                'patient_id'
            ).annotate(
                method_count= Count('payment__paymentrecord__method'),
                total_price = Sum('payment__paymentrecord__paid')
            ).order_by('-total_price')[:20]

    datas_top20 = []
    for data in query_top20:
        patient = Patient.objects.get(id = data['patient_id'])
        last_visit = Reception.objects.filter(patient_id = patient.id, ).exclude(progress='deleted').order_by('-recorded_date').first()
        datas_top20.append({
            'name':patient.name_kor + ' / ' + patient.name_eng,
            'date_of_birth':patient.date_of_birth.strftime('%Y-%m-%d'),
            'gender':patient.gender,
            'phone':patient.phone, 
            'last_visit':last_visit.recorded_date.strftime('%Y-%m-%d'),
            'total_amount':data['total_price'],
            })



    return JsonResponse({
        'result':True,
        'datas_gender':data_list_gender,
        'datas_nation':data_list_nation,
        'datas_payment_method':data_list_payment_method,
        'datas_age':data_list_age,
        'datas_top20':datas_top20,

        })

@login_required
def payment_debit(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })
    #의사 정보 ? 
    doctor = Doctor.objects.values('name_short','id')
    doctors = []
    for d in doctor:
        print(d['name_short'], d['id'])
        if d['id'] not in [19, 20, 24 , 25, 26, 27, 29 , 31, 33, 35, 38, 41,47,48,50 ,52,54]:
            doctors.append(d)

    #결제 방법
    payment_method = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_METHOD',upper_commcode ='000014' ).annotate(code = F('commcode'),name = f_name ).values('code','name')

    #결제 상태
    payment_status = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_STATUS',upper_commcode ='000014' ).annotate(code = F('commcode'),name = f_name ).values('code','name')

    return render(request,
        'Manage/Manage_debit.html',
            {
                'depart_medical':depart_medical,
                'doctor' : doctors,
                'payment_method':payment_method,
                'payment_status':payment_status,
            }
        )

@login_required
def recovery_debit(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })
    #의사 정보 ? 
    doctor = Doctor.objects.values('name_short','id')

    #결제 방법
    payment_method = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_METHOD',upper_commcode ='000014' ).annotate(code = F('commcode'),name = f_name ).values('code','name')

    #결제 상태
    payment_status = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_STATUS',upper_commcode ='000014' ).annotate(code = F('commcode'),name = f_name ).values('code','name')

    return render(request,
        'Manage/Manage_debit.html',
            {
                'depart_medical':depart_medical,
                'doctor' : doctor,
                'payment_method':payment_method,
                'payment_status':payment_status,
            }
        )

@login_required
def search_payment_debit(request):
    page_context = request.POST.get('page_context',10) # 페이지 컨텐츠 
    page = request.POST.get('page',1)

    date_type = request.POST.get('date_type')

    date_start = request.POST.get('start')
    date_end = request.POST.get('end')

    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')

    payment_method = request.POST.get('payment_method')
    payment_status = request.POST.get('payment_status')

    category = request.POST.get('patient_type','')
    string = request.POST.get('patient_search','')

    is_vaccine = request.POST.get('is_vaccine','false')
    is_red_invoice = True if request.POST.get('is_red_invoice','false') == 'true' else False

    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
        fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')
    
    
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    #지불 정보
    dict_payment_method = {}
    query_payment_method= COMMCODE.objects.filter(commcode_grp='PAYMENT_METHOD',upper_commcode='000014').annotate(code = F('commcode'),name = fname).values('code','name')
    for data in query_payment_method:
        dict_payment_method.update({
            data['code'] : data['name']
            })


    kwargs = {}

    argument_list = []
    datas = []

    # if category=='':
    #     argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
    #     argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    #     argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    #     argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 
    #     argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    # elif category=='name':
    #     argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
    #     argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    # elif category=='chart':
    #     argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    # elif category=='date_of_birth':
    #     argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    # elif category=='phone':
    #     argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 


    # kwargs['progress'] = 'done'
    # if depart != '':
    #     kwargs['depart_id'] = depart
    # if doctor != '':
    #     kwargs['doctor_id'] = doctor
    # if is_red_invoice == True:
    #     kwargs['need_invoice'] = is_red_invoice
    if category=='':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'payment__reception__patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'payment__reception__patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'payment__reception__patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'payment__reception__patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'payment__reception__patient__phone__icontains':string} ) ) 

    kwargs['payment__reception__progress'] = 'done'
    if depart != '':
        kwargs['payment__reception__depart_id'] = depart
    if doctor != '':
        kwargs['payment__reception__doctor_id'] = doctor
    if is_red_invoice ==True:
        kwargs['payment__reception__need_invoice'] = is_red_invoice    

    tmp_receptions = PaymentRecord.objects.filter(
            functools.reduce(operator.or_, argument_list),
            **kwargs ,
            date__range = (date_min, date_max), 
        ).order_by("-id")

    new_argument_list = []
    tmp_list=[]
    count = 0

    for tmp_data in tmp_receptions:
        if tmp_data.payment.reception.id in tmp_list:
            continue
        tmp_list.append(tmp_data.payment.reception.id)
        new_argument_list.append( Q(**{'id':tmp_data.payment.reception.id} ) ) 
    new_argument_list.append( Q(**{'id':0} ) ) 

    # receptions = Reception.objects.filter(
    #         functools.reduce(operator.or_, new_argument_list),
    #         # **kwargs ,
    #         # recorded_date__range = (date_min, date_max), 
    #     ).order_by("-id")      


    count = 0

    payment_subtotal = 0
    payment_total = 0
    payment_total_discount = 0
    payment_total_unpaid = 0
    payment_paid_1 = 0
    payment_paid_2 = 0
    payment_paid_3 = 0
    payment_paid_4 = 0
    payment_paid_5 = 0
    payment_record_unpaid = 0

    for roof_cnt in range(len(new_argument_list) // 500  + 1):
        receptions = Reception.objects.filter(
                functools.reduce(operator.or_, new_argument_list[ 500 * roof_cnt : 500 * (roof_cnt+1)]),
            ).order_by("-id") 

        argument_list = []
        argument_list.append( Q(**{'id': 0000000} ) )
        for reception in receptions:
            count +=1
            payment_records = reception.payment.paymentrecord_set.filter(status='paid')
            if payment_records.count() != 0 :
                if payment_records.first().date.date() == reception.recorded_date.date():
                    if payment_records.first().paid == reception.payment.total:
                        continue

            argument_list.append( Q(**{'id': reception.id} ) )

        receptions = Reception.objects.filter(
                functools.reduce(operator.or_, argument_list),
            ).prefetch_related().order_by("recorded_date")

        for reception in receptions:
            payment_records = reception.payment.paymentrecord_set.filter(status='paid')
            payment_record_count = 1
            for payment_record in payment_records:            
                # if payment_record.paid == 0:
                #     continue
                if payment_record_count == 1:
                    payment_paid_1 += payment_record.paid
                elif payment_record_count == 2:
                    payment_paid_2 += payment_record.paid
                elif payment_record_count == 3:
                    payment_paid_3 += payment_record.paid
                elif payment_record_count == 4:
                    payment_paid_4 += payment_record.paid
                elif payment_record_count == 5:
                    payment_paid_5 += payment_record.paid                                

                payment_record_count += 1  

            if reception.payment.reception.payment.discounted != 0:
                discounted_amount = reception.payment.sub_total / 100 * reception.payment.discounted
            else:
                discounted_amount = reception.payment.discounted_amount

            payment_subtotal += reception.payment.sub_total
            payment_total += reception.payment.total
            payment_total_discount += discounted_amount



    payment_record_unpaid = payment_total - payment_paid_1 - payment_paid_2 - payment_paid_3 - payment_paid_4 - payment_paid_5
    payment_total_unpaid += payment_record_unpaid              
          
    receptions = Reception.objects.filter(
            functools.reduce(operator.or_, argument_list),
        ).prefetch_related().order_by("recorded_date").values_list('id',flat=True)
        
    list_id = tmp_list
    list_id = list(set(list_id))


    argument_list = []
    for id in list_id:
        argument_list.append( Q(**{'id':id} ) )
    if len(list_id)==0:
        argument_list.append( Q(**{'id':0} ) )
    count_page = list_id

    receptions = Reception.objects.filter(
                functools.reduce(operator.or_, argument_list[int(page_context) * (int(page) -1): int(page_context) * int(page)]),
            ).prefetch_related().order_by("-id")

    datas = []

    for reception in receptions:
        data = {
            'no': reception.id,
            'date': reception.recorded_date.strftime('%Y-%m-%d'),
            'depart': reception.depart.name,
            'chart': "{:06d}".format(reception.patient.id),
            'Patient': reception.patient.get_name_kor_eng(),
            'phone': reception.patient.phone,
            'precedure': '',
            'Doctor_kor':reception.doctor.name_kor,
            'Doctor_eng': reception.doctor.name_eng,            
            }

        payment_record_date_1 = 0
        payment_record_paid_1 = 0
        payment_record_date_2 = 0
        payment_record_paid_2 = 0
        payment_record_date_3 = 0
        payment_record_paid_3 = 0
        payment_record_date_4 = 0
        payment_record_paid_4 = 0
        payment_record_date_5 = 0
        payment_record_paid_5 = 0
        payment_record_unpaid = 0
        payment_records = reception.payment.paymentrecord_set.filter(status='paid')

        payment_record_count = 1
        for payment_record in payment_records:
            # if payment_record.paid == 0:
            #     continue
            if payment_record_count == 1:
                payment_record_date_1 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_1 = payment_record.paid
            elif payment_record_count == 2:
                payment_record_date_2 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_2 = payment_record.paid
            elif payment_record_count == 3:
                payment_record_date_3 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_3 = payment_record.paid
            elif payment_record_count == 4:
                payment_record_date_4 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_4 = payment_record.paid
            elif payment_record_count == 5:
                payment_record_date_5 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_5 = payment_record.paid                              

            payment_record_count += 1         

        payment_record_unpaid = reception.payment.total - payment_record_paid_1 - payment_record_paid_2 - payment_record_paid_3 - payment_record_paid_4 - payment_record_paid_5

        if reception.payment.reception.payment.discounted != 0:
            discounted_amount = reception.payment.sub_total / 100 * reception.payment.discounted
        else:
            discounted_amount = reception.payment.discounted_amount

        data.update({
            'sub_total': reception.payment.sub_total,
            'discounted_amount': discounted_amount,
            'total': reception.payment.total,
            'payment_record_date_1': payment_record_date_1,
            'payment_record_paid_1': payment_record_paid_1,
            'payment_record_date_2': payment_record_date_2,
            'payment_record_paid_2': payment_record_paid_2,
            'payment_record_date_3': payment_record_date_3,
            'payment_record_paid_3': payment_record_paid_3,
            'payment_record_date_4': payment_record_date_4,
            'payment_record_paid_4': payment_record_paid_4,
            'payment_record_date_5': payment_record_date_5,
            'payment_record_paid_5': payment_record_paid_5,
            'payment_record_unpaid': payment_record_unpaid,
            })

        datas.append(data)

    paginator = Paginator(count_page, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    context = {
            'datas':datas,
            'page_range_start':paging_data.paginator.page_range.start,
            'page_range_stop':paging_data.paginator.page_range.stop,
            'page_number':paging_data.number,
            'has_previous':paging_data.has_previous(),
            'has_next':paging_data.has_next(),

            'payment_total_subtotal':payment_subtotal,
            'payment_total_discount':payment_total_discount,
            'payment_total_total':payment_total,
            'payment_total_unpaid':payment_total_unpaid,
            'payment_paid_1': payment_paid_1,
            'payment_paid_2': payment_paid_2,
            'payment_paid_3': payment_paid_3,
            'payment_paid_4': payment_paid_4,
            'payment_paid_5': payment_paid_5,
            }
    
    return JsonResponse(context)

@login_required
def search_recovery_debit(request):
    page_context = request.POST.get('page_context',10) # 페이지 컨텐츠 
    page = request.POST.get('page',1)

    date_type = request.POST.get('date_type')

    date_start = request.POST.get('start')
    date_end = request.POST.get('end')

    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')

    payment_method = request.POST.get('payment_method')
    payment_status = request.POST.get('payment_status')

    category = request.POST.get('patient_type','')
    string = request.POST.get('patient_search','')

    is_vaccine = request.POST.get('is_vaccine','false')
    is_red_invoice = True if request.POST.get('is_red_invoice','false') == 'true' else False

    current_language = request.session[translation.LANGUAGE_SESSION_KEY]
    if current_language == 'ko':
        fname = F('commcode_name_ko')
    elif current_language == 'en':
        fname = F('commcode_name_en')
    elif current_language == 'vi':
        fname = F('commcode_name_vi')
    
    
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    #지불 정보
    dict_payment_method = {}
    query_payment_method= COMMCODE.objects.filter(commcode_grp='PAYMENT_METHOD',upper_commcode='000014').annotate(code = F('commcode'),name = fname).values('code','name')
    for data in query_payment_method:
        dict_payment_method.update({
            data['code'] : data['name']
            })


    kwargs = {}

    argument_list = []
    datas = []

    if category=='':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 


    kwargs['progress'] = 'done'
    if depart != '':
        kwargs['depart_id'] = depart
    if doctor != '':
        kwargs['doctor_id'] = doctor
    # if payment_method !='':
    #     kwargs['payment__paymentrecord__method'] = payment_method
    # if payment_status !='':
    #     kwargs['payment__progress'] = payment_status
    if is_red_invoice == True:
        kwargs['need_invoice'] = is_red_invoice

    receptions = Reception.objects.filter(
            functools.reduce(operator.or_, argument_list),
            **kwargs ,
            # recorded_date__range = (date_min, date_max), 
        ).prefetch_related(
            'payment__paymentrecord_set'
        ).filter(payment__paymentrecord__date__range = (date_min, date_max)).order_by("-id")      

    argument_list = []
    argument_list.append( Q(**{'id': 0000000} ) )
    count = 0

    payment_subtotal = 0
    payment_total = 0
    payment_total_discount = 0
    payment_total_unpaid = 0
    payment_paid_1 = 0
    payment_paid_2 = 0
    payment_paid_3 = 0
    payment_paid_4 = 0
    payment_paid_5 = 0
    payment_record_unpaid = 0


    for reception in receptions:
        count +=1
        payment_records = reception.payment.paymentrecord_set.filter(status='paid')
        if payment_records.count() != 0 :
            if payment_records.first().date.date() == reception.recorded_date.date():
                if payment_records.first().paid == reception.payment.total:
                    continue

        argument_list.append( Q(**{'id': reception.id} ) )

    receptions = Reception.objects.filter(
            functools.reduce(operator.or_, argument_list),
        ).prefetch_related().order_by("recorded_date")

    for reception in receptions:
        payment_records = reception.payment.paymentrecord_set.filter(status='paid')
        payment_record_count = 1
        for payment_record in payment_records:            
            # if payment_record.paid == 0:
            #     continue
            if payment_record_count == 1:
                payment_paid_1 += payment_record.paid
            elif payment_record_count == 2:
                payment_paid_2 += payment_record.paid
            elif payment_record_count == 3:
                payment_paid_3 += payment_record.paid
            elif payment_record_count == 4:
                payment_paid_4 += payment_record.paid
            elif payment_record_count == 5:
                payment_paid_5 += payment_record.paid                                

            payment_record_count += 1  

        if reception.payment.reception.payment.discounted != 0:
            discounted_amount = reception.payment.sub_total / 100 * reception.payment.discounted
        else:
            discounted_amount = reception.payment.discounted_amount

        payment_subtotal += reception.payment.sub_total
        payment_total += reception.payment.total
        payment_total_discount += discounted_amount

    payment_record_unpaid = payment_total - payment_paid_1 - payment_paid_2 - payment_paid_3 - payment_paid_4 - payment_paid_5
    payment_total_unpaid += payment_record_unpaid              
          
    receptions = Reception.objects.filter(
            functools.reduce(operator.or_, argument_list),
        ).prefetch_related().order_by("recorded_date").values_list('id',flat=True)
        
    list_id = list(receptions)
    list_id = list(set(list_id))


    argument_list = []
    for id in list_id:
        argument_list.append( Q(**{'id':id} ) )
    if len(list_id)==0:
        argument_list.append( Q(**{'id':0} ) )
    count_page = list_id


    receptions = Reception.objects.filter(
                functools.reduce(operator.or_, argument_list[int(page_context) * (int(page) -1): int(page_context) * int(page)]),
            ).prefetch_related().order_by("-id")

    datas = []

    for reception in receptions:
        data = {
            'no': reception.id,
            'date': reception.recorded_date.strftime('%Y-%m-%d'),
            'depart': reception.depart.name,
            'chart': "{:06d}".format(reception.patient.id),
            'Patient': reception.patient.get_name_kor_eng(),
            'phone': reception.patient.phone,
            'precedure': '',
            'Doctor_kor':reception.doctor.name_kor,
            'Doctor_eng': reception.doctor.name_eng,            
            }

        payment_record_date_1 = 0
        payment_record_paid_1 = 0
        payment_record_date_2 = 0
        payment_record_paid_2 = 0
        payment_record_date_3 = 0
        payment_record_paid_3 = 0
        payment_record_date_4 = 0
        payment_record_paid_4 = 0
        payment_record_date_5 = 0
        payment_record_paid_5 = 0
        payment_record_unpaid = 0
        payment_records = reception.payment.paymentrecord_set.filter(status='paid')

        payment_record_count = 1
        for payment_record in payment_records:
            # if payment_record.paid == 0:
            #     continue
            if payment_record_count == 1:
                payment_record_date_1 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_1 = payment_record.paid
            elif payment_record_count == 2:
                payment_record_date_2 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_2 = payment_record.paid
            elif payment_record_count == 3:
                payment_record_date_3 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_3 = payment_record.paid
            elif payment_record_count == 4:
                payment_record_date_4 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_4 = payment_record.paid
            elif payment_record_count == 5:
                payment_record_date_5 = payment_record.date.strftime('%Y-%m-%d')
                payment_record_paid_5 = payment_record.paid                              

            payment_record_count += 1         

        payment_record_unpaid = reception.payment.total - payment_record_paid_1 - payment_record_paid_2 - payment_record_paid_3 - payment_record_paid_4 - payment_record_paid_5

        if reception.payment.reception.payment.discounted != 0:
            discounted_amount = reception.payment.sub_total / 100 * reception.payment.discounted
        else:
            discounted_amount = reception.payment.discounted_amount

        data.update({
            'sub_total': reception.payment.sub_total,
            'discounted_amount': discounted_amount,
            'total': reception.payment.total,
            'payment_record_date_1': payment_record_date_1,
            'payment_record_paid_1': payment_record_paid_1,
            'payment_record_date_2': payment_record_date_2,
            'payment_record_paid_2': payment_record_paid_2,
            'payment_record_date_3': payment_record_date_3,
            'payment_record_paid_3': payment_record_paid_3,
            'payment_record_date_4': payment_record_date_4,
            'payment_record_paid_4': payment_record_paid_4,
            'payment_record_date_5': payment_record_date_5,
            'payment_record_paid_5': payment_record_paid_5,
            'payment_record_unpaid': payment_record_unpaid,
            })

        datas.append(data)

    paginator = Paginator(count_page, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    context = {
            'datas':datas,
            'page_range_start':paging_data.paginator.page_range.start,
            'page_range_stop':paging_data.paginator.page_range.stop,
            'page_number':paging_data.number,
            'has_previous':paging_data.has_previous(),
            'has_next':paging_data.has_next(),

            'payment_total_subtotal':payment_subtotal,
            'payment_total_discount':payment_total_discount,
            'payment_total_total':payment_total,
            'payment_total_unpaid':payment_total_unpaid,
            'payment_paid_1': payment_paid_1,
            'payment_paid_2': payment_paid_2,
            'payment_paid_3': payment_paid_3,
            'payment_paid_4': payment_paid_4,
            'payment_paid_5': payment_paid_5,
            }
    
    return JsonResponse(context)

def statistics_ymw(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    
    
    return render(request, 
    'statistics/statistics_ymw.html',
        {
            'depart_medical':depart_medical,
        }
    )

def search_ymw(request):


    year = request.POST.get('year')
    depart = request.POST.get('depart')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    start = datetime.date(year = int(year), month=1,day=1)
    end = datetime.date(year = int(year), month=12,day=31)

    date_min = datetime.datetime.combine(start, datetime.time.min)
    date_max = datetime.datetime.combine(end, datetime.time.max)


    data_list_year = []
    data_list_monthly = []
    data_list_week = []
    data_list_hour = []

    query = Reception.objects.filter(
            **kwargs ,
            recorded_date__range = (date_min, date_max), 
            payment__paymentrecord__status='paid',
        ).exclude(
            progress='deleted'
        ).prefetch_related(
            'payment__paymentrecord_set'
        )

    #년도별
    year_query = query.aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

    data_list_year.append({
            'name':year ,
            'count':year_query['count'],
            'price_sum':0 if year_query['total_price'] is None else year_query['total_price'],
        })


    #월 별
    list_month = [
        {
            'name':'1',
            'date_start':datetime.datetime(year=int(year), month=1,day=1),
            'date_end': datetime.datetime(year=int(year), month=2,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'2',
            'date_start':datetime.datetime(year=int(year), month=2,day=1),
            'date_end': datetime.datetime(year=int(year), month=3,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'3',
            'date_start':datetime.datetime(year=int(year), month=3,day=1),
            'date_end': datetime.datetime(year=int(year), month=4,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'4',
            'date_start':datetime.datetime(year=int(year), month=4,day=1),
            'date_end': datetime.datetime(year=int(year), month=5,day=1) - datetime.timedelta(seconds = 1),
        },       
        {
            'name':'5',
            'date_start':datetime.datetime(year=int(year), month=5,day=1),
            'date_end': datetime.datetime(year=int(year), month=6,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'6',
            'date_start':datetime.datetime(year=int(year), month=6,day=1),
            'date_end': datetime.datetime(year=int(year), month=7,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'7',
            'date_start':datetime.datetime(year=int(year), month=7,day=1),
            'date_end': datetime.datetime(year=int(year), month=8,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'8',
            'date_start':datetime.datetime(year=int(year), month=8,day=1),
            'date_end': datetime.datetime(year=int(year), month=9,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'9',
            'date_start':datetime.datetime(year=int(year), month=9,day=1),
            'date_end': datetime.datetime(year=int(year), month=10,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'10',
            'date_start':datetime.datetime(year=int(year), month=10,day=1),
            'date_end': datetime.datetime(year=int(year), month=11,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'11',
            'date_start':datetime.datetime(year=int(year), month=11,day=1),
            'date_end': datetime.datetime(year=int(year), month=12,day=1) - datetime.timedelta(seconds = 1),
        },
        {
            'name':'12',
            'date_start':datetime.datetime(year=int(year), month=12,day=1),
            'date_end': datetime.datetime(year=int(year) + 1, month=1,day=1) - datetime.timedelta(seconds = 1),
        },
    ]


    
    for month in list_month: 
        year_query = query.filter(
            recorded_date__gte = month['date_start'],
            recorded_date__lt = month['date_end'],
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        data_list_monthly.append({
            'name':month['name'],
            'count':year_query['count'],
            'price_sum':0 if year_query['total_price'] is None else year_query['total_price'],

            })







    #요일별: 
    list_week= [
        {
            'name':_('Monday'),
            'value':2
        },
        {
            'name':_('Tuesday'),
            'value':3,
        },
        {
            'name':_('Wednesday'),
            'value':4,
        },
        {
            'name':_('Thursday'),
            'value':5,
        },       
        {
            'name':_('Friday'),
            'value':6,
        },
        {
            'name':_('Saturday'),
            'value':7,
        },
        {
            'name':_('Sunday'),
            'value':1,
        },
    ]

    for data in list_week:
        week_query = query.filter(
            recorded_date__week_day = data['value'] ,
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        data_list_week.append({
                'name':data['name'],
                'count':week_query['count'],
                'price_sum':0 if week_query['total_price'] is None else week_query['total_price'],
            })



    #시간별: 
    list_hour= [
        {'name':'8:00 ~ 8:29','hour':8,'min_min':00,'min_max':29},
        {'name':'8:30 ~ 8:59','hour':8,'min_min':30,'min_max':59},
        {'name':'9:00 ~ 9:29','hour':9,'min_min':00,'min_max':29},
        {'name':'9:30 ~ 9:59','hour':9,'min_min':30,'min_max':59},
        {'name':'10:00 ~ 10:29','hour':10,'min_min':00,'min_max':29},
        {'name':'10:30 ~ 10:59','hour':10,'min_min':30,'min_max':59},
        {'name':'11:00 ~ 11:29','hour':11,'min_min':00,'min_max':29},
        {'name':'11:30 ~ 11:59','hour':11,'min_min':30,'min_max':59},
        {'name':'12:00 ~ 12:29','hour':12,'min_min':00,'min_max':29},
        {'name':'12:30 ~ 12:59','hour':12,'min_min':30,'min_max':59},
        {'name':'13:00 ~ 13:29','hour':13,'min_min':00,'min_max':29},
        {'name':'13:30 ~ 13:59','hour':13,'min_min':30,'min_max':59},
        {'name':'14:00 ~ 14:29','hour':14,'min_min':00,'min_max':29},
        {'name':'14:30 ~ 14:59','hour':14,'min_min':30,'min_max':59},
        {'name':'15:00 ~ 15:29','hour':15,'min_min':00,'min_max':29},
        {'name':'15:30 ~ 15:59','hour':15,'min_min':30,'min_max':59},
        {'name':'16:00 ~ 16:29','hour':16,'min_min':00,'min_max':29},
        {'name':'16:30 ~ 16:59','hour':16,'min_min':30,'min_max':59},
        {'name':'17:00 ~ 17:29','hour':17,'min_min':00,'min_max':29},
        {'name':'17:30 ~ 17:59','hour':17,'min_min':30,'min_max':59},
        {'name':'18:00 ~ 18:29','hour':18,'min_min':00,'min_max':29},
        {'name':'18:30 ~ 18:59','hour':18,'min_min':30,'min_max':59},
        {'name':'19:00 ~ 19:29','hour':19,'min_min':00,'min_max':29},
        {'name':'19:30 ~ 19:59','hour':19,'min_min':30,'min_max':59},

        {'name':'20:00 ~ 20:29','hour':20,'min_min':00,'min_max':29},
        {'name':'20:30 ~ 20:59','hour':20,'min_min':30,'min_max':59},
        #{'name':'19:30 ~ 19:59','hour':19,'min_min':30,'min_max':59},        
    ]

    
    for data in list_hour:
        hour_query = query.filter(
            recorded_date__hour = data['hour'] ,
            recorded_date__minute__range = (data['min_min'],data['min_max']) ,
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        data_list_hour.append({
                'name':data['name'],
                'count':hour_query['count'],
                'price_sum':0 if hour_query['total_price'] is None else hour_query['total_price'],
            })
    new = datetime.datetime.now()
    current_month = new.month
    current_year = new.year

    date_start = datetime.datetime(year=int(current_year) , month=current_month, day=1) 
    # date_end = datetime.datetime(year=int(current_year) , month=current_month + 1, day=1 )- datetime.timedelta(seconds = 1)
    olds = Reception.objects.filter(
            **kwargs ,
            recorded_date__range = (date_min, date_max), 
            payment__paymentrecord__status='paid',
        ).exclude(
            progress='deleted'
        ).values_list('patient').distinct()

    news = Reception.objects.filter(
            **kwargs ,
            recorded_date__range = (date_min, date_max), 
            payment__paymentrecord__status='paid',
            patient__date_registered__gte = date_start
        ).exclude(
            progress='deleted'
        ).values_list('patient').distinct()

    list_patient = []
    list_patient.append({
        'old': olds.count(),
        'new': news.count()
    })

    return JsonResponse({
        'result':True,
        'datas_year':data_list_year,
        'datas_monthly':data_list_monthly,
        'datas_week':data_list_week,
        'datas_hour':data_list_hour,
        'old_new': list_patient
        })




def statistics_daily(request):


    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')
    

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })

    today_year = datetime.datetime.now().year
    today_month= datetime.datetime.now().month

    return render(request, 
    'statistics/statistics_daily.html',
        {
            'depart_medical':depart_medical,
            'today_year':today_year,
            'today_month':today_month,
        }
    )

def search_daily(request):

    year = int( request.POST.get('year') )
    month = int(request.POST.get('month') )
    depart = request.POST.get('depart')

    kwargs = {}
    if depart!='':
        kwargs['depart'] = depart # 기본 


    date_count = calendar.monthrange(int(year), int(month))[1]



    #date_min = datetime.datetime.combine(datetime.datetime.strptime(start,"%Y-%m-%d").date(), datetime.time.min)
    #date_max = datetime.datetime.combine(datetime.datetime.strptime(end,"%Y-%m-%d").date(), datetime.time.max)


    query = Reception.objects.filter(
            **kwargs ,
            recorded_date__year = year,
            recorded_date__month= month,
            payment__paymentrecord__status='paid',
        ).exclude(
            progress='deleted'
        ).prefetch_related(
            'payment__paymentrecord_set'
        )    

    list_date_count = []
    list_date_revenue = []
    Total_Due_Amount = 0      
    Total_Actual_Amount = 0
    Total_Debit_Payment = 0
    Total_paid = 0
    Total_unPaid = 0
    Total_visit = 0
    for count in range(1,date_count + 1):
        payment_total_paid_amount = 0
        payment_total_subtotal = 0
        payment_total_additional = 0
        payment_total_total = 0
        payment_total_discount = 0
        payment_total_unpaid = 0  
        Due_Amount = 0      
        Actual_Amount = 0
        Debit_Payment = 0
        tmp_date = datetime.datetime(year = year, month = month, day = count)


        tmp_receptions = Reception.objects.filter(
                **kwargs ,
                payment__paymentrecord__date__date = tmp_date,
                payment__paymentrecord__status='paid',
            ).exclude(
                progress='deleted'
            ).order_by("-id")     

        rec_argument_list = []
        rec_argument_list.append( Q(**{'id': 0000000} ) )
        for reception in tmp_receptions:
            payment_records = reception.payment.paymentrecord_set.filter(status='paid')
            # if payment_records.count() != 0 :
            #     if payment_records.first().date.date() == reception.recorded_date.date():
            #         if payment_records.first().paid == reception.payment.total:
            #             continue 
            # rec_argument_list.append( Q(**{'id': payment_record.id} ) )           
            if payment_records.count() != 0 :
                for payment_record in payment_records:
                    # print('tmp_date: ',tmp_date.date())
                    # print('reception.recorded_date.date(): ',reception.recorded_date.date())
                    if tmp_date.date() != reception.recorded_date.date():
                        # if payment_record.paid == reception.payment.total:
                        # continue
                        rec_argument_list.append( Q(**{'id': payment_record.id} ) )
        # print('rec_argument_list: ',rec_argument_list)
        debit_real_paid_total = PaymentRecord.objects.filter(
            functools.reduce(operator.or_, rec_argument_list),
                date__date = tmp_date,
                status = 'paid',
                ).aggregate(
            Count('id'),
            Sum('paid'),
            )      

        # debit_real_paid_total = PaymentRecord.objects.filter(
        #         date__date = tmp_date,
        #         status = 'paid',
        #         ).exclude(
        #         recorded_date__date= tmp_date
        #         ).aggregate(
        #     Count('id'),
        #     Sum('paid'),
        #     )               

        Debit_Payment =  0 if debit_real_paid_total['paid__sum'] is None else debit_real_paid_total['paid__sum']

        receptions = Reception.objects.filter(
                **kwargs ,
                recorded_date__date = tmp_date,
                payment__paymentrecord__status='paid',
            ).exclude(
                progress='deleted'
            )          
        payment_total = receptions.filter().aggregate(
            Sum('payment__sub_total'),
            Sum('payment__total'),
            Sum('payment__additional'),
            )
            

        # real_paid_total = query.filter( Q(payment__paymentrecord__status='paid'), recorded_date__date = tmp_date, ).aggregate(
        #     Sum('payment__paymentrecord__paid'),
        #     )            
        date_query = query.filter(
            recorded_date__date = tmp_date,
            ).aggregate(
                count = Count('id'),
                total_price = Sum('payment__paymentrecord__paid')
            )

        payment_total_paid_amount =  0 if date_query['total_price'] is None else date_query['total_price']

        payment_total_subtotal = 0 if payment_total['payment__sub_total__sum'] is None else payment_total['payment__sub_total__sum']
        payment_total_additional = 0 if payment_total['payment__additional__sum'] is None else payment_total['payment__additional__sum']
        payment_total_total = 0 if payment_total['payment__total__sum'] is None else payment_total['payment__total__sum']
        payment_total_discount = payment_total_subtotal + payment_total_additional - payment_total_total
        payment_total_unpaid = payment_total_total - payment_total_paid_amount
        Due_Amount = payment_total_subtotal + payment_total_additional - payment_total_discount
        Actual_Amount = Due_Amount + Debit_Payment


        if date_query['count'] is 0 and Debit_Payment is 0:
            continue

        visit =  0 if date_query['count'] is None else date_query['count']
        paid = 0 if date_query['total_price'] is None else date_query['total_price'] + Debit_Payment

        Total_Due_Amount += Due_Amount
        Total_Actual_Amount += Actual_Amount
        Total_Debit_Payment += Debit_Payment
        Total_paid += paid + Debit_Payment
        Total_unPaid += payment_total_unpaid
        Total_visit += visit

        date_query.update({ 'Due_Amount':Due_Amount })
        date_query.update({ 'payment_total_unpaid':payment_total_unpaid })
        date_query.update({ 'payment_total_subtotal':payment_total_subtotal })
        date_query.update({ 'Actual_Amount':Actual_Amount })
        date_query.update({ 'Debit_Payment':Debit_Payment })
        date_query.update({ 'paid':paid })


        list_date_count.append({
            'id':1,
            'start':tmp_date.strftime('%Y-%m-%d'),
            'title':'4.Visits : ' + str(date_query['count']),
            'backgroundColor':'rgb(254,154,202)',
            'borderColor':'#FFA500',
            })                 

        list_date_count.append({
            'id':4,
            'start':tmp_date.strftime('%Y-%m-%d'),
            'title':'3.Actual Amount : ' + str(0 if date_query['Actual_Amount'] is None else "{:,}".format(date_query['Actual_Amount'])) + " VND",
            'backgroundColor':'rgb(254,154,202)',
            'borderColor':'rgb(254,154,202)',
            }) 
        list_date_count.append({
            'id':5,
            'start':tmp_date.strftime('%Y-%m-%d'),
            'title':'2.Debit payment : ' + str(0 if date_query['Debit_Payment'] is None else "{:,}".format(date_query['Debit_Payment'])) + " VND",
            'backgroundColor':'rgb(254,154,202)',
            'borderColor':'rgb(254,154,202)',
            })              
        list_date_count.append({
            'id':2,
            'start':tmp_date.strftime('%Y-%m-%d'),
            'title':'3.2.Unpaid : ' + str(0 if date_query['payment_total_unpaid'] is None else "{:,}".format(date_query['payment_total_unpaid'])) + " VND",
            'backgroundColor':'rgb(254,154,202)',
            'borderColor':'blue',
            })                 
        list_date_count.append({
            'id':3,
            'start':tmp_date.strftime('%Y-%m-%d'),
            'title':'3.1.Total Paid : ' + str(0 if date_query['paid'] is None else "{:,}".format(date_query['paid'])) + " VND",
            'backgroundColor':'rgb(254,154,202)',
            'borderColor':'blue',
            })
        list_date_count.append({
            'id':6,
            'start':tmp_date.strftime('%Y-%m-%d'),
            'title':'1.Due Amount : ' + str(0 if date_query['Due_Amount'] is None else "{:,}".format(date_query['Due_Amount'])) + " VND",
            'backgroundColor':'rgb(254,154,202)',
            'borderColor':'rgb(254,154,202)',
            })   

    return JsonResponse({
        'result':True,
        'datas_date_count':list_date_count,
        'datas_date_revenue':list_date_revenue,
        'Total_Due_Amount':Total_Due_Amount,
        'Total_Actual_Amount':Total_Actual_Amount,
        'Total_Debit_Payment':Total_Debit_Payment,
        'Total_paid':Total_paid,
        'Total_unPaid':Total_unPaid,
        'Total_visit':Total_visit,
        })



@login_required
def code_setting(request):

    
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')

    
    list_upper=[]
    upper_query = COMMCODE.objects.values('upper_commcode','upper_commcode_name').annotate(Count('upper_commcode'))
    for data in upper_query:

        list_upper.append({
            'name':data['upper_commcode_name'],
            'code':data['upper_commcode'],
            })


    return render(request,
    'Manage/code_setting.html',
            {
                
            'list_upper':list_upper,
            },
        )

@login_required
def code_search(request):


    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')

 

    category=request.POST.get("category")
    string=request.POST.get("string")

    kwargs = {}
    if category != '':
        kwargs['upper_commcode_name']=category
    argument_list = [] 
    #if string != '':
    #argument_list.append( Q(**{'name_kor__icontains':string} ) )
    argument_list.append( Q(**{'commcode_name_ko__icontains':string} ) )
    argument_list.append( Q(**{'commcode_name_en__icontains':string} ) )
    argument_list.append( Q(**{'commcode_name_vi__icontains':string} ) )

    #if project_type != '':
    #    kwargs['type'] = project_type



    datas = []
    query = COMMCODE.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs,
        use_yn = 'Y')



     



    for data in query:

        datas.append({
            'id':data.id,

            'upper_commcode':data.upper_commcode,
            'upper_commcode_name':data.upper_commcode_name,
            'commcode_grp':data.commcode_grp,
            'commcode_grp_name':data.commcode_grp_name,
            'commcode':data.commcode,
            'commcode_name_ko':data.commcode_name_ko,
            'commcode_name_en':data.commcode_name_en,
            'commcode_name_vi':data.commcode_name_vi,
            'se1':data.se1,
            'se2':data.se2,
            'se3':data.se3,
            'se4':data.se4,
            'se5':data.se5,
            'se6':data.se6,
            'se7':data.se7,
            'se8':data.se8,
            'seq':data.seq,
            'registrerer':data.registrerer,
            'date_of_registered':data.date_of_registered.strftime('%Y-%m-%d'),
            })


    page_context = request.POST.get('page_context',10) # 페이지 컨텐츠 
    page = request.POST.get('page',1)

    paginator = Paginator(datas, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    context = {
            'datas':list(paging_data),
            'page_range_start':paging_data.paginator.page_range.start,
            'page_range_stop':paging_data.paginator.page_range.stop,
            'page_number':paging_data.number,
            'has_previous':paging_data.has_previous(),
            'has_next':paging_data.has_next(),


            }



    return JsonResponse(context)


@login_required
def code_save(request):

    
    id=request.POST.get("id")

    code_upper_commcode=request.POST.get("code_upper_commcode")
    code_upper_commcode_name=request.POST.get("code_upper_commcode_name")
    code_commcode_grp=request.POST.get("code_commcode_grp")
    code_commcode_grp_name=request.POST.get("code_commcode_grp_name")
    code_commcode=request.POST.get("code_commcode")
    code_commcode_name_ko=request.POST.get("code_commcode_name_ko")
    code_commcode_name_en=request.POST.get("code_commcode_name_en")
    code_commcode_name_vi=request.POST.get("code_commcode_name_vi")
    code_se1=request.POST.get("code_se1")
    code_se2=request.POST.get("code_se2")
    code_se3=request.POST.get("code_se3")
    code_se4=request.POST.get("code_se4")
    code_se5=request.POST.get("code_se5")
    code_se6=request.POST.get("code_se6")
    code_se7=request.POST.get("code_se7")
    code_se8=request.POST.get("code_se8")
    code_seq=request.POST.get("code_seq")
    code_use_yn=request.POST.get("code_use_yn")


    if id == '':
        commcode = COMMCODE()

        commcode.registrerer = request.user
    else:
        commcode = COMMCODE.objects.get(id = id)


    commcode.upper_commcode = code_upper_commcode
    commcode.upper_commcode_name = code_upper_commcode_name
    commcode.commcode_grp = code_commcode_grp
    commcode.commcode_grp_name = code_commcode_grp_name
    commcode.commcode = code_commcode
    commcode.commcode_name_ko = code_commcode_name_ko
    commcode.commcode_name_en = code_commcode_name_en
    commcode.commcode_name_vi = code_commcode_name_vi
    commcode.se1 = code_se1
    commcode.se2 = code_se2
    commcode.se3 = code_se3
    commcode.se4 = code_se4
    commcode.se5 = code_se5
    commcode.se6 = code_se6
    commcode.se7 = code_se7
    commcode.se8 = code_se8
    commcode.seq = code_seq

    commcode.use_yn = code_use_yn

    commcode.lastest_modifier = request.user.id
    commcode.lastest_modified_date = datetime.datetime.now()

    commcode.save()



    return JsonResponse({
        'result':True,        
        })


@login_required
def code_get(request):

    
    id=request.POST.get("id")

    code = COMMCODE.objects.get(id = id)



    return JsonResponse({

        "code_upper_commcode":code.upper_commcode,
        "code_upper_commcode_name":code.upper_commcode_name,
        "code_commcode_grp":code.commcode_grp,
        "code_commcode_grp_name":code.commcode_grp_name,
        "code_commcode":code.commcode,
        "code_commcode_name_ko":code.commcode_name_ko,
        "code_commcode_name_en":code.commcode_name_en,
        "code_commcode_name_vi":code.commcode_name_vi,
        "code_se1":code.se1,
        "code_se2":code.se2,
        "code_se3":code.se3,
        "code_se4":code.se4,
        "code_se5":code.se5,
        "code_se6":code.se6,
        "code_se7":code.se7,
        "code_se8":code.se8,
        "code_seq":code.seq,
        "code_use_yn":code.use_yn,



        })




@login_required
def code_delete(request):

    id=request.POST.get("id")

    code = COMMCODE.objects.get(id = id)
    code.use_yn = 'N'

    code.lastest_modifier = request.user.id
    code.lastest_modified_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    code.save()


    return JsonResponse({
        'result':True,    
        })

@login_required
def get_user_by_depart(request):
    f_name = F('name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('name_vi')



    depart=request.POST.get("depart")

    if 'DOCTOR' in depart:
        doctor_depart = depart.split('_')
        users = User.objects.filter(depart = doctor_depart[0],depart_doctor=doctor_depart[1]).annotate(name = f_name).values('name','id','user_id')
    else:
        users = User.objects.filter(depart = depart).annotate(name = f_name).values('name','id','user_id')


    datas = []
    for data in users:
        datas.append({
            'id':data['id'],
            'name':data['name'],
            'user_id':data['user_id'],
            })

    return JsonResponse({
        'result':True,    
        'datas':datas,
        })




def get_alert(request):
    one_week_before= datetime.datetime.today() - datetime.timedelta(days=7)
    one_week_after= datetime.datetime.today() + datetime.timedelta(days=7)
    date_min = datetime.datetime.combine(datetime.datetime.today(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.today() + datetime.timedelta(days=7), datetime.time.max)

    #협업보드 / 지목
    cowork_point = AlertLog.objects.filter(
        page_type="COWORK",
        content_type="COMMENT",
        user_id=request.user.id,
        created_date__gte = one_week_before,
        use_yn = 'Y',
        check_yn = 'N',
        )

    #협업보드 / 완료 예정일
    cowork_expected = AlertLog.objects.filter(
        page_type="COWORK",
        content_type="COMMENT",
        user_id=request.user.id,
        pointed_date__range = (date_min, date_max), 
        use_yn = 'Y'
        ).exclude(
            Q(status='DONE') | Q(status='CANCEL')
            )

    #기안서 / 대기
    draft_alert = AlertLog.objects.filter(
            page_type="DRAFT",
            content_type="CONTENT",
            creator = request.user.id,
            use_yn = 'Y'
            ).exclude(
            Q(status='DONE') | Q(status='CANCEL')
            )

    #프로젝트 ( KBL )
    project_main_mng = AlertLog.objects.filter(
        page_type="PROJECT",
        content_type="MAIN",
        user_id = request.user.id,
        use_yn = 'Y',
        check_yn = 'N'
        ).exclude(
            Q( status = 'CANCLE') | Q( status='DONE'),
            )

    project_comment_mng = AlertLog.objects.filter(
        page_type="PROJECT",
        content_type="COMMENT",
        user_id = request.user.id,
        use_yn = 'Y',
        check_yn = 'N'
        ).exclude(
            Q( status = 'CANCLE') | Q( status='DONE'),
            )

    list_reservation = []
    if 'ALERT' in  request.session:
        if 'vaccine' in request.session['ALERT']:
            print(date_min)
            print(date_max)

            argument_list = []
            argument_list.append( Q(**{'reservation_date__range':(date_min, date_max) } ) ) 
            argument_list.append( Q(**{'re_reservation_date__range':(date_min, date_max) } ) ) 

            reservations = Reservation.objects.filter(
                functools.reduce(operator.or_, argument_list),
                division = 'VACCIN',
                )

            list_reservation = []
            for reservation in reservations:
                if reservation.re_reservation_date != None:
                    if reservation.re_reservation_date >= date_min and reservation.re_reservation_date < one_week_after:
                        list_reservation.append(reservation.id)
                else:
                    list_reservation.append(reservation.id)


    return JsonResponse({

        'cowork_count':cowork_point.count() + cowork_expected.count(),
        'cowork_point':cowork_point.count(),
        'cowork_expected':cowork_expected.count(),

        'draft_count':draft_alert.count(),
        'draft_requested':draft_alert.filter(status = 'REQUEST').count(),
        'draft_waiting':draft_alert.filter(status = 'WAITING').count(),
        'draft_pending':draft_alert.filter(status = 'PENDING').count(),

        'project_count':project_main_mng.count()+project_comment_mng.count(),
        'project_main':project_main_mng.count(),
        'project_comment':project_comment_mng.count(),

        'vaccine_reserv_count':len(list_reservation),
        })



