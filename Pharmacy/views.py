from django.shortcuts import render 
from django.http import JsonResponse
import datetime
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage
from django.contrib.auth.decorators import login_required
from django.forms.models import model_to_dict

from .models import *
from .forms import *
from Doctor.models import *
from Receptionist.models import *
from app.models import *
# Create your views here.
from django.utils import timezone, translation
from django.utils.translation import gettext as _
from django.db.models import Q, Count, F, Min,Sum
from django.db.models.query import QuerySet
import operator
import functools
import shutil
from openpyxl import Workbook,load_workbook
from django.db.models.functions import Lower

@login_required
def index(request):

    waiting_search_form = WaitingSearchForm()
    medicine_search_form = MedicineSearchForm()
    medicine_control_form = MedicineControl()
    depart = Depart.objects.all()

    return render(request,
    'Pharmacy/index.html',
            {
                'waiting_search':waiting_search_form,
                'medicinesearch':medicine_search_form,
                'medicine_control':medicine_control_form,
                'depart':depart,
            },
        )

def waiting_selected(request):
    #check
    diagnosis_id = request.POST.get('diagnosis_id')

    diagnosis = Diagnosis.objects.get(pk = diagnosis_id)
    medicine_set = MedicineManager.objects.filter(diagnosis_id = diagnosis_id).order_by(Lower('medicine__unit'))

    try:
        medicine_manage = MedicineManage.objects.get(diagnosis_id = diagnosis_id)
        if request.user.depart == 'PHARMACY':
            if medicine_manage.progress == 'changed' or medicine_manage.progress == 'new':
                medicine_manage.progress = 'checked'
                medicine_manage.save()
    except MedicineManage.DoesNotExist:
        pass



    medicines = []
    for data in medicine_set:
        medicine = {}
        medicine.update({
            'code':data.medicine.id,
            'name':data.medicine.name,
            'depart':diagnosis.reception.depart.name,
            'doctor':diagnosis.reception.doctor.name_kor,
            'unit':'' if data.medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else data.medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
            'amount':data.amount,
            'days':data.days,
            'total':data.amount * data.days,
            'memo':data.memo,
            'price': data.medicine.get_price()
            })
        medicines.append(medicine)

    context = {
        'datas':medicines,
        'status':diagnosis.medicinemanage.progress,
        'patient_name':diagnosis.reception.patient.get_name_kor_eng(),

        'need_invoice':diagnosis.reception.need_invoice,
        'need_insurance':diagnosis.reception.need_insurance,


        'chart':diagnosis.reception.patient.get_chart_no(),
        'Name':diagnosis.reception.patient.name_kor + ' / ' + diagnosis.reception.patient.name_eng,
        'Depart':diagnosis.reception.depart.name + ' ( ' + diagnosis.reception.doctor.name_kor + ' )',
        'Date_of_Birth': diagnosis.reception.patient.date_of_birth.strftime('%Y-%m-%d'),
        'gender': diagnosis.reception.patient.gender,
        'phone': diagnosis.reception.patient.phone,
        'diagnosis': diagnosis.diagnosis,
        'address': diagnosis.reception.patient.address
               }
    return JsonResponse(context)

def waiting_list(request):
    date_start = request.POST.get('start_date')
    date_end = request.POST.get('end_date')
    depart_id = request.POST.get('depart_id')
    
    filter = request.POST.get('filter','')
    string = request.POST.get('string','')



    argument_list = [] 

    kwargs={}
    if filter == '':
        argument_list.append(Q(**{'diagnosis__reception__patient__name_kor__icontains':string} ))
        argument_list.append(Q(**{'diagnosis__reception__patient__name_eng__icontains':string} ))
        argument_list.append(Q(**{'diagnosis__reception__patient__id__icontains':string} ))
    elif filter =='name':
        argument_list.append(Q(**{'diagnosis__reception__patient__name_kor__icontains':string} ))
        argument_list.append(Q(**{'diagnosis__reception__patient__name_eng__icontains':string} ))
    elif filter =='chart':
        argument_list.append(Q(**{'diagnosis__reception__patient__id__icontains':string} ))
    if depart_id != '' :
        kwargs['diagnosis__reception__depart_id'] = depart_id

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)


    medicine_manages = MedicineManage.objects.select_related(
        'diagnosis__reception',
        'diagnosis__reception__patient',
        ).filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs,
        date_ordered__range = (date_min, date_max),
        ).order_by('date_ordered')

    datas=[]
    today = datetime.date.today()

    #약국의 상태가 hold / done 인것은 상관없이 출력.
    #진료 상태가 완료인것만 출력


    for medicine_manage in medicine_manages:

        if medicine_manage.progress =='new' or medicine_manage.progress =='changed' or medicine_manage.progress =='checked':
            if medicine_manage.diagnosis.reception.progress != 'done':
                continue

        data={}
        data.update({
            'diagnosis_id':medicine_manage.diagnosis.id,
            'chart':medicine_manage.diagnosis.reception.patient.get_chart_no(),
            'Name':medicine_manage.diagnosis.reception.patient.name_kor + ' / ' + 
                medicine_manage.diagnosis.reception.patient.name_eng,
            'Date_of_Birth':medicine_manage.diagnosis.reception.patient.date_of_birth.strftime('%Y-%m-%d') + ' ( ' + 
                str(medicine_manage.diagnosis.reception.patient.get_age()) + ' / ' + 
                medicine_manage.diagnosis.reception.patient.get_gender_simple() + ' )',
            'Depart':medicine_manage.diagnosis.reception.depart.name + ' ( ' + medicine_manage.diagnosis.reception.doctor.name_kor +' )',
            #'DateTime':medicine_manage.date_ordered.strftime('%Y-%m-%d %H:%M:%S'),
            'status':medicine_manage.progress,
            'ordered':'' if medicine_manage.date_ordered is None else medicine_manage.date_ordered.strftime("%Y-%m-%d %H:%M"),
            'received':'' if medicine_manage.date_received is None else medicine_manage.date_received.strftime("%Y-%m-%d %H:%M"),
            'total': medicine_manage.diagnosis.reception.payment.total
            })
        datas.append(data)

    datas.reverse()
    context = {'datas':datas}
    return JsonResponse(context)


def save(request):
    diagnosis_id = request.POST.get('diagnosis_id')
    status = request.POST.get('status',None)

    medicinmanage = MedicineManage.objects.get(diagnosis_id = diagnosis_id)
    if status == 'done':
        medicine_set = MedicineManager.objects.filter(diagnosis_id = diagnosis_id)

        medicinmanage.date_received = datetime.datetime.now()

        for data in medicine_set:
            log = MedicineLog(
                type='dec',
                )
            medicine = Medicine.objects.get(pk = data.medicine_id)
            medicine.inventory_count -= data.amount * data.days
            medicine.save()

            changes = data.amount * data.days

            log.diagnosis_id = diagnosis_id
            log.changes = changes
            log.medicine = medicine
            log.save()

            tmp_log = MedicineLog.objects.filter(medicine = medicine, type='add' ).exclude(tmp_count=0).order_by('expiry_date')
            for tmp in tmp_log:
                if tmp.tmp_count is not None:
                    if tmp.tmp_count < changes :
                        changes -= tmp.tmp_count
                        tmp.tmp_count = 0 
                        tmp.save()
                    else:
                        tmp.tmp_count -= changes
                        tmp.save()
                        break

    
    medicinmanage.progress = status
    medicinmanage.save()

    context = {'result':True}
    return JsonResponse(context)



def withdraw(request):
    res = True
    msg = ''

    diagnosis_id = request.POST.get('diagnosis_id')
    diagnosis = Diagnosis.objects.get(id = diagnosis_id)
  
    
    medicinmanage = MedicineManage.objects.get(diagnosis_id = diagnosis_id)

    if medicinmanage.progress == 'done':
        medicine_set = MedicineManager.objects.filter(diagnosis_id = diagnosis_id)

        medicinmanage.date_received = None
        medicinmanage.progress = 'checked'

        for data in medicine_set:
            log = MedicineLog(
                type='add',
                tmp_count=0
                )

            changes = data.amount * data.days
            medicine = Medicine.objects.get(pk = data.medicine_id)
            medicine.inventory_count += changes
            medicine.save()

            log.diagnosis_id = diagnosis_id
            log.changes = changes
            log.medicine = medicine
            log.memo = 'withdraw'
            log.save()

            tmp_log = MedicineLog.objects.filter(medicine = medicine, type='add' ).exclude(tmp_count=0).order_by('expiry_date')
            for tmp in tmp_log:
                if tmp.tmp_count is not None:
                    tmp.tmp_count += changes
                    tmp.save()
                    break

        medicinmanage.save()
    else:
        res = False
        msg = _('Only available in done status.')

    return JsonResponse({
        'result':res,
        'msg':msg,
        })


def medicine_search(request):
    string = request.POST.get('string')
    filter = request.POST.get('filter')
    class_id = request.POST.get('class_id')
    

    kwargs = {}
    if class_id != '':
        kwargs.update({
            'medicine_class_id':class_id,
            })

    datas=[]
    argument_list = []
    if string == '' and class_id == '':

        expiry_date = datetime.datetime.now() + datetime.timedelta(days=180)
        #medicine_tmp = MedicineLog.objects.filter(type='add',expiry_date__lte = expiry_date ).select_related('medicine').exclude(medicine__use_yn='N' ,expiry_date=None,tmp_count=0).order_by('expiry_date')
        medicine_tmp= MedicineLog.objects.filter( medicine__type="PHARM", type='add', expiry_date__lte = expiry_date, tmp_count__gte= 0 ).select_related('medicine').values(
            'medicine_id',
            ).annotate(Count('medicine_id')).order_by('expiry_date')
       
        #medicine_tmp = MedicineLog.objects.filter( type='add', expiry_date__lte = expiry_date).exclude(tmp_count__lt = 0,medicine__use_yn='N',expiry_date=None).order_by('expiry_date').select_related('medicine')
        for tmp in medicine_tmp:
            try:
                medicine = Medicine.objects.get(id = tmp['medicine_id'])
                data = {
                        'id' : medicine.id,
                        'code': medicine.code,
                        'tax': medicine.tax_rate,
                        'name' : medicine.name,
                        'company' : '' if medicine.company is None else medicine.company,
                        'country' : '' if medicine.get_country_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_country_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                        'ingredient' : '' if medicine.get_ingredient_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_ingredient_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                        'unit' : '' if medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                        'price' : medicine.get_price(),
                        'count' : medicine.inventory_count,
                        'alaert_expiry':True,
                    }
                datas.append(data)
            except Medicine.DoesNotExist:
                pass

        medicines = Medicine.objects.filter( type="PHARM", **kwargs).exclude(use_yn='N').order_by("id")#.select_related('medicine_class').exclude(use_yn = 'N').order_by("name")
            
    else:
        #argument_list = [] 
        argument_list.append( Q(**{'name__icontains':string} ) )
        argument_list.append( Q(**{'name_vie__icontains':string} ) )
        argument_list.append( Q(**{'name_display__icontains':string} ) )

        medicines = Medicine.objects.filter(functools.reduce(operator.or_, argument_list), type="PHARM", **kwargs).exclude(use_yn='N').order_by("id")#.select_related('medicine_class').exclude(use_yn = 'N').order_by("name")


    
    for medicine in medicines:
        data = {
                'id' : medicine.id,
                'code': medicine.code,
                'name' : medicine.name,
                'tax' : medicine.tax_rate,
                'company' : '' if medicine.company is None else medicine.company,
                'country' : '' if medicine.get_country_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_country_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'ingredient' : '' if medicine.get_ingredient_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_ingredient_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'unit' : '' if medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]) is None else medicine.get_unit_lang(request.session[translation.LANGUAGE_SESSION_KEY]),
                'price' : medicine.get_price(),
                'count' : medicine.inventory_count,
            }

        datas.append(data)


    page = request.POST.get('page',1)
    context_in_page = request.POST.get('context_in_page');
    paginator = Paginator(datas, context_in_page)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    

    context = {
        #'datas':datas,
        'datas':list(paging_data),
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        
        }
    return JsonResponse(context)


def set_data_control(request):
    medicine_id = request.POST.get('medicine_id');


    medicine = Medicine.objects.get(pk = medicine_id)

    context = {
        'name':medicine.name,
        'company':medicine.company,
        'country':medicine.country,
        'ingredient':medicine.ingredient,
        'unit':medicine.unit,
        'price':medicine.get_price(),
        }
    return JsonResponse(context)


def save_data_control(request):
    selected_option = request.POST.get('selected_option');
    name = request.POST.get('name');
    price = request.POST.get('price');
    company = request.POST.get('company');
    ingredient = request.POST.get('ingredient');
    unit = request.POST.get('unit');
    changes = request.POST.get('changes') if request.POST.get('changes') is not '' else 0;


    if selected_option == 'new':
        medicine = Medicine()
        log = MedicineLog(type='new')
        code = Medicine.objects.all().last()
        code = 'M{:04d}'.format(code.id)
        medicine.price = price
    else:
        medicine = Medicine.objects.get(pk = selected_option )
        if int(changes) < 0:
            log = MedicineLog(type='dec')
        else:
            log = MedicineLog(type='add')

    medicine.name = name
    #medicine.price = price
    medicine.company = company
    medicine.ingredient = ingredient
    medicine.unit = unit

    medicine.inventory_count += int(changes)
    medicine.save()
    log.medicine = medicine
    log.changes = int(changes)
    log.save()
     
    context = {'result':True}
    return JsonResponse(context)

def inventory(request):
    waiting_search_form = WaitingSearchForm()
    medicine_search_form = MedicineSearchForm()
    medicine_control_form = MedicineControl()

    price_multiple_level = COMMCODE.objects.filter(commcode_grp = 'MED_MULTI_CODE').values('commcode','se1','se2').order_by('commcode_grp')


    if request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        medicine_class = MedicineClass.objects.filter(use_yn='Y').annotate(name_display = F('name_vie')).values('id','name_display')
    else:
        medicine_class = MedicineClass.objects.filter(use_yn='Y').annotate(name_display = F('name')).values('id','name_display')

    type = ["Medicine","Injection","Vaccine", "Equipment"]
    
    return render(request,
    'Pharmacy/inventory.html',
            {
                'waiting_search':waiting_search_form,
                'medicinesearch':medicine_search_form,
                'medicine_control':medicine_control_form,
                'price_multiple_level':price_multiple_level,
                'medicine_class':medicine_class,
                'type':type,

            },
        )


def medicine_add_edit_get(request):
    id = request.POST.get('id')

    try:
        medicine = Medicine.objects.get(id=id)

        context = {
            'result':True,
            'id':medicine.id,
            'code':medicine.code,
            'name':medicine.name,
            'name_vie':medicine.name_vie,
            'name_display':medicine.name_display,
            'country':medicine.country,
            'country_vie':medicine.country_vie,
            'ingredient':medicine.ingredient,
            'ingredient_vie':medicine.ingredient_vie,
            'vaccine_code':medicine.vaccine_code,
            'vaccine_recommend_time':medicine.vaccine_recommend_time,
            'unit':medicine.unit,
            'unit_vie':medicine.unit_vie,
            'company':medicine.company,

            'type':'Medicine' if 'M' in medicine.code else 'Injection',

            'price':medicine.get_price(),
            'price_input':medicine.get_price_input(),
            'price_dollar':medicine.get_price_dollar(),
            'multiple_level':medicine.multiple_level,

            'inventory_count':medicine.inventory_count,
            'medicine_class_id':medicine.medicine_class_id,

            'red_invoice':medicine.red_invoice,
            'tax':medicine.tax_rate,
            }
    except Medicine.DoesNotExist:
        context = {'result':False}




    return JsonResponse(context)


def medicine_add_edit_set(request):
    id = int(request.POST.get('id'))
    code = request.POST.get('code')
    type = request.POST.get('type')
    medicine_class = request.POST.get('medicine_class')
    name = request.POST.get('name')
    name_vie = request.POST.get('name_vie')
    ingredient = request.POST.get('ingredient')
    ingredient_vie = request.POST.get('ingredient_vie')
    vaccine_code = request.POST.get('vaccine_code')
    vaccine_recommend_time = request.POST.get('vaccine_recommend_time')
    unit = request.POST.get('unit')
    unit_vie = request.POST.get('unit_vie')
    country = request.POST.get('country')
    country_vie = request.POST.get('country_vie')
    company = request.POST.get('company')
    name_display = request.POST.get('name_display')
    price_input = request.POST.get('price_input')
    multiple_level = request.POST.get('multiple_level')
    price = request.POST.get('price')   
    price_dollar = request.POST.get('price_dollar')
    red_invoice = request.POST.get('red_invoice')
    tax = request.POST.get('tax')

    if int(id) == 0 :
        data = Medicine()
        if type in 'Medicine':
            last_code =Medicine.objects.filter(code__icontains="M",type='PHARM').last()
            if last_code == None:
                temp_code = 'M0000'.split('M')
            else:
                temp_code = last_code.code.split('M')
            code = 'M' + str('%04d' % (int(temp_code[1]) + 1))

        elif type in 'Injection':
            last_code =Medicine.objects.filter(code__icontains="I",type='PHARM').last()
            if last_code == None:
                temp_code = 'I0000'.split('I')
            else:
                temp_code = last_code.code.split('I')
            code = 'I' + str('%04d' % (int(temp_code[1]) + 1))

        elif type in "Vaccine":
            last_code =Medicine.objects.filter(code__icontains="VC",type='PHARM').last()
            if last_code == None:
                temp_code = 'VC0000'.split('VC')
            else:
                temp_code = last_code.code.split('VC')
            code = 'VC' + str('%04d' % (int(temp_code[1]) + 1))

        elif type in "Equipment":
            last_code = Medicine.objects.filter(code__icontains="E",type='PHARM').last()
            if last_code == None:
                temp_code = 'E0000'.split('E')
            else:
                temp_code = last_code.code.split('E')
            code = 'E' + str('%04d' % (int(temp_code[1]) + 1))



        data.code = code
        data.price = int(price)
        price_input = int(price_input)
        data.price_dollar = int(price_dollar)
       
    else:
        data = Medicine.objects.get(id=id)
        str_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')

        try: 
            old_price = Pricechange.objects.get(type="Medicine",country='VI',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price.price != int(price):
                old_price.date_end = str_now
                old_price.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price != int(price):
                new_price = Pricechange(type="Medicine",country='VI',type2='OUTPUT',code=data.code)
                new_price.price = price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()


        try:
            old_price_input = Pricechange.objects.get(type="Medicine",country='VI',type2='INPUT',code=data.code, date_end="99999999999999")
            
            if old_price_input.price != int(price_input):
                old_price_input.date_end = str_now
                old_price_input.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=data.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price_input != int(price_input):
                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=data.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        try:
            old_price_dollar = Pricechange.objects.get(type="Medicine",country='US',type2='OUTPUT',code=data.code, date_end="99999999999999")
            
            if old_price_dollar.price != int(price_dollar):
                old_price_dollar.date_end = str_now
                old_price_dollar.save()

                new_price = Pricechange(type="Medicine",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if data.price_dollar != int(price_dollar):
                new_price = Pricechange(type="Medicine",country='US',type2='OUTPUT',code=data.code)
                new_price.price = price_dollar
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()


    data.medicine_class_id = medicine_class
    data.name = name
    data.name_vie = name_vie
    data.ingredient = ingredient
    data.ingredient_vie = ingredient_vie
    data.vaccine_code = vaccine_code
    data.vaccine_recommend_time = vaccine_recommend_time
    data.unit = unit
    data.unit_vie = unit_vie
    data.country = country
    data.country_vie = country_vie
    data.company = company
    data.name_display = name_display
    data.multiple_level = multiple_level
    data.red_invoice = red_invoice
    data.tax_rate = tax

    data.save()

    context = {'result':True}




    return JsonResponse(context)


def medicine_add_edit_check_code(request):
    code = request.POST.get('code')
    id = request.POST.get('id')
    try:
        medicine = Medicine.objects.get(code=code)
        res = "N"
        
        if medicine.id == int(id):
            res = 'Same'
    except Medicine.DoesNotExist:
        res = "Y"

    context = {'result':res}

    return JsonResponse(context)


def medicine_add_edit_delete(request):
    id = request.POST.get('id')

    medicine = Medicine.objects.get(id=id)
    medicine.use_yn = 'N'
    medicine.save()

    log = MedicineLog()
    log.medicine = medicine
    log.type='del'
    log.save()



    return JsonResponse({'result':True})


def list_database_medicine_class_get(request):
        
    datas = []
    query = MedicineClass.objects.filter(use_yn='Y')
    for data in query:
        print(data.use_yn)
        datas.append({
            'id':data.id,
            'name':data.name,
            'name_vie':data.name_vie,
            })


    return JsonResponse({
        'result':True,
        'datas':datas,
        })


def add_edit_medicine_class_menu_get(request):
    id = request.POST.get('id')

    query = MedicineClass.objects.get(pk = id)

    return JsonResponse({
        'result':True,
        'name':query.name,
        'name_vie':query.name_vie,
        })


def add_edit_medicine_class_menu_save(request):
    id = request.POST.get('id','')
    name = request.POST.get('class_name')
    name_vie = request.POST.get('class_name_vie')

    if id == '':
        query = MedicineClass()
    else:
        query = MedicineClass.objects.get(pk = id)
        

    query.name = name
    query.name_vie = name_vie
    query.save()

    return JsonResponse({
        'result':True,
        })


def get_inventory_history(request):
    id = request.POST.get('id')
    date = request.POST.get('date')


    #date_min = datetime.datetime.combine(datetime.datetime.strptime(date, "%Y-%m-%d").date(), datetime.time.min)
    #date_max = datetime.datetime.combine(datetime.datetime.strptime(date, "%Y-%m-%d").date(), datetime.time.max)

    #medicine_logs = MedicineLog.objects.filter(date__range = (date_min, date_max), medicine_id = id).values('date','changes','type','memo').order_by('-date')
    medicine_logs = MedicineLog.objects.filter(medicine_id = id).values('id','diagnosis_id','date','changes','type','memo').order_by('-date')
    datas = []
    for medicine_log in medicine_logs:
        if medicine_log['diagnosis_id'] is not None:
            reception = Reception.objects.get(diagnosis = medicine_log['diagnosis_id'])
        else:
            reception = None
        
        data = {
            'date':medicine_log['date'].strftime('%Y-%m-%d %H:%M:%S'),
            'changes':medicine_log['changes'],
            'type':medicine_log['type'],
            'depart': '' if reception is None else reception.depart.name,
            'memo':'' if medicine_log['memo'] is None else medicine_log['memo'],
            }
        datas.append(data)

    medicine = Medicine.objects.get(id=id)
    return JsonResponse({
        'datas':datas,
        'count':medicine.inventory_count,
        })

def save_database_add_medicine(request):
    id = request.POST.get('id')
    registration_date= request.POST.get('registration_date')
    expiry_date= request.POST.get('expiry_date')
    changes= request.POST.get('changes')
    memo= request.POST.get('memo')
    check= request.POST.get('check')
    
    if int(check)==0 :
        medicine = Medicine.objects.get(id=id)
        count = medicine.inventory_count
        medicine.inventory_count =  count + int(changes)
        medicine.save()

        medicine_logs = MedicineLog()
        medicine_logs.changes = changes
        medicine_logs.medicine = medicine
    else:
        medicine_logs = MedicineLog.objects.get(id = check)
        medicine_logs.date = datetime.datetime.strptime(registration_date, '%Y-%m-%d')

    medicine_logs.memo = memo
    medicine_logs.expiry_date = datetime.datetime.strptime(expiry_date, '%Y-%m-%d')
 
    medicine_logs.tmp_count = changes
    medicine_logs.type='add'
    medicine_logs.save()


    input_price = request.POST.get('input_price',None)
    if input_price:
        try:
            str_now = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
            old_price_input = Pricechange.objects.get(type="Medicine",country='VI',type2='INPUT',code=medicine.code, date_end="99999999999999")
            
            if old_price_input.price != int(input_price):
                old_price_input.date_end = str_now
                old_price_input.save()

                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=medicine.code)
                new_price.price = price_input
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

        except Pricechange.DoesNotExist:
            if medicine.price_input != int(input_price):
                new_price = Pricechange(type="Medicine",country='VI',type2='INPUT',code=medicine.code)
                new_price.price = input_price
                new_price.date_start = str_now
                new_price.date_end = "99999999999999"
                new_price.save()

    return JsonResponse({'result':True})


def get_expiry_date(request):

    id = request.POST.get('id')
    medicine = Medicine.objects.get(id=id)

    medicine_logs = MedicineLog.objects.filter(medicine = medicine, type='add').exclude(tmp_count__lte = 0).order_by('expiry_date').values('id','date','tmp_count','expiry_date')
    
    datas = []
    for medicine_log in medicine_logs:
        
        data = {
            'id':medicine_log['id'],
            'date':medicine_log['date'].strftime('%Y-%m-%d'),
            'expiry_date':0 if medicine_log['expiry_date'] is None else medicine_log['expiry_date'].strftime('%Y-%m-%d'),
            'tmp_count':medicine_log['tmp_count'] if medicine_log['tmp_count'] is not None else 0,
            }
        datas.append(data)

    return JsonResponse({
        'result':True,
        'datas':datas,
        })


def get_edit_database_add_medicine(request):
    id = request.POST.get('id')

    data = MedicineLog.objects.values('id','date','expiry_date','memo','tmp_count').get(id=id)

    return JsonResponse({
        'result':True,
        'data':{
            'id':data['id'],
            'date':data['date'].strftime('%Y-%m-%d'),
            'expiry_date':data['expiry_date'].strftime('%Y-%m-%d'),
            'memo':data['memo'],
            'tmp_count':data['tmp_count'],
            },
        })


def save_database_disposal_medicine(request):
    id = request.POST.get('id')
    disposial = request.POST.get('disposial')
    memo = request.POST.get('memo')


    disposal_data = MedicineLog.objects.get(id=id)
    if disposal_data.tmp_count < int(disposial):
        return JsonResponse({
            'result':False,
            'msg':1, # tmp 카운트가 더 적을 경우
                             })
    else:
        medicine = Medicine.objects.get(id = disposal_data.medicine_id)
        disposal_data.tmp_count -= int(disposial)
        
        disposal_data.save()

        medicine_log = MedicineLog()
        medicine_log.type = 'dec'
        medicine_log.changes = disposial
        medicine_log.memo = memo
        medicine_log.medicine_id = disposal_data.medicine_id
        medicine_log.save()

        
        medicine.inventory_count -= int(disposial)
        medicine.save()



    return JsonResponse({'result':True,})
    
@login_required
def upload_file(request):

    # instance = Patient.objects.filter(pk=patient_id).first()
    file_path = '/home/imedicare/Cofee/Pharmacy/static/pharm_data1.xlsx'
    # file_path = '/home/light/Desktop/Projects/imedicare2/pharm_data1.xlsx'
    try:
        file = request.FILES.getlist('file')[0]
        
        if request.method == 'POST':
            with open(file_path, 'wb+') as destination:
                for chunk in file.chunks():
                    destination.write(chunk)
            wb = load_workbook(file_path) #Workbook()
            ws = wb.active# grab the active worksheet
            for i in range(6,2000):
                code = ws[f'B{i}']
                if code.value != '' and code.value != None:
                    # print('===',code.value)
                    medicine = Medicine.objects.filter(code = code.value).first()
                    if code.value == 'M0122':
                        print(code.value)
                        print(medicine)
                        print(int(ws[f'K{i}'].value))
                    if medicine:
                        # print(medicine.name)
                        medicine.inventory_count = int(ws[f'K{i}'].value)
                        medicine.save()
                        print(medicine)
                    
    except Exception as e:
        print('====>>>>',e)
        return JsonResponse({'url':str(e)})
        
    return JsonResponse({'url':'ok'})