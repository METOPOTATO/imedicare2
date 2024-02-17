from ast import Tuple
import json
from time import time
from django.shortcuts import render, redirect
import datetime ,calendar
from django.utils import timezone
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage
from django.db.models import F,Q,Case,When, CharField,Count,Sum
import operator
import functools
import itertools
from django.http import JsonResponse
from django.forms.models import model_to_dict
from django.contrib.auth.decorators import login_required
from django.utils import timezone, translation

from .forms import *
from .models import *
from Patient.forms import PatientForm, HistoryForm
from Patient.models import Patient,History
from Account.models import User
from Doctor.models import *
from app.models import *
from Laboratory.models import *
from django.utils.translation import gettext as _
from django.http import JsonResponse, HttpResponseRedirect,HttpResponse
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Color, Font,Border,Side,Alignment,PatternFill
import dateutil.relativedelta
# Create your views here.


def suffix(d):
    return 'th' if 11<=d<=13 else {1:'st',2:'nd',3:'rd'}.get(d%10, 'th')

def custom_strftime(format, t):
    return t.strftime(format).replace('{S}', str(t.day) + suffix(t.day))

def notlast(itr):
    itr = iter(itr)  # ensure we have an iterator
    prev = itr.next()
    for item in itr:
        yield prev
        prev = item



@login_required
def index(request):
    
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')

    #reception
    patient_form = PatientForm()
    reception_form = ReceptionForm()
    history_form = HistoryForm()
    #search
    patientsearch_form = PatientSearchForm()
    receptionsearch_form = SearchReceptionStatusForm()
    payment_form = PaymentSearchForm()
    reservation_form = ReservationSearchForm()

    initial_report_q2_option = []
    initial_report_q2 = COMMCODE.objects.filter(commcode_grp = 'PM_IRQ2').values('id','seq','se1','se2','se4','se5')

    initial_report_q2_title = COMMCODE.objects.filter(commcode_grp = 'PM_IRQ2').values('se2','seq','se6','se7','se8').annotate(Count('se2')).extra(select={'tmp_seq':'CAST(seq AS INTERGER)'}).order_by('tmp_seq')
    for title in initial_report_q2_title:
        item = COMMCODE.objects.filter(commcode_grp = 'PM_IRQ2').values('se2').annotate(Count('se2'))
                
    list_depart = Depart.objects.all().values('id','name')
        
    patient_mark = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PT_INFO',use_yn="Y").values('commcode','se1').order_by('seq')

    list_funnels = []
    funnels = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PATIENTS_FUNNELS',use_yn="Y").annotate(name = f_name ).values('commcode','name',)
    for data in funnels:
        list_funnels.append({
            'code':data['commcode'],
            'name':data['name'],
            })



    return render(request,
    'Receptionist/index.html',
            {
                'patient': patient_form,
                'reception':reception_form,
                'history':history_form,
                'patientsearch':patientsearch_form,
                'receptionsearch':receptionsearch_form,
                'payment':payment_form,
                'reservation':reservation_form,

                'patient_mark':patient_mark,
                'list_funnels':list_funnels,

                'initial_report_q2':initial_report_q2,
                'initial_report_q2_title':initial_report_q2_title,
                'list_depart':list_depart,
            },
        )

def set_new_patient(request):

    
    #last = Patient.objects.last()
    #if last == None:
    #    chart_no = 1
    #else:
    #    chart_no = last.id + 1

    last = Patient.objects.last()
    if last is None: #or last.id <= 17000:
        chart_no =1
    else:
        chart_no = last.id + 1

    context = {'chart':"{:06d}".format(chart_no)}
    return JsonResponse(context)


def save_patient(request):
    
    cahrt_no = request.POST.get('cahrt_no')
    name_kor = request.POST.get('name_kor','')
    name_eng = request.POST.get('name_eng','')
    date_of_birth = request.POST.get('date_of_birth','')
    phone = request.POST.get('phone','')
    gender = request.POST.get('gender','')
    address = request.POST.get('address','')
    passport = request.POST.get('passport','')
    
    email = request.POST.get('email','')
    nationality = request.POST.get('nationality','')
    memo = request.POST.get('memo','')
    marking = request.POST.get('marking','')
    funnel = request.POST.get('funnel','')
    funnel_etc = request.POST.get('funnel_etc','')

    chief_complaint = request.POST.get('chief_complaint','')
    past_history = request.POST.get('past_history','')
    family_history = request.POST.get('family_history','')
    reservation_id = request.POST.get('reservation_id','')

    need_invoice = request.POST.get('need_invoice',False)
    need_insurance = request.POST.get('need_insurance',False)

    tax_invoice_number = request.POST.get('tax_invoice_number','')
    tax_invoice_company_name = request.POST.get('tax_invoice_company_name','')
    tax_invoice_address = request.POST.get('tax_invoice_address','')

    id = request.POST.get('id')
    if request.POST.get('id') is None or request.POST.get('id') is '':
        patient = Patient()
    else:
        patient = Patient.objects.get(pk = id)
    #try:
    #    patient = Patient.objects.get(pk = id)
    #except Patient.DoesNotExist:
    #    patient = Patient(pk = id)

    if name_kor != '' and name_kor != None:
        patient.name_kor = name_kor
    if name_eng != '' and name_eng != None:
        patient.name_eng = name_eng
    patient.date_of_birth = date_of_birth
    patient.phone = phone
    patient.gender = gender
    patient.address = address
    patient.funnel = funnel
    patient.funnel_etc = funnel_etc
    if passport != '' and passport != None:
        patient.passport = passport
    if nationality != '' and nationality != None:
        patient.nationality = nationality
    if email != '' and email != None:
        patient.email = email
    if memo != '' and memo != None:
        patient.memo = memo
    if marking != '' and marking != None:
        patient.marking = marking

    patient.save()

    print('====================')
    print(patient.passport)

    try:
        history = History.objects.get(patient=patient)
    except History.DoesNotExist:
        history = History(patient = patient)

    if past_history != '' and past_history != None:
        history.past_history = past_history
    if family_history != '' and family_history != None:
        history.family_history = family_history
    history.save()

    if tax_invoice_number != '' or tax_invoice_company_name != '' or tax_invoice_address != '':
        try:
            taxinvoice = TaxInvoice.objects.get(patient=patient)
        except TaxInvoice.DoesNotExist:
            taxinvoice = TaxInvoice(patient = patient)

        taxinvoice.number = '' if tax_invoice_number == '' else tax_invoice_number
        taxinvoice.company_name = '' if tax_invoice_number == '' else tax_invoice_company_name
        taxinvoice.address = '' if tax_invoice_number == '' else tax_invoice_address
        taxinvoice.save()

    if reservation_id != '':
        try:
            reservation = Reservation.objects.get(pk=reservation_id)
        except Reservation.DoesNotExist:
            reservation = Reservation(patient = patient)

        reservation.patient = patient
        reservation.save()


    result = True







    patient = Patient.objects.get(pk = patient.id)
    context = {'result':result,
                'id':patient.id,
                'chart':patient.get_chart_no(),
                'name_kor':patient.name_kor,
                'name_eng':patient.name_eng,
                'gender':patient.get_gender_simple(),
                'date_of_birth':patient.date_of_birth,
                'phonenumber':patient.phone,
                'age' : patient.get_age(),
                'address':patient.address,
                'passport':patient.passport
               }
    return JsonResponse(context)


def set_patient_data(request):
    patient_id = request.POST.get('patient_id')

    context={}
    patient = Patient.objects.get(pk=int(patient_id))
    try:
        history = History.objects.get(patient = patient)
        context.update({
            'history_past':history.past_history,
            'history_family':history.family_history,
            })
    except History.DoesNotExist:
        context.update({
            'history_past':'',
            'history_family':'',
            })

    try:
        taxinvoice = TaxInvoice.objects.get(patient = patient)
    except TaxInvoice.DoesNotExist:
        taxinvoice = None

    try:
        rec = Reception.objects.filter(patient_id = patient_id,).last()
    except TaxInvoice.DoesNotExist:
        rec = None


    print('=====', rec.chief_complaint)    

    context.update({
        'id':patient.id,
        'chart':patient.get_chart_no(),
        'name_kor':patient.name_kor,
        'name_eng':patient.name_eng,
        'date_of_birth':patient.date_of_birth,
        'gender':patient.gender,
        'email':patient.email,
        'nationality':patient.nationality,
        'phone':patient.phone,
        'address':patient.address,
        # 'address':patient.passport,
        'memo':patient.memo,
        'marking':patient.marking,
        'funnel':patient.funnel,
        'funnel_etc':patient.funnel_etc,
        'passport':patient.passport,
        'tax_invoice_number':'' if taxinvoice is None else taxinvoice.number,
        'tax_invoice_company_name':'' if taxinvoice is None else taxinvoice.company_name,
        'tax_invoice_address':'' if taxinvoice is None else taxinvoice.address,

        'invoice':'' if rec is None else rec.need_invoice,
        'insurance':'' if rec is None else rec.need_insurance,
        'chief_complaint':' ',
        })
    return JsonResponse(context)

def Question(request,patient_id):

    return render(request,
    'Receptionist/Question.html',
            {
                'patient_id': patient_id,
            },
        )

def Question_save(request):
    patient_id = request.POST.get('patient_id')
    try:
        query = FIRST_VISIT_SURVEY.objects.get(PT_ID = patient_id)
    except FIRST_VISIT_SURVEY.DoesNotExist:
        query = FIRST_VISIT_SURVEY()

    query.PT_ID = patient_id
    query.pain_posi_text = request.POST.get('pain_posi_text')
    query.sick_date = request.POST.get('sick_date')
    query.cure_yn = request.POST.get('cure_yn')
    query.cure_phy_yn = request.POST.get('cure_phy_yn')
    query.cure_phy_cnt = request.POST.get('cure_phy_cnt')
    query.cure_inject_yn = request.POST.get('cure_inject_yn')
    query.cure_inject_cnt = request.POST.get('cure_inject_cnt')
    query.cure_medi_yn = request.POST.get('cure_medi_yn')
    query.cure_medi_cnt = request.POST.get('cure_medi_cnt')
    query.cure_needle_yn = request.POST.get('cure_needle_yn')
    query.cure_needle_cnt = request.POST.get('cure_needle_cnt')
    query.pain_level = request.POST.get('pain_level')
    query.surgery_yn = request.POST.get('surgery_yn')
    query.surgery_year = request.POST.get('surgery_year')
    query.surgery_name = request.POST.get('surgery_name')
    query.exam_kind = request.POST.get('exam_kind')
    query.exam_etc = request.POST.get('exam_etc')
    query.cd_film_yn = request.POST.get('cd_film_yn')
    query.disease_kind = request.POST.get('disease_kind')
    query.disease_etc = request.POST.get('disease_etc')
    query.medication = request.POST.get('medication')
    query.side_effect_yn = request.POST.get('side_effect_yn')
    query.pregnant_yn = request.POST.get('pregnant_yn')
    query.visit_motiv_item = request.POST.get('visit_motiv_item')
    query.visit_motiv_friend = request.POST.get('visit_motiv_friend')
    query.visit_motiv_etc = request.POST.get('visit_motiv_etc')

    if query.PT_vital is None:
        vital = Vital()
        vital.patient_id = patient_id
    else:
        vital = Vital.objects.get(id = query.PT_vital)

    vital.height = request.POST.get('vital_height')
    vital.weight = request.POST.get('vital_weight')
    vital.BMI = request.POST.get('vital_bmi')
    vital.blood_pressure = request.POST.get('vital_bp')
    vital.blood_temperature = request.POST.get('vital_bt')
    vital.save()

    query.PT_vital = vital.id
    query.save()


    result = True

    context = {'result':result}
    return JsonResponse(context)


def Question_get(request):
    patient_id = request.POST.get('patient_id')
    context = {}
    try:
        query = FIRST_VISIT_SURVEY.objects.get(PT_ID = patient_id)
        if query.PT_vital is None or query.PT_vital is '':
            context.update({
                'vital_height':'',
                'vital_weight':'',
                'vital_bmi':'',
                'vital_bp':'',
                'vital_bt':'',
                })
        else:
            vital = Vital.objects.get(id = query.PT_vital)
            context.update({
                'vital_height':vital.height,
                'vital_weight':vital.weight,
                'vital_bmi':vital.BMI,
                'vital_bp':vital.blood_pressure,
                'vital_bt':vital.blood_temperature,
                })



        context.update({
            
                'pain_posi_text':query.pain_posi_text,
                'sick_date':query.sick_date,
                'cure_yn':query.cure_yn,
                'cure_phy_yn':query.cure_phy_yn,
                'cure_phy_cnt':query.cure_phy_cnt,
                'cure_inject_yn':query.cure_inject_yn,
                'cure_inject_cnt':query.cure_inject_cnt,
                'cure_medi_yn':query.cure_medi_yn,
                'cure_medi_cnt':query.cure_medi_cnt,
                'cure_needle_yn':query.cure_needle_yn,
                'cure_needle_cnt':query.cure_needle_cnt,
                'pain_level':query.pain_level,
                'surgery_yn':query.surgery_yn,
                'surgery_year':query.surgery_year,
                'surgery_name':query.surgery_name,
                'exam_kind':query.exam_kind,
                'exam_etc':query.exam_etc,
                'cd_film_yn':query.cd_film_yn,
                'disease_kind':query.disease_kind,
                'disease_etc':query.disease_etc,
                'medication':query.medication,
                'side_effect_yn':query.side_effect_yn,
                'pregnant_yn':query.pregnant_yn,
                'visit_motiv_item':query.visit_motiv_item,
                'visit_motiv_friend':query.visit_motiv_friend,
                'visit_motiv_etc':query.visit_motiv_etc,
                 
            })

        context.update({'result':True})
    except FIRST_VISIT_SURVEY.DoesNotExist:
        context.update({'result':False})


    return JsonResponse(context)


def save_reception(request):
    chart_no = request.POST.get('chart_no')
    reservation_id = request.POST.get('reservation_id')
    name_kor = request.POST.get('name_kor')
    name_eng = request.POST.get('name_eng')
    date_of_birth = request.POST.get('date_of_birth')
    phone = request.POST.get('phone')
    gender = request.POST.get('gender')
    address = request.POST.get('address')
    email = request.POST.get('email')
    passport = request.POST.get('passport')
    nationality = request.POST.get('nationality','')
    memo = request.POST.get('memo','')
    marking = request.POST.get('marking','')
    funnel = request.POST.get('funnel','')
    funnel_etc = request.POST.get('funnel_etc','')

    past_history = request.POST.get('past_history','')
    family_history = request.POST.get('family_history','')

    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')
    chief_complaint = request.POST.get('chief_complaint')

    tax_invoice_number = request.POST.get('tax_invoice_number','')
    tax_invoice_company_name = request.POST.get('tax_invoice_company_name','')
    tax_invoice_address = request.POST.get('tax_invoice_address','')

    need_medical_report = request.POST.get('need_medical_report',False)
    need_invoice = request.POST.get('need_invoice',False)
    need_insurance = request.POST.get('need_insurance',False)
    is_vaccine = request.POST.get('is_vaccine',False)

    id = request.POST.get('id')
    if request.POST.get('id') is None or request.POST.get('id') is '':
        patient = Patient()
    else:
        patient = Patient.objects.get(pk = id)
    #try:
    #    patient = Patient.objects.get(pk = int(id))
    #except Patient.DoesNotExist:
    #    patient = Patient(pk = int(id))


    if name_kor != '' and name_kor != None:
        patient.name_kor = name_kor
    if name_eng != '' and name_eng != None:
        patient.name_eng = name_eng
    patient.date_of_birth = date_of_birth
    patient.phone = phone
    patient.gender = gender
    patient.address = address
    patient.funnel = funnel
    patient.funnel_etc = funnel_etc
    patient.passport = passport

    if nationality != '' and nationality != None:
        patient.nationality = nationality
    if email != '' and email != None:
        patient.email = email
    if memo != '' and memo != None:
        patient.memo = memo
    if marking != '' and marking != None:
        patient.marking = marking

    patient.save()

    #try:
    #    history = History.objects.get(patient=patient)
    #except History.DoesNotExist:
    #    history = History(patient = patient)
    #
    #    
    #history.past_historyf = past_history
    #history.family_history = family_history
    #history.save()
    
    reception = Reception(patient = patient)
    reception.depart = Depart.objects.get(pk = depart)
    reception.doctor = Doctor.objects.get(pk = doctor)
    reception.chief_complaint = chief_complaint
    

    try:
       reservation = Reservation.objects.get(pk = reservation_id)
       reservation.patient = patient
       reservation.save()
       reception.reservation = reservation
    except Reservation.DoesNotExist:
       reservation = Reservation(None)

    if need_medical_report == 'true':
        reception.need_medical_report = True
    if need_invoice == 'true':
        reception.need_invoice = True
    if need_insurance == 'true':
        reception.need_insurance = True
    if is_vaccine == 'true':
        reception.is_vaccine = True

    reception.save()


    if tax_invoice_number != '' or tax_invoice_company_name != '' or tax_invoice_address != '':
        try:
            taxinvoice = TaxInvoice.objects.get(patient=patient)
        except TaxInvoice.DoesNotExist:
            taxinvoice = TaxInvoice(patient = patient)

        taxinvoice.number = '' if tax_invoice_number == '' else tax_invoice_number
        taxinvoice.company_name = '' if tax_invoice_number == '' else tax_invoice_company_name
        taxinvoice.address = '' if tax_invoice_number == '' else tax_invoice_address
        taxinvoice.save()

    #33333
    vital_ht = request.POST.get('patient_table_vital_ht',None)
    vital_wt = request.POST.get('patient_table_vital_wt',None)
    vital_bp = request.POST.get('patient_table_vital_bp',None)
    vital_bt = request.POST.get('patient_table_vital_bt',None)
    vital_pr = request.POST.get('patient_table_vital_pr',None)
    vital_breath = request.POST.get('patient_table_vital_breath',None)


    if vital_ht is '' and vital_wt is '' and vital_bp is '' and vital_bt is '' and vital_pr is '' and vital_breath is '':
        pass
    else:
        vital = Vital()
        vital.patient = patient
        vital.weight = vital_wt
        vital.height = vital_ht
        vital.blood_pressure = vital_bp
        vital.blood_temperature = vital_bt
        vital.breath = vital_breath
        vital.pulse_rate = vital_breath
        vital.save()


    result = True

    patient = Patient.objects.get(pk = patient.id)
    context = {'result':result,
                'id':patient.id,
                'chart':patient.get_chart_no(),
                'name_kor':patient.name_kor,
                'name_eng':patient.name_eng,
                'gender':patient.get_gender_simple(),
                'date_of_birth':patient.date_of_birth,
                'phonenumber':patient.phone,
                'age' : patient.get_age(),
                'address':patient.address,
                }
    return JsonResponse(context)

def patient_search(request):
    category = request.POST.get('category')
    string = request.POST.get('string')
    memo = request.POST.get('memo_string')
    print(string)
    print(memo)
    # print(string)
    # print(category)
    argument_list = [] 
    if category=='':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__phone__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
        argument_list.append( Q(**{'patient__memo__icontains':string} ) ) 
    elif category=='name':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    elif category=='chart':
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    elif category=='date_of_birth':
        argument_list.append( Q(**{'patient__date_of_birth__icontains':string} ) ) 
    elif category=='phone':
        argument_list.append( Q(**{'patient__phone__icontains':string} ) )
    elif category=='memo':
        argument_list.append( Q(**{'patient__memo__icontains':string} ) )
     

    # patient = Patient.objects.filter(name_kor=string).first()
    # print(patient)
    datas=[]
    if string and string != '':
        receptions = Reception.objects.select_related('patient').values('patient_id','depart_id').filter( functools.reduce(operator.or_, argument_list) ).exclude(progress='deleted').annotate(c_pt=Count('patient_id'),c_dp=Count('depart_id'))
        # print(receptions)
        patient_memo = []
        if memo:
            patient_memo = DetailMemo.objects.filter(memo__icontains=memo).values_list('patient_id', flat=True);
            patient_memo = list(patient_memo)
        for reception in receptions:
            if memo and memo != '':
                if reception['patient_id'] in patient_memo:
                    reception_last = Reception.objects.filter(patient = reception['patient_id'], depart = reception['depart_id']).last()
                    if reception_last:
                        data = {}
                        data.update({
                            'id':reception_last.patient.id,
                            'chart':reception_last.patient.get_chart_no(),
                            'name_kor':reception_last.patient.name_kor,
                            'name_eng':reception_last.patient.name_eng,
                            'gender':reception_last.patient.gender,
                            'date_of_birth':reception_last.patient.date_of_birth.strftime('%Y-%m-%d'),
                            'phonenumber':reception_last.patient.phone,
                            'age' : reception_last.patient.get_age(),
                            'address':reception_last.patient.address,
                            'has_unpaid':reception_last.patient.has_unpaid(),
                            'depart':reception_last.depart.name,
                            'last_visit':reception_last.recorded_date.strftime('%Y-%m-%d'),
                            
                            'nationality':reception_last.patient.nationality,
                            'passport':reception_last.patient.passport,
                            'email':reception_last.patient.email,
                            'category':reception_last.patient.category


                            })
                        datas.append(data)
                print(1)
            else:
                print(2)
                reception_last = Reception.objects.filter(patient = reception['patient_id'], depart = reception['depart_id']).last()
                data = {}
                data.update({
                    'id':reception_last.patient.id,
                    'chart':reception_last.patient.get_chart_no(),
                    'name_kor':reception_last.patient.name_kor,
                    'name_eng':reception_last.patient.name_eng,
                    'gender':reception_last.patient.gender,
                    'date_of_birth':reception_last.patient.date_of_birth.strftime('%Y-%m-%d'),
                    'phonenumber':reception_last.patient.phone,
                    'age' : reception_last.patient.get_age(),
                    'address':reception_last.patient.address,
                    'has_unpaid':reception_last.patient.has_unpaid(),
                    'depart':reception_last.depart.name,
                    'last_visit':reception_last.recorded_date.strftime('%Y-%m-%d'),
                    
                    'nationality':reception_last.patient.nationality,
                    'passport':reception_last.patient.passport,
                    'email':reception_last.patient.email,
                    'category':reception_last.patient.category


                    })
                datas.append(data)
    else:
        print(3)
        if memo:
            patient_memo = DetailMemo.objects.filter(memo__icontains=memo).values_list('patient_id', flat=True);
            patient_memo = list(patient_memo)
            receptions = Reception.objects.select_related('patient').values('patient_id','depart_id').filter( patient__in=patient_memo ).exclude(progress='deleted').annotate(c_pt=Count('patient_id'),c_dp=Count('depart_id'))
            for reception in receptions:
                reception_last = Reception.objects.filter(patient = reception['patient_id'], depart = reception['depart_id']).last()
                data = {}
                data.update({
                    'id':reception_last.patient.id,
                    'chart':reception_last.patient.get_chart_no(),
                    'name_kor':reception_last.patient.name_kor,
                    'name_eng':reception_last.patient.name_eng,
                    'gender':reception_last.patient.gender,
                    'date_of_birth':reception_last.patient.date_of_birth.strftime('%Y-%m-%d'),
                    'phonenumber':reception_last.patient.phone,
                    'age' : reception_last.patient.get_age(),
                    'address':reception_last.patient.address,
                    'has_unpaid':reception_last.patient.has_unpaid(),
                    'depart':reception_last.depart.name,
                    'last_visit':reception_last.recorded_date.strftime('%Y-%m-%d'),
                    
                    'nationality':reception_last.patient.nationality,
                    'passport':reception_last.patient.passport,
                    'email':reception_last.patient.email,
                    'category':reception_last.patient.category


                    })
                datas.append(data)
    # print(datas)
    context = {'datas':datas}
    return JsonResponse(context)


def reception_search(request):

    date_start = request.POST.get('date_start')
    date_end = request.POST.get('date_end')
    depart_id = request.POST.get('depart')
    doctor_id = request.POST.get('doctor')

    patient_name = request.POST.get('patient_name')
    # patient_name = 'JUHWANG'
    kwargs={}
    if depart_id != '':
        depart = Depart.objects.get(id = depart_id)
        kwargs['depart_id'] = depart

    #if doctor_id != '':
    #    doctor = Doctor.objects.get(id = doctor_id)
    #    kwargs['doctor'] = doctor
    

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)
    
    argument_list2 = []
    argument_list2.append( Q(**{'patient__name_kor__icontains':patient_name } ) ) 
    argument_list2.append( Q(**{'patient__name_eng__icontains':patient_name } ) )  
    # if patient_name != '' and patient_name is not None:
    #     argument_list2.append( Q(**{'name__icontains':patient_name } ) )
    try:
        receptions = Reception.objects.filter(functools.reduce(operator.or_, argument_list2),recorded_date__range = (date_min, date_max),**kwargs,).exclude(progress='deleted')
    except Exception as e:
        print(e)
        receptions = Reception.objects.filter(recorded_date__range = (date_min, date_max),**kwargs).exclude(progress='deleted')

    datas=[]
    today = datetime.date.today()
    
    
    for reception in receptions:
        data={}
        
        is_new = Reception.objects.filter(patient = reception.patient, depart_id = reception.depart_id).count()
        if is_new == 1:
            data.update({'is_new':'N'})
        else:
            tmp_rec = Reception.objects.filter(patient = reception.patient, depart_id = reception.depart_id).first()
            if tmp_rec.id == reception.id:
                data.update({'is_new':'N'})
            else:
                data.update({'is_new':'R'})

        #패키지 유무
        #try:
        #    package = Package_Manage.objects.get(reception_id = reception.id).id
        #except Package_Manage.DoesNotExist:
        #    package = None

        data.update({
            'id':reception.id,
            'patient_id': reception.patient.id,
            'chart':reception.patient.get_chart_no(),
            'name_kor':reception.patient.name_kor,
            'name_eng':reception.patient.name_eng,
            'age':reception.patient.get_age(),
            'gender':reception.patient.get_gender_simple(),
            'date_of_birth':reception.patient.date_of_birth.strftime('%Y-%m-%d'),
            'reception_time':reception.recorded_date.strftime('%Y-%m-%d %H:%M:%S'),
            'depart':reception.depart.name,
            'doctor':reception.doctor.name_kor,
            'has_unpaid':reception.patient.has_unpaid(),
            'date':reception.recorded_date.strftime('%m-%d'),
            'time':reception.recorded_date.strftime('%H:%M'),

            'is_vaccine':reception.is_vaccine,
            'status':reception.progress,
            #'package':package,

            })

        

        datas.append(data)

    datas.reverse()
    context = {'datas':datas}
    return JsonResponse(context)


def payment_search(request):
    date = request.POST.get('date')
    status = request.POST.get('status')
    show_all_unpaid = request.POST.get('show_all_unpaid')
    kwargs={}

    if show_all_unpaid == 'true':
        payments = Payment.objects.filter(progress = 'unpaid') 
    else:
        date_min = datetime.datetime.combine(datetime.datetime.strptime(date, "%Y-%m-%d").date(), datetime.time.min)
        date_max = datetime.datetime.combine(datetime.datetime.strptime(date, "%Y-%m-%d").date(), datetime.time.max)

        reception = Reception.objects.filter(recorded_date__range = (date_min, date_max))
        temp_list=[]
        for i in reception:
            temp_list.append(i.id)
        
        if status == 'all':
            payments = Payment.objects.filter(reception_id__in=temp_list)
        else:
            payments = Payment.objects.filter(reception_id__in=temp_list, progress=status)
        

    datas=[]
    for payment in payments:
        data.update({
            'chart':payment.reception.patient.get_chart_no(),
            'name_kor':payment.reception.patient.name_kor,
            'name_eng':payment.reception.patient.name_eng,
            'depart':payment.reception.depart.name,
            'doctor':payment.reception.doctor.name_kor,
            'test':'test',
            'precedure':'precedure',
            'medicine':'medicine',
            'total':payment.total,
            'status':payment.progress,
            'incomplete':'incomplete',
            })
        datas.append(data)

    context = {'datas':datas}
    return JsonResponse(context)


def reservation_search(request):
    date_start = request.POST.get('date_start')
    date_end = request.POST.get('date_end')
    string = request.POST.get('string')
    kwargs={}
    depart_id = request.POST.get('depart')
    
    # if string != '' and string is not None:
    #     kwargs.update({ 'name__icontains':string} )    
        

    if depart_id != '':
        depart = Depart.objects.get(id = depart_id)
        kwargs['depart_id'] = depart

    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)
    
    argument_list = []
    argument_list.append( Q(**{'reservation_date__range':(date_min, date_max) } ) ) 
    argument_list.append( Q(**{'re_reservation_date__range':(date_min, date_max) } ) )   

    argument_list2 = []
    argument_list2.append( Q(**{'patient__name_kor__icontains':string } ) ) 
    argument_list2.append( Q(**{'patient__name_eng__icontains':string } ) )  
    if string != '' and string is not None:
        argument_list2.append( Q(**{'name__icontains':string } ) )
        # kwargs.update({ '':string} )      

    try:
        reservations = Reservation.objects.filter(
                functools.reduce(operator.or_, argument_list),
                functools.reduce(operator.or_, argument_list2),
                **kwargs,
                
                # name__icontains = string
            ).order_by('reservation_date')
    except:
        reservations = Reservation.objects.filter(
                functools.reduce(operator.or_, argument_list),
                **kwargs,
                
                # name__icontains = string
            ).order_by('reservation_date')

    datas=[]

    for reservation in reservations:
        data = {} 
        if reservation.re_reservation_date != None:
            
            if reservation.re_reservation_date >= date_min and reservation.re_reservation_date < date_max:
                data.update({ 
                    'date':reservation.re_reservation_date.strftime('%Y-%m-%d'), 
                    
                    'time':reservation.re_reservation_date.strftime('%H:%M:00'),
                    })
                
            else:
                continue
        else:
            data.update({ 
                'date':reservation.reservation_date.strftime('%Y-%m-%d'),
                
                'time':reservation.reservation_date.strftime('%H:%M:00'),
                })

        data.update({ 
            'id':reservation.id, 
            'start':reservation.reservation_date.strftime('%Y-%m-%d %H:%M:00'),
            'depart': reservation.depart.name,
            'doctor': reservation.doctor.name_kor,
            #'time':reservation.reservation_date.strftime('%H:%M:00'),
            'division':reservation.division,
            'memo':reservation.memo,
            })
      
        try:
            patient = Patient.objects.get(pk = reservation.patient_id)
            data.update({
                    'chart':reservation.patient.get_chart_no(),
                    'name':reservation.patient.name_kor + '<br />' + reservation.patient.name_eng,
                    'date_of_birth':reservation.patient.date_of_birth.strftime('%Y-%m-%d'),
                    'phone':reservation.patient.phone,
                    'has_unpaid':reservation.patient.has_unpaid(),
                })
            
        except Patient.DoesNotExist:
            data.update({
                    'chart':'',
                    'name':reservation.name,
                    'date_of_birth':reservation.date_of_birth.strftime('%Y-%m-%d'),
                    'phone':reservation.phone,
                })
            


        datas.append(data)


    context = {'datas': sorted( datas, key=lambda i:( i['date'],i['time'])) }
    return JsonResponse(context)


@login_required
def apointment_search(request):
    kwargs ={}

    # date_start = request.POST.get('date_start')
    # date_end = request.POST.get('date_end')
    pick_up = request.POST.get('pick_up')
    name = request.POST.get('string')
    drop_off = request.POST.get('drop_off','')
    depart_id = request.POST.get('depart')
    need_pick_up = True if request.POST.get('need_pick_up','false') == 'true' else False

    date_start = request.POST.get('date_start').split('T')[0]
    date_end = request.POST.get('date_end').split('T')[0]
    
    if depart_id != '' and depart_id is not None:
        kwargs.update({ 'depart_id':depart_id })

    if drop_off != '' and drop_off is not None:
        kwargs.update({ 'drop_off':drop_off} )
    
    if pick_up != '' and pick_up is not None:
        kwargs.update({ 'pick_up':pick_up} )

    if need_pick_up == True:
        kwargs.update({ 'need_pick_up':need_pick_up} )    

    if name != '' and name is not None:
        kwargs.update({ 'name__icontains':name} )           
        
         
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    
    argument_list = []
    argument_list.append( Q(**{'reservation_date__range':(date_min, date_max) } ) ) 
    argument_list.append( Q(**{'re_reservation_date__range':(date_min, date_max) } ) ) 

    reservations = Reservation.objects.filter(
        functools.reduce(operator.or_, argument_list),
        **kwargs,
        ).order_by('reservation_date')
    datas=[]
    count = 0
    count_page = reservations
    total = reservations.count()
    count_ = 0
    # for reservation in reservations[int(page_context) * (int(page) -1): int(page_context) * int(page)]:
    for reservation in reservations:
        data = {} 
        if reservation.re_reservation_date is None or (reservation.re_reservation_date != None and reservation.re_reservation_date >= date_min and reservation.re_reservation_date < date_max):
            
        #     if reservation.re_reservation_date >= date_min and reservation.re_reservation_date < date_max:
        #         data.update({ 
        #             'date':reservation.re_reservation_date.strftime('%Y-%m-%d'), 
                    
        #             'time':reservation.re_reservation_date.strftime('%H:%M:00'),
        #             })
        #         count_ += 1                    
        #     else:
        #         continue
        # else: 
            
            count_ += 1
            

            data.update({ 
                'num':count_, 
                'id':reservation.id, 
                'start':reservation.reservation_date.strftime('%Y-%m-%d %H:%M:00') if reservation.re_reservation_date == None else reservation.re_reservation_date.strftime('%Y-%m-%d %H:%M:00'),
                'depart': reservation.depart.name,
                'doctor': reservation.doctor.name_kor,
                'division':reservation.division,
                'memo':reservation.memo,
                'follower':'' if reservation.follower == None else reservation.follower,
                'pick_up_time': '' if reservation.pick_up_time == None else reservation.pick_up_time.strftime('%Y-%m-%d %H:%M:%S'),
                'drop_off_time': '' if reservation.drop_off_time == None else reservation.drop_off_time.strftime('%Y-%m-%d %H:%M:%S'),
                'pick_up_addr': '' if reservation.pick_up_addr == None else reservation.pick_up_addr,
                'drop_off_addr':'' if reservation.drop_off_addr == None else reservation.drop_off_addr,
                'apointment_memo':'' if reservation.apointment_memo == None else reservation.apointment_memo,
                'pick_up_vehicle': '' if reservation.pick_up_vehicle == None else reservation.pick_up_vehicle,
                'drop_off_vehicle': '' if reservation.drop_off_vehicle == None else reservation.drop_off_vehicle,   
                'pick_up_status': '' if reservation.pick_up_status == None else reservation.pick_up_status,
                'drop_off_status': '' if reservation.drop_off_status == None else reservation.drop_off_status,              
                'date':reservation.reservation_date.strftime('%Y-%m-%d') if reservation.re_reservation_date == None else reservation.re_reservation_date.strftime('%Y-%m-%d'),
                'time':reservation.reservation_date.strftime('%H:%M:00') if reservation.re_reservation_date == None else reservation.re_reservation_date.strftime('%H:%M:00'),     
                'need_pick_up': reservation.need_pick_up,
                })
            try:
                reception = Reception.objects.get(reservation_id = reservation.id)
                try:
                    payment = Payment.objects.get(reception_id = reception.id)
                    try:
                        paymentRecord = PaymentRecord.objects.filter(payment__id = payment.id, status = 'paid').first()
                        data.update({
                                'payment_time':paymentRecord.date.strftime('%H:%M:00'),
                            })
                        
                    except PaymentRecord.DoesNotExist:
                        data.update({
                                'payment_time':'',
                            })                 
                    
                except Payment.DoesNotExist:
                    data.update({
                            'payment_time':'',
                        })            
                
            except Reception.DoesNotExist:
                data.update({
                        'payment_time':'',
                    })

            try:
                patient = Patient.objects.get(pk = reservation.patient_id)
                data.update({
                        'chart':patient.get_chart_no(),
                        'patient_id':reservation.patient_id,
                        'name':reservation.patient.name_kor + '<br />' + reservation.patient.name_eng,
                        'date_of_birth':reservation.patient.date_of_birth.strftime('%Y-%m-%d'),
                        'phone':reservation.patient.phone,
                        'has_unpaid':reservation.patient.has_unpaid(),
                        'address':patient.address,
                    })
                
            except Patient.DoesNotExist:
                data.update({
                        'chart':'',
                        'name':reservation.name,
                        'date_of_birth':reservation.date_of_birth.strftime('%Y-%m-%d'),
                        'phone':reservation.phone,
                        'address':'',
                    })
                


            datas.append(data)
        # paginator = Paginator(count_page , page_context)
        # try:
        #     paging_data = paginator.page(page)
        # except PageNotAnInteger:
        #     paging_data = paginator.page(1)
        # except EmptyPage:
        #     paging_data = paginator.page(paginator.num_pages)

    return JsonResponse({
                        'datas': sorted( datas, key=lambda i:( i['date'],i['time'])),  
                        # 'page_range_start':paging_data.paginator.page_range.start,
                        # 'page_range_stop':paging_data.paginator.page_range.stop,
                        # 'page_number':paging_data.number,
                        # 'has_previous':paging_data.has_previous(),
                        # 'has_next':paging_data.has_next(),
                        'total': count_,
                        })

@login_required
def storage_page(request):
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'en':
        f_name = F('commcode_name_en')


    today_filter = ReceptionForm()
    storage_search_form = StorageSearchForm()
    storage_form = StorageForm()
        
    depart = Depart.objects.all()

    #depart_medical= COMMCODE.objects.filter(use_yn = 'Y',commcode = 'DOCTOR', commcode_grp='DEPART_CLICINC',upper_commcode ='000002' ).annotate(code = F('se1'),name = f_name ).values('code','name')
    depart_medical = []
    depart_medical_query = Depart.objects.all()
    for data in depart_medical_query:
        depart_medical.append({
            'code':data.id,
            'name':data.name
            })
    #의사 정보 ? 
    doctor = Doctor.objects.values('name_short','id')

    payment_method = COMMCODE.objects.filter(use_yn = 'Y', commcode_grp='PAYMENT_METHOD',upper_commcode ='000014' ).annotate(code = F('commcode'),name = f_name ).values('code','name')



    return render(request,
    'Receptionist/storage_page.html',
            {
                'today_filter': today_filter,
                'storage_search':storage_search_form,
                'storage':storage_form,
                'depart':depart,
                'depart_medical':depart_medical,
                'doctor' : doctor,
                'payment_method':payment_method,
            },
        )

def waiting_list(request):
    date_start = request.POST.get('start_date')
    date_end = request.POST.get('end_date')
    depart_id = request.POST.get('depart')
    filter = request.POST.get('filter')
    string = request.POST.get('string')
 
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    print(date_max)
    today = datetime.datetime.combine(datetime.datetime.now(), datetime.time.max)
    last_2_month = datetime.datetime.combine(datetime.datetime.now() - datetime.timedelta(hours=24*60), datetime.time.min)
    print(last_2_month)
    argument_list = [] 
    if filter=='':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 
    elif filter=='name':
        argument_list.append( Q(**{'patient__name_kor__icontains':string} ) )
        argument_list.append( Q(**{'patient__name_eng__icontains':string} ) )
    elif filter=='chart':
        argument_list.append( Q(**{'patient__id__icontains':string} ) ) 



    kwargs={}
    
    if depart_id != '' :
        kwargs['depart_id'] = depart_id

    # receptions = Reception.objects.select_related('patient').exclude(progress='deleted').filter( functools.reduce(operator.or_, argument_list), **kwargs, recorded_date__range = (date_min, date_max) )
    #if filter == 'name':
    #    filter_string.append(Q( ** {'patient__name_kor__icontains' : string } ))
    #    filter_string.append(Q( ** {'patient__name_eng__icontains' : string } ))
    #    receptions = Reception.objects.select_related('patient').exclude(progress='deleted').filter( functools.reduce(operator.or_, filter_string), recorded_date__range = (date_min, date_max) )
    ##if doctor_id != '':
    ##    doctor = Doctor.objects.get(id = doctor_id)
    ##    kwargs['doctor'] = doctor
    #else:
    #    receptions = Reception.objects.select_related('patient').filter(  recorded_date__range = (date_min, date_max) )
    #


    # list_reception = []
    # if string:
    #     receptions = Reception.objects.select_related('patient').exclude(progress='deleted').filter( functools.reduce(operator.or_, argument_list), **kwargs, recorded_date__range = (date_min, date_max) )
    # else:
    #     receptions = Reception.objects.select_related('patient').exclude(progress='deleted').filter( functools.reduce(operator.or_, argument_list), **kwargs, )
    # for reception in receptions:
    #     if hasattr(reception,'payment'):
    #         payment_set = PaymentRecord.objects.filter(payment_id = reception.payment.id, status='paid', date__range = (date_min, date_max))
    #         for payment in payment_set:
    #             if payment.payment.reception not in list_reception:
    #                 list_reception.append(payment.payment.reception)

    # list_reception = []
    # payment_set =  PaymentRecord.objects.filter( status='paid', date__range = (date_min, date_max))     
    # for payment in payment_set:
    #     if payment.payment.reception not in list_reception:
    #         list_reception.append(payment.payment.reception)

    # receptions = Reception.objects.select_related('patient').exclude(progress='deleted').filter( functools.reduce(operator.or_, argument_list), **kwargs, recorded_date__range = (date_min, date_max) )    
    
    # for reception in receptions:
    #     if reception not in list_reception:
    #         list_reception.append(reception)
    datas=[]

    receptions = Reception.objects.select_related('patient').exclude(progress='deleted').filter( functools.reduce(operator.or_, argument_list), **kwargs, recorded_date__range = (date_min, date_max) )
    for reception in receptions:
        if hasattr(reception,'payment'):
            if hasattr(reception.payment,'paymentrecord_set'):
                payment_set = PaymentRecord.objects.filter(payment_id = reception.payment.id, status='paid', date__range = (date_min, date_max))
                try:
                    query = COMMCODE.objects.get(upper_commcode = '000006',commcode_grp='PT_INFO',commcode = reception.patient.marking)
                    marking = query.se1
                except COMMCODE.DoesNotExist:
                    marking=''

                if payment_set.count() is 0:
                    if reception.recorded_date.strftime('%Y%m%d') == datetime.datetime.today().strftime('%Y%m%d'):
                        continue
                    record = {
                        'reception_id':reception.id,
                        'chart':reception.patient.get_chart_no(),
                        'name_kor':reception.patient.name_kor,
                        'name_eng':reception.patient.name_eng,
                        'Depart':reception.depart.name,
                        'Doctor':reception.doctor.name_kor,
                        'unpaid_total': reception.payment.sub_total,
                        'paid':0,
                        'date_visited':reception.recorded_date.strftime('%Y-%m-%d'),
                        'date_paid':'',
                        'status':'paid' if reception.payment.progress=='paid' else 'unpaid',
                        'has_unpaid':reception.patient.has_unpaid(),
                        #'is_unpaid':pay_record.payment.reception.patient.has_unpaid(),
                        'marking':marking,
                        }  
                    datas.append(record)
                else:
                    for pay_record in payment_set:
                        record = {
                            'paymentrecord_id':pay_record.id,
                            'chart':pay_record.payment.reception.patient.get_chart_no(),
                            'name_kor':pay_record.payment.reception.patient.name_kor,
                            'name_eng':pay_record.payment.reception.patient.name_eng,
                            'Depart':pay_record.payment.reception.depart.name,
                            'Doctor':pay_record.payment.reception.doctor.name_kor,
                            'unpaid_total': pay_record.get_rest_total(),
                            'paid':pay_record.paid,
                            'date_visited':reception.recorded_date.strftime('%Y-%m-%d'),
                            'date_paid':pay_record.date.strftime('%Y-%m-%d'),
                            'status':'paid' if reception.payment.progress=='paid' else 'unpaid',
                            'has_unpaid':reception.patient.has_unpaid(),

                            'marking':marking,
                            }

                        datas.append(record)



    datas.reverse()
    context = {'datas':datas}
    return JsonResponse(context)


def get_today_list(request):
    kwargs ={}
    doctor_id = request.GET.get('doctor')
    depart_id = request.POST.get('depart')

    patient = request.POST.get('patient')
    kwargs={}
    if depart_id != '' and depart_id != None:
        depart = Depart.objects.get(pk = depart_id)
        kwargs['depart_id'] = depart

    if doctor_id != '' and doctor_id != None:
        doctor = Doctor.objects.get(id = doctor_id)
        kwargs['doctor'] = doctor

    
    argument_list2 = []
    if patient != '' and patient != None:
        argument_list2.append( Q(**{'patient__name_kor__icontains':patient } ) ) 
        argument_list2.append( Q(**{'patient__name_eng__icontains':patient } ) ) 

    if len(argument_list2) > 0:

        receptions = Reception.objects.filter(  functools.reduce(operator.or_, argument_list2),
            recorded_date__date = datetime.date.today(), progress = 'done',**kwargs,)
    else:
        receptions = Reception.objects.filter(
            recorded_date__date = datetime.date.today(), progress = 'done',**kwargs,)

    

    datas=[]
    for reception in receptions:
        try:
            query = COMMCODE.objects.get(upper_commcode = '000006',commcode_grp='PT_INFO',commcode = reception.patient.marking)
            marking = query.se1
        except COMMCODE.DoesNotExist:
            marking= ''


        count = PaymentRecord.objects.filter(payment = reception.payment,status='paid').count()
        if count is not 0:
            continue
        data={
            'reception_id':reception.id,
            'chart':reception.patient.get_chart_no(),
            'name_kor':reception.patient.name_kor,
            'name_eng':reception.patient.name_eng,
            'Depart':reception.depart.name,
            'Doctor':reception.doctor.name_kor,
            'status':reception.payment.progress,
            'total_amount':reception.payment.total,
            'DateTime':reception.recorded_date.strftime('%H:%M'),
            'has_unpaid':reception.patient.has_unpaid(),

            'is_vaccine':reception.is_vaccine, 

            'marking':marking,
            }
        datas.append(data)
        
    datas.reverse()
    context = {'datas':datas}
    return JsonResponse(context)

def get_today_selected(request):
    reception_id = request.POST.get('reception_id')

    reception = Reception.objects.get(pk = reception_id)
    diagnosis = Diagnosis.objects.get(reception_id = reception_id)
    payment = Payment.objects.get(reception_id = reception_id)

    exam_set = ExamManager.objects.filter(diagnosis_id = diagnosis.id)
    test_set = TestManager.objects.filter(diagnosis_id = diagnosis.id, test__parent_test = None)
    precedure_set = PrecedureManager.objects.filter(diagnosis_id = diagnosis.id)
    medicine_set = MedicineManager.objects.filter(diagnosis_id = diagnosis.id)

    exams = []
    for data in exam_set:
        exam = {}
        exam.update({
            'manager_id':data.id,
            'code':data.exam.code,
            'name':data.exam.name,
            'price':data.exam.get_price(),
            })
        exams.append(exam)

    tests = []
    for data in test_set:
        test = {}
        test.update({
            'manager_id':data.id,
            'code':data.test.code,
            'name':data.test.name,
            'price':data.test.get_price(),
            })
        tests.append(test)

    precedures = []
    for data in precedure_set:
        precedure = {}
        precedure.update({
            'manager_id':data.id,
            'code':data.precedure.code,
            'name':data.precedure.name,
            'amount': data.amount,
            'price':data.precedure.get_price(),
            })


        precedures.append(precedure)

    medicines = []
    for data in medicine_set:
        medicine = {}
        quantity = int(data.days) * int(data.amount)
        unit = data.medicine.get_price()
        price = quantity * int(data.medicine.get_price())
        medicine.update({
            'manager_id':data.id,
            'code':data.medicine.code,
            'name':data.medicine.name,
            'quantity':quantity,
            'price':price,
            'unit':unit,
            })
        medicines.append(medicine)

    paid = 0
    records = PaymentRecord.objects.filter(payment = payment,status = 'paid')
    for record in records:
        paid += record.paid

    print(payment.total)
    datas = {
        'chart':reception.patient.get_chart_no(),
        'name_kor':reception.patient.name_kor,
        'name_eng':reception.patient.name_eng,
        'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),
        'gender':reception.patient.gender,
        'phone':reception.patient.phone,
        'address':reception.patient.address,

        'recommendation':diagnosis.recommendation,

        'exams':exams,
        'tests':tests,
        'precedures':precedures,
        'medicines':medicines,

        'doctor_kor':reception.doctor.name_kor,
        'doctor_eng':reception.doctor.name_eng,
        'depart':reception.doctor.depart.name,

        'sub_total': payment.sub_total,
        'total_amount': payment.total,
        'paid':paid,
        'unpaid':payment.total - paid,

        'discount':'' if payment.discounted is None else payment.discounted,
        'discount_amount':'' if payment.discounted_amount is None else payment.discounted_amount,

        'is_emergency':payment.is_emergency,
        'additional':0 if payment.additional is None else payment.additional,

        'date':reception.recorded_date.strftime('%d/%m/%Y'),


    }
    if reception.reservation:
        datas.update({
            'reservation': reception.reservation.reservation_date.strftime('%Y-%m-%d %H:%M:00'),
            })

    #try:
    #    temp = Report.objects.get(reception_id = reception.id)
    #    report = temp.id
    #except Report.DoesNotExist:
    #    report= None;

    try:
        taxinvoice = TaxInvoice.objects.get(patient = reception.patient)
    except TaxInvoice.DoesNotExist:
        taxinvoice = None


    context = {
        'datas':datas,
        #'report':report,
        'tax_invoice_number':'' if taxinvoice is None else taxinvoice.number,
        'tax_invoice_company_name':'' if taxinvoice is None else taxinvoice.company_name,
        'tax_invoice_address':'' if taxinvoice is None else taxinvoice.address,
        'tax_recommendation':'' if taxinvoice is None else taxinvoice.recommend,
                
        'need_invoice':reception.need_invoice,
        'need_insurance':reception.need_insurance,


        }
    return JsonResponse(context)


def waiting_selected(request):
    record_id = request.POST.get('paymentrecord_id')

    payment_record = PaymentRecord.objects.get(pk = record_id)
    
    payment = payment_record.payment
    reception = Reception.objects.get(pk = payment.reception_id)
    diagnosis = Diagnosis.objects.get(reception_id = reception.id)

    exam_set = ExamManager.objects.filter(diagnosis_id = diagnosis.id)
    test_set = TestManager.objects.filter(diagnosis_id = diagnosis.id, test__parent_test = None)
    precedure_set = PrecedureManager.objects.filter(diagnosis_id = diagnosis.id)
    medicine_set = MedicineManager.objects.filter(diagnosis_id = diagnosis.id)

    standard_date = reception.payment.paymentrecord_set.first().date
    print('reception.need_invoice: ', reception.need_invoice)
    print('reception.need_insurance: ', reception.need_insurance)

    exams = []
    for data in exam_set:
        exam = {}
        exam.update({
            'manager_id':data.id,
            'is_checked':data.is_checked_discount,
            'code':data.exam.code,
            'name':data.exam.name,
            'price':data.exam.get_price(standard_date),
            })
        exams.append(exam)

    tests = []
    for data in test_set:
        test = {}
        test.update({
            'manager_id':data.id,
            'is_checked':data.is_checked_discount,
            'code':data.test.code,
            'name':data.test.name,
            'price':data.test.get_price(standard_date),
            })
        tests.append(test)

    precedures = []
    for data in precedure_set:
        precedure = {}
        precedure.update({
            'manager_id':data.id,
            'is_checked':data.is_checked_discount,
            'code':data.precedure.code,
            'name':data.precedure.name,
            'amount':data.amount,
            'price':data.precedure.get_price(standard_date),
            })
        precedures.append(precedure)

    medicines = []
    for data in medicine_set:
        medicine = {}
        quantity = int(data.days) * int(data.amount)
        unit = data.medicine.get_price(standard_date)
        price = quantity * int(data.medicine.get_price(standard_date))
        medicine.update({
            'manager_id':data.id,
            'is_checked':data.is_checked_discount,
            'code':data.medicine.code,
            'name':data.medicine.name,
            'quantity':quantity,
            'price':price,
            'unit':unit,
            })
        medicines.append(medicine)



    datas = {
        'chart':reception.patient.get_chart_no(),
        'name_kor':reception.patient.name_kor,
        'name_eng':reception.patient.name_eng,
        'date_of_birth':reception.patient.date_of_birth.strftime('%Y-%m-%d '),
        'gender':reception.patient.gender,
        'phone':reception.patient.phone,
        'address':reception.patient.address, 
        'passport':reception.patient.passport,
        'exams':exams,
        'tests':tests,
        'precedures':precedures,
        'medicines':medicines,

        'doctor_kor':reception.doctor.name_kor,
        'doctor_eng':reception.doctor.name_eng,
        'depart':reception.doctor.depart.name,

        'sub_total':payment.sub_total,
        'unpaid_total': payment_record.get_rest_total(),
        'paid':payment_record.paid,
        'total_payment':payment.total,
        
        'paid_by':'' if payment.memo is None else payment.memo,
        'payment_memo':'' if payment.memo is None else payment.memo,
       
        'emergency_amount':payment.sub_total * 0.3 if payment.is_emergency is True else 0,
        'additional':0 if payment.additional is None else payment.additional,

        'discount':'' if payment.discounted is None else payment.discounted,
        'discount_amount':'' if payment.discounted_amount is None else payment.discounted_amount,

        'date':reception.recorded_date.strftime('%d/%m/%Y'),
        
        
    }
    if reception.reservation:
        datas.update({
            'reservation': reception.reservation.reservation_date.strftime('%Y-%m-%d %H:%M:00'),
            })

    try:
        taxinvoice = TaxInvoice.objects.get(patient = reception.patient)
    except TaxInvoice.DoesNotExist:
        taxinvoice = None


    context = {
        'datas':datas,
        'tax_invoice_number':'' if taxinvoice is None else taxinvoice.number,
        'tax_invoice_company_name':'' if taxinvoice is None else taxinvoice.company_name,
        'tax_invoice_address':'' if taxinvoice is None else taxinvoice.address,
        'tax_recommendation': '' if taxinvoice is None else taxinvoice.recommend,
        'reception_id':reception.id,

        'need_invoice':reception.need_invoice,
        'need_insurance':reception.need_insurance,
        'method':payment_record.method,
        }
    return JsonResponse(context)


def storage_page_save(request):
    reception_id = request.POST.get('reception_id')
    paid = request.POST.get('paid')
    paid = int(paid)
    method = request.POST.get('method')
    
    memo = request.POST.get('payment_memo')

    tax_number = request.POST.get('tax_number')
    company = request.POST.get('company')
    address = request.POST.get('address')
    recommendation = request.POST.get('recommendation')
    need_invoice = request.POST.get('need_invoice')
    need_insurance = request.POST.get('need_insurance')    
    

    discount = request.POST.get('discount',0)
    discount_amount = request.POST.get('discount_amount',0)
    total = request.POST.get('total').split(' ')[0].replace(',','')

    is_emergency = request.POST.get('is_emergency',False)
    additional = request.POST.get('additional',0)

    payment = Payment.objects.get(reception_id = reception_id)
    payment.discounted = 0 if discount is '' else discount
    payment.discounted_amount = 0 if discount_amount is '' else discount_amount
    payment.is_emergency = True if is_emergency == 'true' else False
    payment.additional = additional
    payment.total = total
    payment.memo = memo
    print('****')
    print(payment.pay_time)
    payment.pay_time = datetime.datetime.now()

    reception = Reception.objects.get(pk = reception_id)
    patient = Patient.objects.get(pk = reception.patient.id)
    try:
        taxinvoice = TaxInvoice.objects.get(patient=patient)        
    except TaxInvoice.DoesNotExist:
        taxinvoice = TaxInvoice(patient = patient)       

    taxinvoice.recommend = recommendation
    taxinvoice.number = '' if tax_number == '' else tax_number
    taxinvoice.company_name = '' if company == '' else company
    taxinvoice.address = '' if address == '' else address    
    taxinvoice.save()

    if payment.progress == 'paid':
        context = {'result':'paid'}
        return JsonResponse(context)


    reception.need_invoice = True if need_invoice.lower() == 'true' else False 
    reception.need_insurance = True if need_insurance.lower() == 'true' else False  
    reception.save()

    payment_recoreds = PaymentRecord.objects.filter(payment = payment,status='paid' )
    res = int(payment.total)
    for payment_recored in payment_recoreds:
        res -= payment_recored.paid
    
    if paid > res:
        context = {'result':'overflowed'}
        return JsonResponse(context)

    res -= paid

    add_record = PaymentRecord(payment = payment)
    add_record.date = datetime.datetime.now()
    add_record.method = method
    add_record.paid = int(paid)

    if res == 0:
        payment.progress = 'paid'
    else:
        payment.progress = 'unpaid'
    payment.save()
    add_record.save()

    list_checked = request.POST.get('list_checked')
    list_checked = json.loads(list_checked)

    
    for data in list_checked:
        if data['type']=='exam':
            query = ExamManager.objects.get(id = data['id'])
        elif data['type']=='test':
            query = TestManager.objects.get(id = data['id'])
        elif data['type']=='precedure':
            query = PrecedureManager.objects.get(id = data['id'])
        elif data['type']=='medicine':
            query = MedicineManager.objects.get(id = data['id'])
        query.is_checked_discount = data['value']
        query.save()


    context = {'result':'done'}
    print('=============')
    print(payment.pay_time)
    return JsonResponse(context)



def get_bill_list(request):
    reception_id = request.POST.get('reception_id')

    tmp_reception = Reception.objects.get(pk = reception_id)
    patient = Patient.objects.get(pk = tmp_reception.patient_id)

    receptions = patient.reception_set.all()

    datas=[]
    
    for reception in receptions:
        data={}
        paymentrecord_list=[]
        unpaid = reception.payment.total
        for paymentrecord in reception.payment.paymentrecord_set.all():
            unpaid -= paymentrecord.paid
            paymentrecord_data={
                    'method':paymentrecord.method,
                    'paid':paymentrecord.paid,
                    'date':paymentrecord.date.strftime('%d-%b-%y'),
                }
            paymentrecord_list.append(paymentrecord_data)
            
        data.update({
            'date':reception.recorded_date.strftime('%d-%b-%y'),
            'total':reception.payment.total,
            'unpaid':unpaid,
            'paymentrecords':paymentrecord_list,
            })
        datas.append(data)

    
    context = {'datas':datas}
    return JsonResponse(context)


def reservation_events(request):
    kwargs ={}
    string = request.POST.get('name')
    date_start = request.POST.get('date_start').split('T')[0]
    date_end = request.POST.get('date_end').split('T')[0]
    
    depart = request.POST.get('depart')
    if depart != '' and depart is not None:
        kwargs.update({ 'depart_id':depart })
    doctor = request.POST.get('doctor')
    if doctor != '' and doctor is not None:
        kwargs.update({ 'doctor_id':doctor} )
    
    division = request.POST.get('division','')
    if division != '' and division is not None:
        kwargs.update({ 'division':division} )

    # if string != '' and string is not None:
    #     kwargs.update({ 'name__icontains':string} )    
    
    argument_list2 = []
    argument_list2.append( Q(**{'patient__name_kor__icontains':string } ) ) 
    argument_list2.append( Q(**{'patient__name_eng__icontains':string } ) )  
    if string != '' and string is not None:
        argument_list2.append( Q(**{'name__icontains':string } ) )
        
         
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)

    
    argument_list = []
    argument_list.append( Q(**{'reservation_date__range':(date_min, date_max) } ) ) 
    argument_list.append( Q(**{'re_reservation_date__range':(date_min, date_max) } ) )  

    try:

        reservations = Reservation.objects.filter(
            functools.reduce(operator.or_, argument_list),
            functools.reduce(operator.or_, argument_list2),
            **kwargs,
            # name__icontains = name
            )
    except:
        reservations = Reservation.objects.filter(
            functools.reduce(operator.or_, argument_list),
            **kwargs,
            # name__icontains = name
            )

    datas=[]
    for reservation in reservations:
        
        reservation_date = reservation.reservation_date.strftime('%Y-%m-%d %H:%M:00')
        if reservation.re_reservation_date != None:
            reservation_date = reservation.re_reservation_date.strftime('%Y-%m-%d %H:%M:00')
        # print('reservation_date: ', reservation_date)
        data = { 
            'id':reservation.id, 
            'start':reservation_date,
            #'backgroundColor':'red',
            }

        if reservation.depart.id == 1:  #1	PED
            pass
            #data.update({
            #    'backgroundColor':'grba(249,167,82,0.5)',
            #    'eventBorderColor':'grba(249,167,82,1)',
            #    })
        elif reservation.depart.id == 2:  #2	IM
            data.update({
                'backgroundColor':'rgb(149 227 169)',
                'borderColor':'rgb(149 227 169)',
                })
        elif reservation.depart.id == 3:  #3	URO
            # pass
            data.update({
               'backgroundColor':'rgb(188,255,0)',
               'borderColor':'rgb(188,255,0)',
               })
        elif reservation.depart.id == 4:  #4	PS
            data.update({
                'backgroundColor':'rgb(255,81,255)',
                'borderColor':'rgb(255,81,255)',
                })
        elif reservation.depart.id == 5:  #5	ENT
            data.update({
                'backgroundColor':'rgb(255,205,100)',
                'borderColor':'rgb(255,205,100)',
                })
        elif reservation.depart.id == 6:  #6	DERM
            data.update({
                'backgroundColor':'rgb(183,164,210)',
                'borderColor':'rgb(183,164,210)',
                })
        elif reservation.depart.id == 7:  #7	PM
            data.update({
                'backgroundColor':'rgb(147,203,249)',
                'borderColor':'rgb(147,203,249)',
                })
        elif reservation.depart.id == 9:  #9	OBGYN
            data.update({
                'backgroundColor':'rgb(254,154,202)',
                'borderColor':'rgb(254,154,202)',
                })
        elif reservation.depart.id == 10: 
            data.update({
                'backgroundColor':'rgb(255,251,0)',
                'borderColor':'rgb(255,251,0)',
                })
        elif reservation.depart.id == 11:  #9	SUGERY
            data.update({
                'backgroundColor':'rgb(70, 209, 61)',
                'borderColor':'rgb(70, 209, 61)',
                })
        elif reservation.depart.id == 12:  #9	EYES
            data.update({
                'backgroundColor':'rgb(166, 55, 163)',
                'borderColor':'rgb(166, 55, 163)',
                })
        name = ''
        depart = ''
        memo = ''
        is_vaccine = ''
        try:
            patient = Patient.objects.get(pk = reservation.patient_id)

            name = reservation.patient.name_kor + '\n' + reservation.patient.name_eng
            depart = '(' + reservation.depart.name + ' / ' + ('' if reservation.doctor is None else reservation.doctor.name_kor) + ')'
            memo = '- ' + reservation.memo
        except Patient.DoesNotExist:
            name = reservation.name
            depart = reservation.depart.name + ' / ' + ('' if reservation.doctor is None else reservation.doctor.name_kor) + ')'
            memo = '- ' + reservation.memo

            
        if reservation.division == 'VACCIN':
            is_vaccine = u'\U0001F489'#💉
        data.update({
                'title': is_vaccine + name  + '\n' + depart + '\n' + reservation.memo,
                })

    
        datas.append(data)

    


    #각 날짜 별로 환자 토탈 표시
    delta = date_max - date_min
    date_list = {}
    argument_list = []
    for i in range(delta.days + 1):
        argument_list = []
        tmp_date = date_min + datetime.timedelta(days=i)
        
        tmp_date_min = datetime.datetime.combine(tmp_date.date(), datetime.time.min)
        tmp_date_max = datetime.datetime.combine(tmp_date.date(), datetime.time.max)

        argument_list.append( Q(**{'reservation_date__range':(tmp_date_min, tmp_date_max) } ) ) 
        argument_list.append( Q(**{'re_reservation_date__range':(tmp_date_min, tmp_date_max) } ) ) 
        if string != '' and string is not None:
            argument_list.append( Q(**{'name__icontains':string} ) )         

        tmp_count = Reservation.objects.filter(
            functools.reduce(operator.or_, argument_list),
            **kwargs,
            # name__icontains = name
            )
        v_count = 0
        for _count in tmp_count:
            if _count.re_reservation_date is None or (_count.re_reservation_date != None and _count.re_reservation_date >= tmp_date_min and _count.re_reservation_date < tmp_date_max):
                v_count += 1
            else:
                continue

        date_list.update({
            tmp_date.strftime('%Y-%m-%d') : {
                'count':v_count,
                'date': tmp_date.strftime('%a %m/%d')
                }
            })


    context = {
        'datas':datas,
        'count_patient':date_list,
        }
    return JsonResponse(context)


def reservation_events_modify(request):
    id = request.POST.get('id')
    date_start = request.POST.get('date_start').split('GMT')[0].strip()
    datetime_object = datetime.datetime.strptime(date_start, '%a %b %d %Y %H:%M:%S')

    try:
        reservation = Reservation.objects.get(pk = id)
        reservation.reservation_date = datetime_object
        reservation.save()
    except:
        pass

    context = {'datas':True}
    return JsonResponse(context)

def reservation_events_delete(request):
    id = request.POST.get('id')

    try:
        reception = Reception.objects.get(reservation_id = id)
        reception.reservation_id = None
        reception.save()
    except Reception.DoesNotExist:
        pass

    reservation = Reservation.objects.get(pk = id)
    reservation.delete()

    context = {'datas':True}
    return JsonResponse(context)

def search(request):
    print('============')
    search_form = SearchPatientForm()
    page_context = 10 # 페이지 컨텐츠 갯수

    if 'search' in request.POST:
        search_form = SearchPatientForm(data = request.POST)
        if search_form.is_valid():

            category = search_form.cleaned_data['select']
            string = search_form.cleaned_data['search_input']
            kwargs = {
                    '{0}__{1}'.format(category, 'icontains'): string,
                }
            datas = Patient.objects.filter(**kwargs ).order_by("-id")
    else:
        datas = Patient.objects.all().order_by("-id")

    page = request.GET.get('page',1)
 
    paginator = Paginator(datas, page_context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)

    return render(request,
    'Receptionist/search.html',
            {
                'search': search_form,
                'datas':paging_data,
            },
        )

def check_reservation(request):
    patient_id = request.POST.get('patient_id')

    try:
        reservation = Reservation.objects.filter(patient_id = patient_id)
    except Exception as ex: # 에러 종류
        print('에러가 발생 했습니다', ex) 
    
    datas={}
    for data in reservation:
        datas.update({
            id:data.id,
            date:data.date,
            doctor:data.doctor,
            patient:data.patient,
            test:data.test,
            })
    
    context = {'datas':datas}
    return JsonResponse(context)




def reception(request,patient_num=None):
    reception_form = ReceptionForm()
    patient_form = PatientForm()
    history_form = HistoryForm()
    

    if 'save' in request.POST:
        patient_form = PatientForm(data = request.POST,)
        reception_form = ReceptionForm(data = request.POST,)
        history_form = HistoryForm(data = request.POST,)

        if patient_form.is_valid() and reception_form.is_valid() and history_form.is_valid():
            if request.POST['is_revisit']:#re-visit
                
                if patient_form.has_changed():
                    patient = Patient.objects.get(pk=request.POST['is_revisit'])
                    Patient.objects.filter(pk=request.POST['is_revisit']).update(
                        name_kor = patient_form.cleaned_data['name_kor'],
                        name_eng = patient_form.cleaned_data['name_eng'],
                        phone = patient_form.cleaned_data['phone'],
                        gender = patient_form.cleaned_data['gender'],
                        date_of_birth = patient_form.cleaned_data['date_of_birth'],
                        )

                if history_form.has_changed():
                    history_form = History.objects.filter(patient = patient).update(
                        past_history = history_form.cleaned_data['past_history'],
                        family_history = history_form.cleaned_data['family_history'],
                        )

                reception = Reception.objects.create(
                    patient = patient,
                    depart = reception_form.cleaned_data['depart'],
                    doctor = reception_form.cleaned_data['doctor'],
                    chief_complaint = reception_form.cleaned_data['chief_complaint'],
                    )
            else:#first-visit
                
                patient = Patient.objects.create(
                    id = int(request.POST['chart_no']),
                    name_kor = patient_form.cleaned_data['name_kor'],
                    name_eng = patient_form.cleaned_data['name_eng'],
                    phone = patient_form.cleaned_data['phone'],
                    gender = patient_form.cleaned_data['gender'],
                    date_of_birth = patient_form.cleaned_data['date_of_birth'],
                    )

                reception = Reception.objects.create(
                    patient = patient,
                    depart = reception_form.cleaned_data['depart'],
                    doctor = reception_form.cleaned_data['doctor'],
                    chief_complaint = reception_form.cleaned_data['chief_complaint'],
                    )

                history = History.objects.create(
                    patient = patient,
                    past_history = history_form.cleaned_data['past_history'],
                    family_history = history_form.cleaned_data['family_history'],
                    )

                #접수 등록

            return redirect('/receptionist')
        else:
            return render(request,
            'Receptionist/Reception.html',
                {
                    'patient' : patient_form,
                    'reception' : reception_form,
                    'history' : history_form,
                },
            )
    else:
        is_revisit = patient_num
        if patient_num is None: # first-visit
            visit_history = None
            is_revisit = ''

            last = Patient.objects.last()
            if last is None:
                chart_no =1
            else:
                chart_no = last.id + 1
        else: # re-visit
            try:
                patient = Patient.objects.get(id = patient_num)
                history = History.objects.get(patient = patient)

                patient_form = PatientForm(instance = patient)
                history_form = HistoryForm(instance = history)

                chart_no = patient.id
                visit_history = Reception.objects.filter(patient = patient)

            except Patient.DoesNotExist:
                pass
            except History.DoesNotExist:
                pass

        return render(request,
        'Receptionist/Reception.html',
            {
                'chart_no':"{:06d}".format(chart_no),
                'is_revisit' : is_revisit,
                'patient' : patient_form,
                'reception' : reception_form,
                'visit_history':visit_history,
                'history' : history_form,
            },
        )


def get_depart_doctor(request):
    depart_id = request.POST.get('depart')
    
    try:
        doctor = Doctor.objects.filter(depart_id = depart_id,user__depart='DOCTOR', user__is_active = True)
    except Exception as ex: # 에러 종류
        print('에러가 발생 했습니다', ex) 
    
    datas={}
    print('depart id', depart_id)
    for data in doctor:
        print('id',data.id)
        print('name',data.get_name())
        print(type(data.id))
        if depart_id == '8' and data.id in [27, 48, 54]:
            pass
        elif data.id in [30, 38, 50, 52, 41 ,43]:
            pass
        else:
            datas.update({data.get_name():data.id})
    
    context = {'datas':datas}
    return JsonResponse(context)


def reception_status(request):
    kwargs={}
    if 'search' in request.POST:
        form = SearchReceptionStatusForm(data = request.POST)
        if form.is_valid():
            if form.cleaned_data['depart'] is not None:
                kwargs['depart'] = form.cleaned_data['depart']
            if form.cleaned_data['doctor'] is not None:
                kwargs['doctor'] = form.cleaned_data['doctor']

            date_min = datetime.datetime.combine(form.cleaned_data['date'], datetime.time.min)
            date_max = datetime.datetime.combine(form.cleaned_data['date'], datetime.time.max)

    else:
        form = SearchReceptionStatusForm()
        date_min = datetime.datetime.combine(datetime.date.today(), datetime.time.min)
        date_max = datetime.datetime.combine(datetime.date.today(), datetime.time.max)
    
    reception = Reception.objects.filter(recorded_date__range = (date_min, date_max),**kwargs)


    return render(request,
    'Receptionist/Reception_Status.html',
            {
                'form':form,
                'datas':reception,
            },
        )



def storage(request,reception_num):

    try:
        reception = Reception.objects.get(pk = reception_num)
    except Reception.DoesNotExist:
        return redirect('/')

    if 'save' in request.POST:
        #저장
        payment_form = PaymentForm(request.POST)
        patient_form = PatientForm(request.POST)
        reservation_form = ReservationForm(request.POST)

        if payment_form.is_valid() and patient_form.is_valid():
            
            payment_form.save()
            pass
        
    else:
        payment_form = PaymentForm( initial={'reception': reception})
        patient_form = PatientForm(instance = reception.patient)
        reservation_form = ReservationForm(reception.follow_update)
        

    diagnosis = Diagnosis.objects.get(reception = reception)

    tests = diagnosis.test.all()
    precedures = diagnosis.precedure.all()
    medicines = diagnosis.medicine.all()

    total_amount = 0
    for data in tests:
        total_amount += data.get_price(reception.recorded_date)
    for data in precedures:
        total_amount += data.get_price(reception.recorded_date)
    for data in medicines:
        total_amount += data.get_price(reception.recorded_date)

    return render(request,
    'Receptionist/storage.html',
            {
                'reception':reception,
                'patient':patient_form,
                'payment':payment_form,
                'tests':tests,
                'precedures':precedures,
                'medicines':medicines,
                'total_amount':total_amount,
                #'reservation':reservation_form,
            },
        )


@login_required
def reservation(request):

    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')

    list_reservation_division= []
    query_reservation_division = COMMCODE.objects.filter(use_yn = 'Y',upper_commcode='000006', commcode_grp='RSRVT_DVSN').annotate(code = F('commcode'),name =f_name ).values('code','name')
    for data in query_reservation_division:
        list_reservation_division.append({
            'code':data['code'],
            'name':data['name']
            })

    today =datetime.datetime.today()
    reservation_dialog_form = ReservationDialogForm()
    reservation_search_form = ReservationSearchControl()
    datas = Reservation.objects.filter(reservation_date__month = today.month).order_by('reservation_date')
    
    reception_form = ReceptionForm()

    list_depart = Depart.objects.all()

    list_funnels = []
    funnels = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PATIENTS_FUNNELS',use_yn="Y").annotate(name = f_name ).values('commcode','name',)
    for data in funnels:
        list_funnels.append({
            'code':data['commcode'],
            'name':data['name'],
            })
    return render(request,
    'Receptionist/reservation.html',
            {
                'datas':datas,
                'reservation_dialog':reservation_dialog_form,
                'reservation_search':reservation_search_form,

                'list_depart':list_depart,
                'list_funnels': list_funnels,
                'list_reservation_division':list_reservation_division,
            },
        )

@login_required
def apointment(request):

    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')

    list_reservation_division= []
    query_reservation_division = COMMCODE.objects.filter(use_yn = 'Y',upper_commcode='000006', commcode_grp='RSRVT_DVSN').annotate(code = F('commcode'),name =f_name ).values('code','name')
    for data in query_reservation_division:
        list_reservation_division.append({
            'code':data['code'],
            'name':data['name']
            })

    today =datetime.datetime.today()
    reservation_dialog_form = ReservationDialogForm()
    reservation_search_form = ReservationSearchControl()
    datas = Reservation.objects.filter(reservation_date__month = today.month).order_by('reservation_date')
    
    reception_form = ReceptionForm()

    list_depart = Depart.objects.all()
    list_funnels = []
    funnels = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PATIENTS_FUNNELS',use_yn="Y").annotate(name = f_name ).values('commcode','name',)
    for data in funnels:
        list_funnels.append({
            'code':data['commcode'],
            'name':data['name'],
            })

    list_driver = []
    query_user = User.objects.filter(
        depart = 'DRIVERS',
        status = 'PRESENT'
        ).values('name_vi','phone_number1','phone_number2',)

    for data in query_user:
        list_driver.append({
            'name':data['name_vi'],
            'phone_number1':data['phone_number1'],
            'phone_number2':data['phone_number2'],
            })

    return render(request,
    'Receptionist/apointment.html',
            {
                'datas':datas,
                'reservation_dialog':reservation_dialog_form,
                'reservation_search':reservation_search_form,
                'list_funnels':list_funnels,
                'list_depart':list_depart,
                'list_driver':list_driver,
                'list_reservation_division':list_reservation_division,
            },
        )

@login_required
def reservation_save(request):
    reception_id = request.POST.get('reception') 
    date =request.POST.get('reservation_date')
    re_reservation_date =request.POST.get('re_reservation_date')

    # update using second address, 
    # address =request.POST.get('address')
    address =request.POST.get('reservation_address')
    need_pick_up = request.POST.get('need_pick_up',False)
    follower=request.POST.get('follower')
    pick_up_time=request.POST.get('pick_up_time')
    funnel = request.POST.get('patient_funnel','')
    funnel_etc = request.POST.get('patient_funnel_etc','')    
    apointment_memo =request.POST.get('apointment_memo')

    tmp_regis_id = request.POST.get('tmp_regis_id', '')
    print('=====>>>>', tmp_regis_id)
    if tmp_regis_id:
        print('heheheh')
        try:
            tmp_patient = DraftPatient.objects.get(pk=tmp_regis_id)
            tmp_patient.is_registed = True
            tmp_patient.save()
        except:
            pass

    if reception_id == '' or reception_id is None:
        patient=request.POST.get('reservation_patient')
        chart=request.POST.get('reservation_chart')

        memo=request.POST.get('reservation_memo')
        passport=request.POST.get('reservation_passport')
        depart=request.POST.get('reservation_depart')
        doctor=request.POST.get('reservation_doctor')
        division=request.POST.get('reservation_division')

        selected_reservation=request.POST.get('selected_reservation')
        if selected_reservation == '' or selected_reservation is None:
            reservation = Reservation()
        else:
            reservation = Reservation.objects.get(pk = selected_reservation)

        category = request.POST.get('category')
        if chart == '':
            print(1)
            patient = Patient()
            name_kor=request.POST.get('reservation_name')
            name_eng=request.POST.get('reservation_name_eng')
            date_of_birth=request.POST.get('reservation_date_of_birth')
            phone=request.POST.get('reservation_phone')
            gender=request.POST.get('reservation_gender')
            nationality=request.POST.get('reservation_nationality')
            
            print(nationality)
            # patient
            patient.name_eng = name_eng
            patient.name_kor = name_kor
            patient.date_of_birth = date_of_birth
            patient.passport = passport
            patient.gender = gender
            patient.address = address
            patient.nationality= nationality
            patient.phone = phone
            patient.category = category
            # patient.email = email
            patient.save()

            reservation.name = name_kor 
            reservation.date_of_birth= datetime.datetime.strptime(date_of_birth,'%Y-%m-%d')
            reservation.phone = phone
            reservation.passport = passport
            reservation.patient = patient
            reservation.pick_up_addr = address
        else:
            print(2)
            patient = Patient.objects.get(pk = chart)
            patient.passport = passport
            patient.category = category
            name_kor=request.POST.get('reservation_name')
            name_eng=request.POST.get('reservation_name_eng')
            patient.name_kor = name_kor
            patient.name_eng = name_eng
            patient.save()
            reservation.patient = patient
            reservation.name= patient.name_eng
            reservation.date_of_birth=patient.date_of_birth
            reservation.phone=patient.phone          
            reservation.passport = patient.passport
            reservation.pick_up_addr = address
        reservation.depart = Depart.objects.get(pk = depart)
        print(patient.category)
        if doctor == '':
            pass
        else:
            reservation.doctor = Doctor.objects.get(pk = doctor)

        try:
            if re_reservation_date != '' and re_reservation_date != None:
                reservation.re_reservation_date = datetime.datetime.strptime(re_reservation_date, "%Y-%m-%d %H:%M:%S")
            else:
                reservation.re_reservation_date = None
        except OSError as err:
            print("OS error: {0}".format(err))

        try:
            if pick_up_time != '' and pick_up_time != None:
                reservation.pick_up_time = datetime.datetime.strptime(pick_up_time, "%Y-%m-%d %H:%M:%S")
            else:
                reservation.pick_up_time = None
        except OSError as err:
            print("OS error: {0}".format(err))            

        if need_pick_up == 'true':
            reservation.need_pick_up = True
        else:
            reservation.need_pick_up = False            
        reservation.pick_up_addr = address
        reservation.follower = follower
        reservation.funnel = funnel
        reservation.funnel_etc = funnel_etc
        reservation.apointment_memo = apointment_memo
        reservation.memo = memo
        reservation.division = division
        reservation.reservation_date = datetime.datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
        reservation.save()


    else:
        print(3)
        reception = Reception.objects.get(pk=reception_id)
        try:
            if reception.reservation is None:
                reservation = Reservation()
            else:
                reservation = reception.reservation
        except ObjectDoesNotExist:# 에러 종류
            reservation = Reservation()

        
        if re_reservation_date != '' and re_reservation_date != None:
            reservation.re_reservation_date = datetime.datetime.strptime(re_reservation_date, "%Y-%m-%d %H:%M:%S")
        else:
            reservation.re_reservation_date = None

        if need_pick_up == 'true':
            reservation.need_pick_up = True
        else:
            reservation.need_pick_up = False
        reservation.pick_up_addr = address    
        reservation.follower = follower         
        reservation.reservation_date = datetime.datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
        reservation.patient = reception.patient
        reservation.depart = reception.depart
        reservation.doctor = reception.doctor
        reservation.save()

        reception.reservation = reservation
        reception.save()


    print(reservation.patient.name_eng)

    context = {'result':True}
    return JsonResponse(context)

@login_required
def pick_up_excel(request):
    
    date_start = request.GET.get('date_start')
    date_end = request.GET.get('date_end')
    pick_up = request.GET.get('pick_up')
    string = request.GET.get('string')
    drop_off = request.GET.get('drop_off','')
    kwargs={}
    depart_id = request.GET.get('depart')
    print('request: ', request)

    if pick_up != '':
        kwargs['pick_up_vehicle']=pick_up

    if drop_off != '':
        kwargs['drop_off_vehicle']=drop_off        

    if depart_id != '':
        depart = Depart.objects.get(id = depart_id)
        kwargs['depart_id'] = depart

    kwargs['need_pick_up']= True
    argument_list = []
    if string !='':
        kwargs['name'] = string
   
    date_min = datetime.datetime.combine(datetime.datetime.strptime(date_start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(date_end, "%Y-%m-%d").date(), datetime.time.max)
    

    argument_list.append( Q(**{'reservation_date__range':(date_min, date_max) } ) ) 
    argument_list.append( Q(**{'re_reservation_date__range':(date_min, date_max) } ) ) 

    reservations = Reservation.objects.filter(
            functools.reduce(operator.or_, argument_list),
            **kwargs
        ).order_by('reservation_date')


    #엑셀
    #이름 설정
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="PICK_UP_LIST' + date_start + '_' + date_end +'.xlsx"'
    

    #엑셀 파일 불러오기
    wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Pick_up_list.xlsx') #Workbook()
    ws = wb.active# grab the active worksheet

    ws['E3'] = 'Ngày : ' + date_start 
    ws['F3'] = 'đến ngày : ' + date_end


    border_thin = Border(top=Side(border_style="thin", color="000000") ,
                        left=Side(border_style="thin", color="000000") ,
                       right=Side(border_style="thin", color="000000") ,
                      bottom=Side(border_style="thin", color="000000") )

    font_base = Font(
        name = 'Arial',
        bold = False,
        size = 11,
        )

    current_row = 6
    writing_number = 1
    for data in reservations:
        try:
            patient = Patient.objects.get(pk = data.patient_id)
            ws['B' + str(current_row)] = data.patient.name_kor + '/' + data.patient.name_eng   
            ws['D' + str(current_row)] = data.patient.phone            
        except Patient.DoesNotExist:
            ws['B' + str(current_row)] = data.name
            ws['D' + str(current_row)] = data.phone

        ws['A' + str(current_row)] = writing_number        
        ws['C' + str(current_row)] = data.pick_up_addr
        ws['E' + str(current_row)] = data.apointment_memo
        ws['F' + str(current_row)] = data.reservation_date
        ws['G' + str(current_row)] = data.pick_up_time
        ws['H' + str(current_row)] = 0 if (data.follower == '' or data.follower is None) else  data.follower
        ws['I' + str(current_row)] = data.pick_up_vehicle

        current_row += 1
        writing_number += 1

    wb.save(response)
    return response
    

@login_required
def apointment_save(request):
    reservation_id = request.POST.get('reservation_id') 
    type = request.POST.get('type') 
    if type == 'pick_up_status':
        pick_up_status = request.POST.get('pick_up_status') 
        reservation = Reservation.objects.get(pk = reservation_id)
        reservation.pick_up_status = pick_up_status
    elif type == 'drop_off_status':
        drop_off_status = request.POST.get('drop_off_status') 
        reservation = Reservation.objects.get(pk = reservation_id)
        reservation.drop_off_status = drop_off_status
    else:
        pick_up_time = request.POST.get('pick_up_time') 
        drop_off_time = request.POST.get('drop_off_time') 
        pick_up_addr = request.POST.get('pick_up_addr') 
        drop_off_addr = request.POST.get('drop_off_addr')   
        apointment_memo = request.POST.get('apointment_memo') 
        pick_up_vehicle = request.POST.get('pick_up_vehicle') 
        drop_off_vehicle = request.POST.get('drop_off_vehicle') 

        reservation = Reservation.objects.get(pk = reservation_id)

        reservation.pick_up_time = None if pick_up_time == '' else datetime.datetime.strptime(pick_up_time, "%Y-%m-%d %H:%M:%S")
        reservation.drop_off_time = None if drop_off_time == '' else datetime.datetime.strptime(drop_off_time, "%Y-%m-%d %H:%M:%S")
        reservation.pick_up_addr = pick_up_addr
        reservation.drop_off_addr = drop_off_addr
        reservation.apointment_memo = apointment_memo
        reservation.pick_up_vehicle = pick_up_vehicle
        reservation.drop_off_vehicle = drop_off_vehicle

    reservation.save()

    context = {'result':True}
    return JsonResponse(context)    

@login_required
def reservation_info(request):
    
    reservation_id = request.POST.get('reservation_id')

    reservation = Reservation.objects.get(pk = reservation_id)

    try:
        reception = Reception.objects.get(reservation = reservation)
        reservation_reception_id = reception.id
    except Reception.DoesNotExist:
        reservation_reception_id =''
        reception = Reception(None)

    try:
        reception_last = Reception.objects.filter(patient = reservation.patient).last()
    except Reception.DoesNotExist:
        reception_last = Reception(None)        
    # print(reservation.patient.nationality)
    print(reservation.pick_up_addr if reservation.pick_up_addr is None else reservation.patient.address)

    context = {
        'reservation_id':reservation.id,
        'reservation_date': reservation.reservation_date.strftime('%Y-%m-%d %H:%M:%S'),
        'reservation_re_date': '' if reservation.re_reservation_date == None else reservation.re_reservation_date.strftime('%Y-%m-%d %H:%M:%S'),
        'reservation_patient': reservation.name if reservation.patient is None else reservation.patient.name_kor,
        'reservation_patient_id':'' if reservation.patient is None else reservation.patient.id,
        'reservation_chart':'' if reservation.patient is None else reservation.patient.get_chart_no(),
        'reservation_date_of_birth':(reservation.date_of_birth if reservation.patient is None else reservation.patient.date_of_birth).strftime('%Y-%m-%d'),
        'reservation_phone':reservation.phone if reservation.patient is None else reservation.patient.phone,
        'reservation_depart':reservation.depart.id,
        'reservation_doctor':reservation.doctor.id,
        'reservation_memo':reservation.memo,
        'reservation_address':reservation.pick_up_addr if reservation.patient is None else reservation.patient.address,  
        'reservation_passport':reservation.passport,
        'reservation_division':reservation.division,

        'patient_id':'' if reservation.patient is None else reservation.patient.id,
        'patient_passport':'' if reservation.patient is None else reservation.patient.passport,
        'patient_chart':'' if reservation.patient is None else reservation.patient.get_chart_no(),
        'patient_name_kor':reservation.name if reservation.patient is None else reservation.patient.name_kor,
        'patient_name_eng':reservation.name if reservation.patient is None else reservation.patient.name_eng,
        'patient_gender':'' if reservation.patient is None else reservation.patient.gender,
        'patient_nationality':'' if reservation.patient is None else reservation.patient.nationality,        
        'patient_address': reservation.pick_up_addr if reservation.patient is None else reservation.patient.address,   
        'patient_email':'' if reservation.patient is None else reservation.patient.email,      
        'need_invoice':'' if reception_last is None else reception_last.need_invoice,   
        'need_insurance':'' if reception_last is None else reception_last.need_insurance, 
        'chief_complaint':'' if reception_last is None else reception_last.chief_complaint, 
        'reservation_reception_id':reservation_reception_id, 
        'address': reservation.pick_up_addr,     
        'need_pick_up': reservation.need_pick_up,     
        'follower': reservation.follower,      
  
        'funnel': reservation.funnel,
        'funnel_etc': reservation.funnel_etc if reservation.patient is None else reservation.patient.funnel_etc,  
        'patient_memo':reservation.memo if reservation.patient is None else reservation.patient.memo,
        'reservation_patient_eng': reservation.name if reservation.patient is None else reservation.patient.name_eng,

        'reservation_memo': "" if reservation.memo is None else reservation.memo,
        }
    return JsonResponse(context)


@login_required
def apointment_info(request):
    
    reservation_id = request.POST.get('reservation_id')

    reservation = Reservation.objects.get(pk = reservation_id)

    context = {
        'reservation_id':reservation_id,
        'pick_up_time': '' if reservation.pick_up_time == None else reservation.pick_up_time.strftime('%Y-%m-%d %H:%M:%S'),
        'drop_off_time': '' if reservation.drop_off_time == None else reservation.drop_off_time.strftime('%Y-%m-%d %H:%M:%S'),
        'pick_up_addr': '' if reservation.pick_up_addr == None else reservation.pick_up_addr,
        'drop_off_addr':'' if reservation.drop_off_addr == None else reservation.drop_off_addr,
        'apointment_memo':'' if reservation.apointment_memo == None else reservation.apointment_memo,
        'pick_up_vehicle': '' if reservation.pick_up_vehicle == None else reservation.pick_up_vehicle,
        'drop_off_vehicle': '' if reservation.drop_off_vehicle == None else reservation.drop_off_vehicle,
        }
    return JsonResponse(context)

@login_required
def reservation_del(request):
    reception_id = request.POST.get('reception')

    try:
        reception = Reception.objects.get(pk = reception_id)

        tmp_reservation = reception.reservation
        reception.reservation = None
        reception.save()
        tmp_reservation.delete()

    except Reception.DoesNotExist:
        pass

    context = {'result':True}
    return JsonResponse(context)


@login_required
def refund_get_patient(request):
    reception_id = request.POST.get('reception_id')

    reception = Reception.objects.get(pk = reception_id)



    return JsonResponse({
        'chart':reception.patient.get_chart_no(),
        'name':reception.patient.get_name_kor_eng(),
        
        })

@login_required
def refund_save(request):
    reception_id = request.POST.get('reception_id')

    reason = request.POST.get('reason')
    amount = request.POST.get('amount')

    reception = Reception.objects.get(pk = reception_id)
    

    PaymentRecord(
        payment_id = reception.payment.id,
        paid = amount,
        memo = reason,
        status = "refund",
        method = ''
        ).save()


    return JsonResponse({
        'result':True,
        
        })


@login_required
def refund_cancel(request):
    record_id = request.POST.get('record_id')

    record = PaymentRecord.objects.get(id = record_id)
    record.status = 'refund_cancel'
    record.save()

    return JsonResponse({
        'result':True,
        })



@login_required
def report_list(request):
    
    reception_id = request.POST.get('reception_id')
    reception = Reception.objects.get(pk = reception_id)

    reports = Report.objects.filter(patient = reception.patient)

    datas=[]
    for report in reports:
        datas.append({
            'id':report.id,
            'chart':report.patient.get_chart_no(),
            'name_eng':report.patient.name_eng,
            'name_kor':report.patient.name_kor,
            'date_of_birth':report.patient.date_of_birth.strftime('%Y-%m-%d'),
            'depart':report.doctor.depart.name,
            'doctor':report.doctor.name_kor,
            'hospitalization':report.date_of_hospitalization.strftime('%Y-%m-%d'),
            })

    datas.reverse()


    page = request.POST.get('page',1)
    context = request.POST.get('context')
    paginator = Paginator(datas, context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)



    context = {
        'datas':datas,
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        }
    return JsonResponse(context)


@login_required
def payment_record_list(request):
    reception_id = request.POST.get('reception_id')
    reception = Reception.objects.get(pk = reception_id)

    payment_records = reception.payment.paymentrecord_set.all().order_by('-date')
    datas = []
    for payment_record in payment_records:
        data = {
            'id':payment_record.id,
            'date':payment_record.date.strftime('%Y-%m-%d %H:%M'),
            'paid':payment_record.paid,
            'method':payment_record.method,
            'name_eng':reception.patient.name_eng,
            'name_kor':reception.patient.name_kor,
            'chart':reception.patient.get_chart_no(),
            'status':payment_record.status,
            }
        datas.append(data)


    
    page = request.POST.get('page',1)
    context = request.POST.get('context')
    paginator = Paginator(datas, context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    context = {
        'datas':datas,
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),
        }
    return JsonResponse(context)


@login_required
def delete_payment(request):
    record_id = request.POST.get('record_id')

    record = PaymentRecord.objects.get(pk = record_id)
    record.status = 'cancel'
    record.save()


    payment = record.payment
    payment.progress = 'unpaid'
    payment.save()

    

    

    return JsonResponse({'result':'done'})


@login_required
def get_patient_past(request):
    reception_id = request.POST.get('reception_id')
    reception = Reception.objects.get(pk = reception_id)
    
    today = datetime.date.today()
    date_min = datetime.datetime.combine(today, datetime.time.min)
    date_max = datetime.datetime.combine(today, datetime.time.max)

    datas=[]
    receptions = Reception.objects.filter(
        patient_id = reception.patient_id,
        progress = 'done',
        )

    for reception in receptions:
        if hasattr(reception,'payment'):
            if hasattr(reception.payment,'paymentrecord_set'):
                payment_set = PaymentRecord.objects.filter(payment_id = reception.payment,status='paid')
                if payment_set.count() is 0:
                    if reception.recorded_date.strftime('%Y%m%d') == datetime.datetime.today().strftime('%Y%m%d'):
                        continue
                    record = {
                        'reception_id':reception.id,
                        'chart':reception.patient.get_chart_no(),
                        'name_kor':reception.patient.name_kor,
                        'name_eng':reception.patient.name_eng,
                        'Depart':reception.depart.name,
                        'Doctor':reception.doctor.name_kor,
                        'unpaid_total': reception.payment.total,
                        'paid':0,
                        'date_visit':reception.recorded_date.strftime('%Y-%m-%d'),
                        'date_paid':'',
                        'status':'paid' if reception.payment.progress=='paid' else 'unpaid',
                        'has_unpaid':reception.patient.has_unpaid()
                        #'is_unpaid':pay_record.payment.reception.patient.has_unpaid(),
                        }
                    datas.append(record)
                    
                else:
                    for pay_record in payment_set:
                        record = {
                            'paymentrecord_id':pay_record.id,
                            'chart':pay_record.payment.reception.patient.get_chart_no(),
                            'name_kor':pay_record.payment.reception.patient.name_kor,
                            'name_eng':pay_record.payment.reception.patient.name_eng,
                            'Depart':pay_record.payment.reception.depart.name,
                            'Doctor':pay_record.payment.reception.doctor.name_kor,
                            'unpaid_total': pay_record.get_rest_total(),
                            'paid':pay_record.paid,
                            'date_visit':pay_record.payment.reception.recorded_date.strftime('%Y-%m-%d'),
                            'date_paid':pay_record.date.strftime('%Y-%m-%d'),
                            'status':'paid' if reception.payment.progress=='paid' else 'unpaid',
                            'has_unpaid':reception.patient.has_unpaid()
                            }

                        datas.append(record)

    datas.reverse()
  

    return JsonResponse({'datas':datas})



@login_required
def Edit_Reception_get(request):

    reception_id = request.POST.get('reception_id')

    reception = Reception.objects.get(id=reception_id)

    context = {
            'id':reception.id,
            'chart':reception.patient.get_chart_no(),
            'name_kor':reception.patient.name_kor,
            'name_eng':reception.patient.name_eng,
            'age':reception.patient.get_age(),
            'gender':reception.patient.get_gender_simple(),
            'date_of_birth':reception.patient.date_of_birth.strftime('%Y-%m-%d'),
            'depart_id':reception.depart.id,

            'doctor':reception.doctor.name_kor,
            'doctor_id':reception.doctor.id,
            'chief_complaint':reception.chief_complaint,
            'medical_report':reception.need_medical_report,
        }



    return JsonResponse(context)

# 5555
@login_required
def Edit_Reception_save(request):
    reception_id = request.POST.get('reception_id')
    depart = request.POST.get('depart')
    doctor = request.POST.get('doctor')
    chief_complaint = request.POST.get('chief_complaint')
    medical_report = request.POST.get('medical_report')

    rec = Reception.objects.get(id = reception_id)
    rec.depart_id = depart
    rec.doctor_id = doctor
    rec.chief_complaint = chief_complaint
    rec.need_medical_report = 1 if medical_report =='true' else 0
    rec.save()
    
    vital_ht = request.POST.get('patient_table_vital_ht',None)
    vital_wt = request.POST.get('patient_table_vital_wt',None)
    vital_bp = request.POST.get('patient_table_vital_bp',None)
    vital_bt = request.POST.get('patient_table_vital_bt',None)
    vital_pr = request.POST.get('patient_table_vital_pr',None)
    vital_breath = request.POST.get('patient_table_vital_breath',None)

    print(vital_ht)
    print(vital_wt)
    print(vital_bp)
    print(vital_bt)
    print(vital_pr)
    print(vital_breath)

    if vital_ht is '' and vital_wt is '' and vital_bp is '' and vital_bt is '' and vital_pr is '' and vital_breath is '':
        pass
    else:
        vital = Vital.objects.filter(patient = rec.patient).first()
        print('****')
        print(vital.weight)
        print(vital_wt)
        print(rec.patient)
        vital.weight = vital_wt
        vital.height = vital_ht
        vital.blood_pressure = vital_bp
        vital.blood_temperature = vital_bt
        vital.breath = vital_breath
        vital.pulse_rate = vital_pr
        vital.save()



    return JsonResponse({'result':True})


@login_required
def Edit_Profile_Status(request):
    reception_id = request.POST.get('reception_id')
    status = request.POST.get('status')

    rec = Reception.objects.get(id = reception_id)
    rec.profile_status = status
    rec.save()



    return JsonResponse({'result':True})


@login_required
def Edit_Reception_delete(request):
    reception_id = request.POST.get('reception_id')
    rec = Reception.objects.get(id = reception_id)
    rec.progress = 'deleted'
    rec.save()

    return JsonResponse({'result':True})



@login_required
def Tax_Invoice_get(request):
    patient_id = request.POST.get('patient_id')

    context = {}

    patient = Patient.objects.select_related('taxinvoice').get(id = patient_id)

    
    context.update({
        'id':patient.id,
        'chart':patient.get_chart_no(),
        'name_kor':patient.name_kor,
        'name_eng':patient.name_eng,
        'age':patient.get_age(),
        'gender':patient.get_gender_simple(),
        'date_of_birth':patient.date_of_birth.strftime('%Y-%m-%d'),
        })


    try:
        context.update({
            'number':patient.taxinvoice.number,
            'company_name':patient.taxinvoice.company_name,
            'address':patient.taxinvoice.address,
            'employee':patient.taxinvoice.employee,
            'contact':patient.taxinvoice.contact,
            'memo':patient.taxinvoice.memo,
            })
    except TaxInvoice.DoesNotExist:
        context.update({
            'number':'',
            'company_name':'',
            'address':'',
            })


    return JsonResponse(context)


@login_required
def Tax_Invoice_save(request):
    patient_id = request.POST.get('patient_id')
    number = request.POST.get('number')
    company_name = request.POST.get('company_name')
    address = request.POST.get('address')
    employee = request.POST.get('employee', '')
    contact = request.POST.get('contact', '')
    memo = request.POST.get('memo', '')


    try:
        tax_invoice = TaxInvoice.objects.get(patient_id = patient_id)
    except:
        tax_invoice = TaxInvoice()
        tax_invoice.patient_id = patient_id

    tax_invoice.number = number
    tax_invoice.company_name = company_name
    tax_invoice.address = address
    tax_invoice.employee = employee
    tax_invoice.contact = contact
    tax_invoice.memo = memo
    tax_invoice.save()

    return JsonResponse({'result':True})


@login_required
def Documents(request):

    departs = Depart.objects.all()

    return render(request,
    'Receptionist/Documents.html',
            {
                'departs':departs,
            },
        )


@login_required
def document_search(request):
    start = request.POST.get('document_control_start')
    end = request.POST.get('document_control_end')
    depart = request.POST.get('document_control_depart')
    input = request.POST.get('document_control_input')

    
    
    kwargs={}
    if depart != '':
        kwargs['depart_id'] = depart
        #argument_list.append( Q(**{'depart_id':depart} ) ) 

    argument_list = [] 

    argument_list.append( Q(**{'patient__name_kor__icontains':input} ) ) 
    argument_list.append( Q(**{'patient__name_eng__icontains':input} ) ) 


    date_min = datetime.datetime.combine(datetime.datetime.strptime(start, "%Y-%m-%d").date(), datetime.time.min)
    date_max = datetime.datetime.combine(datetime.datetime.strptime(end, "%Y-%m-%d").date(), datetime.time.max)

    datas=[]
    receptions = Reception.objects.select_related(
                    'patient'
                ).select_related(
                    'depart'
                ).select_related(
                    'doctor'
                ).select_related(
                    'diagnosis'
                ).select_related(
                    'payment'
                ).prefetch_related(
                    'diagnosis__medicinemanager_set'
                ).prefetch_related(
                    'diagnosis__testmanager_set'
                ).order_by(
                    '-recorded_date'
                ).filter(
                    functools.reduce(operator.or_, argument_list),
                    **kwargs,
                    recorded_date__range = (date_min, date_max) 
                    ,progress = 'done'
                ).exclude(progress='deleted')
    for reception in receptions:
        data= {
            'id':reception.id,
            'chart':reception.patient.get_chart_no(),
            'name':reception.patient.get_name_kor_eng(),
            'date_of_birth':reception.patient.date_of_birth.strftime('%Y-%m-%d'),   
            'age':reception.patient.get_age(),
            'gender':reception.patient.get_gender_simple(),
            'depart':reception.depart.name,
            'doctor':reception.doctor.name_short,
            'address':reception.patient.address,
            'phone':reception.patient.phone,
            'date_time':reception.recorded_date.strftime('%Y-%m-%d %H:%M'),   
            'passport':reception.patient.passport,      
            }
        diagnosis = True
        try:
            if reception.diagnosis.medicinemanager_set.count() !=0:
                data.update({
                    'prescription':True,
                    'medicine_receipt':True,
                    })
                is_vac = reception.diagnosis.medicinemanager_set.filter(medicine__code__icontains = 'VC')
                if is_vac.count() != 0:
                    data.update({
                        'vaccine_certificate':True,
                        })
        except:
            diagnosis = False
            
        try:    
            if reception.diagnosis.testmanager_set.count() !=0:
                data.update({
                    'lab_report':True,
                    })
        except:
            diagnosis = False

        check = False

        if diagnosis == True:
            for check in reception.diagnosis.preceduremanager_set.all():
                #2,4,5,6,8
                class_id = check.precedure.precedure_class_id 
                if class_id is 2 or class_id is 4 or class_id is 5 or class_id is 6 or class_id is 8 :
                    check = True
                elif class_id is 10:
                    if 'R' in check.precedure.code:
                        check = True

            if reception.diagnosis.testmanager_set.count() !=0 or check is True:
                data.update({
                    'subclinical':True,
                    })
            try:
                report = Report.objects.filter(reception_id = reception.id).last()
                data.update({
                    'medical_report':True,
                    })

            except Report.DoesNotExist:
                pass
        try:
            if reception.payment is not None:
                data.update({
                    'medical_receipt':True,
                    })
        except:
            pass

        datas.append(data)
    

    print('==========================')
    return JsonResponse({
        'result':True,
        'datas':datas,
        })



@login_required
def document_lab(request,reception_id):
    reception = Reception.objects.get(id = reception_id)

    
    test_res = []
    manager_set = reception.diagnosis.testmanager_set.all()

    no = 0
    for lab in manager_set:
        test = TestManage.objects.get(manager_id = lab.id)
        reference_query = TestReferenceInterval.objects.filter(test_id = lab.test_id,use_yn='Y')
        list_interval = []
        unit = ''
        unit_vie = ''
        for reference in reference_query:
            list_interval.append({
                'normal_range':reference.get_range(),
                'minimum':reference.minimum,
                'maximum':reference.maximum,
                'name':'' if reference.name is None else reference.name + ' : ',
                })
            unit = reference.unit
            unit_vie = reference.unit_vie

        no+=1
        test_res.append({
            'no':no,
            'name':lab.test.name,
            'name_vie':lab.test.name_vie,
            'specimens':'',
            'result':test.result,
            'normal_range':list_interval,
            'procedure_method':'',
            'unit':unit,
            'unit_vie':unit_vie,
            })


    diagnostic = reception.diagnosis.diagnosis
    # print(reception.diagnosis)
    # print(diagnostic)
    # print(test_res)

    now = datetime.datetime.now()
    collection_time = now - datetime.timedelta(minutes=20)

    receive_time = now - datetime.timedelta(minutes=15)
    return_time = now + datetime.timedelta(minutes=10)

    # check gender
    nationality = reception.patient.nationality
    if nationality == 'Korea':
        nationality = 'HÀN QUỐC / REPUBLIC OF KOREA'
    
    gender = reception.patient.gender
    if gender == 'F' or gender == 'Female':
        gender = 'Female (Nữ)'
    elif gender == 'M' or gender == 'Male':
        gender = 'Male (Nam)'

    # print(reception.patient.passport)
    name = reception.patient.get_name_kor_eng()
    name_kor = name.split('/')[0]
    name_eng = name.split('/')[1]

    check_type = 0
    for test in test_res:
        if test['name'] == 'Covid-19 Ag_2 (Quick test)' or test['name'] == 'Covid-19 Ag (Quick test)' or test['name'] == 'Real Time PCR Covid19':
            check_type = 1
            break

    file_name = 'Receptionist/form_medical_lab2.html'
    if check_type == 1:
        file_name = 'Receptionist/form_medical_lab.html'

    return render(request,
            file_name,
                {
                    'chart':reception.patient.get_chart_no(),
                    'name_kor':name_kor,
                    'name_eng':name_eng,
                    'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                    'age':reception.patient.get_age(),
                    'gender': gender,
                    'depart_full':reception.depart.full_name,
                    'depart_full_vie':reception.depart.full_name_vie,
                    'doctor':reception.doctor.name_short,
                    'address':reception.patient.address,
                    'phone':reception.patient.phone,
                    'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                    'test_res':test_res,
                    'nationality':nationality,
                    'date_today':datetime.datetime.now().strftime('%d/%m/%Y'),
                    'diagnostic':diagnostic,
                    'passport':reception.patient.passport,
                    'collection_time':collection_time.strftime('%H:%M'),
                    'receive_time':receive_time.strftime('%H:%M'),
                    'return_time':return_time.strftime('%H:%M'),
                },
            )
    


@login_required
def document_lab2(request,reception_id):
    reception = Reception.objects.get(id = reception_id)

    
    test_res = []
    manager_set = reception.diagnosis.testmanager_set.all()
    no = 0
    for lab in manager_set:
        test = TestManage.objects.get(manager_id = lab.id)
        reference_query = TestReferenceInterval.objects.filter(test_id = lab.test_id,use_yn='Y')
        list_interval = []
        unit = ''
        unit_vie = ''
        for reference in reference_query:
            list_interval.append({
                'normal_range':reference.get_range(),
                'minimum':reference.minimum,
                'maximum':reference.maximum,
                'name':'' if reference.name is None else reference.name + ' : ',
                })
            unit = reference.unit
            unit_vie = reference.unit_vie

        no+=1
        test_res.append({
            'no':no,
            'name':lab.test.name,
            'name_vie':lab.test.name_vie,
            'specimens':'',
            'result':test.result,
            'normal_range':list_interval,
            'procedure_method':'',
            'unit':unit,
            'unit_vie':unit_vie,
            })


    diagnostic = reception.diagnosis.diagnosis
    print(diagnostic)

    now = datetime.datetime.now()
    collection_time = now - datetime.timedelta(minutes=20)

    receive_time = now - datetime.timedelta(minutes=15)
    return_time = now + datetime.timedelta(minutes=10)

    # check gender
    nationality = reception.patient.nationality
    if nationality == 'Korea':
        nationality = 'HÀN QUỐC / REPUBLIC OF KOREA'
    
    gender = reception.patient.gender
    if gender == 'F' or gender == 'Female':
        gender = 'Female (Nữ)'
    elif gender == 'M' or gender == 'Male':
        gender = 'Male (Nam)'

    # print(reception.patient.passport)
    name = reception.patient.get_name_kor_eng()
    name_kor = name.split('/')[0]
    name_eng = name.split('/')[1]
    return render(request,
    'Receptionist/form_medical_lab2.html',
            {
                'chart':reception.patient.get_chart_no(),
                'name_kor':name_kor,
                'name_eng':name_eng,
                'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                'age':reception.patient.get_age(),
                'gender': gender,
                'depart_full':reception.depart.full_name,
                'depart_full_vie':reception.depart.full_name_vie,
                'doctor':reception.doctor.name_short,
                'address':reception.patient.address,
                'phone':reception.patient.phone,
                'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                'test_res':test_res,
                'nationality':nationality,
                'date_today':datetime.datetime.now().strftime('%d/%m/%Y'),
                'diagnostic':diagnostic,
                'passport':reception.patient.passport,
                'collection_time':collection_time.strftime('%H:%M'),
                'receive_time':receive_time.strftime('%H:%M'),
                'return_time':return_time.strftime('%H:%M'),
            },
        )

    
@login_required
def document_prescription(request,reception_id):
    reception = Reception.objects.get(id = reception_id)

    
    medicine_res = []
    manager_set = reception.diagnosis.medicinemanager_set.all()
    no = 0
    for manager in manager_set:
        no+=1
        medicine_res.append({
            'no':no,
            'name':manager.medicine.name,
            'name_vie':manager.medicine.name_vie,
            'unit':'' if manager.medicine.unit is None else manager.medicine.unit,
            'unit_vie':'' if manager.medicine.unit_vie is None else manager.medicine.unit_vie,
            'quantity':manager.amount * manager.days,
            'direction_for_use':manager.memo,
            'note':'',
            })


    diagnostic = reception.diagnosis.diagnosis
    return render(request,
    'Receptionist/form_prescription.html',
            {
                'chart':reception.patient.get_chart_no(),
                'name':reception.patient.get_name_kor_eng(),
                'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                'age':reception.patient.get_age(),
                'gender':reception.patient.get_gender_simple(),
                'depart_full':reception.depart.full_name,
                'depart_full_vie':reception.depart.full_name_vie,
                'doctor':reception.doctor.name_short,
                'address':reception.patient.address,
                'phone':reception.patient.phone,
                'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                'medicine_res':medicine_res,
                'reservation_date':'' if reception.reservation_id is None else reception.reservation.reservation_date.strftime('%H:%M %d/%m/%Y'),
                'doctor':reception.doctor.name_eng,
                'nationality':reception.patient.nationality,


                'date_today':datetime.datetime.now().strftime('%d/%m/%Y'),

                'diagnostic':diagnostic
            },
        )


@login_required
def document_excel(request, reception_id):

    receptions = Reception.objects.select_related(
                    'patient'
                ).select_related(
                    'depart'
                ).select_related(
                    'doctor'
                ).select_related(
                    'diagnosis'
                ).select_related(
                    'payment'
                ).prefetch_related(
                    'diagnosis__medicinemanager_set'
                ).prefetch_related(
                    'diagnosis__testmanager_set'
                ).order_by(
                    '-recorded_date'
                ).filter(
                    id = reception_id,
                    progress = 'done'
                ).exclude(progress='deleted')          
    prescription = False
    medicine_receipt = False
    vaccine_certificate = False
    lab_report = False
    subclinical = False
    medical_report = False
    medical_receipt = False
    response = HttpResponse(content_type='application/ms-excel')
    
    diagnosis = True

    for reception in receptions:
        response['Content-Disposition'] = 'attachment; filename="' + reception.recorded_date.strftime('%d/%m/%Y') + '_' +reception.depart.name + '_'+ reception.patient.name_eng + '_' + str(reception.patient.get_chart_no()) + '.xlsx"'
        try:
            if reception.diagnosis.medicinemanager_set.count() !=0:
                prescription = True
                medicine_receipt = True
                is_vac = reception.diagnosis.medicinemanager_set.filter(medicine__code__icontains = 'VC')
                if is_vac.count() != 0:
                    vaccine_certificate = True
        except:
            diagnosis = False 
        try:    
            if reception.diagnosis.testmanager_set.count() !=0:
                lab_report = True
        except:
            diagnosis = False
        check = False
        if diagnosis == True:
            for check in reception.diagnosis.preceduremanager_set.all():
                #2,4,5,6,8
                class_id = check.precedure.precedure_class_id 
                if class_id is 2 or class_id is 4 or class_id is 5 or class_id is 6 or class_id is 8 :
                    check = True
                elif class_id is 10:
                    if 'R' in check.precedure.code:
                        check = True

            if reception.diagnosis.testmanager_set.count() !=0 or check is True:
                subclinical = True
            try:
                report = Report.objects.get(reception_id = reception.id)
                medical_report = True
            except:
                pass
        try:
            if reception.payment is not None:
                medical_receipt = True
        except:
            pass

        name = reception.depart.name
        if name == 'IM':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_im.xlsx') #Workbook()
            # wb = load_workbook('/home/light/Desktop/Projects/imedicare2/static/excel_form/document_report_im.xlsx')
        elif name == 'DENTAL':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_dental.xlsx') #Workbook()
        elif name == 'DERM':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_derm.xlsx') #Workbook()
        elif name == 'ENT':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_ent.xlsx') #Workbook()
        elif name == 'OBGYN':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_obgyn.xlsx') #Workbook()
        elif name == 'PM':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_pm.xlsx') #Workbook()  
        elif name == 'OPH':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_oph.xlsx') #Workbook()
        elif name == 'OBGYN':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_obgyn.xlsx') #Workbook()
        elif name == 'PM':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_pm.xlsx') #Workbook()
        elif name == 'DENTAL':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_dental.xlsx') #Workbook()
        elif name == 'VACCINE':
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/document_report_vaccine.xlsx') #Workbook()
        else:
            wb = load_workbook('/home/imedicare/Cofee/static/excel_form/Document_report.xlsx') #Workbook()
        # wb = load_workbook('/Users/light/Desktop/work/imdc/imedicare2/static/excel_form/Document_report.xlsx')
        border_thin = Border(top=Side(border_style="thin", color="000000") ,
                            left=Side(border_style="thin", color="000000") ,
                        right=Side(border_style="thin", color="000000") ,
                        bottom=Side(border_style="thin", color="000000") )

        font_base = Font(
            name = 'Times New Roman',
            bold = False,
            size = 9,
            )
        ws = wb.get_sheet_by_name('Medical_Receipt')# grab the active worksheet
        ws['A9'] = '1. Số hồ sơ/ PID : ' + reception.patient.get_chart_no() 
        ws['A10'] = '3. Họ và tên/ Fullname :  ' + reception.patient.get_name_kor_eng()
        ws['A11'] = '5. Địa chỉ/ Address : ' + reception.patient.address 
        ws['A12'] = '7. Số điện thoại/Phone number : ' + reception.patient.phone
        ws['A13'] = '8. Chẩn đoán/ Diagnostic : ' + reception.diagnosis.diagnosis 
        vie_name = reception.depart.full_name_vie if reception.depart.full_name_vie else ''
        full_name = reception.depart.full_name if reception.depart.full_name else ''
        ws['A14'] = '9. Khoa Khám bệnh/ Department : ' + vie_name + '/' + full_name
        ws['C9'] = '2. Quốc tịch/ Nationality : ' + reception.patient.nationality 
        ws['C10'] = '4. Giới tính/ Gender : ' + reception.patient.get_gender_simple()                
        ws['C11'] = '6. Ngày sinh/ D.O.B : ' + reception.patient.date_of_birth.strftime('%d/%m/%Y')           
        if medical_receipt == True:

            

            exam_set = ExamManager.objects.filter(diagnosis_id = reception.diagnosis.id)
            test_set = TestManager.objects.filter(diagnosis_id = reception.diagnosis.id, test__parent_test = None)
            precedure_set = PrecedureManager.objects.filter(diagnosis_id = reception.diagnosis.id)
            medicine_set = MedicineManager.objects.filter(diagnosis_id = reception.diagnosis.id)


            amount_consult = 0
            amount_image = 0
            amount_test = 0
            aount_other_exam = 0

            total = 0
            sub_total = 0
            discount = 0 
            no = 1
            current_row = 18
            for data in exam_set:
                amount_consult += data.exam.get_price(reception.recorded_date)

                ws['A' + str(current_row)] = no
                ws['B' + str(current_row)] = data.exam.name
                ws['D' + str(current_row)] = data.exam.get_price(reception.recorded_date)

                current_row +=1
                no += 1                       
            for data in test_set:
                amount_test += data.test.get_price(reception.recorded_date)    

                ws['A' + str(current_row)] = no
                ws['B' + str(current_row)] = data.test.name
                ws['D' + str(current_row)] = data.test.get_price(reception.recorded_date)
                
                current_row +=1
                no += 1                               
            for precedure in precedure_set:
                if precedure.precedure.precedure_class_id == 2 or precedure.precedure.precedure_class_id == 4 or precedure.precedure.precedure_class_id == 5 or precedure.precedure.precedure_class_id == 6 or precedure.precedure.precedure_class_id == 8:
                    amount_image+= precedure.precedure.get_price(reception.recorded_date) * precedure.amount

                    ws['A' + str(current_row)] = no
                    ws['B' + str(current_row)] = precedure.precedure.name
                    ws['D' + str(current_row)] = precedure.precedure.get_price(reception.recorded_date) * precedure.amount
                    
                    current_row +=1
                    no += 1                      
                elif precedure.precedure.precedure_class_id == 10:
                    if 'R' in precedure.precedure.code:
                        amount_image+= precedure.precedure.get_price(reception.recorded_date) * precedure.amount

                        ws['A' + str(current_row)] = no
                        ws['B' + str(current_row)] = precedure.precedure.name
                        ws['D' + str(current_row)] = precedure.precedure.get_price(reception.recorded_date) * precedure.amount
                        
                        current_row +=1
                        no += 1                          
                    else:
                        aount_other_exam += precedure.precedure.get_price(reception.recorded_date)

                        ws['A' + str(current_row)] = no
                        ws['B' + str(current_row)] = precedure.precedure.name
                        ws['D' + str(current_row)] = precedure.precedure.get_price(reception.recorded_date)
                        
                        current_row +=1
                        no += 1                           
                else:
                    aount_other_exam += precedure.precedure.get_price(reception.recorded_date)

                    ws['A' + str(current_row)] = no
                    ws['B' + str(current_row)] = precedure.precedure.name
                    ws['D' + str(current_row)] = precedure.precedure.get_price(reception.recorded_date)
                    
                    current_row +=1
                    no += 1                        

            sub_total +=amount_consult
            sub_total +=amount_image
            sub_total +=amount_test
            sub_total +=aount_other_exam                    
            if reception.payment.discounted is not None :
                discount = (reception.payment.discounted / 100) * reception.payment.sub_total
            elif reception.payment.discounted_amount is not None:
                discount = reception.payment.discounted_amount
            else:
                discount = 0

            total = sub_total - discount

            ws.delete_rows(current_row , 78 - (current_row))
            current_row +=2
            #당시 지불한 금액
            # is_paid = PaymentRecord.objects.filter(
            #     payment_id = reception.payment.id
            #     ).exists()

            # paid_total = 0
            payment_query = Payment.objects.get(reception_id = reception.id)
            # paymentrecord_query = PaymentRecord.objects.filter(payment_id = payment_query.pk).aggregate(total_price=Sum('paid'))
            
            
            # paid_total = 0 if paymentrecord_query['total_price'] is None else paymentrecord_query['total_price']
            # remain_amount = sub_total - paid_total


            # current_row = 18     

            # ws['A18'] = 1
            # ws['A19'] = 2
            # ws['A20'] = 3
            # ws['A21'] = 4
            # ws['B' + str(18)] = 'Phí khám & tư vấn/ Consultation fee'
            # ws['B' + str(19)] = 'Chẩn đoán hình ảnh/ Image Analysis'
            # ws['B' + str(20)] = 'Xét nghiệm/ Test'
            # ws['B' + str(21)] = 'Chỉ định khác/ Other examination'
            # ws['D' + str(18)] = f"{amount_consult:,}"
            # ws['D' + str(19)] = f"{amount_image:,}"
            # ws['D' + str(20)] = f"{amount_test:,}"
            # ws['D' + str(21)] = f"{aount_other_exam:,}"

            # rows = ws['A18:D' + str(current_row + 4)]
            # for row in rows:
            #     for cell in row:
            #         cell.border = border_thin
            #         cell.font = font_base

            ws['B' + str(current_row)] = 'Tổng tiền/Total '
            ws['B' + str(current_row)].font = Font(bold = True)
            ws['D' + str(current_row)] = sub_total
            current_row +=1

            ws['B' + str(current_row)] = 'Giảm trừ/Discount'
            ws['B' + str(current_row)].font = Font(bold = True)
            ws['D' + str(current_row)] = discount
            current_row +=1

            # ws['B' + str(24)] = 'Thanh toán / Paid'
            # ws['B' + str(24)].font = Font(bold = True)
            # ws['D' + str(24)] = f"{paid_total:,}"

            # ws['B' + str(25)] = 'Chưa thanh toán / Unpaid '
            # ws['B' + str(25)].font = Font(bold = True)
            # ws['D' + str(25)] = f"{remain_amount:,}"

            ws['B' + str(current_row)] = 'Amount due / Thành tiền '
            ws['B' + str(current_row)].font = Font(bold = True)
            ws['D' + str(current_row)] = total

            ws['D' + str(current_row + 3)] = reception.recorded_date.strftime('%d/%m/%Y') 
            ws['B' + str(current_row + 3)] = 'Ngày/Date:'
            ws['B' + str(current_row + 4)] = 'Bác sĩ/ Doctor'
            ws['B' + str(current_row + 5)] = '(Ký và ghi rõ họ tên)/(Sign)'

            ws['B' + str(current_row + 3)].alignment = Alignment(horizontal='right')
            ws['B' + str(current_row + 4)].alignment = Alignment(horizontal='right')
            ws['B' + str(current_row + 5)].alignment = Alignment(horizontal='right')
            # ws['C' + str(current_row + 3)].font = Font(bold = True)                   

        ws = wb.get_sheet_by_name('Medicine_Receipt')# grab the active worksheet
        ws['A9'] = '1. Số hồ sơ/ PID : ' + reception.patient.get_chart_no() 
        ws['A10'] = '3. Họ và tên/ Fullname :  ' + reception.patient.get_name_kor_eng()
        ws['A11'] = '5. Địa chỉ/ Address : ' + reception.patient.address 
        ws['A12'] = '7. Số điện thoại/Phone number : ' + reception.patient.phone
        vie_name = reception.depart.full_name_vie if reception.depart.full_name_vie else ''
        full_name = reception.depart.full_name if reception.depart.full_name else ''
        ws['A14'] = '9. Khoa Khám bệnh/ Department : ' + vie_name + '/' + full_name
        ws['E9'] = '2. Quốc tịch/ Nationality : ' + reception.patient.nationality 
        ws['E10'] = '4. Giới tính/ Gender : ' + reception.patient.get_gender_simple()                
        ws['E11'] = '6. Ngày sinh/ D.O.B : ' + reception.patient.date_of_birth.strftime('%d/%m/%Y')          
        if medicine_receipt == True:            

            medicine_set = MedicineManager.objects.filter(diagnosis_id = reception.diagnosis.id)
        
            sub_total = 0
            vat = 0
            total = 0
            diagnostic = reception.diagnosis.diagnosis
            
            current_row = 18
            writing_number = 1
            ws['A13'] = '8. Chẩn đoán/ Diagnostic : ' + diagnostic       

            for data in medicine_set:
                medicine = {}
                quantity = int(data.days) * int(data.amount)
                price = quantity * int(data.medicine.get_price(reception.recorded_date))


                ws['A' + str(current_row)] = writing_number
                ws['B' + str(current_row)] = data.medicine.name
                ws['D' + str(current_row)] = data.medicine.unit_vie + '/' + data.medicine.unit
                ws['E' + str(current_row)] = quantity
                ws['F' + str(current_row)] = f"{data.medicine.get_price(reception.recorded_date):,}"
                ws['G' + str(current_row)] = f"{price:,}"

                writing_number +=1 
                current_row +=1
                sub_total += price
            total = sub_total + (sub_total * vat / 100 )
            rows = ws['A18:G' + str(current_row + 3)]
            # for row in rows:
            #     for cell in row:
            #         cell.border = border_thin
            #         cell.font = font_base

            ws['B' + str(48)] = 'Tổng tiền đã bao gồm thuế/Sub Total including tax'
            ws['B' + str(48)].font = Font(bold = True)
            ws['G' + str(48)] = f"{sub_total:,}"

            # ws['B' + str(49)] = 'Thuế GTGT/VAT ( %)'
            # ws['B' + str(49)].font = Font(bold = True)
            # ws['G' + str(49)] = vat

            ws['B' + str(49)] = 'Tổng tiền/Total'
            ws['B' + str(49)].font = Font(bold = True)
            ws['G' + str(49)] = f"{int(total):,}"

        ws = wb.get_sheet_by_name('Subclinical')# grab the active worksheet
        ws['A9'] = '1. Số hồ sơ/ PID : ' + reception.patient.get_chart_no()
        ws['A10'] = '3. Họ và tên/ Fullname :  ' + reception.patient.get_name_kor_eng()
        ws['A11'] = '5. Địa chỉ/ Address : ' + reception.patient.address 
        ws['A12'] = '7. Số điện thoại/Phone number : ' + reception.patient.phone
        ws['A13'] = '8. Chẩn đoán/ Diagnostic : ' + reception.diagnosis.diagnosis 
        vie_name = reception.depart.full_name_vie if reception.depart.full_name_vie else ''
        full_name = reception.depart.full_name if reception.depart.full_name else ''
        ws['A14'] = '9. Khoa Khám bệnh/ Department : ' + vie_name + '/' + full_name
        ws['D9'] = '2. Quốc tịch/ Nationality : ' + reception.patient.nationality 
        ws['D10'] = '4. Giới tính/ Gender : ' + reception.patient.get_gender_simple()                
        ws['D11'] = '6. Ngày sinh/ D.O.B : ' + reception.patient.date_of_birth.strftime('%d/%m/%Y') 
        if name == 'IM':
            if subclinical == True:            

                test_set = TestManager.objects.filter(diagnosis_id = reception.diagnosis.id, test__parent_test = None)
                precedure_set = PrecedureManager.objects.filter(diagnosis_id = reception.diagnosis.id)
                
                current_row = 19
                writing_number = 1      

                no = 1
                current_row = 41
                for test in test_set:
                    ws['A' + str(current_row)] = no
                    try:
                        ws['B' + str(current_row)] = test.test.name
                    except:
                        pass
                    ws['D' + str(current_row)] = 1
                    ws['E' + str(current_row)] = ''
                    current_row +=1
                    no += 1         

                no = 1
                no_other = 1
                current_row = 67
                for precedure in precedure_set:
                    if precedure.precedure.precedure_class_id == 8:
                        ws['A' + str(current_row)] = no_other
                        ws['B' + str(current_row)] = precedure.precedure.name
                        ws['D' + str(current_row)] = precedure.amount
                        ws['E' + str(current_row)] = ''
                        current_row +=1
                        no_other += 1
                current_row = 19
                for precedure_other in precedure_set:
                    if precedure_other.precedure.precedure_class_id == 2 or precedure_other.precedure.precedure_class_id == 4 or precedure_other.precedure.precedure_class_id == 5 or precedure_other.precedure.precedure_class_id == 6 or precedure_other.precedure.precedure_class_id == 8:
                        ws['A' + str(current_row)] = no
                        ws['B' + str(current_row)] = precedure.precedure.name
                        ws['D' + str(current_row)] = precedure.amount
                        ws['E' + str(current_row)] = ''
                        current_row +=1
                        no += 1
                    elif precedure_other.precedure.precedure_class_id == 10:
                        if 'R' in precedure.precedure.code:
                            ws['A' + str(current_row)] = no
                            ws['B' + str(current_row)] = precedure.precedure.name
                            ws['D' + str(current_row)] = precedure.amount
                            ws['E' + str(current_row)] = ''
                            current_row +=1
                            no += 1           
                ws['D78'] = 'Ngàyyyy/Date: ' + reception.recorded_date.strftime('%d/%m/%Y') 
                # ws['D' + str(67)].font = Font(bold = True)  
        else:
            if subclinical == True:            

                test_set = TestManager.objects.filter(diagnosis_id = reception.diagnosis.id, test__parent_test = None)
                precedure_set = PrecedureManager.objects.filter(diagnosis_id = reception.diagnosis.id)
                
                current_row = 19
                writing_number = 1      

                no = 1
                current_row = 41
                for test in test_set:
                    ws['A' + str(current_row)] = no
                    try:
                        ws['B' + str(current_row)] = test.test.name
                    except:
                        pass
                    ws['D' + str(current_row)] = 1
                    ws['E' + str(current_row)] = ''
                    current_row +=1
                    no += 1         

                no = 1
                no_other = 1
                current_row = 63
                for precedure in precedure_set:
                    if precedure.precedure.precedure_class_id == 8:
                        ws['A' + str(current_row)] = no_other
                        ws['B' + str(current_row)] = precedure.precedure.name
                        ws['D' + str(current_row)] = precedure.amount
                        ws['E' + str(current_row)] = ''
                        current_row +=1
                        no_other += 1
                current_row = 19
                for precedure_other in precedure_set:
                    if precedure_other.precedure.precedure_class_id == 2 or precedure_other.precedure.precedure_class_id == 4 or precedure_other.precedure.precedure_class_id == 5 or precedure_other.precedure.precedure_class_id == 6 or precedure_other.precedure.precedure_class_id == 8:
                        ws['A' + str(current_row)] = no
                        ws['B' + str(current_row)] = precedure.precedure.name
                        ws['D' + str(current_row)] = precedure.amount
                        ws['E' + str(current_row)] = ''
                        current_row +=1
                        no += 1
                    elif precedure_other.precedure.precedure_class_id == 10:
                        if 'R' in precedure.precedure.code:
                            ws['A' + str(current_row)] = no
                            ws['B' + str(current_row)] = precedure.precedure.name
                            ws['D' + str(current_row)] = precedure.amount
                            ws['E' + str(current_row)] = ''
                            current_row +=1
                            no += 1           
                ws['D78'] = 'Ngàyyyy/Date: ' + reception.recorded_date.strftime('%d/%m/%Y') 
                # ws['D' + str(67)].font = Font(bold = True)
        ws = wb.get_sheet_by_name('Medical_Report')# grab the active worksheet
        ws['A9'] = '1. Số hồ sơ/ PID : ' + reception.patient.get_chart_no() 
        ws['A10'] = '3. Họ và tên/ Fullname :  ' + reception.patient.get_name_kor_eng()
        ws['A11'] = '5. Địa chỉ/ Address : ' + reception.patient.address 
        ws['A12'] = '7. Số điện thoại/Phone number : ' + reception.patient.phone
        ws['A13'] = '8. Chẩn đoán/ Diagnostic : ' + reception.diagnosis.diagnosis 
        vie_name = reception.depart.full_name_vie if reception.depart.full_name_vie else ''
        full_name = reception.depart.full_name if reception.depart.full_name else ''
        ws['A14'] = '9. Khoa Khám bệnh/ Department : ' + vie_name + '/' + full_name
        ws['B9'] = '2. Quốc tịch/ Nationality : ' + reception.patient.nationality 
        ws['B10'] = '4. Giới tính/ Gender : ' + reception.patient.get_gender_simple()                
        ws['B11'] = '6. Ngày sinh/ D.O.B : ' + reception.patient.date_of_birth.strftime('%d/%m/%Y')          
        if medical_report == True:            

            reception = Reception.objects.get(id = reception_id)
            report = Report.objects.get(reception_id = reception_id)

            next_visit = '' if reception.reservation is None else reception.reservation.reservation_date.strftime("%H:%M %d/%m/%Y")
            ICD_code_vie =''
            ICD_ =''
            ICD_code_en =''
            try:
                icd_code = ICD.objects.get(code = reception.diagnosis.ICD_code )
                ICD_code_vie =icd_code.name_vie
                ICD_ =icd_code.code
                ICD_code_en =icd_code.name

            except ICD.DoesNotExist:
                ICD_code_vie =''
                ICD_ =''
                ICD_code_en =''
      

            ws['A18'] = ' ' + '' if reception.chief_complaint is None else reception.chief_complaint
            ws['A20'] = ' ' +  reception.patient.history.past_history
            ws['A22'] = ' ' + ' ' if reception.diagnosis.assessment is None else reception.diagnosis.assessment
            ws['A24'] = ' ' + reception.diagnosis.sub_clinical_test
            ws['A26'] = ' ' + ' ' if reception.diagnosis.diagnosis is None else reception.diagnosis.diagnosis
            ws['A28'] = ' ' + ICD_ + ' '+ ICD_code_en
            ws['A29'] =  '  - ' + ICD_code_vie
            ws['A31'] = ' ' + ''if reception.diagnosis.plan is None else reception.diagnosis.plan
            ws['A33'] = ' ' + report.report
            ws['A34'] = '9.Ngày hẹn tái khám / Date of re-examination :  ' + next_visit
            # ws.Range("A34").GetCharacters(1,19).Font.Bold = True

            ws['B' + str(38)] = 'Ngày/Date: ' + reception.recorded_date.strftime('%d/%m/%Y') 
            #ws['C' + str(38)].font = Font(bold = True)    
            # row = ws[38]
            # for row in rows:
            #     for cell in row:
            #         cell.border = border_thin
            #         cell.font = font_base

        ws = wb.get_sheet_by_name('Prescription_Letter')# grab the active worksheet
        ws['A9'] = '1. Số hồ sơ/ PID : ' + reception.patient.get_chart_no() 
        ws['A10'] = '3. Họ và tên/ Fullname :  ' + reception.patient.get_name_kor_eng()
        ws['A11'] = '5. Địa chỉ/ Address : ' + reception.patient.address 
        ws['A12'] = '7. Số điện thoại/Phone number : ' + reception.patient.phone
        ws['A13'] = '8. Chẩn đoán/ Diagnostic : ' + reception.diagnosis.diagnosis  
        vie_name = reception.depart.full_name_vie if reception.depart.full_name_vie else ''
        full_name = reception.depart.full_name if reception.depart.full_name else ''
        ws['A14'] = '9. Khoa Khám bệnh/ Department : ' + vie_name + '/' + full_name
        ws['E9'] = '2. Quốc tịch/ Nationality : ' + reception.patient.nationality 
        ws['E10'] = '4. Giới tính/ Gender : ' + reception.patient.get_gender_simple()                
        ws['E11'] = '6. Ngày sinh/ D.O.B : ' + reception.patient.date_of_birth.strftime('%d/%m/%Y')           
        if prescription == True:            

            reception = Reception.objects.get(id = reception_id)
            manager_set = reception.diagnosis.medicinemanager_set.all()
            no = 0
            
            current_row = 18
            writing_number = 1     

            for manager in manager_set:
                ws['A' + str(current_row)] = writing_number
                ws['B' + str(current_row)] = manager.medicine.name

                # ws['D' + str(current_row)].style.alignment.wrap_text = True
                ws['D' + str(current_row)] = manager.medicine.unit_vie + '' if manager.medicine.unit is None or manager.medicine.unit_vie is None else '\n' + manager.medicine.unit

                ws['E' + str(current_row)] = manager.amount * manager.days
                ws['F' + str(current_row)] = manager.memo
                ws['G' + str(current_row)] = ''

                writing_number +=1 
                current_row +=1

            ws['B' + str(40)] = 'Lời dặn của bác sĩ / Advice of Doctor: '
            # ws['B' + str(37)].font = Font(bold = True)

            ws['B' + str(41)] = 'Ngày tái khám / Date of re-examination: ' + '' if reception.reservation_id is None else reception.reservation.reservation_date.strftime('%H:%M %d/%m/%Y')
            # ws['B' + str(38)].font = Font(bold = True)

            ws['F' + str(42)] = 'Ngày/Date: ' + reception.recorded_date.strftime('%d/%m/%Y')       

        ws = wb.get_sheet_by_name('Lab_Result')# grab the active worksheet
        ws['A9'] = '1. Số hồ sơ/ PID : ' + reception.patient.get_chart_no() 
        ws['A10'] = '3. Họ và tên/ Fullname :  ' + reception.patient.get_name_kor_eng()
        ws['A11'] = '5. Địa chỉ/ Address : ' + reception.patient.address 
        ws['A12'] = '7. Số điện thoại/Phone number : ' + reception.patient.phone
        ws['A13'] = '8. Chẩn đoán/ Diagnostic : ' + reception.diagnosis.diagnosis   
        vie_name = reception.depart.full_name_vie if reception.depart.full_name_vie else ''
        full_name = reception.depart.full_name if reception.depart.full_name else ''
        ws['A14'] = '9. Khoa Khám bệnh/ Department : ' + vie_name + '/' + full_name
        ws['E9'] = '2. Quốc tịch/ Nationality : ' + reception.patient.nationality 
        ws['E10'] = '4. Giới tính/ Gender : ' + reception.patient.get_gender_simple()                
        ws['E11'] = '6. Ngày sinh/ D.O.B : ' + reception.patient.date_of_birth.strftime('%d/%m/%Y')         
        if lab_report == True:            

            reception = Reception.objects.get(id = reception_id)
            manager_set = reception.diagnosis.testmanager_set.filter(test__parent_test = None)
            diagnostic = reception.diagnosis.diagnosis
            
            no = 0
            
            current_row = 18
            writing_number = 1       

            for lab in manager_set:
                test = TestManage.objects.get(manager_id = lab.id)
                reference_query = TestReferenceInterval.objects.filter(test_id = lab.test_id,use_yn='Y')
                list_interval = []
                unit = ''
                unit_vie = ''   
                interval = ''
                for reference in reference_query:
                    list_interval.append({
                        'normal_range':reference.get_range(),
                        'minimum':reference.minimum,
                        'maximum':reference.maximum,
                        'name':'' if reference.name is None else reference.name + ' : ',
                        })
                    interval += '' if reference.name is None else reference.name + ' : ' + str(reference.minimum) + ' < ' + str(reference.maximum) + '\n'
                    unit = reference.unit
                    unit_vie = reference.unit_vie                             
                ws['A' + str(current_row)] = writing_number
                ws['B' + str(current_row)] = lab.test.name_vie + '\n' + lab.test.name
                ws['D' + str(current_row)] = test.result
                ws['E' + str(current_row)] = unit if unit_vie is None else unit_vie + '\n' + unit
                ws['F' + str(current_row)] = '' if interval == '' else interval
                ws['G' + str(current_row)] = ''

                writing_number +=1 
                current_row +=1

            ws['F' + str(60)] = 'Ngày/Date: ' + datetime.datetime.now().strftime('%d/%m/%Y')  

        reception = Reception.objects.get(id = reception_id)
        ws = wb.get_sheet_by_name('Vaccine_Certification')# grab the active worksheet
        ws['A10'] = '1. Số hồ sơ/ PID : ' + reception.patient.get_chart_no() 
        ws['A11'] = '3. Họ và tên/ Fullname :  ' + reception.patient.get_name_kor_eng()
        ws['A12'] = '5. Địa chỉ/ Address : ' + reception.patient.address 
        ws['A13'] = '7. Số điện thoại/Phone number : ' + reception.patient.phone
        ws['C10'] = '2. Quốc tịch/ Nationality : ' + reception.patient.nationality 
        ws['C11'] = '4. Giới tính/ Gender : ' + reception.patient.get_gender_simple()                
        ws['C12'] = '6. Ngày sinh/ D.O.B : ' + reception.patient.date_of_birth.strftime('%d/%m/%Y') 
        if vaccine_certificate == True:
            
            

            vac_list_str = ''
            vac_list_query = reception.diagnosis.medicinemanager_set.filter(medicine__code__icontains = 'VC')

            no = 0
            
            current_row = 17
            writing_number = 1       

            for data in vac_list_query:
                vac_list_str += data.medicine.vaccine_code + "(" + data.medicine.name + ")"
                if vac_list_query[vac_list_query.count()-1].id != data.id:#라스트 체크
                    vac_list_str += ', '

            str_date = custom_strftime('%Y %B {S}',reception.recorded_date)
            ws['B' + str(16)] = 'The customer has received the ' + vac_list_str + ' vaccine on ' + str_date

            ws['C' + str(20)] = 'Ngày/Date: ' + reception.recorded_date.strftime('%d/%m/%Y')             

    wb.save(response)
    return response

@login_required
def apointment_letter(request,reservation_id):

    reservation = Reservation.objects.get(pk = reservation_id)

    return render(request,
    'Receptionist/form_apointment_letter.html',
            {
                'chart': reservation.id,
                'name': reservation.name if reservation.patient is None else reservation.patient.get_name_kor_eng(),
                'date_of_birth': (reservation.date_of_birth if reservation.patient is None else reservation.patient.date_of_birth).strftime('%Y-%m-%d'),
                'age': '',
                'gender': '' if reservation.patient is None else reservation.patient.get_gender_simple(),
                'depart_full': reservation.depart.full_name,
                'depart_full_vie': reservation.depart.full_name_vie,
                'doctor': '',
                'address':'' if reservation.patient is None else reservation.patient.address,
                'phone': '' if reservation.patient is None else reservation.patient.phone,
                'date_time': reservation.reservation_date.strftime('%H:%M %d/%m/%Y'),    
                'res_time': reservation.reservation_date.strftime('%H:%M'),    
                'res_date': reservation.reservation_date.strftime('%d'),    
                'res_month': reservation.reservation_date.strftime('%m'),    
                'res_year': reservation.reservation_date.strftime('%Y'),
                'medicine_res':'',
                'reservation_date': reservation.reservation_date.strftime('%H:%M %d/%m/%Y'),
                'nationality':  '' if reservation.patient is None else reservation.patient.nationality,
                'date_today':datetime.datetime.now().strftime('%d/%m/%Y'),
            },
        )

@login_required
def document_medical_receipt(request,reception_id):

    reception = Reception.objects.get(id = reception_id)

    exam_set = ExamManager.objects.filter(diagnosis_id = reception.diagnosis.id)
    test_set = TestManager.objects.filter(diagnosis_id = reception.diagnosis.id, test__parent_test = None)
    precedure_set = PrecedureManager.objects.filter(diagnosis_id = reception.diagnosis.id)
    medicine_set = MedicineManager.objects.filter(diagnosis_id = reception.diagnosis.id)


    amount_consult = 0
    amount_image = 0
    amount_test = 0
    aount_other_exam = 0

    total = 0
    sub_total = 0
    discount = 0 

    exams = []
    no = 1
    for data in exam_set:
        #exam = {}
        #exam.update({
        #    'no':no,
        #    'name':data.exam.name,
        #    'price':f"{data.exam.get_price(reception.recorded_date):,}",
        #    })
        #no += 1
        #exams.append(exam)
        amount_consult += data.exam.get_price(reception.recorded_date)

    tests = []
    for data in test_set:
        #test = {}
        #test.update({
        #    'no':no,
        #    'name':data.test.name,
        #    'price':f"{data.test.get_price(reception.recorded_date):,}",
        #    })
        #no += 1
        #tests.append(test)
        amount_test += data.test.get_price(reception.recorded_date)

    #2,4,5,6,10(pm 의 R)
    no = 1
    no_other = 1
    for precedure in precedure_set:
        #if precedure.precedure.precedure_class_id == 8:
        #    other_tests.append({
        #        'no':no_other,
        #        'name':precedure.precedure.name,
        #        'amount':f"{precedure.precedure.get_price(reception.recorded_date):,}",
        #        'waiting_time':''
        #        })
        #    no_other += 1
        #8
        if precedure.precedure.precedure_class_id == 2 or precedure.precedure.precedure_class_id == 4 or precedure.precedure.precedure_class_id == 5 or precedure.precedure.precedure_class_id == 6 or precedure.precedure.precedure_class_id == 8:
            #image_analysations.append({
            #    'no':no,
            #    'name':precedure.precedure.name,
            #    'amount':f"{precedure.precedure.get_price(reception.recorded_date):,}",
            #    'waiting_time':''
            #    })
            #no += 1
            amount_image+= precedure.precedure.get_price(reception.recorded_date) * precedure.amount
        elif precedure.precedure.precedure_class_id == 10:
            if 'R' in precedure.precedure.code:
               #image_analysations.append({
               #'no':no,
               #'name':precedure.precedure.name,
               #'amount':f"{precedure.precedure.get_price(reception.recorded_date):,}",
               #'waiting_time':''
               #})
               #no += 1
               amount_image+= precedure.precedure.get_price(reception.recorded_date) * precedure.amount
            else:
                aount_other_exam += precedure.precedure.get_price(reception.recorded_date)
        else:
            aount_other_exam += precedure.precedure.get_price(reception.recorded_date)

    sub_total +=amount_consult
    sub_total +=amount_image
    sub_total +=amount_test
    sub_total +=aount_other_exam

    
    

    if reception.payment.discounted is not None :
        discount = (reception.payment.discounted / 100) * reception.payment.sub_total
    elif reception.payment.discounted_amount is not None:
        discount = reception.payment.discounted_amount
    else:
        discount = 0

    total = sub_total - discount

       #당시 지불한 금액
    is_paid = PaymentRecord.objects.filter(
        payment_id = reception.payment.id
        ).exists()




    #잔액
    paid_total = 0
    payment_query = Payment.objects.get(reception_id = reception.id)
    paymentrecord_query = PaymentRecord.objects.filter(payment_id = payment_query.pk).aggregate(total_price=Sum('paid'))
    
    
    paid_total = 0 if paymentrecord_query['total_price'] is None else paymentrecord_query['total_price']
    remain_amount = sub_total - paid_total



    return render(request,
    'Receptionist/form_medical_receipt.html',
            {
                'chart':reception.patient.get_chart_no(),
                'name':reception.patient.get_name_kor_eng(),
                'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                'age':reception.patient.get_age(),
                'gender':reception.patient.get_gender_simple(),
                'depart_full':reception.depart.full_name,
                'depart_full_vie':reception.depart.full_name_vie,
                'doctor':reception.doctor.name_short,
                'address':reception.patient.address,
                'phone':reception.patient.phone,
                'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                'doctor':reception.doctor.name_eng,
                'diagnostic':reception.diagnosis.diagnosis,
                'nationality':reception.patient.nationality,


                #'exams':exams,
                #'tests':tests,
                #'precedures':precedures,
                #'medicines':medicines,
                'amount_consult':f"{amount_consult:,}",
                'amount_image':f"{amount_image:,}",
                'amount_test':f"{amount_test:,}",
                'aount_other_exam':f"{aount_other_exam:,}",

                'sub_total':f"{sub_total:,}",
                #'discount':discount,#'' if reception.payment.discounted is None else reception.payment.discounted,

                'discount':f"{discount:,}",
                'total_payment':f"{total:,}",

                'date_today':reception.recorded_date.strftime('%d/%m/%Y'),
                                
                'is_paid':is_paid,
                'paid':f"{paid_total:,}",
                'remain':f"{remain_amount:,}",
            },
        )

@login_required
def document_medical_receipt_old(request,reception_id,):

    reception = Reception.objects.get(id = reception_id)

    exam_set = ExamManager.objects.filter(diagnosis_id = reception.diagnosis.id)
    test_set = TestManager.objects.filter(diagnosis_id = reception.diagnosis.id, test__parent_test = None)
    precedure_set = PrecedureManager.objects.filter(diagnosis_id = reception.diagnosis.id)
    medicine_set = MedicineManager.objects.filter(diagnosis_id = reception.diagnosis.id)
    print(medicine_set)
    exams = []
    no = 1
    
    is_paid = False
    paid_amount = 0
    remain_amount = 0
    record_id= request.GET.get('record_id','')

    type= request.GET.get('type')
    if type == 'bf':
        discount_input= request.GET.get('discount_input',0) 
        discount_amount= request.GET.get('discount_amount',0) 
        additional_amount= int( request.GET.get('additional_amount',0) )
        total_amount= request.GET.get('total_amount',0)
        total_amount = int(total_amount.replace(',', ''))

        if discount_input is not 0 and discount_input is not '' and discount_input is not '0':
    
            discount = round( ( int( discount_input) / 100) * total_amount )
        elif discount_amount is not 0 and discount_amount is not '':

            discount = int( discount_amount )
        else:
   
            discount = 0
        #temp_today= datetime.datetime.strptime('2020-06-13 10:00',"%Y-%m-%d %H:%M")

        additional = additional_amount
        sub_total = total_amount + additional
        total = total_amount - discount + additional
        
        for data in exam_set:
            exam = {}
            exam.update({
                'no':no,
                'name':data.exam.name,
                'price':f"{data.exam.get_price():,}",
                })
            no += 1
            exams.append(exam)

        tests = []
        for data in test_set:
            test = {}
            test.update({
                'no':no,
                'name':data.test.name,
                'price':f"{data.test.get_price():,}",
                })
            no += 1
            tests.append(test)

        precedures = []
        for data in precedure_set:
            precedure = {}
            precedure.update({
                'no':no,
                'name':data.precedure.name,
                'amount':data.amount,
                'price':f"{data.precedure.get_price():,}",
                'sub_total':f"{data.precedure.get_price() * data.amount:,}",
                })
            no += 1
            precedures.append(precedure)

        medicine_show_no = no
        total_medicine_fee = 0
        medicines= []
        for data in medicine_set:
            print(data)
            medicine = {}
            medicine.update({
                'no':no,
                'name':data.medicine.name,
                'amount':data.amount * data.days,
                'price':f"{data.medicine.get_price():,}",
                'sub_total':f"{data.medicine.get_price() * data.amount * data.days:,}",
                })
            no += 1
            medicines.append(medicine)
            total_medicine_fee += data.amount * data.days * data.medicine.get_price()


    else:
        print('h2')

        if reception.payment.discounted is not 0 :
            discount = (reception.payment.discounted / 100) * reception.payment.sub_total
        elif reception.payment.discounted_amount is not 0:
            discount = reception.payment.discounted_amount
        else:
            discount = 0

        additional = reception.payment.additional
        sub_total = reception.payment.sub_total + additional
        total = reception.payment.sub_total - discount + additional

        for data in exam_set:
            exam = {}
            exam.update({
                'no':no,
                'name':data.exam.name,
                'price':f"{data.exam.get_price(reception.recorded_date):,}",
                })
            no += 1
            exams.append(exam)

        tests = []
        for data in test_set:
            test = {}
            test.update({
                'no':no,
                'name':data.test.name,
                'price':f"{data.test.get_price(reception.recorded_date):,}",
                })
            no += 1
            tests.append(test)

        precedures = []
        for data in precedure_set:
            precedure = {}
            precedure.update({
                'no':no,
                'name':data.precedure.name,
                'amount':data.amount,
                'price':f"{data.precedure.get_price(reception.recorded_date):,}",
                'sub_total':f"{data.precedure.get_price(reception.recorded_date) * data.amount:,}",
                })
            no += 1
            precedures.append(precedure)

        medicine_show_no = no
        total_medicine_fee = 0
        medicines= []
        for data in medicine_set:
            print(data)
            medicine = {}
            medicine.update({
                'no':no,
                'name':data.medicine.name,
                'amount':data.amount * data.days,
                'price':f"{data.medicine.get_price(reception.recorded_date):,}",
                'sub_total':f"{data.medicine.get_price(reception.recorded_date) * data.amount * data.days:,}",
                })
            no += 1
            medicines.append(medicine)
            total_medicine_fee += data.amount * data.days * data.medicine.get_price(reception.recorded_date)

        
        #당시 지불한 금액

        paid_amount = 0
        if(record_id != ''):
            paid_amount = PaymentRecord.objects.get(id = record_id).paid
            is_paid = True

        #잔액
        paid_total = 0
        payment_query = Payment.objects.get(reception_id = reception.id)
        paymentrecord_query = PaymentRecord.objects.filter(payment_id = payment_query.pk,status='paid').aggregate(total_price=Sum('paid'))
    
        paid_total = paymentrecord_query['total_price']
        print(paid_total)
        remain_amount = int(total) - int(paid_total)

    
    is_medicine_show = request.GET.get('is_medicine_show')
    if is_medicine_show== 'false':
        no = medicine_show_no +1

    print('=======')
    print(discount)
    print(total)
    return render(request,
    'Receptionist/form_medical_receipt_old.html',
            {
                'chart':reception.patient.get_chart_no(),
                'name':reception.patient.get_name_kor_eng(),
                'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                'age':reception.patient.get_age(),
                'email': reception.patient.email,
                'gender':reception.patient.get_gender_simple(),
                'depart_full':reception.depart.full_name,
                'depart_full_vie':reception.depart.full_name_vie,
                'doctor':reception.doctor.name_short,
                'address':reception.patient.address,
                'phone':reception.patient.phone,
                'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                'doctor':reception.doctor.name_eng,
                'diagnostic':reception.diagnosis.diagnosis,
                'nationality':reception.patient.nationality,
                'date_today':reception.recorded_date.strftime('%d/%m/%Y'),

                'sub_total':f"{sub_total:,}",
                'total':f"{total:,}",
                'discount':f"{discount:,}",
                'additional':f"{additional:,}",
                'additional_no':no,

                'is_paid':is_paid,
                'paid':f"{paid_amount:,}",
                'remain':f"{remain_amount:,}",

                'exams':exams,
                'tests':tests,
                'precedures':precedures,
                'medicines':medicines,

                'is_medicine_show':'true',
                'medicine_show_no':medicine_show_no,
                'total_medicine_fee':f"{total_medicine_fee:,}",
                }
    )


@login_required
def document_medicine_receipt(request,reception_id):
    reception = Reception.objects.get(id = reception_id)

    medicine_set = MedicineManager.objects.filter(diagnosis_id = reception.diagnosis.id)

    no=1
    medicines = []
    
    
    sub_total = 0
    vat = 0
    total = 0

    diagnostic = reception.diagnosis.diagnosis
    for data in medicine_set:
        medicine = {}
        quantity = int(data.days) * int(data.amount)
        price = quantity * int(data.medicine.get_price(reception.recorded_date))
        medicine.update({
            'no':no,
            'name':data.medicine.name,
            'quantity':quantity,
            'price':f"{price:,}",
            'unit':data.medicine.unit,
            'unit_vie':data.medicine.unit_vie,
            'unit_price':f"{data.medicine.get_price(reception.recorded_date):,}",
            })
        no += 1
        sub_total += price

        medicines.append(medicine)

    total = sub_total + (sub_total * vat / 100 )
    return render(request,
    'Receptionist/form_medicine_receipt.html',
            {
                'chart':reception.patient.get_chart_no(),
                'name':reception.patient.get_name_kor_eng(),
                'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                'age':reception.patient.get_age(),
                'gender':reception.patient.get_gender_simple(),
                'depart_full':reception.depart.full_name,
                'depart_full_vie':reception.depart.full_name_vie,
                'doctor':reception.doctor.name_short,
                'address':reception.patient.address,
                'phone':reception.patient.phone,
                'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                'doctor':reception.doctor.name_eng,
                'nationality':reception.patient.nationality,


                'date_reception':reception.recorded_date.strftime('%d/%m/%Y'),

                'medicines':medicines,
                
                'sub_total':f"{sub_total:,}",
                'vat':vat,
                'total':f"{int(total):,}",

                'diagnostic':diagnostic
            },
        )


@login_required
def document_subclinical(request,reception_id):
    reception = Reception.objects.get(id = reception_id)

    # test_set = TestManager.objects.filter(diagnosis_id = reception.diagnosis.id)
    test_set = TestManager.objects.filter(diagnosis_id = reception.diagnosis.id, test__parent_test = None)
    precedure_set = PrecedureManager.objects.filter(diagnosis_id = reception.diagnosis.id)

    

    tests = []

    no = 1
    for test in test_set:
        tests.append({
            'no':no,
            'name':test.test.name,
            'amount':1,
            'waiting_time':''
            })
        no += 1



    image_analysations = []
    other_tests = []
    #2,4,5,6,10(pm 의 R)
    no = 1
    no_other = 1
    for precedure in precedure_set:
        if precedure.precedure.precedure_class_id == 8:
            other_tests.append({
                'no':no_other,
                'name':precedure.precedure.name,
                'amount':precedure.amount,
                'waiting_time':''
                })
            no_other += 1
        #8
        elif precedure.precedure.precedure_class_id == 2 or precedure.precedure.precedure_class_id == 4 or precedure.precedure.precedure_class_id == 5 or precedure.precedure.precedure_class_id == 6 or precedure.precedure.precedure_class_id == 8:
            image_analysations.append({
                'no':no,
                'name':precedure.precedure.name,
                'amount':precedure.amount,
                'waiting_time':''
                })
            no += 1
        elif precedure.precedure.precedure_class_id == 10:
            if 'R' in precedure.precedure.code:
                image_analysations.append({
                'no':no,
                'name':precedure.precedure.name,
                'amount':precedure.amount,
                'waiting_time':''
                })
                no += 1


    return render(request,
    'Receptionist/form_subclinical.html',
            {
                'chart':reception.patient.get_chart_no(),
                'name':reception.patient.get_name_kor_eng(),
                'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                'age':reception.patient.get_age(),
                'gender':reception.patient.get_gender_simple(),
                'depart_full':reception.depart.full_name,
                'depart_full_vie':reception.depart.full_name_vie,
                'doctor':reception.doctor.name_short,
                'address':reception.patient.address,
                'phone':reception.patient.phone,
                'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                'doctor':reception.doctor.name_eng,
                'diagnostic':reception.diagnosis.diagnosis,
                'nationality':reception.patient.nationality,


                'date_today':reception.recorded_date.strftime('%d/%m/%Y'),

                'tests':tests,
                'image_analysations':image_analysations,
                'other_tests':other_tests,
            },
        )


@login_required
def document_medical_report(request,reception_id):
    reception = Reception.objects.get(id = reception_id)
    report = Report.objects.get(reception_id = reception_id)

    next_visit = '' if reception.reservation is None else reception.reservation.reservation_date.strftime("%H:%M %d/%m/%Y")
    ICD_code_vie =''
    ICD_ =''
    ICD_code_en =''
    try:
        icd_code = ICD.objects.get(code = reception.diagnosis.ICD_code )
        ICD_code_vie =icd_code.name_vie
        ICD_ =icd_code.code
        ICD_code_en =icd_code.name

    except ICD.DoesNotExist:
        ICD_code_vie =''
        ICD_ =''
        ICD_code_en =''


    return render(request,
    'Receptionist/form_medical_report.html',
            {
                'chart':reception.patient.get_chart_no(),
                'name':reception.patient.get_name_kor_eng(),
                'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                'age':reception.patient.get_age(),
                'gender':reception.patient.get_gender_simple(),
                'depart_full':reception.depart.full_name,
                'depart_full_vie':reception.depart.full_name_vie,
                'doctor':reception.doctor.name_short,
                'address':reception.patient.address,
                'phone':reception.patient.phone,
                'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                'doctor':reception.doctor.name_eng,
                'nationality':reception.patient.nationality,
                'date_today':reception.recorded_date.strftime('%d/%m/%Y'),

                'chief_complaint':'<br />' if reception.chief_complaint is None else reception.chief_complaint,
                'past_history':reception.patient.history.past_history,
                'assessment':'<br />' if reception.diagnosis.assessment is None else reception.diagnosis.assessment,
                'object':'<br />' if reception.diagnosis.objective_data is None else reception.diagnosis.objective_data,
                'diagnosis':'<br />' if reception.diagnosis.diagnosis is None else reception.diagnosis.diagnosis,
                'sub_clinical_test':reception.diagnosis.sub_clinical_test,
                'icd_code':ICD_ ,
                'ICD_code_vie':ICD_code_vie,
                'ICD_code_en':ICD_code_en,


                'plan':'<br />'if reception.diagnosis.plan is None else reception.diagnosis.plan,
                'doctor_reommend':report.report,

                'recorded_date':reception.recorded_date.strftime('%d/%m/%Y'),
                'next_visit':next_visit,
            },
        )


@login_required
def document_vaccine_certificate(request,reception_id):

    reception = Reception.objects.get(id = reception_id)

    vac_list_str = ''
    vac_list_query = reception.diagnosis.medicinemanager_set.filter(medicine__code__icontains = 'VC')


    for data in vac_list_query:
        vac_list_str += data.medicine.vaccine_code + "(" + data.medicine.name + ")"
        if vac_list_query[vac_list_query.count()-1].id != data.id:#라스트 체크
            vac_list_str += ', '
        

    str_date = custom_strftime('%Y %B {S}',reception.recorded_date)

    return render(request,
    'Receptionist/form_vaccine_certificate.html',
            {
                'chart':reception.patient.get_chart_no(),
                'name':reception.patient.get_name_kor_eng(),
                'date_of_birth':reception.patient.date_of_birth.strftime('%d/%m/%Y'),   
                'age':reception.patient.get_age(),
                'gender':reception.patient.get_gender_simple(),
                'depart_full':reception.depart.full_name,
                'depart_full_vie':reception.depart.full_name_vie,
                'doctor':reception.doctor.name_short,
                'address':reception.patient.address,
                'phone':reception.patient.phone,
                'date_time':reception.recorded_date.strftime('%H:%M %d/%m/%Y'),    
                'doctor':reception.doctor.name_eng,
                'nationality':reception.patient.nationality,
                'date_today':reception.recorded_date.strftime('%d/%m/%Y'),

                'str_date':str_date,
                'vac_list_str':vac_list_str,
            },
        )


@login_required
def package_list(request):

    filter = request.POST.get('filter')
    string = request.POST.get('string')

    
    query = Precedure.objects.filter(type='PKG',use_yn='Y')

    datas = []

    for data in query:
        datas.append({
            'id':data.id,
            'code':data.code,
            'name':data.id,
            'price':data.get_price(),
            'count':data.count,
            })


        
    page = request.POST.get('page',1)
    context = request.POST.get('context_in_page')

    paginator = Paginator(datas, context)
    try:
        paging_data = paginator.page(page)
    except PageNotAnInteger:
        paging_data = paginator.page(1)
    except EmptyPage:
        paging_data = paginator.page(paginator.num_pages)


    context = {
        'datas':datas,
        'page_range_start':paging_data.paginator.page_range.start,
        'page_range_stop':paging_data.paginator.page_range.stop,
        'page_number':paging_data.number,
        'has_previous':paging_data.has_previous(),
        'has_next':paging_data.has_next(),

        }
    return JsonResponse(context)


@login_required
def patient_package_list(request):

    patient_id = request.POST.get('patient_id')

    datas = []
    patient = Patient.objects.get(id = patient_id)
    query = Package_Manage.objects.filter( patient_id = patient_id).values('grouping').annotate(Sum('grouping')).order_by('-grouping')


    for data in query:
        patient_package = Package_Manage.objects.filter( 
            patient_id = patient_id, 
            grouping = data['grouping'],
            )

        count_max = Package_Manage.objects.filter( 
            patient_id = patient_id, 
            grouping = data['grouping'],
            ).last()

        depart = Depart.objects.get(id = patient_package[0].depart)

        datas.append({
            'id':patient_package[0].id,
            'name':patient_package[0].precedure_name,
            'depart':depart.name,
            'depart_id':depart.id,
            'doctor':patient_package[0].doctor,
            'count_now':patient_package.exclude(date_used = '0000-00-00 00:00:00').count(),
            'count_max':patient_package.last().itme_round,
            })


    return JsonResponse({
        'result':True,
        'datas':datas,

        'chart':patient.get_chart_no(),
        'name_kor':patient.name_kor,
        'name_eng':patient.name_eng,
        'age':patient.get_age(),
        'gender':patient.get_gender_simple(),
        'date_of_birth':patient.date_of_birth.strftime('%Y-%m-%d'),
        })







@login_required
def set_package_to_patient(request):


    id = request.POST.get('id')
    patient_id = request.POST.get('patient_id')
    depart_id = request.POST.get('depart_id')
    doctor_id = request.POST.get('doctor_id')

    package = Precedure.objects.get(id = id)

    

    tmp_query = Package_Manage.objects.filter(patient_id = patient_id).order_by('grouping').last()
    if tmp_query is None:
        tmp_group_id = 1
    else:
        tmp_group_id = int(tmp_query.grouping) + 1


    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for i in range(1, int(package.count) + 1):
        pack_mng = Package_Manage()

        pack_mng.patient_id = patient_id
        pack_mng.depart = depart_id
        pack_mng.doctor = doctor_id
        #pack_mng.reception_id = 1
        pack_mng.precedure = package
        pack_mng.precedure_name = package.name
        pack_mng.itme_round = i
        #pack_mng.memo 
        pack_mng.date_bought = now
        #pack_mng.date_used
        #pack_mng.date_refund
        pack_mng.grouping = tmp_group_id
        pack_mng.registrant = request.user.id
        pack_mng.date_register = now
        pack_mng.modifier = request.user.id
        pack_mng.date_modify = now


        pack_mng.save()

    reception = Reception()
    reception.depart_id = depart_id
    reception.doctor_id = doctor_id
    reception.patient_id = patient_id
    reception.progress = 'done'
    reception.chief_complaint = ''
    reception.save()
    diagnosis = Diagnosis()
    diagnosis.assessment =''
    diagnosis.objective_data =''
    diagnosis.diagnosis =''
    diagnosis.plan =''
    diagnosis.reception = reception
    diagnosis.save()
    precedure_mng = PrecedureManager()
    precedure_mng.diagnosis = diagnosis
    precedure_mng.precedure = package
    precedure_mng.amount = 1
    precedure_mng.save()
    payment = Payment()
    payment.sub_total = package.get_price()
    payment.total = package.get_price()
    payment.reception = reception
    payment.progress = 'unpaid'
    payment.save()







    return JsonResponse({
        'result':True,
        })


def patient_package_reception(request):

    id = request.POST.get('id')
    patient_id = request.POST.get('patient_id')
    depart_id = request.POST.get('depart_id')
    doctor_id = request.POST.get('doctor_id')


    reception = Reception(patient_id = patient_id)
    reception.depart_id = depart_id
    reception.doctor_id = doctor_id
    reception.chief_complaint = ''

    reception.save()

    package = Package_Manage.objects.get(id = id)

    package_now = Package_Manage.objects.filter(
        patient_id = patient_id,
        precedure = package.precedure,
        grouping = package.grouping,
        date_used = '0000-00-00 00:00:00',
        ).order_by('itme_round').first()

    package_now.reception_id = reception.id
    package_now.save()

    return JsonResponse({
        'result':True,
        })



def patient_package_history_modal(request):

    id = request.POST.get('id')

    package = Package_Manage.objects.get(id = id)

    datas = []
    query_package = Package_Manage.objects.filter(
        patient = package.patient, 
        precedure = package.precedure, 
        grouping = package.grouping
        ).order_by('itme_round')


    for data in query_package:
        datas.append({
            'id':data.id,
            'patient_name':data.patient.name_kor + '<br/>' + data.patient.name_eng,
            'precedure_name':data.precedure_name,
            'round':data.itme_round,
            'date_bought':data.date_bought[0:16],
            'date_used':'' if data.date_used == '0000-00-00 00:00:00' else data.date_used[0:16],
            })

    return JsonResponse({
        'result':True,
        'datas':datas,
        })


def list_agreement(request):

    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')

    list_agreement = []
    query_agreement = COMMCODE.objects.filter(use_yn = 'Y',commcode_grp='PT_SIGN').annotate(code = F('commcode'),name = F('commcode_name_en') ).values('code','name')
    for data in query_agreement:
        list_agreement.append({
            'id':data['code'],
            'name':data['name']
            })
    
    reception_id = request.POST.get('reception_id')
    depart =request.user.doctor.depart.name

    #reception
    #depart
    #patient_id
    #patient_name
    #type
    #document
    #is_sign
    #sign_data

    datas = []
    #sign_query = Sign_Manage.objects.filter(reception = reception_id, depart = depart)
    for data in query_agreement:
        try:
            sign_query = Sign_Manage.objects.get(reception = reception_id,document = data['code'],use_yn = 'Y')
        except Sign_Manage.DoesNotExist:
            sign_query = None

        datas.append({
            'id':'' if sign_query is None else sign_query.id,
            'name':data['name'],
            'date':'' if sign_query is None else sign_query.sign_date[0:10],
            'depart':depart,

            'type':data['code'],
            'is_sign':'N' if sign_query is None else sign_query.is_sign,
            }) 


    return JsonResponse({
        'result':True,
        'datas':datas,

        })



def save_agreement(request):

    reception_id = request.POST.get('reception_id','')
    sign_pad_id = request.POST.get('sign_pad_id')
    sign_pad_type = request.POST.get('sign_pad_type')
    sign_pad_data = request.POST.get('sign_pad_data')



    if sign_pad_id == '':
        sign_pad = Sign_Manage()

        sign_pad.registrant = request.user.id
        sign_pad.date_register = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    else:
        sign_pad = Sign_Manage.objects.get(id = sign_pad_id)

    reception = Reception.objects.get(id = reception_id)

    sign_pad.reception_id = reception_id
    sign_pad.depart = reception.depart_id
    sign_pad.patient_id = reception.patient_id
    sign_pad.patient_name = reception.patient.name_kor + '/' +  reception.patient.name_eng

    sign_pad.document = sign_pad_type
    sign_pad.is_sign = 'Y'
    sign_pad.sign_data = sign_pad_data
    sign_pad.sign_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sign_pad.modifier = request.user.id
    sign_pad.date_modify = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sign_pad.save()





    return JsonResponse({
        'result':True,

        })


def get_agreement(request):

    id = request.POST.get('id')

    sign_pad = Sign_Manage.objects.get(id = id)


    return JsonResponse({
        'result':True,
        'sign_pad_data':sign_pad.sign_data,
        })
    

def delete_agreement(request):
    id = request.POST.get('id')

    sign_pad = Sign_Manage.objects.get(id = id)
    sign_pad.use_yn = 'N'

    sign_pad.modifier = request.user.id
    sign_pad.date_modify = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    sign_pad.save()

    return JsonResponse({
        'result':True,
        })


def get_memo_detail(request):

    patient_id = request.POST.get('patient_id')

    data = []
    data_note = {
            'memo_detail_company': '',
            'memo_detail_order': '',
            'memo_detail_insurance': '',
            'memo_detail_disease': '',
    }
    data_relative = []
    try:
        patient = Patient.objects.filter(pk=patient_id).first()
        memo_detail = DetailMemo.objects.filter(patient=patient)

        for val in memo_detail:
            data.append({
                'detail_memo_id': val.pk,
                'creator': val.creator.name_en,
                'depart': val.creator.depart,
                'memo': val.memo,
                'memo_depart': val.memo_depart
            })

        patient_note = PatientNotes.objects.filter(patient_id=patient_id).first()
        if patient_note:
            data_note = {
                'memo_detail_company': patient_note.company_name,
                'memo_detail_order': patient_note.order,
                'memo_detail_insurance': patient_note.insurance,
                'memo_detail_disease': patient_note.disease,
            }

        list_relative = PatientRelative.objects.filter(patient_id=patient_id)
        print('=====',list_relative)
        for relative in list_relative:
            data_relative.append({
                'relative_id': relative.id,
                'person_name': relative.name,
                'relative_name': relative.relative,
                'person_id': relative.person_id
            })
        
    except Exception as e:
        print('**', e)
    return JsonResponse({
        'result':True,
        'datas': data,
        'data_note': data_note,
        'data_relative': data_relative,
        'marking': patient.marking
    })

    
def create_memo_detail(request):
    user = request.user

    patient_id = request.POST.get('patient_id')
    memo_depart = request.POST.get('memo_depart')
    memo = request.POST.get('memo')

    data = []
    if request.user.is_authenticated:
        try:
            patient = Patient.objects.filter(pk=patient_id).first()
            DetailMemo.objects.create(creator=user, patient=patient, memo=memo, memo_depart=memo_depart)
            memo_detail = DetailMemo.objects.filter(patient=patient)
            for val in memo_detail:
                data.append({
                    'detail_memo_id': val.pk,
                    'creator': val.creator.name_en,
                    'depart': val.creator.depart,
                    'memo': val.memo,
                    'memo_depart': val.memo_depart
                })
        except Exception as e:
            print(e)

    return JsonResponse({
        'result':True,
        'datas': data
    }) 


def delete_memo_detail(request):
    try:
        memo_id = request.POST.get('memo_id')
        patient = request.POST.get('patient_id')
        DetailMemo.objects.filter(pk=memo_id).first().delete()
        memo_detail = DetailMemo.objects.filter(patient=patient)

        data = []
        for val in memo_detail:
            data.append({
                'detail_memo_id': val.pk,
                'creator': val.creator.name_en,
                'depart': val.creator.depart,
                'memo': val.memo,
                'memo_depart': val.memo_depart
            })
    except Exception as e:
        print(e)
        return JsonResponse({
        'result': False
    }) 
    
    return JsonResponse({
        'result':True,
        'datas': data
    }) 


def update_memo_detail(request):
    try:
        memo_id = request.POST.get('memo_id')
        patient = request.POST.get('patient_id')
        memo = request.POST.get('memo')
        
        memo_detail = DetailMemo.objects.filter(pk=memo_id).first()
        memo_detail.memo = memo
        memo_detail.save()

        memo_detail = DetailMemo.objects.filter(patient=patient)
        data = []
        for val in memo_detail:
            data.append({
                'detail_memo_id': val.pk,
                'creator': val.creator.name_en,
                'depart': val.creator.depart,
                'memo': val.memo,
                'memo_depart': val.memo_depart
            })
    except Exception as e:
        print(e)
        return JsonResponse({
        'result': False
    }) 
    
    return JsonResponse({
        'result':True,
        'datas': data
    }) 


def update_patient_notes(request):
    data = {}
    try:
        patient_id = request.POST.get('patient_id')
        # memo_detail_company = request.POST.get('memo_detail_company')
        memo_detail_order = request.POST.get('memo_detail_order')
        # memo_detail_insurance = request.POST.get('memo_detail_insurance')
        memo_detail_disease = request.POST.get('memo_detail_disease')

        patient_note = PatientNotes.objects.filter(patient_id=patient_id).first()
        if not patient_note:
            patient_note = PatientNotes(patient_id=patient_id)
        # patient_note.company_name = memo_detail_company
        patient_note.order = memo_detail_order
        # patient_note.insurance = memo_detail_insurance
        patient_note.disease = memo_detail_disease
        patient_note.save()
        data = {
            # 'memo_detail_company': memo_detail_company,
            'memo_detail_order': memo_detail_order,
            # 'memo_detail_insurance': memo_detail_insurance,
            'memo_detail_disease': memo_detail_disease,
        }
    except Exception as e:
        print(e)
        return JsonResponse({
        'result': False
    }) 
    
    return JsonResponse({
        'result':True,
        'datas': data
    }) 


def create_patient_relative(request):
    data = []
    try:
        patient_id = request.POST.get('patient_id')
        print(patient_id)
        person_id = int(request.POST.get('person_id'))
        print(person_id)
        relative_name = request.POST.get('relative_name')
        person = Patient.objects.get(id=person_id)
        patient = Patient.objects.get(id=patient_id)
        PatientRelative.objects.create(
            patient_id=patient_id,
            name=person.name_kor,
            relative=relative_name,
            person_id=str(person_id),
        )
        PatientRelative.objects.create(
            patient_id=int(person_id),
            name=patient.name_kor,
            relative=relative_name,
            person_id=str(patient_id),
        )

        list_relative = PatientRelative.objects.filter(patient_id=patient_id)
        for relative in list_relative:
            data.append({
                'relative_id': relative.id,
                'person_name': relative.name,
                'relative_name': relative.relative,
                'person_id': relative.person_id
            })
    except Exception as e:
        print(e)
        return JsonResponse({
        'result': False
    }) 
    
    return JsonResponse({
        'result':True,
        'datas': data
    }) 


def delete_patient_relative(request):
    data = []
    try:
        relative_id = request.POST.get('relative_id')
        patient_id = request.POST.get('patient_id')

        PatientRelative.objects.get(id=relative_id).delete()
        
        list_relative = PatientRelative.objects.filter(patient_id=patient_id)
        for relative in list_relative:
            data.append({
                'relative_id': relative.id,
                'person_name': relative.name,
                'relative_name': relative.relative,
                'person_id': relative.person_id
            })
    except Exception as e:
        print(e)
        return JsonResponse({
        'result': False
    }) 
    
    return JsonResponse({
        'result':True,
        'datas': data
    }) 


def patient_search2(request):
    print("=======")
    memo = request.POST.get('memo')
    name = request.POST.get('name')
    chart = request.POST.get('chart')
    print(chart)
    email = request.POST.get('email')
    phone = request.POST.get('phone')
    dob = request.POST.get('dob')
    memo_detail = request.POST.get('memo_detail')
    argument_list = [] 
    if memo and memo != '':
        argument_list.append( Q(**{'patient__memo__icontains':memo} ) )
    
    if name and name != '':
        argument_list.append( Q(**{'patient__name_kor__icontains':name}) | Q(**{'patient__name_eng__icontains':name}) )
    
    if chart and chart != '':
        argument_list.append( Q(**{'patient__id__icontains':chart} ) ) 
    
    if email and email != '':
        argument_list.append( Q(**{'patient__email__icontains':email} ) )
    
    if phone and phone != '':
        argument_list.append( Q(**{'patient__phone__icontains':phone} ) )

    if dob and dob != '':
        argument_list.append( Q(**{'patient__date_of_birth__icontains':dob} ) )

    # patient = Patient.objects.filter(name_kor=string).first()
    # print(patient)
    datas=[]
    if (memo and memo != '') or (name and name != '') or (chart and chart != '') or (email and email != '') or (phone and phone != '') or (dob and dob != ''):
        receptions = Reception.objects.select_related('patient').values('patient_id','depart_id').filter( functools.reduce(operator.and_, argument_list) ).exclude(progress='deleted').annotate(c_pt=Count('patient_id'),c_dp=Count('depart_id'))
        # print(receptions)
        patient_memo = []
        if memo_detail:
            patient_memo = DetailMemo.objects.filter(memo__icontains=memo_detail).values_list('patient_id', flat=True);
            patient_memo = list(patient_memo)
        for reception in receptions:
            if memo_detail and memo_detail != '':
                if reception['patient_id'] in patient_memo:
                    reception_last = Reception.objects.filter(patient = reception['patient_id'], depart = reception['depart_id']).last()
                    if reception_last:
                        data = {}
                        data.update({
                            'id':reception_last.patient.id,
                            'chart':reception_last.patient.get_chart_no(),
                            'name_kor':reception_last.patient.name_kor,
                            'name_eng':reception_last.patient.name_eng,
                            'gender':reception_last.patient.gender,
                            'date_of_birth':reception_last.patient.date_of_birth.strftime('%Y-%m-%d'),
                            'phonenumber':reception_last.patient.phone,
                            'age' : reception_last.patient.get_age(),
                            'address':reception_last.patient.address,
                            'has_unpaid':reception_last.patient.has_unpaid(),
                            'depart':reception_last.depart.name,
                            'last_visit':reception_last.recorded_date.strftime('%Y-%m-%d'),
                            
                            'nationality':reception_last.patient.nationality,
                            'passport':reception_last.patient.passport,
                            'email':reception_last.patient.email,
                            'category':reception_last.patient.category


                            })
                        datas.append(data)
                print(1)
            else:
                print(2)
                reception_last = Reception.objects.filter(patient = reception['patient_id'], depart = reception['depart_id']).last()
                data = {}
                data.update({
                    'id':reception_last.patient.id,
                    'chart':reception_last.patient.get_chart_no(),
                    'name_kor':reception_last.patient.name_kor,
                    'name_eng':reception_last.patient.name_eng,
                    'gender':reception_last.patient.gender,
                    'date_of_birth':reception_last.patient.date_of_birth.strftime('%Y-%m-%d'),
                    'phonenumber':reception_last.patient.phone,
                    'age' : reception_last.patient.get_age(),
                    'address':reception_last.patient.address,
                    'has_unpaid':reception_last.patient.has_unpaid(),
                    'depart':reception_last.depart.name,
                    'last_visit':reception_last.recorded_date.strftime('%Y-%m-%d'),
                    
                    'nationality':reception_last.patient.nationality,
                    'passport':reception_last.patient.passport,
                    'email':reception_last.patient.email,
                    'category':reception_last.patient.category


                    })
                datas.append(data)
    else:
        print(3)
        if memo_detail:
            patient_memo = DetailMemo.objects.filter(memo__icontains=memo_detail).values_list('patient_id', flat=True);
            patient_memo = list(patient_memo)
            receptions = Reception.objects.select_related('patient').values('patient_id','depart_id').filter( patient__in=patient_memo ).exclude(progress='deleted').annotate(c_pt=Count('patient_id'),c_dp=Count('depart_id'))
            for reception in receptions:
                reception_last = Reception.objects.filter(patient = reception['patient_id'], depart = reception['depart_id']).last()
                data = {}
                data.update({
                    'id':reception_last.patient.id,
                    'chart':reception_last.patient.get_chart_no(),
                    'name_kor':reception_last.patient.name_kor,
                    'name_eng':reception_last.patient.name_eng,
                    'gender':reception_last.patient.gender,
                    'date_of_birth':reception_last.patient.date_of_birth.strftime('%Y-%m-%d'),
                    'phonenumber':reception_last.patient.phone,
                    'age' : reception_last.patient.get_age(),
                    'address':reception_last.patient.address,
                    'has_unpaid':reception_last.patient.has_unpaid(),
                    'depart':reception_last.depart.name,
                    'last_visit':reception_last.recorded_date.strftime('%Y-%m-%d'),
                    
                    'nationality':reception_last.patient.nationality,
                    'passport':reception_last.patient.passport,
                    'email':reception_last.patient.email,
                    'category':reception_last.patient.category


                    })
                datas.append(data)
    # print(datas)
    context = {'datas':datas}
    return JsonResponse(context)

def patient_search3(request):
    # category = request.POST.get('category')
    string = request.POST.get('string')

    # kwargs = {
    #     '{0}__{1}'.format(category, 'icontains'): string,
    #     }
    patients = Patient.objects.filter(name_kor__icontains = string).order_by("-id")

    datas=[]
    for patient in patients:
        data = {}
        data.update({
            'id':patient.id,
            'chart':patient.get_chart_no(),
            'name_kor':patient.name_kor,
            'name_eng':patient.name_eng,
            'gender':patient.gender,
            'date_of_birth':patient.date_of_birth,
            'phonenumber':patient.phone,
            'email': patient.email
            })
        datas.append(data)

    context = {'datas':datas}
    return JsonResponse(context)


def upload_file_patient(request):
    file_path = '/home/imedicare/Cofee/Receptionist/static/draft_patient.xlsx'
    try:
        file = request.FILES.getlist('file')[0]
        
        if request.method == 'POST':
            with open(file_path, 'wb+') as f:
                for chunk in file.chunks():
                    f.write(chunk)
            wb = load_workbook(file_path)

            ws = wb.get_sheet_by_name('Sheet1')

            for i in range(7,200):
                eng_name = ws[f'D{i}'].value
                phone = ws[f'G{i}'].value

                founded_phone = False
                founded_name = False
                if eng_name:
                    eng_name = eng_name.upper()
                    dob = ""
                    try:
                        dob = str(ws[f'E{i}'].value.date())
                    except Exception as e:
                        
                        dob = ws[f'E{i}'].value
                    try:
                        find_patients = Patient.objects.filter(phone = phone)
                        if len(find_patients) > 0:
                            founded_phone = True

                        find_name = Patient.objects.filter(name_eng__icontains = eng_name)
                        if len(find_name) > 0:
                            founded_name = True

                        DraftPatient.objects.create(
                            kor_name = ws[f'C{i}'].value,
                            eng_name = eng_name,
                            dob = dob,
                            gender = ws[f'F{i}'].value,
                            phone = ws[f'G{i}'].value,
                            address = ws[f'H{i}'].value,
                            email = ws[f'I{i}'].value,
                            founded_phone = founded_phone,
                            founded_eng_name = founded_name,
                        )
                    except:
                        pass
                # DraftPatient.objects.all().delete()  
    except Exception as e:
        print(e)
        return JsonResponse({'url':str(e)}, status=400)
        
    return JsonResponse({'url':'ok'})


def pre_regis(request):
    f_name = F('commcode_name_en')
    if request.session[translation.LANGUAGE_SESSION_KEY] == 'ko':
        f_name = F('commcode_name_ko')
    elif request.session[translation.LANGUAGE_SESSION_KEY] == 'vi':
        f_name = F('commcode_name_vi')

    list_depart = Depart.objects.all()
    reservation_dialog_form = ReservationDialogForm()
    reservation_search_form = ReservationSearchControl()

    list_funnels = []
    funnels = COMMCODE.objects.filter(upper_commcode = '000006',commcode_grp = 'PATIENTS_FUNNELS',use_yn="Y").annotate(name = f_name ).values('commcode','name',)
    for data in funnels:
        list_funnels.append({
            'code':data['commcode'],
            'name':data['name'],
            })
        
        
    list_reservation_division= []
    query_reservation_division = COMMCODE.objects.filter(use_yn = 'Y',upper_commcode='000006', commcode_grp='RSRVT_DVSN').annotate(code = F('commcode'),name =f_name ).values('code','name')
    for data in query_reservation_division:
        list_reservation_division.append({
            'code':data['code'],
            'name':data['name']
            })

    return render(request,'Receptionist/pre_regis.html' , {
        'reservation_dialog':reservation_dialog_form,
        'reservation_search':reservation_search_form,
        'list_depart':list_depart,
        'list_funnels': list_funnels,
        'list_reservation_division':list_reservation_division,
    })


def draft_patient_list(request):
    string = request.GET.get('string', '')
    today = datetime.datetime.today().date()
    if string == '':
        today_list = DraftPatient.objects.filter(added_date__date = today, is_deleted = False)
    else:
        argument_list = []
        argument_list.append( Q(**{'eng_name__icontains':string } ) ) 
        argument_list.append( Q(**{'kor__icontains':string } ) ) 
        today_list = DraftPatient.objects.filter( functools.reduce(operator.or_, argument_list),added_date__date = today, is_deleted = False)

    datas = []
    number = 1
    
    for patient in today_list:
        datas.append({
            'no': number,
            'draft_id': patient.id,
            'kor_name': patient.kor_name,
            'eng_name': patient.eng_name,
            'dob': patient.dob,
            'gender': patient.gender,
            'phone': patient.phone,
            'email': patient.email,
            'address': patient.address,
            'company': patient.company,
            'founded_phone': patient.founded_phone,
            'founded_eng_name': patient.founded_eng_name,
            'is_registed': patient.is_registed,
        })

        number +=1 
    
    
    context = {'datas':datas}
    return JsonResponse(context)


def remove_draft_patient(request):
    p_id = request.POST.get('id')
    
    DraftPatient.objects.get(pk = p_id).delete()

    today = datetime.datetime.today().date()
    today_list = DraftPatient.objects.filter(added_date__date = today, is_deleted = False)
    datas = []
    number = 1
    for patient in today_list:
        datas.append({
            'no': number,
            'draft_id': patient.id,
            'kor_name': patient.kor_name,
            'eng_name': patient.eng_name,
            'dob': patient.dob,
            'gender': patient.gender,
            'phone': patient.phone,
            'email': patient.email,
            'address': patient.address,
            'company': patient.company,
            'founded_phone': patient.founded_phone,
            'founded_eng_name': patient.founded_eng_name,
            'is_registed': patient.is_registed,
        })

        number +=1 
    
    context = {'datas':datas}
    return JsonResponse(context)